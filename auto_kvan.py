import os
import time
import json
from dataclasses import dataclass
from pathlib import Path
from typing import List
from datetime import datetime

from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC

# 서버(Railway 등)에서 실행 시 헤드리스 + 자동 종료
def _is_server_env() -> bool:
    return bool(os.environ.get("RAILWAY_ENVIRONMENT") or os.environ.get("RUN_HEADLESS"))


SIGN_IN_URL = "https://store.k-van.app/sign-in"
FACE_TO_FACE_URL = "https://store.k-van.app/face-to-face-payment"

# 코드와 데이터 경로 분리: SISA_DATA_DIR (없으면 ./data)
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get("SISA_DATA_DIR") or (BASE_DIR / "data"))
DATA_DIR.mkdir(parents=True, exist_ok=True)

# 입력 데이터 JSON 파일 (web_form.py 가 생성)
ORDER_JSON_PATH = DATA_DIR / "current_order.json"

# 결제 결과를 저장할 엑셀 / JSON 파일
RESULT_EXCEL_PATH = DATA_DIR / "kvan_results.xlsx"
RESULT_JSON_PATH = DATA_DIR / "last_result.json"
HQ_STATE_PATH = DATA_DIR / "hq_state.json"

# 세션별 주문/결과 디렉토리 (동시 여러 세션용)
SESSION_ORDER_DIR = DATA_DIR / "sessions" / "orders"
SESSION_RESULT_DIR = DATA_DIR / "sessions" / "results"
SESSION_ORDER_DIR.mkdir(parents=True, exist_ok=True)
SESSION_RESULT_DIR.mkdir(parents=True, exist_ok=True)

# JSON / 결과 엑셀에서 공통으로 사용하는 필드 목록
HEADERS: List[str] = [
    "login_id",
    "login_password",
    "login_pin",
    "card_type",  # personal / business
    "card_number",
    "expiry_mm",
    "expiry_yy",
    "card_password",
    "installment_months",
    "phone_number",
    "customer_name",
    "resident_front",
    "amount",
    "product_name",
]


# 로그인 페이지 셀렉터
# 제공해주신 로그인 화면 HTML 기준:
# - 아이디 / 비밀번호 input 은 name 이 없고 label 텍스트로만 구분
SIGN_IN_SELECTORS = {
    # <label>아이디</label> 바로 아래 첫 번째 input
    "id": (
        By.XPATH,
        "//label[normalize-space(text())='아이디']/following::input[1]",
    ),
    # <label>비밀번호</label> 바로 아래 첫 번째 input
    "password": (
        By.XPATH,
        "//label[normalize-space(text())='비밀번호']/following::input[1]",
    ),
    # 로그인 버튼 (텍스트가 '로그인' 인 submit 버튼)
    "submit": (
        By.XPATH,
        "//button[@type='submit' and contains(normalize-space(.), '로그인')]",
    ),
}

# PIN 팝업 셀렉터 (로그인 이후 2차 인증)
PIN_POPUP_SELECTORS = {
    # "PIN을 입력해 주세요" 문구가 포함된 영역 안의 첫 번째 input
    "input": (
        By.XPATH,
        "//*[contains(text(), 'PIN') and contains(text(), '입력')]/ancestor::div[1]//input",
    ),
    # 확인 버튼
    "confirm": (
        By.XPATH,
        "//button[contains(normalize-space(.), '확인')]",
    ),
}


# 대면결제 페이지 셀렉터 (필요 시 사이트 구조에 맞게 수정)
FACE_TO_FACE_SELECTORS = {
    # 결제 정보
    "card_personal_checkbox": (
        By.XPATH,
        "//span[normalize-space(text())='개인카드']/preceding-sibling::div[1]",
    ),
    "card_business_checkbox": (
        By.XPATH,
        "//span[contains(normalize-space(text()), '사업자')]/preceding-sibling::div[1]",
    ),
    # 카드번호: label 텍스트 기준으로 바로 아래 input
    "card_number": (
        By.XPATH,
        "//label[normalize-space(text())='카드번호']/following::input[1]",
    ),
    # MM / YY 는 label 텍스트 기준으로 select 요소를 찾는다
    "expiry_mm": (
        By.XPATH,
        "//label[normalize-space(text())='MM']/following::select[1]",
    ),
    "expiry_yy": (
        By.XPATH,
        "//label[normalize-space(text())='YY']/following::select[1]",
    ),
    # 비밀번호 앞 두 자리
    "card_password": (
        By.XPATH,
        "//label[normalize-space(text())='비밀번호']/following::input[1]",
    ),
    # 할부개월 select
    "installment_select": (
        By.XPATH,
        "//label[contains(normalize-space(text()), '할부개월')]/following::select[1]",
    ),
    # 연락처 / 이름 / 주민번호 앞자리
    "phone_prefix": (By.XPATH, "//select[contains(@name, 'phone') or contains(., '010')]"),
    "phone_number": (By.XPATH, "//input[@placeholder='연락처를 입력해주세요.']"),
    "customer_name": (
        By.XPATH,
        "//label[normalize-space(text())='이름']/following::input[1]",
    ),
    "resident_front": (
        By.XPATH,
        "//label[contains(normalize-space(text()), '주민등록번호 앞자리')]/following::input[1]",
    ),
    "business_reg_no": (
        By.XPATH,
        "//input[@placeholder='10자리 사업자등록번호를 입력해주세요.']",
    ),
    # 상품정보
    "product_name": (By.XPATH, "//label[contains(., '상품명')]/following::input[1]"),
    "product_price": (By.XPATH, "//label[contains(., '판매가격')]/following::input[1]"),
    # 버튼
    "submit_button": (By.XPATH, "//button[contains(., '결제하기')]"),
}


@dataclass
class PaymentRow:
    login_id: str
    login_password: str
    login_pin: str
    card_type: str
    card_number: str
    expiry_mm: str
    expiry_yy: str
    card_password: str
    installment_months: str
    phone_number: str
    customer_name: str
    resident_front: str
    amount: int
    product_name: str


def load_order_from_json(path: str) -> PaymentRow:
    """web_form.py 가 생성한 JSON 파일에서 1건의 주문 데이터를 읽어온다."""
    if not Path(path).exists():
        raise FileNotFoundError(
            f"{path} 파일을 찾을 수 없습니다. web_form.py 의 웹 폼에서 먼저 데이터를 저장해 주세요."
        )

    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    missing = [k for k in HEADERS if k not in raw or raw[k] in ("", None)]
    if missing:
        raise ValueError(f"JSON 데이터에 누락된 필드가 있습니다: {missing}")

    try:
        amount_int = int(raw["amount"])
    except (TypeError, ValueError) as e:
        raise ValueError(f"amount 값이 숫자가 아닙니다: {raw['amount']!r}") from e

    card_type = str(raw.get("card_type", "personal")).strip().lower()
    if card_type not in ("personal", "business"):
        card_type = "personal"

    return PaymentRow(
        login_id=str(raw["login_id"]).strip(),
        login_password=str(raw["login_password"]).strip(),
        login_pin=str(raw["login_pin"]).strip(),
        card_type=card_type,
        card_number=str(raw["card_number"]).strip(),
        expiry_mm=str(raw["expiry_mm"]).strip(),
        expiry_yy=str(raw["expiry_yy"]).strip(),
        card_password=str(raw["card_password"]).strip(),
        installment_months=str(raw["installment_months"]).strip(),
        phone_number=str(raw["phone_number"]).strip(),
        customer_name=str(raw["customer_name"]).strip(),
        resident_front=str(raw["resident_front"]).strip(),
        amount=amount_int,
        product_name=str(raw["product_name"]).strip(),
    )


def _get_chromedriver_path() -> str | None:
    """로컬 tool 폴더 또는 환경변수에 지정된 ChromeDriver 경로."""
    env_path = os.environ.get("CHROMEDRIVER_PATH", "").strip()
    if env_path and Path(env_path).exists():
        return env_path
    base = Path(__file__).resolve().parent / "tool"
    for name in ("chromedriver.exe", "chromedriver"):
        p = base / name
        if p.exists():
            return str(p)
    return None


def create_driver(headless: bool | None = None) -> webdriver.Chrome:
    if headless is None:
        headless = _is_server_env()
    options = webdriver.ChromeOptions()
    if headless:
        options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-setuid-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920,1080")
    else:
        options.add_argument("--start-maximized")
    # Railway 등 서버: 시스템에 설치된 Chrome 사용
    if _is_server_env():
        for binary in ("/usr/bin/google-chrome-stable", "/usr/bin/google-chrome", "/usr/bin/chromium"):
            if Path(binary).exists():
                options.binary_location = binary
                break
    service = None
    driver_path = _get_chromedriver_path()
    if driver_path:
        service = Service(executable_path=driver_path)
    driver = webdriver.Chrome(service=service, options=options) if service else webdriver.Chrome(options=options)
    driver.implicitly_wait(5)
    return driver


def sign_in(driver: webdriver.Chrome, row: PaymentRow) -> None:
    driver.get(SIGN_IN_URL)
    wait = WebDriverWait(driver, 20)

    id_input = wait.until(EC.visibility_of_element_located(SIGN_IN_SELECTORS["id"]))
    id_input.clear()
    id_input.send_keys(row.login_id)

    pw_input = driver.find_element(*SIGN_IN_SELECTORS["password"])
    pw_input.clear()
    pw_input.send_keys(row.login_password)

    submit_btn = driver.find_element(*SIGN_IN_SELECTORS["submit"])
    submit_btn.click()

    # 2차 인증 PIN 팝업 처리 (있으면 입력, 없으면 통과)
    try:
        # 로그인 버튼 클릭 후 사람이 화면을 보는 것처럼 약간 대기
        time.sleep(1.5)
        pin_wait = WebDriverWait(driver, 20)
        pin_input_container = pin_wait.until(
            EC.visibility_of_element_located(PIN_POPUP_SELECTORS["input"])
        )
        # 위에서 찾은 컨테이너 안의 input 이나 자신이 input 일 수 있음
        if pin_input_container.tag_name.lower() == "input":
            pin_input = pin_input_container
        else:
            pin_input = pin_input_container.find_element(By.XPATH, ".//input")

        pin_input.clear()
        pin_input.send_keys(row.login_pin)
        # 사람이 확인 내용을 읽는 것처럼 잠시 대기
        time.sleep(0.8)

        confirm_btn = driver.find_element(*PIN_POPUP_SELECTORS["confirm"])
        confirm_btn.click()
        # 서버가 토큰을 발급하고 홈으로 이동할 시간을 충분히 준다
        time.sleep(3.0)
    except TimeoutException:
        # PIN 팝업이 없는 계정/환경일 수 있으므로 조용히 통과
        pass

    # 로그인 완료까지 잠시 대기 (홈 또는 다른 보호된 페이지로 진입)
    wait.until(EC.url_contains("store.k-van.app"))


def _click_box(driver: webdriver.Chrome, locator: tuple) -> None:
    """단순 클릭 (div 기반 토글 박스용)."""
    el = driver.find_element(*locator)
    el.click()


def _select_dropdown_any(driver: webdriver.Chrome, locator: tuple, text: str) -> None:
    el = driver.find_element(*locator)
    tag = el.tag_name.lower()

    if tag == "select":
        Select(el).select_by_visible_text(text)
        return

    el.click()
    el.clear()
    el.send_keys(text)
    el.send_keys(Keys.ENTER)


def fill_face_to_face_form(driver: webdriver.Chrome, row: PaymentRow) -> None:
    driver.get(FACE_TO_FACE_URL)
    wait = WebDriverWait(driver, 30)

    # 페이지가 완전히 렌더링될 수 있도록 사람처럼 잠시 대기
    time.sleep(2.0)

    # 1) 상품명 먼저 입력
    product_input = wait.until(
        EC.visibility_of_element_located(FACE_TO_FACE_SELECTORS["product_name"])
    )
    driver.execute_script(
        "arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
        product_input,
    )
    time.sleep(0.8)
    product_input.clear()
    product_input.send_keys(row.product_name or "잡화")
    time.sleep(0.6)

    # 2) 판매가격(결제금액) 입력
    try:
        price_input = wait.until(
            EC.visibility_of_element_located(FACE_TO_FACE_SELECTORS["product_price"])
        )
    except TimeoutException:
        # 기본 셀렉터로 못 찾으면 몇 가지 대체 XPATH 를 시도
        alt_xpaths = [
            "//*[contains(text(), '판매가격') or contains(text(), '결제금액')]/following::input[1]",
            "//input[@inputmode='decimal' or @type='number']",
        ]
        price_input = None
        for xp in alt_xpaths:
            try:
                price_input = wait.until(
                    EC.visibility_of_element_located((By.XPATH, xp))
                )
                break
            except TimeoutException:
                continue
        if price_input is None:
            raise RuntimeError("판매가격(결제금액) 입력창을 찾지 못했습니다.")

    driver.execute_script(
        "arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
        price_input,
    )
    time.sleep(0.8)
    price_input.clear()
    price_input.send_keys(str(row.amount))
    time.sleep(0.8)

    # 3) 아래 결제 정보 입력창들이 보이도록 스크롤을 조금 더 내림
    try:
        scroll_container = wait.until(
            EC.presence_of_element_located(
                (
                    By.XPATH,
                    "//div[contains(@class,'overflow-y-auto') and contains(@class,'min-h-screen')]",
                )
            )
        )
        driver.execute_script(
            "arguments[0].scrollTo({top: arguments[0].scrollTop + 600, behavior: 'smooth'});",
            scroll_container,
        )
    except TimeoutException:
        # 컨테이너를 못 찾으면 윈도우 스크롤로 폴백
        driver.execute_script("window.scrollBy(0, 600);")
    time.sleep(1.5)

    # 결제 정보 - 카드 종류 체크 (개인 / 법인)
    # 페이지 기본값이 개인카드이므로, 개인카드는 굳이 클릭하지 않는다.
    if row.card_type == "business":
        try:
            _click_box(driver, FACE_TO_FACE_SELECTORS["card_business_checkbox"])
            time.sleep(0.8)
        except Exception:
            pass

    # 카드번호
    card_input = wait.until(
        EC.visibility_of_element_located(FACE_TO_FACE_SELECTORS["card_number"])
    )
    card_input.clear()
    card_input.send_keys(row.card_number)

    # 유효기간 MM / YY
    try:
        _select_dropdown_any(driver, FACE_TO_FACE_SELECTORS["expiry_mm"], row.expiry_mm)
    except Exception:
        mm_input = driver.find_element(*FACE_TO_FACE_SELECTORS["expiry_mm"])
        mm_input.clear()
        mm_input.send_keys(row.expiry_mm)

    try:
        _select_dropdown_any(driver, FACE_TO_FACE_SELECTORS["expiry_yy"], row.expiry_yy)
    except Exception:
        yy_input = driver.find_element(*FACE_TO_FACE_SELECTORS["expiry_yy"])
        yy_input.clear()
        yy_input.send_keys(row.expiry_yy)

    # 비밀번호 앞 두 자리
    pw2_input = wait.until(
        EC.visibility_of_element_located(FACE_TO_FACE_SELECTORS["card_password"])
    )
    pw2_input.clear()
    pw2_input.send_keys(row.card_password)

    # 할부개월
    try:
        _select_dropdown_any(
            driver, FACE_TO_FACE_SELECTORS["installment_select"], row.installment_months
        )
    except Exception:
        pass

    # 연락처
    try:
        _select_dropdown_any(driver, FACE_TO_FACE_SELECTORS["phone_prefix"], "010")
    except Exception:
        pass

    phone_input = wait.until(
        EC.visibility_of_element_located(FACE_TO_FACE_SELECTORS["phone_number"])
    )
    phone_input.clear()
    phone_input.send_keys(row.phone_number)

    # 이름
    name_input = wait.until(
        EC.visibility_of_element_located(FACE_TO_FACE_SELECTORS["customer_name"])
    )
    name_input.clear()
    name_input.send_keys(row.customer_name)

    # 주민등록번호 앞자리 / 사업자등록번호
    if row.card_type == "business":
        # 사업자 카드인 경우: 사업자등록번호 입력칸이 나타남
        try:
            biz_input = wait.until(
                EC.visibility_of_element_located(
                    FACE_TO_FACE_SELECTORS["business_reg_no"]
                )
            )
            biz_input.clear()
            biz_input.send_keys(row.resident_front)
        except Exception:
            # 셀렉터가 맞지 않거나 필드가 안 보이면 조용히 통과
            pass
    else:
        # 개인카드인 경우: 주민등록번호 앞자리 입력
        res_input = wait.until(
            EC.visibility_of_element_located(FACE_TO_FACE_SELECTORS["resident_front"])
        )
        res_input.clear()
        res_input.send_keys(row.resident_front)

    # 버튼이 "결제하기" 로 바뀔 때까지 기다렸다가 클릭
    try:
        submit_btn = wait.until(
            EC.element_to_be_clickable(FACE_TO_FACE_SELECTORS["submit_button"])
        )
    except TimeoutException:
        raise RuntimeError("결제하기 버튼을 찾지 못했습니다. 셀렉터를 확인하세요.")

    submit_btn.click()


def confirm_popup_and_get_result(driver: webdriver.Chrome) -> tuple[str, str]:
    wait = WebDriverWait(driver, 30)

    # 결제 확인 팝업에서 "결제" 또는 "확인" 버튼 클릭 (텍스트 기준으로 찾기)
    clicked_confirm = False
    time.sleep(1.0)

    # 팝업 타이틀(합계 ~ 결제)이 보일 때까지 먼저 대기 (있으면)
    try:
        wait.until(
            EC.visibility_of_element_located(
                (
                    By.XPATH,
                    "//*[contains(normalize-space(.), '합계') and contains(normalize-space(.), '결제')]",
                )
            )
        )
    except TimeoutException:
        # 팝업 타이틀을 못 찾더라도 계속 진행 (환경에 따라 구조가 다를 수 있음)
        pass

    # 결제 버튼 → 확인 버튼 순서로 여러 패턴을 시도
    button_xpaths = [
        # SweetAlert2 팝업 전용 (가장 우선)
        "//div[contains(@class,'swal2-actions')]//button[contains(@class,'swal2-confirm')]",
        "//button[contains(@class,'swal2-confirm') and contains(normalize-space(.), '결제')]",
        # 버튼 자체 텍스트에 "결제"
        "//button[contains(normalize-space(.), '결제')]",
        # 버튼 내부 span 등에 "결제"
        "//button[.//span[contains(normalize-space(normalize-space(.)), '결제')]]",
        # 일부 환경에서는 '확인'만 표시될 수 있음
        "//button[contains(normalize-space(.), '확인')]",
        "//button[.//span[contains(normalize-space(.), '확인')]]",
    ]

    for xp in button_xpaths:
        try:
            btn = wait.until(
                EC.element_to_be_clickable(
                    (
                        By.XPATH,
                        xp,
                    )
                )
            )
            # 버튼이 가려져 있을 수 있으므로 가운데로 스크롤
            try:
                driver.execute_script(
                    "arguments[0].scrollIntoView({behavior:'smooth', block:'center'});",
                    btn,
                )
                time.sleep(0.5)
            except Exception:
                pass

            try:
                btn.click()
            except Exception:
                # 일반 click 이 안 먹히면 JS 로 강제 클릭
                driver.execute_script("arguments[0].click();", btn)

            clicked_confirm = True
            break
        except TimeoutException:
            continue

    if not clicked_confirm:
        print("확인/결제 팝업 버튼을 찾지 못했습니다. 바로 결과 화면이 떴을 수 있습니다.")

    # 결과 화면의 h1 / p 텍스트 읽기
    try:
        title_el = wait.until(
            EC.visibility_of_element_located(
                (By.XPATH, "//h1[contains(., '결제')]")
            )
        )
        title_text = title_el.text.strip()
    except TimeoutException:
        return "unknown", "결과 제목을 찾지 못했습니다."

    try:
        message_el = driver.find_element(
            By.XPATH,
            "//p[contains(@class, 'text-red-500') or contains(@class, 'text-green-500')]",
        )
        message_text = message_el.text.strip()
    except Exception:
        message_text = ""

    status = "success" if "완료" in title_text else "fail" if "실패" in title_text else "unknown"
    return status, f"{title_text} / {message_text}"


def save_result_to_excel(
    path: str, row: PaymentRow, status: str, message: str
) -> None:
    """결제 결과를 별도 엑셀 파일(RESULT_EXCEL_PATH)에 누적 저장.

    기존 버전에서 생성된 엑셀은 card_type 열이 없을 수 있으므로,
    필요하면 결과 시트의 헤더를 최신 HEADERS 구조로 자동 마이그레이션한다.
    """

    def _ensure_latest_header(ws) -> None:
        """기존 결과 시트의 헤더를 최신 HEADERS 구조로 맞춘다."""
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if not any(header_row):
            # 헤더가 비어 있으면 새로 작성
            ws.append(HEADERS + ["result_status", "result_message"])
            return

        # 이전 버전: card_type 이 없던 경우 (열 개수로 판단)
        old_headers = [h for h in HEADERS if h != "card_type"]
        if header_row[: len(old_headers)] == old_headers and len(header_row) == len(
            old_headers
        ) + 2:
            # 4번째 열에 card_type 열 삽입 (기존 데이터는 기본값 없음)
            ws.insert_cols(4)
            ws.cell(row=1, column=4, value="card_type")

    path_obj = Path(path)
    if path_obj.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "results"
        ws.append(HEADERS + ["result_status", "result_message"])

    # 기존 파일인 경우 헤더를 최신 구조로 보정
    _ensure_latest_header(ws)

    data = [getattr(row, key) for key in HEADERS]
    ws.append(data + [status, message])

    try:
        wb.save(path)
    except PermissionError:
        # 파일이 엑셀 등에서 열려 있어서 저장이 안 되는 경우
        print("kvan_results.xlsx 저장에 실패했습니다. 엑셀 파일을 닫은 뒤 다시 실행해 주세요.")


def save_result_to_json(path: str, status: str, message: str) -> None:
    """web_form.py 에서 읽어갈 수 있도록 마지막 결제 결과를 JSON 으로도 저장."""
    payload = {
        "status": status,
        "message": message,
    }
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except OSError:
        # 파일 시스템 오류는 치명적이지 않으므로 조용히 무시
        pass


def _load_hq_state() -> dict:
    """본사 어드민 상태(hq_state.json)를 로드 (transactions 포함)."""
    state = {"applications": [], "agencies": [], "transactions": []}
    path = Path(HQ_STATE_PATH)
    if path.exists():
        try:
            with path.open("r", encoding="utf-8") as f:
                loaded = json.load(f)
            if isinstance(loaded, dict):
                for key in state.keys():
                    if key in loaded and isinstance(loaded[key], list):
                        state[key] = loaded[key]
        except Exception:
            pass
    return state


def _save_hq_state(state: dict) -> None:
    try:
        with Path(HQ_STATE_PATH).open("w", encoding="utf-8") as f:
            json.dump(state, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def append_transaction_to_hq(
    row,
    status: str,
    message: str,
    session_id: str,
) -> None:
    """결제 결과를 본사 어드민용 거래 리스트에 추가."""
    try:
        state = _load_hq_state()
        txs = state.get("transactions") or []
        now = datetime.utcnow().isoformat()
        tx_id = datetime.utcnow().strftime("TX%Y%m%d%H%M%S%f")

        # 세션 ID 로부터 실제 대행사 ID 를 추적 (admin_state.json 참조)
        agency_id = ""
        if session_id:
            admin_path = Path("admin_state.json")
            if admin_path.exists():
                try:
                    with admin_path.open("r", encoding="utf-8") as f:
                        admin_state = json.load(f)
                except Exception:
                    admin_state = {}
                sessions = admin_state.get("sessions") or []
                history = admin_state.get("history") or []
                for s in sessions:
                    if str(s.get("id")) == str(session_id) and s.get("agency_id"):
                        agency_id = s["agency_id"]
                        break
                if not agency_id:
                    for h in history:
                        if str(h.get("id")) == str(session_id) and h.get("agency_id"):
                            agency_id = h["agency_id"]
                            break

        # 금액은 정수 변환 시도 (실패 시 0)
        try:
            amount_int = int(getattr(row, "amount", 0) or 0)
        except (ValueError, TypeError):
            amount_int = 0

        tx = {
            "id": tx_id,
            "created_at": now,
            "agency_id": agency_id or "",
            "amount": amount_int,
            "customer_name": getattr(row, "customer_name", ""),
            "phone_number": getattr(row, "phone_number", ""),
            "card_type": getattr(row, "card_type", ""),
            "resident_front": getattr(row, "resident_front", ""),
            "status": status,
            "message": message,
            "settlement_status": "미정산",
        }
        txs.append(tx)
        state["transactions"] = txs
        _save_hq_state(state)
    except Exception:
        # HQ 집계 실패는 결제 자체에는 영향을 주지 않으므로 조용히 무시
        pass


def main() -> None:
    # 명령행 인자로 세션 ID 를 받을 수 있게 함:
    # python auto_kvan.py            -> 기본 단일 모드 (current_order.json 사용)
    # python auto_kvan.py SESSIONID  -> 세션별 orders/SESSIONID.json 사용
    import sys

    session_id = sys.argv[1].strip() if len(sys.argv) > 1 else ""

    if session_id:
        SESSION_ORDER_DIR.mkdir(parents=True, exist_ok=True)
        SESSION_RESULT_DIR.mkdir(parents=True, exist_ok=True)
        order_path = SESSION_ORDER_DIR / f"{session_id}.json"
        result_json_path = SESSION_RESULT_DIR / f"{session_id}.json"
    else:
        order_path = Path(ORDER_JSON_PATH)
        result_json_path = Path(RESULT_JSON_PATH)

    print("JSON 주문 데이터를 읽습니다...")
    try:
        row = load_order_from_json(str(order_path))
    except FileNotFoundError as e:
        print(e)
        return
    except ValueError as e:
        print(f"입력 데이터 오류: {e}")
        return

    driver = create_driver()
    try:
        print("K-VAN 가맹점 페이지에 로그인 중...")
        sign_in(driver, row)
        print("로그인 후, 대면결제 폼 자동 입력 시작...")

        try:
            fill_face_to_face_form(driver, row)
            status, msg = confirm_popup_and_get_result(driver)
            print(f"결제 결과: {status} - {msg}")
        except Exception as e:  # noqa: BLE001
            status = "error"
            msg = str(e)
            print(f"결제 처리 중 오류 발생: {e}")

        save_result_to_excel(RESULT_EXCEL_PATH, row, status, msg)
        save_result_to_json(str(result_json_path), status, msg)
        append_transaction_to_hq(row, status, msg, session_id=session_id)

        # 세션 모드인 경우, 어드민 상태에 세션 결과를 반영하고 세션을 히스토리로 이동
        if session_id:
            admin_path = Path("admin_state.json")
            if admin_path.exists():
                try:
                    with open(admin_path, "r", encoding="utf-8") as f:
                        admin_state = json.load(f)
                except Exception:
                    admin_state = {}
                sessions = admin_state.get("sessions") or []
                history = admin_state.get("history") or []

                new_sessions: list = []
                found = False
                for s in sessions:
                    if str(s.get("id")) == str(session_id):
                        found = True
                        human_status = (
                            "결제완료"
                            if status == "success"
                            else "결제실패"
                            if status in ("fail", "error")
                            else "알수없음"
                        )
                        entry = {
                            "id": session_id,
                            "amount": str(row.amount),
                            "installment": row.installment_months,
                            "status": human_status,
                            "created_at": s.get("created_at")
                            or datetime.utcnow().isoformat(),
                            "finished_at": datetime.utcnow().isoformat(),
                            "result_message": msg,
                            "customer_name": row.customer_name,
                            "phone_number": row.phone_number,
                            "product_name": row.product_name,
                            "settled": "정산전",
                        }
                        history.append(entry)
                    else:
                        new_sessions.append(s)

                if found:
                    admin_state["sessions"] = new_sessions
                    admin_state["history"] = history
                    try:
                        with open(admin_path, "w", encoding="utf-8") as f:
                            json.dump(admin_state, f, ensure_ascii=False, indent=2)
                    except OSError:
                        pass
        print(f"결과가 {RESULT_EXCEL_PATH} 파일에 기록되었습니다.")
        if not _is_server_env():
            input("브라우저를 확인한 뒤, 종료하려면 Enter 키를 누르세요...")
    finally:
        driver.quit()


if __name__ == "__main__":
    main()


