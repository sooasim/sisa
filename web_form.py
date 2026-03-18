from __future__ import annotations

import json
import time
from typing import List
from pathlib import Path
from datetime import datetime, timedelta
import os
from io import BytesIO
import subprocess
import sys

import pymysql
from flask import (
    Flask,
    render_template_string,
    redirect,
    url_for,
    request,
    flash,
    jsonify,
    session,
    send_file,
)

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from openpyxl import Workbook

# 코드와 데이터 경로 분리: SISA_DATA_DIR (없으면 ./data)
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get("SISA_DATA_DIR") or (BASE_DIR / "data"))
DATA_DIR.mkdir(parents=True, exist_ok=True)

ORDER_JSON_PATH = str(DATA_DIR / "current_order.json")
RESULT_JSON_PATH = str(DATA_DIR / "last_result.json")
ADMIN_STATE_PATH = str(DATA_DIR / "admin_state.json")
HQ_STATE_PATH = str(DATA_DIR / "hq_state.json")
ADMIN_LOG_PATH = DATA_DIR / "hq_logs.log"
SESSION_ORDER_DIR = DATA_DIR / "sessions" / "orders"
SESSION_RESULT_DIR = DATA_DIR / "sessions" / "results"
SESSION_ORDER_DIR.mkdir(parents=True, exist_ok=True)
SESSION_RESULT_DIR.mkdir(parents=True, exist_ok=True)

# MySQL 환경 변수 (Railway 용)
DB_HOST = os.environ.get("MYSQLHOST") or os.environ.get("MYSQL_HOST") or "localhost"
DB_PORT = int(os.environ.get("MYSQLPORT") or os.environ.get("MYSQL_PORT") or "3306")
DB_USER = os.environ.get("MYSQLUSER") or os.environ.get("MYSQL_USER") or "root"
DB_PASSWORD = os.environ.get("MYSQLPASSWORD") or os.environ.get("MYSQL_PASSWORD") or ""
DB_NAME = (
    os.environ.get("MYSQL_DATABASE")
    or os.environ.get("MYSQLDATABASE")
    or os.environ.get("MYSQL_DB")
    or "railway"
)


def _append_hq_log(source: str, message: str) -> None:
    """HQ 어드민 로그 파일(hq_logs.log)에 한 줄 추가."""
    try:
        ADMIN_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(ADMIN_LOG_PATH, "a", encoding="utf-8") as f:
            ts = datetime.utcnow().isoformat()
            f.write(f"{ts} [{source}] {message}\n")
    except Exception:
        pass


def get_db():
    """
    공용 MySQL 커넥션 헬퍼.

    - connect_timeout 을 짧게 두어(2초) DB 연결 문제로 크롤러/서버가 오래 멈추지 않게 한다.
    - read/write_timeout 도 3초로 제한해, 쿼리 중 연결이 끊기면 빠르게 예외를 발생시킨다.
    """
    return pymysql.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=False,
        connect_timeout=2,
        read_timeout=3,
        write_timeout=3,
    )


def cleanup_history_files() -> dict:
    """
    세션 JSON 히스토리 파일을 3개월 기준으로 정리하고,
    7일/3일 뒤 삭제 예정인 파일 이름 목록을 반환한다.

    - 삭제 대상: SESSION_ORDER_DIR, SESSION_RESULT_DIR 의 *.json
    - created_at 기준이 없으므로 파일의 mtime 을 사용한다.
    """
    now = datetime.utcnow()
    warn_7: list[str] = []
    warn_3: list[str] = []

    targets: list[Path] = []
    for d in [SESSION_ORDER_DIR, SESSION_RESULT_DIR]:
        if d.exists():
            targets.extend(list(d.glob("*.json")))

    for path in targets:
        try:
            st = path.stat()
            created = datetime.utcfromtimestamp(st.st_mtime)
            delete_at = created + timedelta(days=90)
            days_left = (delete_at - now).days

            if days_left <= 0:
                # 실제 삭제
                try:
                    path.unlink()
                except OSError:
                    # 삭제 실패는 치명적이지 않으므로 경고만 남긴다.
                    print(f"[WARN] 히스토리 파일 삭제 실패: {path}")
            else:
                name = path.name
                if days_left <= 3:
                    warn_3.append(name)
                elif days_left <= 7:
                    warn_7.append(name)
        except Exception as e:
            print(f"[WARN] cleanup_history_files 처리 중 오류: {e}")

    return {"warn_7_days": warn_7, "warn_3_days": warn_3}


def init_db() -> None:
    """신청 / 대행사 / 거래 테이블 생성 (없으면)."""
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS applications (
                  id            VARCHAR(32) PRIMARY KEY,
                  company_name  VARCHAR(255) NOT NULL,
                  domain        VARCHAR(255) NOT NULL,
                  phone         VARCHAR(50),
                  bank_name     VARCHAR(100),
                  account_number VARCHAR(100),
                  email_or_sheet TEXT,
                  login_id      VARCHAR(100),
                  login_password VARCHAR(255),
                  fee_percent   INT DEFAULT 10,
                  created_at    DATETIME,
                  status        VARCHAR(20)
                ) CHARACTER SET utf8mb4
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS agencies (
                  id             VARCHAR(32) PRIMARY KEY,
                  company_name   VARCHAR(255) NOT NULL,
                  domain         VARCHAR(255) NOT NULL,
                  phone          VARCHAR(50),
                  bank_name      VARCHAR(100),
                  account_number VARCHAR(100),
                  email_or_sheet TEXT,
                  login_id       VARCHAR(100) UNIQUE,
                  login_password VARCHAR(255),
                  fee_percent    INT DEFAULT 10,
                  kvan_mid       VARCHAR(100),
                  created_at     DATETIME,
                  status         VARCHAR(20)
                ) CHARACTER SET utf8mb4
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS transactions (
                  id                VARCHAR(32) PRIMARY KEY,
                  created_at        DATETIME,
                  agency_id         VARCHAR(32),
                  amount            INT,
                  customer_name     VARCHAR(255),
                  phone_number      VARCHAR(50),
                  card_type         VARCHAR(20),
                  resident_front    VARCHAR(6),
                  status            VARCHAR(20),
                  message           TEXT,
                  settlement_status VARCHAR(20),
                  settled_at        DATETIME,
                  -- K-VAN 연동을 위한 보조 필드들 (있으면 사용)
                  kvan_mid          VARCHAR(100),
                  kvan_approval_no  VARCHAR(100),
                  kvan_tx_type      VARCHAR(50),
                  kvan_registered_at VARCHAR(50),
                  FOREIGN KEY (agency_id) REFERENCES agencies(id)
                ) CHARACTER SET utf8mb4
                """
            )
            # K-VAN 대시보드 요약 정보 저장용 테이블
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS kvan_dashboard (
                  id                     INT AUTO_INCREMENT PRIMARY KEY,
                  captured_at            DATETIME,
                  monthly_sales_amount   BIGINT,
                  monthly_approved_count INT,
                  monthly_approved_amount BIGINT,
                  monthly_canceled_count INT,
                  monthly_canceled_amount BIGINT,
                  yesterday_sales_amount BIGINT,
                  yesterday_approved_count INT,
                  yesterday_approved_amount BIGINT,
                  yesterday_canceled_count INT,
                  yesterday_canceled_amount BIGINT,
                  settlement_expected_amount BIGINT,
                  today_settlement_expected_amount BIGINT,
                  credit_amount          BIGINT,
                  recent_tx_summary      TEXT
                ) CHARACTER SET utf8mb4
                """
            )
            # K-VAN 결제링크 목록 저장 테이블
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS kvan_links (
                  id BIGINT UNSIGNED NOT NULL AUTO_INCREMENT PRIMARY KEY,
                  captured_at DATETIME NOT NULL,
                  title VARCHAR(255) DEFAULT '',
                  amount BIGINT DEFAULT 0,
                  ttl_label VARCHAR(100) DEFAULT '',
                  status VARCHAR(100) DEFAULT '',
                  kvan_link VARCHAR(512) DEFAULT '',
                  mid VARCHAR(100) DEFAULT '',
                  kvan_session_id VARCHAR(100) DEFAULT '',
                  raw_text TEXT
                ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
                """
            )
            # K-VAN 거래내역 저장 테이블
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS kvan_transactions (
                  id BIGINT UNSIGNED NOT NULL AUTO_INCREMENT PRIMARY KEY,
                  captured_at DATETIME NOT NULL,
                  merchant_name VARCHAR(255) DEFAULT '',
                  pg_name VARCHAR(100) DEFAULT '',
                  mid VARCHAR(100) DEFAULT '',
                  fee_rate VARCHAR(50) DEFAULT '',
                  tx_type VARCHAR(50) DEFAULT '',
                  amount BIGINT DEFAULT 0,
                  cancel_amount BIGINT DEFAULT 0,
                  payable_amount BIGINT DEFAULT 0,
                  card_company VARCHAR(100) DEFAULT '',
                  card_number VARCHAR(64) DEFAULT '',
                  installment VARCHAR(50) DEFAULT '',
                  approval_no VARCHAR(100) DEFAULT '',
                  registered_at VARCHAR(50) DEFAULT '',
                  raw_text TEXT
                ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
                """
            )
            # 기존 DB에 이미 생성된 테이블들이 있는 경우를 위해
            # 필요한 열이 없으면 추가 (에러는 무시)
            try:
                cur.execute("ALTER TABLE agencies ADD COLUMN kvan_mid VARCHAR(100)")
            except Exception:
                pass
            for col, typ in [
                ("kvan_login_id", "VARCHAR(100)"),
                ("kvan_login_password", "VARCHAR(255)"),
                ("kvan_login_pin", "VARCHAR(20)"),
            ]:
                try:
                    cur.execute(f"ALTER TABLE agencies ADD COLUMN {col} {typ}")
                except Exception:
                    pass
            try:
                cur.execute("ALTER TABLE transactions ADD COLUMN kvan_mid VARCHAR(100)")
            except Exception:
                pass
            try:
                cur.execute("ALTER TABLE transactions ADD COLUMN kvan_approval_no VARCHAR(100)")
            except Exception:
                pass
            try:
                cur.execute("ALTER TABLE transactions ADD COLUMN kvan_tx_type VARCHAR(50)")
            except Exception:
                pass
            try:
                cur.execute("ALTER TABLE transactions ADD COLUMN kvan_registered_at VARCHAR(50)")
            except Exception:
                pass
        conn.commit()
        conn.close()
    except Exception as e:  # noqa: BLE001
        print(f"[WARN] DB 초기화 실패: {e}")


def ensure_runtime_files() -> None:
    """런타임 필수 디렉토리/파일을 항상 준비한다."""
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        SESSION_ORDER_DIR.mkdir(parents=True, exist_ok=True)
        SESSION_RESULT_DIR.mkdir(parents=True, exist_ok=True)

        state_path = Path(ADMIN_STATE_PATH)
        if not state_path.exists():
            with open(state_path, "w", encoding="utf-8") as f:
                json.dump({"sessions": [], "history": []}, f, ensure_ascii=False, indent=2)

        ADMIN_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(ADMIN_LOG_PATH, "a", encoding="utf-8"):
            pass
    except Exception as e:  # noqa: BLE001
        print(f"[WARN] runtime 파일 준비 실패: {e}")


def _create_code_backup() -> None:
    """앱 시작 시 핵심 파이썬 파일을 날짜별 스냅샷으로 백업한다."""
    try:
        backup_dir = DATA_DIR / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        today = datetime.utcnow().strftime("%Y%m%d")
        for src in [BASE_DIR / "web_form.py", BASE_DIR / "wsisa" / "auto_kvan.py"]:
            if not src.exists():
                continue
            dest = backup_dir / f"{src.stem}_{today}.py"
            if dest.exists():
                continue  # 하루 1회만 백업
            import shutil
            shutil.copy2(str(src), str(dest))
            print(f"[BACKUP] {src.name} → {dest}")
    except Exception as e:  # noqa: BLE001
        print(f"[WARN] 코드 백업 실패: {e}")


# K-VAN 은 동일 계정 동시 로그인을 허용하지 않는다.
# 여러 세션이 동시에 링크 생성을 요청하면 한 번에 하나씩 직렬 처리해야 한다.
# 큐 파일: DATA_DIR/kvan_queue.json  (session_id 목록)
# 락 파일: DATA_DIR/kvan_running.lock (현재 실행 중인 session_id)
KVAN_QUEUE_PATH = DATA_DIR / "kvan_queue.json"
KVAN_LOCK_PATH = DATA_DIR / "kvan_running.lock"
PAYMENT_NOTIFICATIONS_PATH = DATA_DIR / "payment_notifications.json"
EXPIRED_WITH_TRANSACTIONS_PATH = DATA_DIR / "expired_with_transactions.json"
KVAN_CRAWLER_LOCK_PATH = DATA_DIR / "kvan_crawler.lock"
# 거래 내역 엑셀형 리스트용 컬럼 순서 (DB transactions 테이블 전체 양식)
TX_EXCEL_COLUMNS = [
    "id", "created_at", "agency_id", "amount", "customer_name", "phone_number",
    "card_type", "resident_front", "status", "message", "settlement_status", "settled_at",
    "kvan_mid", "kvan_approval_no", "kvan_tx_type", "kvan_registered_at",
]
KVAN_CRAWLER_WAKEUP_PATH = DATA_DIR / "crawler_wakeup.flag"
KVAN_CRAWLER_HEARTBEAT_PATH = DATA_DIR / "kvan_crawler.heartbeat"


def _load_payment_notifications(agency_id: str | None = None) -> list[dict]:
    """미확인 결제 알림 목록. agency_id 가 None 이면 전체, 아니면 해당 대행사만."""
    try:
        if not PAYMENT_NOTIFICATIONS_PATH.exists():
            return []
        items = json.loads(PAYMENT_NOTIFICATIONS_PATH.read_text(encoding="utf-8"))
        if not isinstance(items, list):
            return []
        out = [x for x in items if x.get("seen") is not True]
        if agency_id is not None and str(agency_id).strip():
            out = [x for x in out if str(x.get("agency_id") or "").strip() == str(agency_id).strip()]
        return out[-100:]
    except Exception:
        return []


def _mark_payment_notifications_seen(agency_id: str | None = None) -> None:
    """결제 알림을 확인 처리. agency_id None 이면 전체, 아니면 해당 대행사만."""
    try:
        if not PAYMENT_NOTIFICATIONS_PATH.exists():
            return
        items = json.loads(PAYMENT_NOTIFICATIONS_PATH.read_text(encoding="utf-8"))
        if not isinstance(items, list):
            return
        for x in items:
            if agency_id is None or str(x.get("agency_id") or "").strip() == str(agency_id).strip():
                x["seen"] = True
        PAYMENT_NOTIFICATIONS_PATH.write_text(
            json.dumps(items[-500:], ensure_ascii=False, indent=2), encoding="utf-8"
        )
    except Exception as e:  # noqa: BLE001
        print(f"[WARN] 결제 알림 확인 처리 실패: {e}")


def _kvan_enqueue(session_id: str) -> None:
    """세션 ID를 K-VAN 실행 큐에 추가한다 (중복 방지)."""
    try:
        KVAN_QUEUE_PATH.parent.mkdir(parents=True, exist_ok=True)
        queue: list[str] = []
        if KVAN_QUEUE_PATH.exists():
            try:
                queue = json.loads(KVAN_QUEUE_PATH.read_text(encoding="utf-8"))
            except Exception:
                queue = []
        if session_id and session_id not in queue:
            queue.append(session_id)
        KVAN_QUEUE_PATH.write_text(json.dumps(queue), encoding="utf-8")
    except Exception as e:  # noqa: BLE001
        print(f"[WARN] kvan_enqueue 실패: {e}")


def _kvan_dequeue() -> str | None:
    """큐의 첫 번째 세션 ID를 꺼낸다."""
    try:
        if not KVAN_QUEUE_PATH.exists():
            return None
        queue: list[str] = json.loads(KVAN_QUEUE_PATH.read_text(encoding="utf-8"))
        if not queue:
            return None
        next_id = queue.pop(0)
        KVAN_QUEUE_PATH.write_text(json.dumps(queue), encoding="utf-8")
        return next_id
    except Exception:
        return None


def _kvan_is_running() -> bool:
    """락 파일이 존재하고 실제 프로세스가 살아있으면 True."""
    try:
        if not KVAN_LOCK_PATH.exists():
            return False
        pid_str = KVAN_LOCK_PATH.read_text(encoding="utf-8").strip()
        if not pid_str:
            return False
        pid = int(pid_str)
        # PID 가 살아있는지 확인 (Unix/Windows 공통: os.kill 0 시그널)
        try:
            os.kill(pid, 0)
            return True
        except (OSError, ProcessLookupError):
            # 프로세스 없음 → 락 파일 제거
            KVAN_LOCK_PATH.unlink(missing_ok=True)
            return False
    except Exception:
        KVAN_LOCK_PATH.unlink(missing_ok=True)
        return False


def _crawler_is_running() -> bool:
    """크롤러 락 파일 PID가 실제로 살아있으면 True."""
    try:
        if not KVAN_CRAWLER_LOCK_PATH.exists():
            return False
        pid_str = KVAN_CRAWLER_LOCK_PATH.read_text(encoding="utf-8").strip()
        if not pid_str:
            return False
        pid = int(pid_str)
        try:
            os.kill(pid, 0)
            # PID는 살아있지만 크롤러 루프가 멈췄을 수 있으므로 heartbeat 신선도도 함께 본다.
            try:
                if KVAN_CRAWLER_HEARTBEAT_PATH.exists():
                    hb_age = time.time() - KVAN_CRAWLER_HEARTBEAT_PATH.stat().st_mtime
                    if hb_age > 180:
                        KVAN_CRAWLER_LOCK_PATH.unlink(missing_ok=True)
                        _append_hq_log("WEB", f"[WARN] crawler heartbeat stale({int(hb_age)}s) - lock reset")
                        return False
            except Exception:
                pass
            return True
        except (OSError, ProcessLookupError):
            KVAN_CRAWLER_LOCK_PATH.unlink(missing_ok=True)
            return False
    except Exception:
        KVAN_CRAWLER_LOCK_PATH.unlink(missing_ok=True)
        return False


def trigger_kvan_crawler_refresh() -> None:
    """
    결제링크/거래내역 크롤러를 즉시 깨운다.

    - 이미 크롤러가 실행 중이면 wakeup.flag만 남겨 다음 사이클 즉시 실행
    - 실행 중이 아니면 크롤러 프로세스를 새로 기동
    """
    try:
        KVAN_CRAWLER_WAKEUP_PATH.parent.mkdir(parents=True, exist_ok=True)
        KVAN_CRAWLER_WAKEUP_PATH.write_text(datetime.utcnow().isoformat(), encoding="utf-8")
        _append_hq_log("WEB", f"crawler wakeup 신호 생성: {KVAN_CRAWLER_WAKEUP_PATH}")
    except Exception as e:  # noqa: BLE001
        _append_hq_log("WEB", f"[WARN] crawler wakeup 신호 생성 실패: {e}")

    if _crawler_is_running():
        _append_hq_log("WEB", "crawler 이미 실행 중 - wakeup만 전송")
        return

    lf = None
    try:
        crawler_path = BASE_DIR / "wsisa" / "kvan_crawler.py"
        if not crawler_path.exists():
            raise FileNotFoundError(f"kvan_crawler.py not found: {crawler_path}")
        ADMIN_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        lf = open(ADMIN_LOG_PATH, "a", encoding="utf-8")  # noqa: SIM115
        cmd = [sys.executable, str(crawler_path)]
        env = os.environ.copy()
        env["PYTHONUNBUFFERED"] = "1"
        p = subprocess.Popen(
            cmd,
            stdout=lf,
            stderr=lf,
            env=env,
            cwd=str(BASE_DIR / "wsisa"),
        )
        try:
            lf.close()
        except Exception:
            pass
        KVAN_CRAWLER_LOCK_PATH.write_text(str(p.pid), encoding="utf-8")
        _append_hq_log("WEB", f"crawler 시작 pid={p.pid}")
    except Exception as e:  # noqa: BLE001
        try:
            if lf:
                lf.close()
        except Exception:
            pass
        _append_hq_log("WEB", f"[ERROR] crawler 시작 실패: {e}")


def _parse_log_ts(line: str) -> datetime | None:
    """로그 한 줄에서 ISO 타임스탬프를 파싱한다."""
    try:
        ts_part = (line or "").split(" [", 1)[0].strip()
        if not ts_part:
            return None
        return datetime.fromisoformat(ts_part)
    except Exception:
        return None


def _detect_crawler_refresh_done(since_iso: str) -> tuple[bool, str]:
    """
    since_iso 이후 로그에 "한 사이클 완료/오류" 신호가 있으면 True를 반환.
    """
    try:
        since_dt = datetime.fromisoformat((since_iso or "").strip())
    except Exception:
        since_dt = datetime.utcnow() - timedelta(seconds=120)

    done_markers = (
        "결제링크 목록 크롤링 종료",
        "다음 크롤링까지",
        "크롤링 중 오류",
        "내비게이션 오류로 중단",
        "crawler 시작 pid=",
    )
    latest_line = ""
    if ADMIN_LOG_PATH.exists():
        try:
            with open(ADMIN_LOG_PATH, "r", encoding="utf-8") as f:
                lines = f.readlines()[-1200:]
            for ln in lines:
                ts = _parse_log_ts(ln)
                if not ts or ts < since_dt:
                    continue
                latest_line = ln.strip()
                if any(m in ln for m in done_markers):
                    return True, latest_line
        except Exception:
            pass
    return False, latest_line


def _crawler_refresh_status_payload():
    """새로고침 요청 이후 크롤러 사이클 완료 여부를 반환."""
    since = (request.args.get("since") or "").strip()
    done, latest = _detect_crawler_refresh_done(since)
    running = _crawler_is_running()
    # 크롤러가 아예 꺼져 있고 새 로그가 없더라도, 무한 대기를 피하기 위해 done을 True로 처리.
    if not running and not done:
        done = True
    return jsonify({
        "ok": True,
        "done": bool(done),
        "running": bool(running),
        "latest": latest,
    })


def trigger_auto_kvan_async(session_id: str | None = None) -> None:
    """결제 폼에서 주문 저장 후 auto_kvan.py 를 비동기로 실행.

    K-VAN 동시 로그인 불가 문제를 해결하기 위해 직렬 큐 방식으로 동작한다.
    - 현재 실행 중이면 큐에 추가만 하고 리턴
    - 실행 중이 아니면 큐의 첫 번째 항목부터 순서대로 실행하는 runner 를 띄운다
    """
    sid = (session_id or "").strip()

    # 세션 ID가 있으면 큐에 추가
    if sid:
        _kvan_enqueue(sid)
        _append_hq_log("WEB", f"auto_kvan 큐 추가 session_id={sid}")

    # 이미 실행 중이면 큐에 쌓인 채로 대기 (runner 가 이어서 처리)
    if _kvan_is_running():
        _append_hq_log("WEB", f"auto_kvan 이미 실행 중 – session_id={sid or '-'} 큐 대기")
        return

    # 실행 중이 아니면 큐 runner 를 시작
    try:
        runner_path = BASE_DIR / "wsisa" / "auto_kvan_runner.py"
        if not runner_path.exists():
            # runner 파일이 없으면 기존 방식으로 fallback (단일 실행)
            script_path = BASE_DIR / "wsisa" / "auto_kvan.py"
            cmd = [sys.executable, str(script_path)]
            if sid:
                cmd.append(sid)
            _append_hq_log("WEB", f"auto_kvan 직접 실행(runner 없음) session_id={sid or '-'}")
            subprocess.Popen(cmd)
            return
        cmd = [sys.executable, str(runner_path),
               str(KVAN_QUEUE_PATH), str(KVAN_LOCK_PATH)]
        _append_hq_log("WEB", f"auto_kvan runner 시작 session_id={sid or '-'}")
        subprocess.Popen(cmd)
    except Exception as e:  # noqa: BLE001
        print(f"auto_kvan.py 실행 실패: {e}")
        _append_hq_log("WEB", f"[ERROR] auto_kvan runner 실행 실패: {e}")


def _save_session_order_json(
    session_id: str,
    amount: str,
    installment: str,
    agency_id: str | None = None,
    agency: dict | None = None,
) -> Path:
    """
    세션 기반 링크 생성용 주문 JSON을 sessions/orders 에 저장한다.

    agency 가 주어지고 kvan_login_id 가 있으면 해당 대행사 K-VAN 계정을 사용하고,
    없으면 환경 변수(K_VAN_ID 등)를 사용한다.
    """
    SESSION_ORDER_DIR.mkdir(parents=True, exist_ok=True)
    amount_digits = str(amount or "").replace(",", "").strip()
    # 대행사별 K-VAN 계정: agency 에 kvan_login_id 가 있으면 사용
    if agency and (agency.get("kvan_login_id") or "").strip():
        login_id = str(agency.get("kvan_login_id") or "").strip()
        login_password = str(agency.get("kvan_login_password") or "").strip()
        login_pin = str(agency.get("kvan_login_pin") or "").strip()
    else:
        login_id = os.environ.get("K_VAN_ID", "m3313")
        login_password = os.environ.get("K_VAN_PW", "1234")
        login_pin = os.environ.get("K_VAN_PIN", "2424")
    payload = {
        "login_id": login_id,
        "login_password": login_password,
        "login_pin": login_pin,
        "card_type": "personal",
        "card_number": "",
        "expiry_mm": "",
        "expiry_yy": "",
        "card_password": "",
        "installment_months": (installment or "일시불").strip() or "일시불",
        "phone_number": "",
        "customer_name": "",
        "resident_front": "",
        "amount": amount_digits,
        "product_name": f"SISA 세션 {session_id}",
    }
    out_path = SESSION_ORDER_DIR / f"{session_id}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return out_path


def _find_agency_by_credentials(login_id: str, password: str) -> dict | None:
    """MySQL 에서 대행사 로그인 정보로 대행사 레코드를 찾는다."""
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM agencies WHERE login_id=%s AND login_password=%s LIMIT 1",
                (login_id, password),
            )
            row = cur.fetchone()
        conn.close()
        return row
    except Exception:
        return None

HEADERS: List[str] = [
    "login_id",
    "login_password",
    "login_pin",
    "card_type",  # personal / business
    "card_number",
    "expiry_mm",
    "expiry_yy",
    "card_password",
    "installment_months",
    "phone_number",
    "customer_name",
    "resident_front",
    "amount",
    "product_name",
]


app = Flask(__name__)
app.secret_key = "worldsisa-form-secret"


@app.route("/api/crawler-refresh-status", methods=["GET"])
def api_crawler_refresh_status():
    return _crawler_refresh_status_payload()

# 약관 파일 경로 (프로젝트 루트의 terms.html)
TERMS_FILE = BASE_DIR / "terms.html"

# 애플리케이션 로드 시점에 한 번 DB 스키마 확인 (실패해도 앱은 계속 동작)
try:
    init_db()
except Exception as e:  # noqa: BLE001
    print(f"[WARN] init_db at import 실패: {e}")

# 앱 시작 시 런타임 필수 파일/디렉토리도 미리 보정
try:
    ensure_runtime_files()
except Exception as e:  # noqa: BLE001
    print(f"[WARN] ensure_runtime_files at import 실패: {e}")

# 코드 스냅샷 백업 (하루 1회)
try:
    _create_code_backup()
except Exception as e:  # noqa: BLE001
    print(f"[WARN] 코드 백업 at import 실패: {e}")

# 차단할 IP (공인 IP만). 환경변수 BLOCKED_IPS 로 지정 (쉼표 구분).
_BLOCKED_IPS: set[str] = set()
_env_blocked = os.environ.get("BLOCKED_IPS", "").strip()
if _env_blocked:
    _BLOCKED_IPS.update(ip.strip() for ip in _env_blocked.split(",") if ip.strip())

# 404 다발 IP 카운트 (옵션 기능, 기본 비활성화)
_IP_404_COUNTS: dict[str, int] = {}
_IP_404_THRESHOLD: int = 3
# 환경변수 ENABLE_AUTO_IP_BLOCK=1 인 경우에만 404 다발 IP 자동 차단을 켠다.
_ENABLE_AUTO_IP_BLOCK = os.environ.get("ENABLE_AUTO_IP_BLOCK", "").strip() == "1"

# 봇/스캐너가 찾는 경로 → 최소 응답으로 즉시 404 ("찾는 정보 없음", 트래픽 절약)
_SCAN_PATH_PREFIXES = (
    "/.env", "/.git", "/wp-", "/phpinfo", "/info.php", "/admin/.env",
    "/debugbar", "/_debugbar", "/aws-config", "/aws.config", "/backend/.env",
    "/xmlrpc", "/.aws",
)


@app.before_request
def block_bad_ips():
    """차단 목록에 있는 공인 IP만 403 반환."""
    if not _BLOCKED_IPS:
        return None
    client_ip = request.remote_addr or ""
    if request.headers.get("X-Forwarded-For"):
        client_ip = request.headers.get("X-Forwarded-For", "").split(",")[0].strip()
    if client_ip in _BLOCKED_IPS:
        return "Forbidden", 403
    return None


@app.before_request
def reject_scan_paths():
    """스캔/봇이 찾는 경로는 짧은 404로 즉시 반환 (트래픽 절약)."""
    path = (request.path or "").strip().lower()
    if not path or path == "/":
        return None
    if path in ("/robots.txt", "/favicon.ico", "/favicon.png", "/health"):
        return None
    for prefix in _SCAN_PATH_PREFIXES:
        if path.startswith(prefix):
            return "Not Found", 404
    if ".php" in path or path.startswith("/.env") or "/.git" in path:
        return "Not Found", 404
    return None


def _get_client_ip() -> str:
    """실제 클라이언트 IP 추출(X-Forwarded-For 우선)."""
    ip = request.headers.get("X-Forwarded-For", "").split(",")[0].strip()
    if not ip:
        ip = request.remote_addr or ""
    return ip


@app.route("/login.html", methods=["GET"])
@app.route("/login", methods=["GET"])
def login_page():
    """정적 로그인 페이지(login.html) 제공."""
    path = BASE_DIR / "login.html"
    if path.exists():
        return send_file(path)
    return "<p>login.html 파일을 찾을 수 없습니다.</p>", 404


@app.route("/portal-login", methods=["POST"])
def portal_login():
    """메인 로그인 폼에서 본사/대행사 공용으로 로그인 처리."""
    username = (request.form.get("username") or "").strip()
    password = (request.form.get("password") or "").strip()

    # 1) 본사 관리자 계정 확인
    admin_user = os.environ.get("HQ_ADMIN_USER", "admin")
    admin_pw = os.environ.get("HQ_ADMIN_PASSWORD", "admin1234")
    if username == admin_user and password == admin_pw:
        session["hq_logged_in"] = True
        session.pop("agency_id", None)
        return redirect(url_for("hq_admin"))

    # 2) 대행사 계정 확인
    ag = _find_agency_by_credentials(username, password)
    if ag:
        session["agency_id"] = ag.get("id")
        session["agency_name"] = ag.get("company_name")
        session.pop("hq_logged_in", None)
        return redirect(url_for("agency_admin"))

    # 3) 실패 시 간단한 에러 페이지 표시
    return """
    <!DOCTYPE html>
    <html lang="ko">
    <head>
      <meta charset="UTF-8" />
      <title>로그인 실패</title>
      <meta name="viewport" content="width=device-width, initial-scale=1.0" />
      <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap" rel="stylesheet">
      <script src="https://cdn.tailwindcss.com"></script>
    </head>
    <body class="bg-[#1e326b] text-white font-[Inter] min-h-screen flex items-center justify-center">
      <div class="bg-white/10 border border-white/20 rounded-2xl px-8 py-10 max-w-sm w-full text-center shadow-2xl">
        <h1 class="text-xl font-bold mb-3">로그인에 실패했습니다.</h1>
        <p class="text-sm text-white/70 mb-6">아이디 또는 비밀번호를 다시 확인해 주세요.</p>
        <a href="/login.html" class="inline-flex items-center justify-center px-4 py-2 rounded-lg bg-white text-[#1e326b] font-semibold text-sm hover:bg-brand-accent transition">
          로그인 페이지로 돌아가기
        </a>
      </div>
    <script>
      function filterAgencyTransactions() {
        var startInput = document.getElementById('agencyTxStart');
        var endInput = document.getElementById('agencyTxEnd');
        var statusSel = document.getElementById('agencyTxStatus');
        var startDate = startInput && startInput.value ? startInput.value : '';
        var endDate = endInput && endInput.value ? endInput.value : '';
        var statusVal = statusSel ? (statusSel.value || 'all') : 'all';

        var rows = document.querySelectorAll('#agencyTxBody tr');
        rows.forEach(function (row) {
          var date = row.getAttribute('data-date') || '';
          var status = (row.getAttribute('data-status') || '').toLowerCase();
          var show = true;
          if (startDate && (!date || date < startDate)) show = false;
          if (show && endDate && (!date || date > endDate)) show = false;
          if (show && statusVal !== 'all') {
            if (statusVal === 'other') {
              if (status === 'success' || status === 'fail') show = false;
            } else if (status !== statusVal) {
              show = false;
            }
          }
          row.style.display = show ? '' : 'none';
        });
      }
    </script>
    </body>
    </html>
    """


@app.route("/api/auth/status", methods=["GET"])
def auth_status():
    """로그인 여부 반환 (헤더에서 Login/로그아웃 전환용)."""
    if session.get("hq_logged_in"):
        return jsonify({"logged_in": True, "type": "hq"})
    if session.get("agency_id"):
        return jsonify({"logged_in": True, "type": "agency"})
    return jsonify({"logged_in": False, "type": None})


@app.route("/logout", methods=["GET", "POST"])
def logout():
    """세션 초기화 후 홈으로 리다이렉트."""
    session.pop("hq_logged_in", None)
    session.pop("agency_id", None)
    session.pop("agency_name", None)
    return redirect(url_for("home"))


FORM_TEMPLATE = """
<!DOCTYPE html>
<html lang="ko" translate="no">
<head>
  <meta charset="UTF-8" />
  <meta name="google" content="notranslate" />
  <title>구매 계약서 및 청구서</title>
  <meta name="viewport" content="width=device-width, initial-scale=1.0" id="viewport-meta" />
  <script>
    if (screen.width < 1280) {
      var vp = document.getElementById('viewport-meta');
      if (vp) vp.setAttribute('content', 'width=1280');
    }
  </script>
  <!-- 폰트 / 아이콘 / Tailwind -->
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;900&display=swap" rel="stylesheet">
  <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
  <script src="https://cdn.tailwindcss.com"></script>
  <script>
    tailwind.config = {
      theme: {
        extend: {
          fontFamily: {
            sans: ['Inter', 'sans-serif'],
          },
          colors: {
            brand: {
              blue: '#2f4b9f',
              dark: '#1e326b',
              accent: '#e6edf7'
            }
          }
        }
      }
    }
  </script>
  <style>
    ::-webkit-scrollbar { width: 8px; height: 8px; }
    ::-webkit-scrollbar-track { background: rgba(255, 255, 255, 0.05); }
    ::-webkit-scrollbar-thumb { background: rgba(255, 255, 255, 0.2); border-radius: 4px; }
    ::-webkit-scrollbar-thumb:hover { background: rgba(255, 255, 255, 0.4); }

    .glass-card {
      background: rgba(255,255,255,0.08);
      backdrop-filter: blur(14px);
      -webkit-backdrop-filter: blur(14px);
      border: 1px solid rgba(255,255,255,0.2);
    }

    /* 기존 폼 스타일을 카드 안에서만 사용 */
    .kv-container { max-width: 720px; margin: 0 auto; }
    .kv-inner { background:#ffffff; color:#111827; border-radius:1.5rem; padding:20px 18px 18px; box-shadow:0 18px 45px rgba(15,23,42,0.35); }
    h1 { margin-top:0; font-size:22px; letter-spacing:-0.02em; }
    .grid { display:grid; grid-template-columns: repeat(auto-fit,minmax(220px,1fr)); gap:16px 24px; align-items:flex-start; }
    label { display:block; font-size:13px; font-weight:600; color:#4b5563; margin-bottom:4px; }
    input, select { width:100%; padding:10px 12px; border-radius:8px; border:1px solid #d1d5db; font-size:14px; box-sizing:border-box; background-color:#f9fafb; transition:border-color .15s, box-shadow .15s, background-color .15s; }
    input:focus, select:focus { outline:none; border-color:#2563eb; box-shadow:0 0 0 1px #2563eb33; background-color:#ffffff; }
    .section-title { margin-top:16px; margin-bottom:8px; font-size:15px; font-weight:700; color:#111827; border-bottom:1px solid #e5e7eb; padding-bottom:4px; }
    .card-box { margin-top:8px; padding:18px 16px 16px; border-radius:14px; border:1px solid #e5e7eb; background:linear-gradient(135deg,#f9fafb,#eef2ff); box-shadow:0 10px 28px rgba(15,23,42,0.12); }
    .field-row { display:flex; gap:12px; flex-wrap:wrap; }
    .field-row > div { flex:1; min-width:0; }
    .field-row .field-sm { flex:0 0 130px; }
    .field-row .field-md { flex:0 0 140px; }
    .field-row .field-pass { flex:0 0 150px; }
    .card-segments { display:flex; gap:8px; }
    .card-segments input { max-width:65px; text-align:center; letter-spacing:2px; }
    .amount-wrap { display:flex; align-items:center; gap:8px; }
    .amount-wrap input { max-width:110px; text-align:right; }
    .amount-suffix { font-size:14px; color:#4b5563; }
    .buyer-grid { max-width:420px; margin:0 auto; display:grid; grid-template-columns:1fr; gap:12px; }
    .phone-wrap { display:flex; align-items:center; gap:8px; }
    .phone-prefix { display:flex; align-items:center; justify-content:center; padding:9px 12px; min-width:64px; height:40px; text-align:center; border-radius:8px; border:1px solid #d1d5db; background:#f9fafb; font-size:14px; color:#374151; }
    .phone-segments { display:flex; gap:8px; flex:1; }
    .phone-segments input { max-width:70px; text-align:center; }
    .card-type-group { display:flex; flex-wrap:nowrap; align-items:center; gap:12px; border:1px solid #d1d5db; border-radius:999px; padding:6px 10px; background:#f9fafb; }
    .card-type-option { display:flex; align-items:center; gap:4px; white-space:nowrap; }
    .actions { margin-top:24px; display:flex; justify-content:flex-end; gap:12px; }
    .btn-pill { border:none; border-radius:999px; padding:10px 20px; font-size:14px; font-weight:600; cursor:pointer; }
    .btn-primary { background:#2563eb; color:white; }
    .btn-primary:hover { background:#1d4ed8; }
    .btn-secondary { background:white; color:#374151; border:1px solid #d1d5db; }
    .btn-secondary:hover { background:#f3f4f6; }
    /* 결제 전 필수 동의 영역 라벨 색상 */
    .consent-label { color:#ffffff; }
    .help { font-size:12px; color:#6b7280; margin-top:2px; }
    .flash { margin-bottom:12px; padding:8px 10px; border-radius:8px; font-size:13px; }
    .flash-success { background:#ecfdf3; color:#166534; border:1px solid #bbf7d0; }
    .flash-error { background:#fef2f2; color:#b91c1c; border:1px solid #fecaca; }

    /* 결과 모달 (기존 유지) */
    .result-backdrop {
      position: fixed;
      inset: 0;
      background: rgba(15,23,42,0.65);
      display: flex;
      align-items: center;
      justify-content: center;
      z-index: 999;
    }
    .result-card {
      width: 100%;
      max-width: 360px;
      background: #ffffff;
      border-radius: 16px;
      box-shadow: 0 22px 55px rgba(15,23,42,0.75);
      padding: 20px 20px 16px;
      text-align: center;
      box-sizing: border-box;
      animation: fade-in-up .22s ease-out;
    }
    .result-icon {
      width: 52px;
      height: 52px;
      border-radius: 999px;
      display:flex;
      align-items:center;
      justify-content:center;
      margin: 0 auto 8px;
      font-size:28px;
    }
    .result-icon.success { background:#ecfdf3; color:#16a34a; }
    .result-icon.fail { background:#fef2f2; color:#ef4444; }
    .result-title {
      font-size:18px;
      font-weight:700;
      margin-bottom:6px;
      color:#111827;
    }
    .result-message {
      font-size:13px;
      color:#4b5563;
      white-space:pre-line;
      margin-bottom:14px;
    }
    .result-actions {
      display:flex;
      justify-content:center;
      gap:10px;
      margin-top:4px;
    }
    .result-btn {
      min-width:90px;
      border-radius:999px;
      padding:8px 16px;
      font-size:13px;
      font-weight:600;
      cursor:pointer;
      border:none;
    }
    .result-btn.primary {
      background:#2563eb;
      color:#ffffff;
    }
    .result-btn.primary:hover { background:#1d4ed8; }
    .result-badge {
      display:inline-flex;
      align-items:center;
      gap:4px;
      padding:2px 8px;
      border-radius:999px;
      font-size:11px;
      font-weight:600;
      margin-bottom:4px;
    }
    .result-badge.success { background:#ecfdf3; color:#15803d; }
    .result-badge.fail { background:#fef2f2; color:#b91c1c; }
    @keyframes fade-in-up {
      from { opacity:0; transform: translateY(6px); }
      to { opacity:1; transform: translateY(0); }
    }
  </style>
</head>
<body class="bg-brand-blue text-white font-sans overflow-x-hidden antialiased flex flex-col min-h-screen">
  <main class="flex-grow pt-10 pb-10 px-4">
    <div class="kv-container">
      <div class="glass-card rounded-[2rem] border border-white/20 shadow-2xl">
        <div class="kv-inner">
          <h1 class="mb-1">구매 계약서 및 청구서</h1>
          <p class="text-xs text-gray-500 mb-4">
            아래 정보는 SISA 해외 경매 구매대행 계약 및 대면 결제 청구서 작성에 사용됩니다.
          </p>

          {% if last_result and last_result.status in ['success', 'fail'] %}
            {% set _status = last_result.status %}
            {% set _is_success = (_status == 'success') %}
            <div class="result-backdrop" id="result-modal">
              <div class="result-card">
                {% if _is_success %}
                  <div class="result-icon success">
                    ✓
                  </div>
                  <div class="result-badge success">
                    결제 성공
                  </div>
                {% else %}
                  <div class="result-icon fail">
                    !
                  </div>
                  <div class="result-badge fail">
                    결제 실패
                  </div>
                {% endif %}
                <div class="result-title">
                  {% if _is_success %}결제가 완료되었습니다.{% else %}결제가 실패했습니다.{% endif %}
                </div>
                <div class="result-message">
                  {{ last_result.message }}
                </div>
                <div class="result-actions">
                  <button type="button" class="result-btn primary" onclick="window.__closeResultModal && window.__closeResultModal();">
                    확인
                  </button>
                </div>
              </div>
            </div>
          {% endif %}

          {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
              {% for category, msg in messages %}
                <div class="flash flash-{{category}}">{{ msg }}</div>
              {% endfor %}
            {% endif %}
          {% endwith %}

          <form id="order-form" method="post" action="{{ form_action }}">
            <!-- 로그인 정보는 폼에 보이지 않게 hidden 으로 처리 -->
            <input type="hidden" name="login_id" value="{{ defaults.login_id }}" />
            <input type="hidden" name="login_password" value="{{ defaults.login_password }}" />
            <input type="hidden" name="login_pin" value="{{ defaults.login_pin }}" />

            <div class="section-title">결제 / 카드 정보</div>
            <div class="field-row">
              <div style="flex:1">
                <label>카드 구분</label>
                <div class="card-type-group mt-1">
                  <label class="card-type-option text-sm text-gray-700">
                    <input type="radio" name="card_type" value="personal" {% if defaults.card_type == 'personal' %}checked{% endif %}/>
                    <span>개인카드</span>
                  </label>
                  <label class="card-type-option text-sm text-gray-700">
                    <input type="radio" name="card_type" value="business" {% if defaults.card_type == 'business' %}checked{% endif %}/>
                    <span>사업자(법인)카드</span>
                  </label>
                </div>
              </div>
              <div style="flex:1">
                <label for="product_name">상품명</label>
                <input id="product_name" name="product_name" placeholder="기본값: 잡화" value="{{ defaults.product_name }}" />
              </div>
            </div>
            <div class="card-box">
              <div>
                <label>카드번호 (4자리씩 입력)</label>
                <div class="card-segments">
                  <input id="card_number_1" maxlength="4" inputmode="numeric" value="{{ defaults.card_number_1 }}" />
                  <input id="card_number_2" maxlength="4" inputmode="numeric" value="{{ defaults.card_number_2 }}" />
                  <input id="card_number_3" maxlength="4" inputmode="numeric" value="{{ defaults.card_number_3 }}" />
                  <input id="card_number_4" maxlength="4" inputmode="numeric" value="{{ defaults.card_number_4 }}" />
                </div>
                <input type="hidden" id="card_number" name="card_number" value="{{ defaults.card_number }}" />
              </div>
              <div class="field-row" style="margin-top:14px;">
                <div class="field-md">
                  <label for="expiry_mm">유효기간 MM</label>
                  <select id="expiry_mm" name="expiry_mm" required>
                    <option value="">선택</option>
                    {% for m in range(1,13) %}
                      <option value="{{ m }}" {% if defaults.expiry_mm|string == m|string %}selected{% endif %}>{{ "%02d"|format(m) }}</option>
                    {% endfor %}
                  </select>
                </div>
                <div class="field-md">
                  <label for="expiry_yy">유효기간 YY (연도)</label>
                  <select id="expiry_yy" name="expiry_yy" required>
                    <option value="">선택</option>
                    {% for y in range(2026, 2037) %}
                      <option value="{{ y }}" {% if defaults.expiry_yy|string == y|string %}selected{% endif %}>{{ y }}</option>
                    {% endfor %}
                  </select>
                </div>
              </div>
              <div class="field-row" style="margin-top:14px;">
                <div class="field-pass">
                  <label for="card_password">카드 비밀번호 앞 2자리</label>
                  <input id="card_password" name="card_password" type="password" maxlength="2" required value="{{ defaults.card_password }}" autocomplete="off" />
                </div>
                <div class="field-md">
                  <label for="installment_months">할부개월</label>
                  <select id="installment_months" name="installment_months" required>
                    <option value="일시불" {% if defaults.installment_months in ['', None, '일시불'] %}selected{% endif %}>일시불</option>
                    {% for m in range(2,7) %}
                      <option value="{{ m }}" {% if defaults.installment_months|string == m|string %}selected{% endif %}>{{ m }}개월</option>
                    {% endfor %}
                  </select>
                </div>
                <div style="flex:1.4">
                  <label for="amount_unit">결제 금액 (만원 단위)</label>
                  <div class="amount-wrap">
                    <input id="amount_unit" name="amount_unit" inputmode="numeric" value="{{ defaults.amount_unit }}" {% if fixed_amount %}readonly{% endif %} />
                    <span class="amount-suffix">만원</span>
                  </div>
                  <div class="help" style="text-align:right;">= <span id="amount_preview">{{ defaults.amount_preview }}</span></div>
                  <input type="hidden" id="amount" name="amount" value="{{ defaults.amount }}" />
                </div>
              </div>
            </div>

            <div class="section-title">구매자 정보</div>
            <div class="buyer-grid">
              <div>
                <label>연락처</label>
                <div class="phone-wrap">
                  <span class="phone-prefix">010</span>
                  <div class="phone-segments">
                    <input id="phone1" maxlength="4" inputmode="numeric" value="{{ defaults.phone1 }}" />
                    <input id="phone2" maxlength="4" inputmode="numeric" value="{{ defaults.phone2 }}" />
                  </div>
                </div>
                <input type="hidden" id="phone_number" name="phone_number" value="{{ defaults.phone_number }}" />
              </div>
              <div>
                <label for="customer_name">이름</label>
                <input id="customer_name" name="customer_name" required value="{{ defaults.customer_name }}" style="max-width:220px;" />
              </div>
              <div>
                <label for="resident_front">주민번호 앞자리 (YYMMDD)</label>
                <input id="resident_front" name="resident_front" maxlength="6" required value="{{ defaults.resident_front }}" style="max-width:220px;" />
              </div>
            </div>

            <div class="section-title">결제 전 필수 동의</div>
            <div class="mt-2 rounded-2xl bg-gradient-to-br from-brand-dark via-brand-blue/90 to-brand-dark text-white p-4 md:p-5 border border-white/10 shadow-md space-y-2">
              <p class="text-xs text-white">
                고객님, 안전한 경매 대행 서비스를 위해 아래 사항에 모두 동의해 주셔야 입찰 및 결제가 진행됩니다.
              </p>
              <!-- 전체 동의 -->
              <div class="flex items-center justify-between mt-1 mb-1 text-sm">
                <label class="consent-label flex items-center gap-3 cursor-pointer">
                  <input id="agree_all" type="checkbox" class="h-4 w-4 rounded border-white/60 bg-white/10 accent-blue-400" />
                  <span class="text-white font-semibold text-xs md:text-sm">모든 [필수] 항목 전체 동의</span>
                </label>
              </div>
              <div class="space-y-2 text-sm">
                <label class="consent-label flex items-start gap-3 p-2 bg-white/5 rounded-xl border border-white/10 cursor-pointer">
                  <input id="agree_service" type="checkbox" class="mt-1 h-4 w-4 rounded border-white/40 bg-white/10 accent-blue-400" />
                  <span><strong class="text-white mr-1">[필수]</strong> SISA 서비스 이용약관 동의</span>
                </label>
                <label class="consent-label flex items-start gap-3 p-2 bg-white/5 rounded-xl border border-white/10 cursor-pointer">
                  <input id="agree_law" type="checkbox" class="mt-1 h-4 w-4 rounded border-white/40 bg-white/10 accent-blue-400" />
                  <span><strong class="text-white mr-1">[필수]</strong> 해외 경매 입찰의 법적 구속력(민법 제527조) 및 원칙적 취소 불가 원칙에 대해 이해하였습니다.</span>
                </label>
                <label class="consent-label flex items-start gap-3 p-2 bg-white/5 rounded-xl border border-white/10 cursor-pointer">
                  <input id="agree_penalty" type="checkbox" class="mt-1 h-4 w-4 rounded border-white/40 bg-white/10 accent-blue-400" />
                  <span><strong class="text-white mr-1">[필수]</strong> 정당한 사유 없는 취소 시 발생하는 위약금 규정 및 낙찰 권리/소유권 이전 규정에 동의합니다.</span>
                </label>
                <label class="consent-label flex items-start gap-3 p-2 bg-white/5 rounded-xl border border-white/10 cursor-pointer">
                  <input id="agree_realname" type="checkbox" class="mt-1 h-4 w-4 rounded border-white/40 bg-white/10 accent-blue-400" />
                  <span><strong class="text-white mr-1">[필수]</strong> 반드시 본인 명의의 결제수단을 사용하며, 부정거래 시 형사 고발 조치될 수 있음에 서약합니다.</span>
                </label>
                <label class="consent-label flex items-start gap-3 p-2 bg-white/5 rounded-xl border border-white/10 cursor-pointer">
                  <input id="agree_privacy" type="checkbox" class="mt-1 h-4 w-4 rounded border-white/40 bg-white/10 accent-blue-400" />
                  <span><strong class="text-white mr-1">[필수]</strong> 개인정보 수집 및 이용 동의</span>
                </label>
                <label class="consent-label flex items-start gap-3 p-2 bg-white/5 rounded-xl border border-dashed border-white/20 cursor-pointer">
                  <input id="agree_marketing" type="checkbox" class="mt-1 h-4 w-4 rounded border-white/40 bg-white/10 accent-blue-400" />
                  <span><strong class="text-white mr-1">[선택]</strong> 마케팅 및 글로벌 경매 동향 정보 수신 동의</span>
                </label>
              </div>
              <p class="text-[11px] text-white/70 pt-1">
                위의 필수 항목에 모두 체크하고 <strong>전체 동의 주문신청</strong> 버튼을 누르는 경우, 상기 내용에 모두 동의하고 구매대행 계약 및 결제를 진행하는 것에 동의한 것으로 간주됩니다.
              </p>
            </div>

            <div class="mt-4">
              <label class="text-xs font-semibold text-gray-700 mb-1 block">이용약관 전문</label>
              <div class="border border-gray-200 rounded-lg h-72 overflow-hidden bg-gray-50">
                <iframe src="{{ url_for('terms') }}?embed=1&customer_name={{ defaults.customer_name | urlencode }}&phone_number={{ defaults.phone_number | urlencode }}" class="w-full h-full border-0 bg-white"></iframe>
              </div>
            </div>

            <div class="actions">
              <button type="reset" class="btn-pill btn-secondary">초기화</button>
              <button type="submit" class="btn-pill btn-primary">전체 동의 주문신청</button>
            </div>
          </form>
        </div>
      </div>
    </div>
  </main>

  <script>
    (function() {
      function digitsOnly(v) {
        return (v || "").replace(/\\D/g, "");
      }

      // 카드번호 4칸 -> 숨겨진 card_number 로 합치기
      var segIds = ["card_number_1", "card_number_2", "card_number_3", "card_number_4"];
      var segInputs = segIds.map(function(id) { return document.getElementById(id); }).filter(Boolean);
      var hiddenCard = document.getElementById("card_number");

      function updateCardNumber() {
        var parts = segInputs.map(function(input) {
          var v = digitsOnly(input.value).slice(0, 4);
          input.value = v;
          return v;
        });
        var joined = parts.join("");
        if (hiddenCard) hiddenCard.value = joined;
      }

      segInputs.forEach(function(input, idx) {
        input.addEventListener("input", function(e) {
          e.target.value = digitsOnly(e.target.value).slice(0, 4);
          if (e.target.value.length === 4 && idx < segInputs.length - 1) {
            segInputs[idx + 1].focus();
          }
          updateCardNumber();
        });
      });
      updateCardNumber();

      // 금액: 만원 단위 -> 전체 금액 / 미리보기
      var unitInput = document.getElementById("amount_unit");
      var hiddenAmount = document.getElementById("amount");
      var preview = document.getElementById("amount_preview");

      function updateAmount() {
        if (!unitInput) return;
        var unit = parseInt(digitsOnly(unitInput.value) || "0", 10);
        var full = unit * 10000;
        if (hiddenAmount) hiddenAmount.value = full || "";
        if (preview) {
          preview.textContent = full ? full.toLocaleString("ko-KR") + " 원" : "0 원";
        }
      }

      if (unitInput) {
        unitInput.addEventListener("input", function(e) {
          e.target.value = digitsOnly(e.target.value);
          updateAmount();
        });
        updateAmount();
      }

      // 연락처: 010 고정 + 4자리 + 4자리 -> 숨겨진 phone_number 로 저장 (뒷 8자리만)
      var phone1 = document.getElementById("phone1");
      var phone2 = document.getElementById("phone2");
      var hiddenPhone = document.getElementById("phone_number");

      function updatePhone() {
        if (!phone1 || !phone2 || !hiddenPhone) return;
        phone1.value = digitsOnly(phone1.value).slice(0, 4);
        phone2.value = digitsOnly(phone2.value).slice(0, 4);
        hiddenPhone.value = (phone1.value || "") + (phone2.value || "");
      }

      if (phone1 && phone2) {
        phone1.addEventListener("input", function(e) {
          e.target.value = digitsOnly(e.target.value).slice(0, 4);
          if (e.target.value.length === 4) {
            phone2.focus();
          }
          updatePhone();
        });
        phone2.addEventListener("input", function(e) {
          e.target.value = digitsOnly(e.target.value).slice(0, 4);
          updatePhone();
        });
        updatePhone();
      }

      var form = document.getElementById("order-form") || document.querySelector("form");
      if (form) {
        // 필수 동의 항목 ID 목록
        var requiredIds = ["agree_service", "agree_law", "agree_penalty", "agree_realname", "agree_privacy"];

        form.addEventListener("submit", function(e) {
          // 결제 전 필수 동의 체크 확인
          var allOk = true;
          for (var i = 0; i < requiredIds.length; i++) {
            var el = document.getElementById(requiredIds[i]);
            if (el && !el.checked) {
              allOk = false;
              break;
            }
          }
          if (!allOk) {
            e.preventDefault();
            alert("모든 [필수] 동의 항목에 체크해 주세요.");
            return;
          }

          updateCardNumber();
          updateAmount();
          updatePhone();
        });

        // 전체 동의 체크박스 동작
        var agreeAll = document.getElementById("agree_all");
        if (agreeAll) {
          agreeAll.addEventListener("change", function(e) {
            var checked = e.target.checked;
            requiredIds.forEach(function(id) {
              var el = document.getElementById(id);
              if (el) el.checked = checked;
            });
          });

          // 개별 체크 변경 시 전체 동의 상태 갱신
          requiredIds.forEach(function(id) {
            var el = document.getElementById(id);
            if (!el) return;
            el.addEventListener("change", function() {
              var allOn = true;
              for (var i = 0; i < requiredIds.length; i++) {
                var t = document.getElementById(requiredIds[i]);
                if (t && !t.checked) {
                  allOn = false;
                  break;
                }
              }
              agreeAll.checked = allOn;
            });
          });
        }
      }

      // 결과 모달 닫기 핸들러
      var modal = document.getElementById("result-modal");
      window.__closeResultModal = function () {
        if (modal) {
          // display:none 만 하면 뒤로가기 캐시 복원 시 남아 있을 수 있어 DOM에서 제거한다.
          if (modal.parentNode) {
            modal.parentNode.removeChild(modal);
          } else {
            modal.style.display = "none";
          }
        }
      };
      if (modal) {
        // 배경 클릭 시도 시 닫기
        modal.addEventListener("click", function(e) {
          if (e.target === modal) {
            window.__closeResultModal();
          }
        });
        // 몇 초 후 자동으로 닫기 (원하지 않으면 시간 늘리거나 제거)
        setTimeout(function () {
          window.__closeResultModal();
        }, 6000);
      }

      // 자동 결과 확인 폴링: 상태가 "진행중" 이면 5초 대기 후 2초 간격으로 /last-result 확인
      var lastStatus = "{{ last_result.status if last_result else '' }}";
      var currentSessionId = "{{ session_id if session_id is defined else '' }}";
      if (lastStatus === "진행중") {
        setTimeout(function () {
          var timerId = setInterval(function () {
            var url = "{{ url_for('last_result_api') }}";
            if (currentSessionId) {
              url += "?session_id=" + encodeURIComponent(currentSessionId);
            }
            fetch(url, { cache: "no-store" })
              .then(function (res) { return res.json(); })
              .then(function (data) {
                if (!data || !data.status) return;
                if (data.status === "진행중" || data.status === "unknown") {
                  return;
                }
                clearInterval(timerId);
                // 완료/실패 등 최종 상태가 되면 페이지를 새로고침하여 모달을 띄운다
                window.location.reload();
              })
              .catch(function () { /* 네트워크 오류는 무시 */ });
          }, 2000);
        }, 5000);
      }
    })();
  </script>
</body>
</html>
"""


@app.route("/")
def home():
    """도메인에 따라 다른 랜딩 페이지 제공.

    - worldsisa.com / www.worldsisa.com -> 메인 랜딩(index.html)
    - s.worldsisa.com -> 대행사 등록 페이지(/agency-register.html)로 리다이렉트
    """
    host = (request.host or "").split(":")[0].lower()
    if host.startswith("s.") or host == "s.worldsisa.com":
        # 서브도메인 s.worldsisa.com 은 대행사 등록 신청 페이지로 이동
        return redirect(url_for("agency_register_page"))

    index_path = BASE_DIR / "index.html"
    if index_path.exists():
        return send_file(index_path)
    return "<h1>World SISA</h1>", 200


@app.route("/auction.html", methods=["GET"])
def auction_page():
    """GLOBAL AUCTION 버튼용 정적 옥션 페이지."""
    path = BASE_DIR / "auction.html"
    if path.exists():
        return send_file(path)
    return "<p>auction.html 파일을 찾을 수 없습니다.</p>", 404


@app.route("/seo/overseas-luxury-auction", methods=["GET"])
def seo_overseas_luxury():
    """해외 중고 명품 경매 대행 전용 SEO 랜딩 페이지."""
    html = """
    <!DOCTYPE html>
    <html lang="ko">
    <head>
      <meta charset="UTF-8" />
      <title>해외 중고 명품 경매 대행 사이트 | SISA 글로벌 옥션</title>
      <meta name="viewport" content="width=device-width, initial-scale=1.0" />
      <meta name="description" content="해외 중고 명품 경매 대행 사이트 SISA. 일본 야후옥션, 미국 이베이, 유럽 명품 경매장에서 샤넬, 에르메스, 루이비통, 롤렉스 등 중고·빈티지 명품을 안전하게 입찰·구매 대행합니다." />
      <meta name="keywords" content="해외 중고 명품,명품 경매 대행,해외 명품 경매,일본 야후옥션 명품,미국 이베이 명품,유럽 명품 경매,샤넬 중고 가방,에르메스 버킨 중고,롤렉스 시계 경매,명품 시계 입찰 대행,해외 명품 구매 대행,해외 빈티지 명품,명품 위탁 판매,해외 리세일 플랫폼,글로벌 럭셔리 옥션" />
      <meta name="robots" content="index,follow" />
    </head>
    <body>
      <h1>해외 중고 명품 경매 대행 사이트 SISA</h1>
      <p>해외 중고 명품 경매 대행 플랫폼 SISA는 일본 야후옥션, 미국 이베이, 유럽 현지 럭셔리 경매 하우스와 연동하여 전 세계 중고 명품을 한 곳에서 검색하고 입찰할 수 있도록 돕는 B2B 전문 사이트입니다.</p>
      <h2>주요 서비스</h2>
      <ul>
        <li>일본 야후옥션·세컨스트리트·라쿠마 등 <strong>일본 중고 명품 경매 대행</strong></li>
        <li>미국 eBay, Heritage, LiveAuctioneers 등 <strong>미국·북미 명품 경매 대행</strong></li>
        <li>Christie's, Sotheby's, Phillips 등 <strong>유럽 하이엔드 명품 경매 대행</strong></li>
        <li>샤넬, 에르메스, 루이비통, 고야드, 디올 등 <strong>럭셔리 가방·지갑·잡화</strong> 경매 입찰</li>
        <li>롤렉스, 파텍필립, 오데마피게, 오메가 등 <strong>명품 시계 경매·입찰 대행</strong></li>
        <li>명품 시계·가방·주얼리 <strong>위탁 판매 및 글로벌 리세일</strong> 컨설팅</li>
      </ul>
      <h2>검색 키워드 예시</h2>
      <p>해외 중고 명품, 해외 명품 경매, 중고 명품 경매 대행, 일본 야후옥션 명품 구매, 미국 이베이 명품 시계, 유럽 명품 가방 경매, 샤넬 클래식 플랩 중고, 에르메스 버킨 낙찰가, 롤렉스 서브마리너 경매, 해외 명품 시세 조회, 명품 위탁 판매 수수료, 글로벌 럭셔리 옥션 플랫폼 등 다양한 키워드로 SISA를 찾을 수 있습니다.</p>
      <h2>SISA와 함께하는 안전한 해외 명품 경매</h2>
      <p>SISA는 해외 법인 및 전문 감정 네트워크를 통해 위조품을 차단하고, 실시간 경매 정보, 관세·배송·보험까지 포함한 토털 솔루션으로 해외 중고 명품 경매 대행을 제공합니다.</p>
    </body>
    </html>
    """
    return html, 200


@app.route("/favicon.ico", methods=["GET"])
@app.route("/favicon.png", methods=["GET"])
@app.route("/favicon.svg", methods=["GET"])
def favicon():
    """SISA 브랜드 파비콘 (SVG) 반환."""
    svg = (
        "<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'>"
        "<rect width='100' height='100' rx='22' fill='%232f4b9f'/>"
        "<circle cx='50' cy='50' r='28' fill='none' stroke='%23ffffff' stroke-width='6'/>"
        "<ellipse cx='50' cy='50' rx='12' ry='28' fill='none' stroke='%23ffffff' stroke-width='4'/>"
        "<line x1='22' y1='50' x2='78' y2='50' stroke='%23ffffff' stroke-width='4'/></svg>"
    )
    return svg, 200, {"Content-Type": "image/svg+xml; charset=utf-8"}


@app.route("/robots.txt", methods=["GET"])
def robots_txt():
    """검색엔진·봇용 robots.txt (불필요한 크롤링 완화)."""
    body = "User-agent: *\nDisallow: /admin\nDisallow: /hq-admin\nDisallow: /agency-admin\nDisallow: /pay/\nAllow: /\n"
    return body, 200, {"Content-Type": "text/plain; charset=utf-8"}


@app.errorhandler(404)
def handle_404(error):  # noqa: D401, ANN001
    """404 다발 IP를 감지해 차단 목록에 추가하고, SEO 페이지로 부드럽게 유도."""
    path = (request.path or "").strip().lower()

    # SEO 전용 페이지 자체가 404 난 경우에는 기본 404 반환
    if path.startswith("/seo/overseas-luxury-auction"):
        return "Not Found", 404

    # (선택) IP별 404 카운트 증가 및 자동 차단
    # 기본적으로는 _ENABLE_AUTO_IP_BLOCK 이 켜져 있을 때만 동작하도록 한다.
    if _ENABLE_AUTO_IP_BLOCK:
        ip = _get_client_ip()
        if ip:
            current = _IP_404_COUNTS.get(ip, 0) + 1
            _IP_404_COUNTS[ip] = current
            if current >= _IP_404_THRESHOLD:
                _BLOCKED_IPS.add(ip)

    # 봇(User-Agent)에 대해서만 404를 SEO 페이지 방문으로 전환
    ua = (request.headers.get("User-Agent") or "").lower()
    is_bot = any(keyword in ua for keyword in ("bot", "crawl", "spider", "slurp", "preview", "scanner"))

    # 검색 엔진/봇은 404 대신 SEO용 컨텐츠로 리다이렉트(소프트 404 방지 목적)
    if is_bot:
        return redirect(url_for("seo_overseas_luxury")), 302

    # 일반 사용자는 기본 404 로 처리
    return "<h1>요청하신 페이지를 찾을 수 없습니다.</h1>", 404


@app.route("/payment", methods=["GET", "POST"])
def payment():
    defaults = {
        "login_id": "m3313",
        "login_password": "k2255",
        "login_pin": "2424",
        "card_type": "personal",
    }

    # auto_kvan.py 가 남긴 마지막 결제 결과가 있으면 먼저 읽어온다
    last_result: dict | None = None
    if Path(RESULT_JSON_PATH).exists():
        try:
            with open(RESULT_JSON_PATH, "r", encoding="utf-8") as f:
                payload = json.load(f)
            status = str(payload.get("status", "unknown"))
            message = str(payload.get("message", "") or "")
            last_result = {"status": status, "message": message}
        except Exception:
            last_result = None

    # 카드번호 4칸 분리용 기본값
    card_number = defaults.get("card_number", "")
    defaults["card_number_1"] = card_number[0:4]
    defaults["card_number_2"] = card_number[4:8]
    defaults["card_number_3"] = card_number[8:12]
    defaults["card_number_4"] = card_number[12:16]

    # 금액: 전체 금액 -> 만원 단위 / 미리보기
    amount_str = defaults.get("amount", "") or "0"
    try:
        amount_int = int(amount_str)
    except ValueError:
        amount_int = 0
    defaults["amount_unit"] = str(amount_int // 10000) if amount_int else ""
    defaults["amount_preview"] = f"{amount_int:,} 원" if amount_int else "0 원"

    # 연락처: 저장된 뒷 8자리를 4-4 로 분할
    phone_suffix = (defaults.get("phone_number") or "").strip()
    phone_digits = "".join(ch for ch in phone_suffix if ch.isdigit())
    phone_digits = phone_digits[-8:] if len(phone_digits) >= 8 else phone_digits.rjust(8, "0")
    defaults["phone1"] = phone_digits[0:4] if len(phone_digits) >= 4 else ""
    defaults["phone2"] = phone_digits[4:8] if len(phone_digits) >= 8 else ""

    if request.method == "POST":
        form = request.form
        try:
            data = {h: form.get(h, "").strip() for h in HEADERS}
            if not data["product_name"]:
                data["product_name"] = "잡화"
            with open(ORDER_JSON_PATH, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            # 결과 상태 JSON 을 "진행중" 으로 초기화
            with open(RESULT_JSON_PATH, "w", encoding="utf-8") as f:
                json.dump(
                    {"status": "진행중", "message": "자동 결제를 대기 중입니다."},
                    f,
                    ensure_ascii=False,
                    indent=2,
                )
            # 주문 저장이 성공하면 auto_kvan.py 를 백그라운드에서 실행
            trigger_auto_kvan_async(session_id=None)
        except Exception as e:  # noqa: BLE001
            flash(f"데이터 저장 중 오류가 발생했습니다: {e}", "error")
        else:
            flash("주문 데이터가 성공적으로 저장되었습니다. 자동 결제를 진행합니다.", "success")
        return redirect(url_for("payment"))

    return render_template_string(
        FORM_TEMPLATE,
        defaults=defaults,
        last_result=last_result,
        form_action=url_for("payment"),
    )


@app.route("/pay/<session_id>", methods=["GET", "POST"])
def pay(session_id: str):
    """관리자가 생성한 단일 결제 링크용 폼 (금액/할부를 고정해서 노출)."""
    defaults = {
        "login_id": "m3313",
        "login_password": "k2255",
        "login_pin": "2424",
        "card_type": "personal",
    }

    # 세션별 주문/결과 파일 경로
    SESSION_ORDER_DIR.mkdir(parents=True, exist_ok=True)
    SESSION_RESULT_DIR.mkdir(parents=True, exist_ok=True)
    order_path = SESSION_ORDER_DIR / f"{session_id}.json"
    result_path = SESSION_RESULT_DIR / f"{session_id}.json"

    # 관리자 상태에서 현재 세션 정보 읽기 (금액/할부 고정용)
    fixed_amount = False
    if Path(ADMIN_STATE_PATH).exists():
        try:
            with open(ADMIN_STATE_PATH, "r", encoding="utf-8") as f:
                admin_state = json.load(f)
            sessions = admin_state.get("sessions") or []
            for s in sessions:
                if str(s.get("id")) == str(session_id):
                    amount_str = str(s.get("amount", "") or "")
                    if amount_str:
                        defaults["amount"] = amount_str
                        fixed_amount = True
                    installment = str(s.get("installment", "") or "")
                    if installment:
                        defaults["installment_months"] = installment
                    break
        except Exception:
            pass

    # 세션별 마지막 결과 읽기
    last_result: dict | None = None
    if result_path.exists():
        try:
            with open(result_path, "r", encoding="utf-8") as f:
                payload = json.load(f)
            status = str(payload.get("status", "unknown"))
            message = str(payload.get("message", "") or "")
            last_result = {"status": status, "message": message}
        except Exception:
            last_result = None

    # 기본 파생 값들 구성 (카드번호 분리, 금액 unit, 전화번호 분리)
    card_number = defaults.get("card_number", "")
    defaults["card_number_1"] = card_number[0:4]
    defaults["card_number_2"] = card_number[4:8]
    defaults["card_number_3"] = card_number[8:12]
    defaults["card_number_4"] = card_number[12:16]

    amount_str = defaults.get("amount", "") or "0"
    try:
        amount_int = int(amount_str)
    except ValueError:
        amount_int = 0
    defaults["amount_unit"] = str(amount_int // 10000) if amount_int else ""
    defaults["amount_preview"] = f"{amount_int:,} 원" if amount_int else "0 원"

    phone_suffix = (defaults.get("phone_number") or "").strip()
    phone_digits = "".join(ch for ch in phone_suffix if ch.isdigit())
    phone_digits = phone_digits[-8:] if len(phone_digits) >= 8 else phone_digits.rjust(8, "0")
    defaults["phone1"] = phone_digits[0:4] if len(phone_digits) >= 4 else ""
    defaults["phone2"] = phone_digits[4:8] if len(phone_digits) >= 8 else ""

    if request.method == "POST":
        form = request.form
        try:
            data = {h: form.get(h, "").strip() for h in HEADERS}
            if not data["product_name"]:
                data["product_name"] = "잡화"
            with open(order_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            # 결과 상태 JSON 을 "진행중" 으로 초기화
            with open(result_path, "w", encoding="utf-8") as f:
                json.dump(
                    {"status": "진행중", "message": "자동 결제를 대기 중입니다."},
                    f,
                    ensure_ascii=False,
                    indent=2,
                )
            # 세션 전용 결제의 경우에도 auto_kvan.py 를 백그라운드에서 실행 (세션 ID 전달)
            trigger_auto_kvan_async(session_id=session_id)
        except Exception as e:  # noqa: BLE001
            flash(f"데이터 저장 중 오류가 발생했습니다: {e}", "error")
        else:
            flash("주문 데이터가 성공적으로 저장되었습니다. 자동 결제를 진행합니다.", "success")
        return redirect(url_for("pay", session_id=session_id))

    return render_template_string(
        FORM_TEMPLATE,
        defaults=defaults,
        last_result=last_result,
        fixed_amount=fixed_amount,
        session_id=session_id,
        form_action=url_for("pay", session_id=session_id),
    )


def _is_same_origin_referer() -> bool:
    """Referer가 우리 사이트에서 온 경우만 True (외부 봇/직접 접근 차단)."""
    ref = (request.headers.get("Referer") or "").strip()
    if not ref:
        return True  # Referer 없으면 허용 (일부 브라우저/환경에서 생략)
    try:
        from urllib.parse import urlparse
        ref_host = urlparse(ref).netloc.split(":")[0].lower()
        req_host = (request.host or "").split(":")[0].lower()
        if not req_host:
            return True
        return ref_host == req_host or ref_host.endswith("." + req_host) or req_host.endswith("." + ref_host)
    except Exception:
        return True


# /last-result 호출 횟수 제한 (IP당 분당 60회 = 2초 폴링 여유)
_last_result_requests: dict[str, list[float]] = {}
_LAST_RESULT_LIMIT = 60
_LAST_RESULT_WINDOW = 60.0  # 초


@app.route("/last-result", methods=["GET"])
def last_result_api():
    """자동 결제 결과를 JSON 으로 반환 (폼에서 폴링용). 우리 사이트에서 온 요청만 허용."""
    same_origin = _is_same_origin_referer()
    if not same_origin:
        return "Forbidden", 403
    # Referer 없이 직접 반복 호출하는 경우만 IP당 분당 60회 제한 (봇/스캔 완화)
    ref = (request.headers.get("Referer") or "").strip()
    if not ref and _LAST_RESULT_LIMIT > 0:
        now = time.time()
        client_ip = request.remote_addr or ""
        if request.headers.get("X-Forwarded-For"):
            client_ip = (request.headers.get("X-Forwarded-For") or "").split(",")[0].strip()
        if client_ip:
            if client_ip not in _last_result_requests:
                _last_result_requests[client_ip] = []
            times = _last_result_requests[client_ip]
            times[:] = [t for t in times if now - t < _LAST_RESULT_WINDOW]
            if len(times) >= _LAST_RESULT_LIMIT:
                return "Too Many Requests", 429
            times.append(now)
    payload = {"status": "unknown", "message": ""}
    session_id = request.args.get("session_id", "").strip()
    if session_id:
        path = SESSION_RESULT_DIR / f"{session_id}.json"
    else:
        path = Path(RESULT_JSON_PATH)

    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                payload["status"] = str(data.get("status", "unknown"))
                payload["message"] = str(data.get("message", "") or "")
        except Exception:
            pass
    return jsonify(payload)


@app.route("/health", methods=["GET"])
def health():
    """간단 헬스 체크 엔드포인트."""
    return jsonify({"status": "ok"}), 200


def _path_diagnostic(path_value: str | Path) -> dict:
    """경로 존재/크기/수정시각을 진단용 dict 로 반환."""
    p = Path(path_value)
    info: dict = {
        "path": str(p),
        "exists": p.exists(),
        "is_dir": p.is_dir(),
        "size": "-",
        "mtime": "-",
        "error": "",
    }
    try:
        if p.exists() and p.is_file():
            st = p.stat()
            info["size"] = st.st_size
            info["mtime"] = datetime.utcfromtimestamp(st.st_mtime).isoformat()
    except Exception as e:  # noqa: BLE001
        info["error"] = str(e)
    return info


def _run_path_self_heal(auto_env_data_dir: Path) -> list[str]:
    """디렉토리/파일 자동 복구와 쓰기 권한 점검을 수행한다."""
    report: list[str] = []
    target_dirs = [
        DATA_DIR,
        SESSION_ORDER_DIR,
        SESSION_RESULT_DIR,
        auto_env_data_dir,
        auto_env_data_dir / "sessions",
        auto_env_data_dir / "sessions" / "orders",
        auto_env_data_dir / "sessions" / "results",
    ]
    seen_dirs: set[str] = set()
    for d in target_dirs:
        d_key = str(d)
        if d_key in seen_dirs:
            continue
        seen_dirs.add(d_key)
        try:
            d.mkdir(parents=True, exist_ok=True)
            report.append(f"[OK] dir ensured: {d}")
        except Exception as e:  # noqa: BLE001
            report.append(f"[ERR] dir create failed: {d} ({e})")

    # 필수 파일이 없으면 기본 구조로 생성
    admin_state_candidates = [
        Path(ADMIN_STATE_PATH),
        auto_env_data_dir / "admin_state.json",
    ]
    for st_path in admin_state_candidates:
        try:
            if not st_path.exists():
                with open(st_path, "w", encoding="utf-8") as f:
                    json.dump({"sessions": [], "history": []}, f, ensure_ascii=False, indent=2)
                report.append(f"[OK] created admin_state: {st_path}")
        except Exception as e:  # noqa: BLE001
            report.append(f"[ERR] admin_state create failed: {st_path} ({e})")

    seen_logs: set[str] = set()
    for p in [ADMIN_LOG_PATH, auto_env_data_dir / "hq_logs.log"]:
        p_key = str(p)
        if p_key in seen_logs:
            continue
        seen_logs.add(p_key)
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
            with open(p, "a", encoding="utf-8"):
                pass
            report.append(f"[OK] log file ensured: {p}")
        except Exception as e:  # noqa: BLE001
            report.append(f"[ERR] log file ensure failed: {p} ({e})")

    # 쓰기 권한 점검 (임시 파일 생성/삭제)
    write_test_dirs = [
        SESSION_ORDER_DIR,
        SESSION_RESULT_DIR,
        auto_env_data_dir / "sessions" / "orders",
        auto_env_data_dir / "sessions" / "results",
    ]
    seen_write_dirs: set[str] = set()
    for d in write_test_dirs:
        d_key = str(d)
        if d_key in seen_write_dirs:
            continue
        seen_write_dirs.add(d_key)
        test_file = d / f".write_test_{int(time.time() * 1000)}.tmp"
        try:
            d.mkdir(parents=True, exist_ok=True)
            with open(test_file, "w", encoding="utf-8") as f:
                f.write("ok")
            test_file.unlink(missing_ok=True)
            report.append(f"[OK] write test passed: {d}")
        except Exception as e:  # noqa: BLE001
            report.append(f"[ERR] write test failed: {d} ({e})")

    # DB 핵심 테이블 자동 생성/보정
    try:
        init_db()
        report.append("[OK] DB schema ensured via init_db()")
    except Exception as e:  # noqa: BLE001
        report.append(f"[ERR] DB schema ensure failed: {e}")

    return report


@app.route("/debug-paths", methods=["GET", "POST"])
def debug_paths():
    """초보 사용자용 경로/연동 상태 점검 페이지 (HQ 로그인 필요)."""
    if not session.get("hq_logged_in"):
        return redirect(url_for("hq_login"))

    env_sisa_data_dir = (os.environ.get("SISA_DATA_DIR") or "").strip()
    auto_base_dir = BASE_DIR / "wsisa"
    auto_env_data_dir = Path(env_sisa_data_dir) if env_sisa_data_dir else (auto_base_dir.parent / "data")
    repair_report: list[str] = []
    repair_message = ""

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()
        if action == "self_heal":
            repair_report = _run_path_self_heal(auto_env_data_dir)
            _append_hq_log(
                "WEB",
                "debug-paths 자동복구 실행: "
                + "; ".join(repair_report[:12])
                + ("; ... (truncated)" if len(repair_report) > 12 else ""),
            )
            err_count = sum(1 for r in repair_report if r.startswith("[ERR]"))
            if err_count:
                repair_message = f"자동복구 완료: 오류 {err_count}건 (아래 보고서 확인)"
            else:
                repair_message = "자동복구 완료: 오류 없이 정상 처리되었습니다."

    path_checks: list[dict] = []
    for p in [
        BASE_DIR,
        DATA_DIR,
        auto_base_dir,
        Path(ORDER_JSON_PATH),
        Path(RESULT_JSON_PATH),
        Path(ADMIN_STATE_PATH),
        ADMIN_LOG_PATH,
        SESSION_ORDER_DIR,
        SESSION_RESULT_DIR,
        auto_env_data_dir / "admin_state.json",
        auto_env_data_dir / "sessions" / "orders",
    ]:
        path_checks.append(_path_diagnostic(p))

    # 최근 세션 기준으로 주문 JSON 존재 여부를 즉시 확인
    recent_sessions: list[dict] = []
    try:
        state_path = Path(ADMIN_STATE_PATH)
        if state_path.exists():
            with open(state_path, "r", encoding="utf-8") as f:
                saved = json.load(f)
            sessions_all = saved.get("sessions") or []
            for s in list(reversed(sessions_all))[:10]:
                sid = str(s.get("id") or "").strip()
                if not sid:
                    continue
                candidates = [
                    DATA_DIR / "sessions" / "orders" / f"{sid}.json",
                    BASE_DIR / "wsisa" / "data" / "sessions" / "orders" / f"{sid}.json",
                ]
                exists_any = any(c.exists() for c in candidates)
                recent_sessions.append(
                    {
                        "id": sid,
                        "amount": str(s.get("amount") or ""),
                        "status": str(s.get("status") or ""),
                        "agency_id": str(s.get("agency_id") or ""),
                        "kvan_link": str(s.get("kvan_link") or ""),
                        "order_exists": exists_any,
                        "order_candidates": [str(c) for c in candidates],
                    }
                )
    except Exception as e:  # noqa: BLE001
        recent_sessions.append({"id": "-", "amount": "-", "status": "-", "agency_id": "-", "kvan_link": "-", "order_exists": False, "order_candidates": [f"admin_state 읽기 실패: {e}"]})

    def _row_pick(row: dict, key: str, default: str = "") -> str:
        if key in row:
            return str(row.get(key) or default)
        key_lower = key.lower()
        for k, v in row.items():
            if str(k).lower() == key_lower:
                return str(v or default)
        return default

    db_info: dict = {"ok": False, "error": "", "counts": {}, "schemas": {}}
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT table_name, column_name
                FROM information_schema.columns
                WHERE table_schema = %s
                  AND table_name IN ('transactions','kvan_links','kvan_transactions','kvan_dashboard')
                ORDER BY table_name, ordinal_position
                """,
                (DB_NAME,),
            )
            rows = cur.fetchall()
            schemas: dict[str, list[str]] = {}
            for r in rows:
                t = _row_pick(r, "table_name")
                c = _row_pick(r, "column_name")
                if not t or not c:
                    continue
                if t not in schemas:
                    schemas[t] = []
                schemas[t].append(c)
            db_info["schemas"] = schemas

            for tname in ["transactions", "kvan_links", "kvan_transactions", "kvan_dashboard"]:
                try:
                    cur.execute(f"SELECT COUNT(*) AS cnt FROM {tname}")
                    one = cur.fetchone() or {}
                    db_info["counts"][tname] = int(one.get("cnt") or 0)
                except Exception as e_cnt:  # noqa: BLE001
                    db_info["counts"][tname] = f"조회 실패: {e_cnt}"
        conn.close()
        db_info["ok"] = True
    except Exception as e:  # noqa: BLE001
        db_info["error"] = str(e)

    # K-VAN 큐/락 상태 (실시간 표시)
    kvan_queue_len = 0
    kvan_queue_ids: list[str] = []
    kvan_lock_pid: str | None = None
    kvan_lock_running = False
    try:
        if KVAN_QUEUE_PATH.exists():
            q = json.loads(KVAN_QUEUE_PATH.read_text(encoding="utf-8"))
            if isinstance(q, list):
                kvan_queue_len = len(q)
                kvan_queue_ids = [str(x) for x in q[:20]]
    except Exception:
        pass
    try:
        if KVAN_LOCK_PATH.exists():
            kvan_lock_pid = KVAN_LOCK_PATH.read_text(encoding="utf-8").strip()
            kvan_lock_running = _kvan_is_running()
    except Exception:
        pass

    template = """
    <!DOCTYPE html>
    <html lang="ko">
    <head>
      <meta charset="UTF-8" />
      <title>SISA Debug Paths</title>
      <meta name="viewport" content="width=device-width, initial-scale=1.0" />
      <script src="https://cdn.tailwindcss.com"></script>
    </head>
    <body class="bg-slate-950 text-slate-100 min-h-screen p-4">
      <div class="max-w-6xl mx-auto space-y-4">
        <div class="flex items-center justify-between">
          <h1 class="text-lg font-semibold">경로/연동 진단</h1>
          <div class="flex items-center gap-2">
            <a href="{{ url_for('hq_admin') }}" class="px-3 py-1.5 rounded bg-slate-700 hover:bg-slate-600 text-sm">HQ로 돌아가기</a>
            <a href="{{ url_for('debug_paths') }}" class="px-3 py-1.5 rounded bg-indigo-600 hover:bg-indigo-500 text-sm">다시 진단</a>
            <form method="post" action="{{ url_for('debug_paths') }}">
              <input type="hidden" name="action" value="self_heal" />
              <button type="submit" class="px-3 py-1.5 rounded bg-emerald-600 hover:bg-emerald-500 text-sm font-semibold">
                원클릭 자동복구
              </button>
            </form>
          </div>
        </div>
        {% if repair_message %}
        <section class="rounded border border-emerald-500/40 p-3 bg-emerald-900/20 text-emerald-200 text-sm">
          {{ repair_message }}
        </section>
        {% endif %}
        {% if repair_report %}
        <section class="rounded border border-slate-700 p-3 bg-slate-900/60">
          <h2 class="text-sm font-semibold mb-2">자동복구 보고서</h2>
          <div class="text-xs font-mono space-y-1">
            {% for line in repair_report %}
            <div>{{ line }}</div>
            {% endfor %}
          </div>
        </section>
        {% endif %}

        <section class="rounded border border-slate-700 p-3 bg-slate-900/60">
          <h2 class="text-sm font-semibold mb-2">환경 변수</h2>
          <div class="text-xs font-mono break-all">SISA_DATA_DIR={{ env_sisa_data_dir or '(비어있음)' }}</div>
          <div class="text-xs font-mono break-all mt-1">WEB DATA_DIR={{ web_data_dir }}</div>
          <div class="text-xs font-mono break-all mt-1">AUTO 예상 DATA_DIR={{ auto_data_dir }}</div>
        </section>

        <section class="rounded border border-slate-700 p-3 bg-slate-900/60">
          <h2 class="text-sm font-semibold mb-2">K-VAN 큐·락 상태</h2>
          <div class="text-xs space-y-1">
            <div>큐 대기: <span class="font-mono font-semibold">{{ kvan_queue_len }}</span>건 {% if kvan_queue_ids %}(세션 ID: {{ kvan_queue_ids|join(', ') }}{% if kvan_queue_len > 20 %} …{% endif %}){% endif %}</div>
            <div>락: {% if kvan_lock_pid %}<span class="font-mono">{{ kvan_lock_pid }}</span> (PID) — {% if kvan_lock_running %}<span class="text-amber-300">실행 중</span>{% else %}<span class="text-slate-400">프로세스 종료됨</span>{% endif %}{% else %}<span class="text-slate-400">대기 중</span>{% endif %}</div>
          </div>
        </section>

        <section class="rounded border border-slate-700 p-3 bg-slate-900/60">
          <h2 class="text-sm font-semibold mb-2">주요 경로 체크</h2>
          <div class="overflow-x-auto">
            <table class="min-w-full text-xs">
              <thead class="text-slate-300">
                <tr>
                  <th class="text-left px-2 py-1">Path</th>
                  <th class="text-left px-2 py-1">Exists</th>
                  <th class="text-left px-2 py-1">Type</th>
                  <th class="text-left px-2 py-1">Size</th>
                  <th class="text-left px-2 py-1">MTime(UTC)</th>
                  <th class="text-left px-2 py-1">Error</th>
                </tr>
              </thead>
              <tbody>
              {% for p in path_checks %}
                <tr class="border-t border-slate-800">
                  <td class="px-2 py-1 font-mono break-all">{{ p.path }}</td>
                  <td class="px-2 py-1">{{ 'Y' if p.exists else 'N' }}</td>
                  <td class="px-2 py-1">{{ 'dir' if p.is_dir else 'file' }}</td>
                  <td class="px-2 py-1">{{ p.size }}</td>
                  <td class="px-2 py-1">{{ p.mtime }}</td>
                  <td class="px-2 py-1 text-rose-300">{{ p.error }}</td>
                </tr>
              {% endfor %}
              </tbody>
            </table>
          </div>
        </section>

        <section class="rounded border border-slate-700 p-3 bg-slate-900/60">
          <h2 class="text-sm font-semibold mb-2">최근 세션 연동 체크 (admin_state 기준)</h2>
          <div class="overflow-x-auto">
            <table class="min-w-full text-xs">
              <thead class="text-slate-300">
                <tr>
                  <th class="text-left px-2 py-1">Session</th>
                  <th class="text-left px-2 py-1">Amount</th>
                  <th class="text-left px-2 py-1">Status</th>
                  <th class="text-left px-2 py-1">Agency</th>
                  <th class="text-left px-2 py-1">Order JSON exists</th>
                  <th class="text-left px-2 py-1">K-VAN Link</th>
                </tr>
              </thead>
              <tbody>
              {% for s in recent_sessions %}
                <tr class="border-t border-slate-800">
                  <td class="px-2 py-1 font-mono">{{ s.id }}</td>
                  <td class="px-2 py-1">{{ s.amount }}</td>
                  <td class="px-2 py-1">{{ s.status }}</td>
                  <td class="px-2 py-1">{{ s.agency_id or '-' }}</td>
                  <td class="px-2 py-1">{{ 'Y' if s.order_exists else 'N' }}</td>
                  <td class="px-2 py-1 font-mono break-all">{{ s.kvan_link or '-' }}</td>
                </tr>
                <tr class="border-b border-slate-800">
                  <td colspan="6" class="px-2 py-1 text-[11px] text-slate-400">
                    order candidates: {{ s.order_candidates|join(' | ') }}
                  </td>
                </tr>
              {% endfor %}
              </tbody>
            </table>
          </div>
        </section>

        <section class="rounded border border-slate-700 p-3 bg-slate-900/60">
          <h2 class="text-sm font-semibold mb-2">DB 구조/건수 체크</h2>
          {% if db_info.ok %}
          <div class="text-xs text-emerald-300 mb-2">DB 연결 성공</div>
          {% else %}
          <div class="text-xs text-rose-300 mb-2">DB 연결 실패: {{ db_info.error }}</div>
          {% endif %}
          <div class="grid grid-cols-1 md:grid-cols-2 gap-2 text-xs">
            <div class="rounded border border-slate-800 p-2">
              <div class="font-semibold mb-1">테이블 건수</div>
              <pre class="whitespace-pre-wrap">{{ db_info.counts }}</pre>
            </div>
            <div class="rounded border border-slate-800 p-2">
              <div class="font-semibold mb-1">컬럼 구조</div>
              <pre class="whitespace-pre-wrap">{{ db_info.schemas }}</pre>
            </div>
          </div>
        </section>
      </div>
    </body>
    </html>
    """
    return render_template_string(
        template,
        env_sisa_data_dir=env_sisa_data_dir,
        web_data_dir=str(DATA_DIR),
        auto_data_dir=str(auto_env_data_dir),
        repair_message=repair_message,
        repair_report=repair_report,
        path_checks=path_checks,
        recent_sessions=recent_sessions,
        db_info=db_info,
        kvan_queue_len=kvan_queue_len,
        kvan_queue_ids=kvan_queue_ids,
        kvan_lock_pid=kvan_lock_pid,
        kvan_lock_running=kvan_lock_running,
    )


def _load_hq_state() -> dict:
    """본사 어드민 상태를 MySQL 에서 로드."""
    state = {"applications": [], "agencies": [], "transactions": []}
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM applications ORDER BY created_at DESC")
            state["applications"] = cur.fetchall()
            cur.execute("SELECT * FROM agencies ORDER BY created_at DESC")
            state["agencies"] = cur.fetchall()
            cur.execute("SELECT * FROM transactions ORDER BY created_at DESC")
            state["transactions"] = cur.fetchall()
        conn.close()
    except Exception as e:  # noqa: BLE001
        print(f"[WARN] _load_hq_state 실패: {e}")
    return state


def _is_recent_duplicate_amount(amount_str: str, window_minutes: int = 5) -> bool:
    """
    admin_state.json 기준으로, 최근 window_minutes 분 이내에
    동일 금액으로 생성된 세션/히스토리가 있는지 확인한다.

    - 금액이 비어 있거나 0 이하이면 중복 검사 대상에서 제외
    - created_at / finished_at 중 존재하는 타임스탬프를 사용
    """
    amount_str = (amount_str or "").replace(",", "").strip()
    if not amount_str:
        return False
    try:
        amt = int(amount_str)
    except ValueError:
        return False
    if amt <= 0:
        return False

    try:
        cutoff = datetime.utcnow() - timedelta(minutes=window_minutes)
        state_path = Path(ADMIN_STATE_PATH)
        if not state_path.exists():
            return False
        with open(state_path, "r", encoding="utf-8") as f:
            saved = json.load(f)
        sessions = saved.get("sessions") or []
        history = saved.get("history") or []
        candidates = list(sessions) + list(history)

        for s in candidates:
            s_amount = str(s.get("amount") or "").replace(",", "").strip()
            try:
                s_amt = int(s_amount)
            except ValueError:
                continue
            if s_amt != amt:
                continue

            ts = s.get("created_at") or s.get("finished_at")
            if not ts:
                continue
            try:
                dt = datetime.fromisoformat(ts)
            except Exception:
                continue
            if dt >= cutoff:
                return True
        return False
    except Exception:
        # 중복 검사에서 문제가 생겨도 결제 자체를 막지는 않는다.
        return False


def _save_hq_state(state: dict) -> None:
    """기존 JSON 기반 코드와의 호환을 위해 전체 상태를 DB에 반영."""
    try:
        conn = get_db()
        with conn.cursor() as cur:
            # applications 동기화
            cur.execute("DELETE FROM applications")
            for a in state.get("applications") or []:
                cur.execute(
                    """
                    INSERT INTO applications
                    (id, company_name, domain, phone, bank_name, account_number,
                     email_or_sheet, login_id, login_password, fee_percent, created_at, status)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                    (
                        a.get("id"),
                        a.get("company_name"),
                        a.get("domain"),
                        a.get("phone"),
                        a.get("bank_name"),
                        a.get("account_number"),
                        a.get("email_or_sheet"),
                        a.get("login_id"),
                        a.get("login_password"),
                        a.get("fee_percent", 10),
                        a.get("created_at"),
                        a.get("status"),
                    ),
                )
            # agencies 동기화
            cur.execute("DELETE FROM agencies")
            for ag in state.get("agencies") or []:
                cur.execute(
                    """
                    INSERT INTO agencies
                    (id, company_name, domain, phone, bank_name, account_number,
                     email_or_sheet, login_id, login_password, fee_percent,
                     kvan_mid, kvan_login_id, kvan_login_password, kvan_login_pin,
                     created_at, status)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                    (
                        ag.get("id"),
                        ag.get("company_name"),
                        ag.get("domain"),
                        ag.get("phone"),
                        ag.get("bank_name"),
                        ag.get("account_number"),
                        ag.get("email_or_sheet"),
                        ag.get("login_id"),
                        ag.get("login_password"),
                        ag.get("fee_percent", 10),
                        ag.get("kvan_mid"),
                        ag.get("kvan_login_id"),
                        ag.get("kvan_login_password"),
                        ag.get("kvan_login_pin"),
                        ag.get("created_at"),
                        ag.get("status"),
                    ),
                )
            # transactions 동기화
            cur.execute("DELETE FROM transactions")
            for t in state.get("transactions") or []:
                cur.execute(
                    """
                    INSERT INTO transactions
                    (id, created_at, agency_id, amount, customer_name, phone_number,
                     card_type, resident_front, status, message, settlement_status, settled_at,
                     kvan_mid, kvan_approval_no, kvan_tx_type, kvan_registered_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                    (
                        t.get("id"),
                        t.get("created_at"),
                        t.get("agency_id"),
                        t.get("amount"),
                        t.get("customer_name"),
                        t.get("phone_number"),
                        t.get("card_type"),
                        t.get("resident_front"),
                        t.get("status"),
                        t.get("message"),
                        t.get("settlement_status"),
                        t.get("settled_at"),
                        t.get("kvan_mid"),
                        t.get("kvan_approval_no"),
                        t.get("kvan_tx_type"),
                        t.get("kvan_registered_at"),
                    ),
                )
        conn.commit()
        conn.close()
    except Exception as e:  # noqa: BLE001
        print(f"[WARN] _save_hq_state 실패: {e}")


@app.route("/agency-apply", methods=["POST"])
def agency_apply():
    """대행사 등록 신청 폼 수신 엔드포인트 (agency-register.html 에서 POST)."""
    form = request.form
    company_name = form.get("업체명", "").strip()
    domain = form.get("도메인(영문)", "").strip()
    phone = form.get("전화번호", "").strip()
    bank_name = form.get("은행명", "").strip()
    account_number = form.get("계좌번호", "").strip()
    email_or_sheet = form.get("이메일_또는_구글시트", "").strip()
    agency_login_id = form.get("대행사아이디", "").strip()
    agency_login_pw = form.get("대행사비밀번호", "").strip()

    app_id = datetime.utcnow().strftime("AG%Y%m%d%H%M%S%f")

    # MySQL applications 테이블에 직접 INSERT (DB 기준 진짜 저장)
    db_ok = False
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO applications
                (id, company_name, domain, phone, bank_name, account_number,
                 email_or_sheet, login_id, login_password, fee_percent, created_at, status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    app_id,
                    company_name,
                    domain,
                    phone,
                    bank_name,
                    account_number,
                    email_or_sheet,
                    agency_login_id,
                    agency_login_pw,
                    10,
                    datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
                    "pending",
                ),
            )
        conn.commit()
        conn.close()
        db_ok = True
    except Exception as e:  # noqa: BLE001
        print(f"[ERROR] agency_apply DB insert 실패: {e}")

    if db_ok:
        status_message = "대행사 신청 정보가 데이터베이스에 정상 저장되었습니다. SISA 본사에서 검토 후 개별 연락을 드립니다."
    else:
        status_message = "신청 접수 과정에서 오류가 발생했을 수 있습니다. 잠시 후 다시 시도하시거나 본사에 문의해 주세요."

    # 간단한 접수 완료 페이지 반환 (SISA 스타일)
    return f"""
<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="UTF-8" />
  <title>대행사 등록 신청 완료</title>
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap" rel="stylesheet">
  <script src="https://cdn.tailwindcss.com"></script>
</head>
<body class="bg-[#2f4b9f] text-white font-[Inter] flex items-center justify-center min-h-screen">
  <div class="bg-white/10 border border-white/20 rounded-2xl px-8 py-10 max-w-md w-full text-center shadow-2xl">
    <div class="w-12 h-12 rounded-full bg-emerald-400/20 border border-emerald-300/50 flex items-center justify-center mx-auto mb-4">
      <span class="text-2xl text-emerald-300">✓</span>
    </div>
    <h1 class="text-2xl font-bold mb-2">대행사 등록 신청이 접수되었습니다.</h1>
    <p class="text-sm text-white/70 mb-4 leading-relaxed">{status_message}</p>
    <p class="text-[11px] text-white/60 mb-6">
      아래 버튼을 누르시면 SISA 메인 페이지로 이동합니다.
    </p>
    <button onclick="window.location.href='/'"
            class="mt-2 inline-flex items-center justify-center px-5 py-2 rounded-full bg-white text-[#2f4b9f] text-sm font-semibold hover:bg-[#e6edf7] transition">
      메인 페이지로 이동
    </button>
  </div>
</body>
</html>
"""


@app.route("/terms", methods=["GET"])
def terms():
    """이용약관 HTML 파일을 iframe/직접 방문 둘 다에서 표시."""
    if TERMS_FILE.exists():
        return send_file(TERMS_FILE)
    return "<!doctype html><html><body><p>이용약관 파일을 불러올 수 없습니다.</p></body></html>"


@app.route("/terms-consent-pdf", methods=["POST"])
def terms_consent_pdf():
    """이용약관 동의 내용을 PDF로 생성하여 다운로드."""
    name = (request.form.get("customer_name") or "").strip()
    phone = (request.form.get("phone_number") or "").strip()

    now = datetime.now()
    date_str = now.strftime("%Y%m%d")

    # 전화번호에서 숫자만 추출 후 뒤 4자리
    digits = "".join(ch for ch in phone if ch.isdigit())
    last4 = digits[-4:] if digits else "0000"

    safe_name = name or "anonymous"
    filename = f"{safe_name}_{last4}_{date_str}.pdf"

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)

    text = c.beginText(40, 800)
    text.setFont("Helvetica-Bold", 14)
    text.textLine("SISA 플랫폼 서비스 이용약관 동의서")
    text.textLine("")
    text.setFont("Helvetica", 11)
    text.textLine(f"이름: {name}")
    text.textLine(f"전화번호: {phone}")
    text.textLine(f"동의 일시: {now.strftime('%Y-%m-%d %H:%M:%S')}")
    text.textLine("")
    text.textLine("위 고객은 SISA 플랫폼 서비스 이용약관 및 결제 전 필수 동의 항목에 모두 동의하였습니다.")

    c.drawText(text)
    c.showPage()
    c.save()
    buf.seek(0)

    return send_file(buf, as_attachment=True, download_name=filename, mimetype="application/pdf")


@app.route("/agency-register.html", methods=["GET"])
def agency_register_page():
    """대행사 등록 신청 정적 페이지 제공."""
    path = BASE_DIR / "agency-register.html"
    if path.exists():
        return send_file(path)
    return "<p>agency-register.html 파일을 찾을 수 없습니다.</p>", 404

@app.route("/admin", methods=["GET", "POST"])
def admin():
    """본사 공용 K-VAN 세션 어드민 (HQ용). 최대 5개 세션 관리."""
    base_url = request.url_root.rstrip("/")

    # 기존 상태 로드 (sessions 리스트 기반)
    sessions: list[dict] = []
    history: list[dict] = []
    message = ""
    crawler_refresh_since = ""
    if Path(ADMIN_STATE_PATH).exists():
        try:
            with open(ADMIN_STATE_PATH, "r", encoding="utf-8") as f:
                saved = json.load(f)
            if isinstance(saved, dict):
                raw_sessions = saved.get("sessions") or []
                raw_history = saved.get("history") or []
                # /admin 페이지에서는 HQ(agency_id 가 비어 있는) 세션/히스토리만 본다.
                sessions = [
                    s for s in raw_sessions if not str(s.get("agency_id") or "").strip()
                ]
                history = [
                    h for h in raw_history if not str(h.get("agency_id") or "").strip()
                ]
                # 이전 단일 세션 포맷에서 마이그레이션
                if saved.get("current_session_id") and not sessions:
                    sessions = [
                        {
                            "id": str(saved.get("current_session_id")),
                            "amount": str(saved.get("amount", "") or ""),
                            "installment": str(saved.get("installment", "") or "일시불"),
                            "status": "결제중",
                            "created_at": saved.get("created_at")
                            or datetime.utcnow().isoformat(),
                        }
                    ]
        except Exception:
            sessions = []

    # 진행 중(결제중만) vs 완료/종료(history + 세션 중 만료·결제완료·실패 등) 구분 표시
    _status = lambda s: (str(s.get("status") or "결제중").strip())
    # K-VAN에서 삭제(deleted_in_kvan=True)되고 거래 내역도 없는 세션은 목록에서 숨긴다.
    # 거래 내역이 있는 만료 세션(has_transaction=True)은 정산 확인을 위해 계속 표시.
    _is_visible = lambda s: not (bool(s.get("deleted_in_kvan")) and not bool(s.get("has_transaction")))
    active_sessions = [s for s in sessions if _status(s) == "결제중" and _is_visible(s)]
    completed_sessions = [h for h in history if _is_visible(h)] + [s for s in sessions if _status(s) != "결제중" and _is_visible(s)]

    # 결제중인데 아직 K-VAN 링크가 없는 세션이 있는지 여부 (자동 새로고침/팝업 트리거 용)
    has_pending_link = any(
        (s.get("status", "결제중") == "결제중") and not s.get("kvan_link")
        for s in sessions
    )

    # 결제중 세션인데 주문 JSON이 없으면 자동 재생성 + 매크로 재트리거
    for s in sessions:
        if s.get("status", "결제중") != "결제중":
            continue
        if s.get("kvan_link"):
            continue
        sid = str(s.get("id") or "")
        if not sid:
            continue
        order_file = SESSION_ORDER_DIR / f"{sid}.json"
        if not order_file.exists():
            try:
                _save_session_order_json(sid, str(s.get("amount") or ""), str(s.get("installment") or "일시불"))
                _append_hq_log("WEB", f"[AUTO-HEAL] 누락된 주문 JSON 재생성 session_id={sid}")
                trigger_auto_kvan_async(session_id=sid)
                _append_hq_log("WEB", f"[AUTO-HEAL] auto_kvan 재트리거 session_id={sid}")
            except Exception as _e:
                _append_hq_log("WEB", f"[AUTO-HEAL][WARN] 주문 JSON 재생성 실패 session_id={sid}: {_e}")

    # 크롤러가 생성한 DB 기반 정보 (참고용 요약)
    recent_links: list[dict] = []
    recent_tx: list[dict] = []
    try:
        conn = get_db()
        with conn.cursor() as cur:
            try:
                cur.execute(
                    """
                    SELECT captured_at, title, amount, ttl_label, status, kvan_link
                    FROM kvan_links
                    ORDER BY captured_at DESC
                    LIMIT 10
                    """
                )
                recent_links = cur.fetchall()
            except Exception:
                recent_links = []
            try:
                cur.execute(
                    """
                    SELECT created_at, amount, customer_name, status, settlement_status
                    FROM transactions
                    ORDER BY created_at DESC
                    LIMIT 10
                    """
                )
                recent_tx = cur.fetchall()
            except Exception:
                recent_tx = []
        conn.close()
    except Exception as e:  # noqa: BLE001
        print(f"[WARN] /admin DB 요약 조회 실패: {e}")

    # /admin 페이지에서는 이제 K-VAN 연동용 거래/링크 리스트를 표시하지 않는다.

    if request.method == "POST":
        action = request.form.get("action", "create").strip()

        if action == "create":
            amount = request.form.get("admin_amount", "").strip()
            installment = request.form.get("admin_installment", "일시불").strip()

            # 결제금액이 비어 있으면 세션/링크를 만들지 않고 안내
            if not amount:
                message = "결제 금액을 입력해 주세요. 금액 없이 결제요청 링크를 생성할 수 없습니다."
            else:
                # 현재 진행 중(결제중) 세션 수 확인
                active_count = sum(
                    1 for s in sessions if s.get("status", "결제중") == "결제중"
                )
                if active_count >= 5:
                    message = "동시에 진행할 수 있는 세션은 최대 5개입니다."
                else:
                    # 새 세션 ID 생성
                    session_id = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")[-12:]
                    session = {
                        "id": session_id,
                        "amount": amount,  # 비어 있으면 '고정 안 됨' 으로 동작
                        "installment": installment or "",
                        "status": "결제중",
                        "created_at": datetime.utcnow().isoformat(),
                        "agency_id": "",  # HQ에서 생성한 세션은 특정 대행사에 속하지 않음
                    }
                    sessions.append(session)
                    admin_state = {"sessions": sessions, "history": history}
                    try:
                        with open(ADMIN_STATE_PATH, "w", encoding="utf-8") as f:
                            json.dump(admin_state, f, ensure_ascii=False, indent=2)
                    except Exception as e:  # noqa: BLE001
                        message = f"상태 저장 중 오류가 발생했습니다: {e}"
                    else:
                        try:
                            order_json = _save_session_order_json(session_id, amount, installment)
                            _append_hq_log("WEB", f"세션 주문 JSON 저장 session_id={session_id}, path={order_json}")
                        except Exception as e_order:  # noqa: BLE001
                            _append_hq_log("WEB", f"[WARN] 세션 주문 JSON 저장 실패 session_id={session_id}: {e_order}")
                        _append_hq_log(
                            "WEB",
                            f"HQ 세션 생성 session_id={session_id}, amount={amount}, installment={installment or '일시불'}",
                        )
                        # HQ에서 링크를 생성한 시점에도 자동 결제 매크로를 준비
                        try:
                            trigger_auto_kvan_async(session_id=session_id)
                        except Exception as e:  # noqa: BLE001
                            print(f"HQ 세션 생성 시 auto_kvan 트리거 실패: {e}")
                        # 중복 생성 방지를 위해 PRG 패턴 적용: 성공 시에는 항상 리다이렉트
                        # new=1 쿼리스트링으로 "새 링크 생성" 플래그를 전달
                        return redirect(url_for("admin", new="1"))

        elif action == "retry_kvan":
            sid = request.form.get("session_id", "").strip()
            if sid:
                # 실패 상태 → 결제중으로 초기화 후 재시도
                state_path = Path(ADMIN_STATE_PATH)
                if state_path.exists():
                    try:
                        with open(state_path, "r", encoding="utf-8") as f:
                            saved = json.load(f)
                        for s in saved.get("sessions") or []:
                            if str(s.get("id")) == sid:
                                s["status"] = "결제중"
                                s.pop("error_reason", None)
                                s.pop("failed_at", None)
                                s.pop("kvan_link", None)
                                break
                        with open(state_path, "w", encoding="utf-8") as f:
                            json.dump(saved, f, ensure_ascii=False, indent=2)
                    except Exception as e:  # noqa: BLE001
                        _append_hq_log("WEB", f"[WARN] retry 상태 초기화 실패: {e}")
                # 주문 JSON 재생성 + 큐에 재등록
                try:
                    for s in sessions:
                        if str(s.get("id")) == sid:
                            _save_session_order_json(sid, str(s.get("amount") or ""), str(s.get("installment") or "일시불"))
                            break
                    trigger_auto_kvan_async(session_id=sid)
                    _append_hq_log("WEB", f"retry_kvan 재요청 session_id={sid}")
                    message = f"세션 {sid} 링크 생성을 다시 요청했습니다."
                except Exception as e:  # noqa: BLE001
                    message = f"재요청 중 오류: {e}"
            return redirect(url_for("admin"))
        elif action == "refresh_kvan":
            try:
                trigger_kvan_crawler_refresh()
                crawler_refresh_since = datetime.utcnow().isoformat()
                message = "K-VAN 크롤러 새로고침 신호를 보냈습니다. 잠시 후 최신 상태가 반영됩니다."
            except Exception as e:  # noqa: BLE001
                message = f"K-VAN 새로고침 중 오류: {e}"

        elif action == "close_session":
            sid = request.form.get("session_id", "").strip()
            memo = request.form.get("memo", "").strip()
            new_sessions: list[dict] = []
            for s in sessions:
                if str(s.get("id")) == sid:
                    entry = {
                        "id": sid,
                        "amount": str(s.get("amount", "") or ""),
                        "installment": str(s.get("installment", "") or "일시불"),
                        "status": "관리자종료",
                        "created_at": s.get("created_at") or datetime.utcnow().isoformat(),
                        "finished_at": datetime.utcnow().isoformat(),
                        "result_message": memo or "관리자가 세션을 종료했습니다.",
                        "customer_name": "",
                        "phone_number": "",
                        "product_name": "",
                        "settled": "정산전",
                        "agency_id": s.get("agency_id", ""),
                    }
                    history.append(entry)
                else:
                    new_sessions.append(s)
            sessions = new_sessions
            admin_state = {"sessions": sessions, "history": history}
            try:
                with open(ADMIN_STATE_PATH, "w", encoding="utf-8") as f:
                    json.dump(admin_state, f, ensure_ascii=False, indent=2)
            except Exception as e:  # noqa: BLE001
                message = f"세션 종료 중 오류가 발생했습니다: {e}"

        elif action == "toggle_settle":
            sid = request.form.get("session_id", "").strip()
            for h in history:
                if str(h.get("id")) == sid:
                    h["settled"] = "정산완료" if h.get("settled") != "정산완료" else "정산전"
                    break
            admin_state = {"sessions": sessions, "history": history}
            try:
                with open(ADMIN_STATE_PATH, "w", encoding="utf-8") as f:
                    json.dump(admin_state, f, ensure_ascii=False, indent=2)
            except Exception as e:  # noqa: BLE001
                message = f"정산 상태 변경 중 오류가 발생했습니다: {e}"

        elif action == "delete_history":
            sid = request.form.get("session_id", "").strip()
            history = [h for h in history if str(h.get("id")) != sid]
            admin_state = {"sessions": sessions, "history": history}
            try:
                with open(ADMIN_STATE_PATH, "w", encoding="utf-8") as f:
                    json.dump(admin_state, f, ensure_ascii=False, indent=2)
            except Exception as e:  # noqa: BLE001
                message = f"기록 삭제 중 오류가 발생했습니다: {e}"

    ADMIN_TEMPLATE = """
    <!DOCTYPE html>
    <html lang="ko">
    <head>
      <meta charset="UTF-8" />
      <title>SISA K-VAN 결제 어드민</title>
      <meta name="viewport" content="width=device-width, initial-scale=1.0" id="viewport-meta" />
      <!-- SISA 브랜드 파비콘 -->
      <link rel="icon" href="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'><rect width='100' height='100' rx='22' fill='%232f4b9f'/><circle cx='50' cy='50' r='28' fill='none' stroke='%23ffffff' stroke-width='6'/><ellipse cx='50' cy='50' rx='12' ry='28' fill='none' stroke='%23ffffff' stroke-width='4'/><line x1='22' y1='50' x2='78' y2='50' stroke='%23ffffff' stroke-width='4'/></svg>">
      <script>
        if (screen.width < 1280) {
          var vp = document.getElementById('viewport-meta');
          if (vp) vp.setAttribute('content', 'width=1280');
        }
      </script>
      <!-- 폰트 / 아이콘 / Tailwind -->
      <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;900&display=swap" rel="stylesheet">
      <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
      <script src="https://cdn.tailwindcss.com"></script>
      <script>
        tailwind.config = {
          theme: {
            extend: {
              fontFamily: {
                sans: ['Inter', 'sans-serif'],
              },
              colors: {
                brand: {
                  blue: '#2f4b9f',
                  dark: '#1e326b',
                  accent: '#e6edf7'
                }
              }
            }
          }
        }
      </script>
      <style>
        ::-webkit-scrollbar { width: 8px; height: 8px; }
        ::-webkit-scrollbar-track { background: rgba(255, 255, 255, 0.05); }
        ::-webkit-scrollbar-thumb { background: rgba(255, 255, 255, 0.2); border-radius: 4px; }
        ::-webkit-scrollbar-thumb:hover { background: rgba(255, 255, 255, 0.4); }

        .glass-card {
          background: rgba(255, 255, 255, 0.06);
          backdrop-filter: blur(14px);
          -webkit-backdrop-filter: blur(14px);
          border: 1px solid rgba(255, 255, 255, 0.22);
        }

        .admin-card-inner {
          background: rgba(15,23,42,0.92);
          border-radius: 1.5rem;
          padding: 18px 18px 20px;
          box-shadow: 0 22px 60px rgba(15,23,42,0.9);
          border: 1px solid #1f2937;
        }

        label { display:block; font-size:13px; font-weight:600; color:#9ca3af; margin-bottom:4px; }
        input, select { width:100%; padding:10px 12px; border-radius:10px; border:1px solid #374151; background:#020617; color:#e5e7eb; box-sizing:border-box; font-size:14px; }
        input:focus, select:focus { outline:none; border-color:#3b82f6; box-shadow:0 0 0 1px #3b82f6; }
        .grid { display:grid; grid-template-columns:2fr 1.5fr; gap:16px; margin-top:8px; }
        .actions { margin-top:16px; display:flex; gap:10px; align-items:center; flex-wrap:wrap; }
        .btn-pill { border:none; border-radius:999px; padding:10px 18px; font-size:14px; font-weight:600; cursor:pointer; }
        .btn-primary { background:#3b82f6; color:white; }
        .btn-primary:hover { background:#2563eb; }
        .btn-secondary { background:transparent; color:#e5e7eb; border:1px solid #4b5563; }
        .btn-secondary:hover { background:#111827; }
        .hint { font-size:12px; color:#9ca3af; margin-top:4px; }
        .status-card { margin-top:18px; padding:14px 12px; border-radius:16px; background:#020617; border:1px dashed #374151; font-size:13px; }
        .box-schema { position:sticky; top:72px; z-index:4; margin:6px 0 10px; padding:6px 8px; border-radius:8px; border:1px solid #334155; background:#0b1220; color:#93c5fd; font-size:10px; line-height:1.35; }
        .box-schema code { color:#bfdbfe; }
        .table-sticky thead th { position: sticky; top: 0; background: #0b1220; z-index: 3; }
        .status-title { font-size:13px; font-weight:600; color:#9ca3af; margin-bottom:6px; display:flex; align-items:center; gap:6px; }
        .status-row { display:flex; justify-content:space-between; margin-bottom:4px; gap:8px; }
        .status-label { color:#9ca3af; font-size:12px; }
        .status-value { color:#e5e7eb; font-size:12px; text-align:right; }
        .link-box { margin-top:8px; padding:8px 10px; border-radius:12px; background:#020617; border:1px solid #1f2937; display:flex; gap:8px; align-items:center; }
        .link-text { flex:1; font-size:12px; color:#e5e7eb; word-break:break-all; }
        .msg { margin-top:10px; font-size:12px; color:#a5b4fc; }
        .pill-btn { border-radius:999px; padding:6px 10px; font-size:11px; border:none; cursor:pointer; }
        .pill-danger { background:#b91c1c; color:#fef2f2; }
        .pill-muted { background:#111827; color:#e5e7eb; border:1px solid #4b5563; }
        .small-input { width:100%; padding:6px 8px; border-radius:8px; border:1px solid #374151; background:#020617; color:#e5e7eb; font-size:12px; box-sizing:border-box; }
        .loading-backdrop { position:fixed; inset:0; background:rgba(2,6,23,0.78); display:none; align-items:center; justify-content:center; z-index:2000; }
        .loading-backdrop.show { display:flex; }
        .loading-card { background:#0f172a; border:1px solid #334155; border-radius:14px; padding:16px 18px; color:#e2e8f0; min-width:240px; text-align:center; box-shadow:0 18px 44px rgba(2,6,23,0.65); }
        .loading-spinner { width:28px; height:28px; border:3px solid #475569; border-top-color:#60a5fa; border-radius:50%; margin:0 auto 10px; animation:spin1 0.8s linear infinite; }
        @keyframes spin1 { to { transform: rotate(360deg); } }
        .pending-popup { position:fixed; left:50%; top:50%; transform:translate(-50%,-50%); z-index:2100; min-width:260px; max-width:90vw; background:#0f172a; border:1px solid #334155; border-radius:14px; padding:14px 16px; text-align:center; color:#e2e8f0; box-shadow:0 18px 48px rgba(2,6,23,.72); display:none; }
        .pending-popup.show { display:block; }
        .pending-top-banner { position:fixed; top:76px; left:50%; transform:translateX(-50%); z-index:2090; display:none; align-items:center; gap:8px; background:#7c2d12; border:1px solid #fdba74; color:#fff7ed; border-radius:999px; padding:8px 14px; font-size:12px; font-weight:700; box-shadow:0 10px 24px rgba(124,45,18,.45); }
        .pending-top-banner.show { display:flex; animation:pulseBanner 1.2s ease-in-out infinite; }
        .pending-inline { margin-top:2px; padding:7px 10px; border-radius:10px; background:#7c2d12; border:1px solid #fdba74; color:#fff7ed; font-size:11px; font-weight:700; line-height:1.45; box-shadow:0 8px 16px rgba(124,45,18,.35); animation:pulseBanner 1.2s ease-in-out infinite; }
        .pending-dot { width:10px; height:10px; border-radius:999px; background:#60a5fa; display:inline-block; animation:pulseDot 1s infinite ease-in-out; }
        .pending-dot:nth-child(2) { animation-delay:.2s; }
        .pending-dot:nth-child(3) { animation-delay:.4s; }
        @keyframes pulseBanner { 0%,100% { opacity:.92; } 50% { opacity:1; } }
        @keyframes pulseDot { 0%,100% { opacity:.2; transform:translateY(0);} 50% { opacity:1; transform:translateY(-2px);} }
      </style>
      <script>
        // /admin 페이지에서: 결제중인데 아직 K-VAN 링크가 없는 세션이 있으면
        // 한 번만 7초 후 자동 새로고침하고, 링크가 생성된 뒤에는 새로고침하지 않는다.
        window.addEventListener('DOMContentLoaded', function () {
          var hasPending = {{ 'true' if has_pending_link else 'false' }};
          var pendingPopup = document.getElementById("pending-create-popup");
          var pendingBanner = document.getElementById("pending-create-banner");
          if (hasPending) {
            if (pendingPopup) pendingPopup.classList.add("show");
            if (pendingBanner) pendingBanner.classList.add("show");
            setTimeout(function () {
              window.location.reload();
            }, 7000);
          } else {
            if (pendingPopup) pendingPopup.classList.remove("show");
            if (pendingBanner) pendingBanner.classList.remove("show");
          }
          // 새 링크 생성 후 리다이렉트된 경우(new=1)에는 팝업으로 한 번 안내
          var params = new URLSearchParams(window.location.search || "");
          var isNew = params.get("new") === "1";
          if (isNew && !hasPending) {
            alert("새 결제 링크 생성 작업이 완료되었습니다. 아래 '결제 요청 링크' 영역에서 복사 버튼을 확인해 주세요.");
          }

          // 수동 새로고침(refresh_kvan) 요청은 완료/오류 신호를 받을 때까지 진행 표시를 유지한다.
          var refreshSince = "{{ crawler_refresh_since or '' }}";
          var refreshOverlay = document.getElementById("crawler-refresh-overlay");
          var refreshText = document.getElementById("crawler-refresh-text");
          if (!refreshSince || !refreshOverlay) return;
          var startedAt = Date.now();
          refreshOverlay.classList.add("show");
          function stopRefreshOverlay(finalText) {
            if (refreshText && finalText) refreshText.textContent = finalText;
            setTimeout(function () {
              refreshOverlay.classList.remove("show");
            }, 400);
          }
          function pollRefresh() {
            fetch("/api/crawler-refresh-status?since=" + encodeURIComponent(refreshSince), { cache: "no-store" })
              .then(function (r) { return r.json(); })
              .then(function (d) {
                if (d && d.ok && d.done) {
                  stopRefreshOverlay("크롤러 상태 확인이 완료되었습니다.");
                  return;
                }
                if (Date.now() - startedAt > 90000) {
                  stopRefreshOverlay("크롤러 확인 시간이 초과되어 표시를 종료합니다.");
                  return;
                }
                setTimeout(pollRefresh, 2500);
              })
              .catch(function () {
                if (Date.now() - startedAt > 90000) {
                  stopRefreshOverlay("네트워크 확인 시간 초과로 표시를 종료합니다.");
                  return;
                }
                setTimeout(pollRefresh, 3000);
              });
          }
          pollRefresh();
        });
      </script>
    </head>
    <body class="bg-brand-blue text-white font-sans overflow-x-hidden antialiased flex flex-col min-h-screen">
      <div id="pending-create-banner" class="pending-top-banner" aria-hidden="true">
        <i class="fa-solid fa-spinner fa-spin"></i>
        <span>K-VAN 링크 생성 중 (1분정도 소요됩니다.)</span>
      </div>
      <div id="pending-create-popup" class="pending-popup" aria-hidden="true">
        <div style="font-size:13px;font-weight:700; margin-bottom:6px;">링크 생성중입니다 (1분정도 소요됩니다.)</div>
        <div style="font-size:11px; color:#94a3b8; margin-bottom:8px;">생성이 완료되면 자동으로 반영됩니다.</div>
        <div style="display:flex; justify-content:center; gap:6px;">
          <span class="pending-dot"></span><span class="pending-dot"></span><span class="pending-dot"></span>
        </div>
      </div>
      <!-- 헤더 -->
      <header class="fixed top-0 left-0 right-0 z-30 glass-card border-b border-white/10">
        <div class="max-w-[96vw] mx-auto px-4 py-3 flex items-center justify-between">
          <div class="flex items-center gap-2">
            <i class="fa-solid fa-globe text-white text-xl drop-shadow-sm"></i>
            <div class="flex flex-col leading-tight">
              <span class="text-xs font-semibold tracking-[0.18em] uppercase text-white/70">SISA</span>
              <span class="text-xs text-white/80">K-VAN Payment Admin</span>
            </div>
          </div>
          <div class="hidden sm:flex items-center gap-2 text-[11px] text-white/70">
            <span class="px-2 py-1 rounded-full bg-black/20 border border-white/20">실시간 결제 세션 관리</span>
          </div>
        </div>
      </header>

      <div id="link-loading-overlay" class="loading-backdrop" aria-hidden="true">
        <div class="loading-card">
          <div class="loading-spinner"></div>
          <div id="link-loading-text" style="font-size:13px;font-weight:600;">링크 생성중입니다... (1분정도 소요됩니다.)</div>
          <div style="margin-top:4px;font-size:11px;color:#94a3b8;">잠시만 기다려 주세요.</div>
        </div>
      </div>
      <div id="crawler-refresh-overlay" class="loading-backdrop" aria-hidden="true">
        <div class="loading-card">
          <div class="loading-spinner"></div>
          <div id="crawler-refresh-text" style="font-size:13px;font-weight:600;">크롤러 작업 진행중...</div>
          <div style="margin-top:4px;font-size:11px;color:#94a3b8;">결과를 확인하면 자동으로 종료됩니다.</div>
        </div>
      </div>
      <main class="flex-grow pt-24 pb-12 px-3 sm:px-4">
        <div class="max-w-[96vw] mx-auto">
          <div class="glass-card rounded-[2rem] border border-white/20 shadow-2xl">
            <div class="admin-card-inner">
              <div class="flex items-center justify-between mb-4">
                <div>
                  <h1 class="text-xl font-semibold text-white mb-1">World SISA 대면결제 세션 어드민</h1>
                  <p class="text-xs text-slate-300">
                    고객에게 보낼 결제 링크를 생성하고, 진행 중인 결제와 완료된 결제를 한 곳에서 확인합니다.
                  </p>
                </div>
              </div>

              <form method="post" action="{{ url_for('admin') }}" data-loading-msg="링크 생성중입니다... (1분정도 소요됩니다.)">
                <div class="grid">
                  <div>
                    <label for="admin_amount">결제 금액 (원 단위)</label>
                    <input id="admin_amount" name="admin_amount" inputmode="numeric" placeholder="예: 20000" />
                    <div class="hint">비워두면 금액이 고정되지 않은 결제 요청 링크가 생성됩니다.</div>
                  </div>
                  <div>
                    <label for="admin_installment">할부개월</label>
                    <select id="admin_installment" name="admin_installment">
                      <option value="일시불" selected>일시불</option>
                      {% for m in range(2,7) %}
                        <option value="{{ m }}">{{ m }}개월</option>
                      {% endfor %}
                    </select>
                  </div>
                </div>
                <div class="actions">
                  <input type="hidden" name="action" value="create" />
                  <button type="submit" class="btn-pill btn-primary">결제창 생성</button>
                  <span class="hint">버튼을 누르면 새로운 결제 요청 링크가 만들어집니다. (동시 최대 5개)</span>
                </div>
              </form>

              <div class="status-card">
                <div class="status-title">
                  <i class="fa-solid fa-circle-play text-emerald-400 text-xs"></i>
                  진행 중인 결제 세션 (최대 5개)
                  <form method="post" action="{{ url_for('admin') }}" style="margin-left:auto;" data-loading-msg="크롤러를 새로고침하는 중입니다...">
                    <input type="hidden" name="action" value="refresh_kvan" />
                    <button type="submit" class="pill-btn pill-muted" style="font-size:11px; padding:4px 10px;">
                      새로고침
                    </button>
                  </form>
                </div>
                <div class="box-schema"><code>admin_state.sessions</code> (진행 중 = 결제중만 표시)</div>
                {% if active_sessions %}
                  {% for s in active_sessions %}
                    <div style="margin:8px 0; padding:10px 11px; border-radius:12px; background:#020617; border:1px solid #111827;">
                      <div class="status-row">
                        <span class="status-label">세션 ID</span>
                        <span class="status-value">{{ s.id }}</span>
                      </div>
                      <div class="status-row">
                        <span class="status-label">결제금액</span>
                        <span class="status-value">{{ s.amount or '고정 안 됨' }}</span>
                      </div>
                      <div class="status-row">
                        <span class="status-label">할부개월</span>
                        <span class="status-value">{{ s.installment }}</span>
                      </div>
                      <div class="status-row">
                        <span class="status-label">상태</span>
                        <span class="status-value">
                          <span style="display:inline-block;padding:2px 8px;border-radius:6px;font-size:11px;font-weight:600;background:#065f46;color:#a7f3d0;border:1px solid #047857;">결제중</span>
                        </span>
                      </div>
                      {% if s.status == '만료' %}
                      <div class="status-row">
                        <span class="status-label">만료/삭제</span>
                        <span class="status-value" style="color:#fca5a5;">
                          만료 감지됨{% if s.deleted_in_kvan %} · K-VAN에서 삭제됨{% endif %}
                        </span>
                      </div>
                      {% endif %}
                      <div class="status-title" style="margin-top:6px;">
                        <i class="fa-solid fa-link text-blue-400 text-xs"></i>
                        결제 요청 링크
                      </div>
                      <div class="link-box">
                        {% set kvan_link = s.kvan_link %}
                        {% if kvan_link %}
                        <div class="link-text" id="pay-link-{{ loop.index }}">{{ kvan_link }}</div>
                        <button type="button" class="btn-pill btn-secondary" onclick="copyPayLink('pay-link-{{ loop.index }}')">복사</button>
                        {% elif s.status == '링크생성실패' %}
                        <div style="background:#450a0a;border:1px solid #7f1d1d;border-radius:8px;padding:8px 10px;margin-bottom:4px;">
                          <div style="color:#fca5a5;font-size:11px;font-weight:600;">⚠ 링크 생성 실패</div>
                          {% if s.error_reason %}
                          <div style="color:#fecaca;font-size:10px;margin-top:2px;">{{ s.error_reason }}</div>
                          {% endif %}
                          <form method="post" action="{{ url_for('admin') }}" style="margin-top:6px;" data-loading-msg="링크 재생성 요청중입니다...">
                            <input type="hidden" name="action" value="retry_kvan" />
                            <input type="hidden" name="session_id" value="{{ s.id }}" />
                            <button type="submit" style="background:#dc2626;color:white;border:none;border-radius:6px;padding:4px 12px;font-size:11px;cursor:pointer;">
                              🔄 다시 링크 생성
                            </button>
                          </form>
                        </div>
                        {% else %}
                        <div class="pending-inline">K-VAN 링크를 생성 중입니다. 1분정도 소요됩니다. 잠시 후 새로고침 해 주세요.</div>
                        {% endif %}
                      </div>
                      <form method="post" action="{{ url_for('admin') }}" style="margin-top:6px; display:flex; gap:6px; align-items:center; flex-wrap:wrap;">
                        <input type="hidden" name="action" value="close_session" />
                        <input type="hidden" name="session_id" value="{{ s.id }}" />
                        <input class="small-input" name="memo" placeholder="종료 메모 (선택)" />
                        <button type="submit" class="pill-btn pill-danger">강제종료</button>
                      </form>
                    </div>
                  {% endfor %}
                {% else %}
                  <div class="hint">아직 생성된 결제 요청 링크가 없습니다.</div>
                {% endif %}
              </div>

              <!-- K-VAN 연동 거래 내역 / 결제링크 관리는 HQ 대시보드에서만 표시하고,
                   /admin 페이지에서는 결제 세션 생성/관리 UI 만 제공한다. -->

              <div class="status-card">
                <div class="status-title">
                  <i class="fa-solid fa-clipboard-list text-indigo-300 text-xs"></i>
                  결제관리 (완료/종료된 세션)
                </div>
                <div class="box-schema">크롤링 결과에 따라 만료·결제완료·실패 등은 이 섹션으로 표시됩니다.</div>
                {% if completed_sessions %}
                  {% for h in completed_sessions %}
                    <div style="margin:8px 0; padding:10px 11px; border-radius:12px; background:#020617; border:1px solid #1f2937;">
                      <div class="status-row">
                        <span class="status-label">세션 ID</span>
                        <span class="status-value">{{ h.id }}</span>
                      </div>
                      <div class="status-row">
                        <span class="status-label">이름</span>
                        <span class="status-value">{{ h.customer_name or '-' }}</span>
                      </div>
                      <div class="status-row">
                        <span class="status-label">전화번호</span>
                        <span class="status-value">{{ h.phone_number or '-' }}</span>
                      </div>
                      <div class="status-row">
                        <span class="status-label">금액</span>
                        <span class="status-value">{{ h.amount or '-' }}</span>
                      </div>
                      <div class="status-row">
                        <span class="status-label">상태</span>
                        <span class="status-value">
                          {% set st = (h.status or '') %}
                          {% if '만료' in st or st == '만료' %}
                            <span style="display:inline-block;padding:2px 8px;border-radius:6px;font-size:11px;font-weight:600;background:#78350f;color:#fde68a;border:1px solid #b45309;">링크만료</span>
                          {% elif st in ['결제완료','성공','success'] or h.get('has_transaction') %}
                            <span style="display:inline-block;padding:2px 8px;border-radius:6px;font-size:11px;font-weight:600;background:#065f46;color:#a7f3d0;border:1px solid #047857;">결제완료</span>
                          {% elif st in ['실패','fail','링크생성실패'] %}
                            <span style="display:inline-block;padding:2px 8px;border-radius:6px;font-size:11px;font-weight:600;background:#7f1d1d;color:#fecaca;border:1px solid #b91c1c;">결제 실패</span>
                          {% else %}
                            <span style="display:inline-block;padding:2px 8px;border-radius:6px;font-size:11px;background:#374151;color:#d1d5db;">{{ h.status or '-' }}</span>
                          {% endif %}
                          {% if h.deleted_in_kvan %}<span style="color:#fca5a5;"> · 삭제됨</span>{% endif %}
                        </span>
                      </div>
                      <div class="status-row">
                        <span class="status-label">정산</span>
                        <span class="status-value">{{ h.settled or '정산전' }}</span>
                      </div>
                      <div class="status-row">
                        <span class="status-label">완료시간</span>
                        <span class="status-value" style="font-size:11px;">{{ h.finished_at or '-' }}</span>
                      </div>
                      <div class="status-title" style="margin-top:6px;">메모 / 실패사유</div>
                      <div style="font-size:12px; color:#e5e7eb; white-space:pre-line; margin-bottom:6px;">
                        {{ h.result_message or '-' }}
                      </div>
                      <div class="status-row" style="gap:6px; margin-top:4px; flex-wrap:wrap;">
                        <form method="post" action="{{ url_for('admin') }}">
                          <input type="hidden" name="action" value="toggle_settle" />
                          <input type="hidden" name="session_id" value="{{ h.id }}" />
                          <button type="submit" class="pill-btn pill-muted">
                            {% if h.settled == '정산완료' %}정산취소{% else %}정산완료{% endif %}
                          </button>
                        </form>
                        <button type="button" class="pill-btn pill-muted" onclick="copyHistory('{{ h.customer_name or '' }}','{{ h.phone_number or '' }}','{{ h.amount or '' }}','{{ h.status or '' }}','{{ (h.result_message or '').replace('\\n',' ') }}')">복사</button>
                        <form method="post" action="{{ url_for('admin') }}">
                          <input type="hidden" name="action" value="delete_history" />
                          <input type="hidden" name="session_id" value="{{ h.id }}" />
                          <button type="submit" class="pill-btn pill-danger">삭제</button>
                        </form>
                      </div>
                    </div>
                  {% endfor %}
                {% else %}
                  <div class="hint">아직 완료/종료된 결제 기록이 없습니다.</div>
                {% endif %}
              </div>

              <!-- K-VAN / 내부 DB 요약 (크롤러/매크로가 생성한 데이터 스냅샷) -->
              <div class="status-card">
                <div class="status-title">
                  <i class="fa-solid fa-database text-sky-300 text-xs"></i>
                  K-VAN 링크 DB 요약 (최근 10건)
                </div>
                <div class="box-schema"><code>kvan_links</code> 항목: <code>captured_at, title, amount, ttl_label, status, kvan_link, mid, kvan_session_id</code></div>
                {% if recent_links %}
                  <div style="max-height:200px; overflow-y:auto; font-size:11px; margin-top:4px;">
                    <table class="table-sticky" style="width:100%; border-collapse:collapse;">
                      <thead>
                        <tr style="border-bottom:1px solid rgba(148,163,184,0.4);">
                          <th style="padding:3px 4px; text-align:left;">생성시각</th>
                          <th style="padding:3px 4px; text-align:left;">제목</th>
                          <th style="padding:3px 4px; text-align:right;">금액</th>
                          <th style="padding:3px 4px; text-align:left;">상태</th>
                          <th style="padding:3px 4px; text-align:left;">링크</th>
                        </tr>
                      </thead>
                      <tbody>
                        {% for l in recent_links %}
                        <tr style="border-bottom:1px dashed rgba(55,65,81,0.8);">
                          <td style="padding:3px 4px; font-size:10px;">{{ l.captured_at }}</td>
                          <td style="padding:3px 4px;">{{ l.title }}</td>
                          <td style="padding:3px 4px; text-align:right;">{{ "{:,}".format(l.amount or 0) }} 원</td>
                          <td style="padding:3px 4px;">{{ l.status }}</td>
                          <td style="padding:3px 4px; font-size:10px; word-break:break-all;">
                            {{ l.kvan_link }}
                          </td>
                        </tr>
                        {% endfor %}
                      </tbody>
                    </table>
                  </div>
                {% else %}
                  <div class="hint">아직 K-VAN 링크 DB(kvan_links)에 저장된 내역이 없습니다.</div>
                {% endif %}
              </div>

              <div class="status-card">
                <div class="status-title">
                  <i class="fa-solid fa-receipt text-emerald-300 text-xs"></i>
                  내부 거래 DB 요약 (최근 10건)
                </div>
                <div class="box-schema"><code>transactions</code> 항목: <code>created_at, amount, customer_name, status, settlement_status</code></div>
                {% if recent_tx %}
                  <div style="max-height:200px; overflow-y:auto; font-size:11px; margin-top:4px;">
                    <table class="table-sticky" style="width:100%; border-collapse:collapse;">
                      <thead>
                        <tr style="border-bottom:1px solid rgba(148,163,184,0.4);">
                          <th style="padding:3px 4px; text-align:left;">시간</th>
                          <th style="padding:3px 4px; text-align:right;">금액</th>
                          <th style="padding:3px 4px; text-align:left;">고객</th>
                          <th style="padding:3px 4px; text-align:left;">결제상태</th>
                          <th style="padding:3px 4px; text-align:left;">정산상태</th>
                        </tr>
                      </thead>
                      <tbody>
                        {% for t in recent_tx %}
                        <tr style="border-bottom:1px dashed rgba(55,65,81,0.8);">
                          <td style="padding:3px 4px; font-size:10px;">{{ t.created_at }}</td>
                          <td style="padding:3px 4px; text-align:right;">{{ "{:,}".format(t.amount or 0) }} 원</td>
                          <td style="padding:3px 4px;">{{ t.customer_name }}</td>
                          <td style="padding:3px 4px;">{{ t.status }}</td>
                          <td style="padding:3px 4px;">{{ t.settlement_status }}</td>
                        </tr>
                        {% endfor %}
                      </tbody>
                    </table>
                  </div>
                {% else %}
                  <div class="hint">아직 내부 거래 DB(transactions)에 저장된 내역이 없습니다.</div>
                {% endif %}
              </div>
              {% if message %}
                <div class="msg">{{ message }}</div>
              {% endif %}
            </div>
          </div>
        </div>
      </main>

      <script>
        function showLinkLoading(msg) {
          var overlay = document.getElementById("link-loading-overlay");
          var textEl = document.getElementById("link-loading-text");
          if (!overlay) return;
          if (textEl && msg) textEl.textContent = msg;
          overlay.classList.add("show");
        }
        (function () {
          var forms = document.querySelectorAll("form[data-loading-msg]");
          forms.forEach(function (f) {
            f.addEventListener("submit", function () {
              var msg = f.getAttribute("data-loading-msg") || "처리중입니다...";
              showLinkLoading(msg);
            });
          });
        })();
        function copyPayLink(id) {
          var el = document.getElementById(id);
          if (!el) return;
          var text = el.textContent || el.innerText || "";
          if (!navigator.clipboard) {
            var ta = document.createElement("textarea");
            ta.value = text;
            document.body.appendChild(ta);
            ta.select();
            document.execCommand("copy");
            document.body.removeChild(ta);
          } else {
            navigator.clipboard.writeText(text).catch(function() {});
          }
          alert("결제요청 페이지 링크가 복사되었습니다.");
        }

        function copyHistory(name, phone, amount, status, memo) {
          var parts = [
            "이름: " + (name || ""),
            "전화: 0" + (phone || ""),
            "금액: " + (amount || ""),
            "상태: " + (status || ""),
            "메모: " + (memo || "")
          ];
          var text = parts.join("\\t");
          if (!navigator.clipboard) {
            var ta = document.createElement("textarea");
            ta.value = text;
            document.body.appendChild(ta);
            ta.select();
            document.execCommand("copy");
            document.body.removeChild(ta);
          } else {
            navigator.clipboard.writeText(text).catch(function () {});
          }
          alert("결제 실패/완료 정보가 복사되었습니다.");
        }
      </script>
    </body>
    </html>
    """

    return render_template_string(
        ADMIN_TEMPLATE,
        sessions=sessions,
        history=history,
        active_sessions=active_sessions,
        completed_sessions=completed_sessions,
        message=message,
        base_url=base_url,
        recent_links=recent_links,
        recent_tx=recent_tx,
        has_pending_link=has_pending_link,
        crawler_refresh_since=crawler_refresh_since,
    )


@app.route("/hq-login", methods=["GET", "POST"])
def hq_login():
    """본사 메인 어드민 로그인 (admin / admin1234 기본값)."""
    error = ""
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        admin_user = os.environ.get("HQ_ADMIN_USER", "admin")
        admin_pw = os.environ.get("HQ_ADMIN_PASSWORD", "admin1234")
        if username == admin_user and password == admin_pw:
            session["hq_logged_in"] = True
            return redirect(url_for("hq_admin"))
        error = "아이디 또는 비밀번호가 올바르지 않습니다."

    template = """
    <!DOCTYPE html>
    <html lang="ko">
    <head>
      <meta charset="UTF-8" />
      <title>SISA HQ 어드민 로그인</title>
      <meta name="viewport" content="width=device-width, initial-scale=1.0" id="viewport-meta" />
      <script>
        if (screen.width < 1280) {
          var vp = document.getElementById('viewport-meta');
          if (vp) vp.setAttribute('content', 'width=1280');
        }
      </script>
      <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap" rel="stylesheet">
      <script src="https://cdn.tailwindcss.com"></script>
      <style>
        body { background-color: #2f4b9f; }
        /* 결제 폼의 결과 모달 오버레이가 남아 있어도 로그인 화면에서는 항상 숨긴다. */
        #result-modal,
        .result-backdrop {
          display: none !important;
        }
      </style>
    </head>
    <body class="bg-[#2f4b9f] text-white font-[Inter] min-h-screen flex items-center justify-center">
      <div class="bg-white/10 border border-white/20 rounded-2xl px-8 py-10 max-w-sm w-full shadow-2xl">
        <h1 class="text-xl font-bold mb-2 text-center">SISA HQ Admin</h1>
        <p class="text-xs text-white/70 text-center mb-6">본사 전용 어드민 로그인</p>
        <form method="post" class="space-y-4">
          <div>
            <label class="block text-xs font-semibold text-white/70 mb-1">아이디</label>
            <input name="username" type="text" required class="w-full bg-black/20 border border-white/20 rounded-lg py-2.5 px-3 text-sm text-white placeholder-white/40 focus:outline-none focus:border-blue-300" placeholder="admin" />
          </div>
          <div>
            <label class="block text-xs font-semibold text-white/70 mb-1">비밀번호</label>
            <input name="password" type="password" required class="w-full bg-black/20 border border-white/20 rounded-lg py-2.5 px-3 text-sm text-white placeholder-white/40 focus:outline-none focus:border-blue-300" placeholder="********" />
          </div>
          {% if error %}
          <p class="text-xs text-red-200">{{ error }}</p>
          {% endif %}
          <button type="submit" class="w-full mt-2 bg-white text-brand-blue font-bold py-2.5 rounded-lg hover:bg-brand-accent transition">
            로그인
          </button>
        </form>
      </div>
    </body>
    </html>
    """
    return render_template_string(template, error=error)


@app.route("/agency-login", methods=["GET", "POST"])
def agency_login():
    """대행사 전용 로그인 페이지."""
    error = ""
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        ag = _find_agency_by_credentials(username, password)
        if ag:
            session["agency_id"] = ag.get("id")
            session["agency_name"] = ag.get("company_name")
            return redirect(url_for("agency_admin"))
        error = "아이디 또는 비밀번호가 올바르지 않습니다."

    template = """
    <!DOCTYPE html>
    <html lang="ko">
    <head>
      <meta charset="UTF-8" />
      <title>SISA 대행사 어드민 로그인</title>
      <meta name="viewport" content="width=device-width, initial-scale=1.0" id="viewport-meta" />
      <script>
        if (screen.width < 1280) {
          var vp = document.getElementById('viewport-meta');
          if (vp) vp.setAttribute('content', 'width=1280');
        }
      </script>
      <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap" rel="stylesheet">
      <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css" />
      <script src="https://cdn.tailwindcss.com"></script>
      <script>
        tailwind.config = {
          theme: {
            extend: {
              fontFamily: { sans: ['Inter', 'sans-serif'] },
              colors: {
                brand: { blue: '#2f4b9f', dark: '#1e326b', accent: '#e6edf7' }
              }
            }
          }
        }
      </script>
      <style>
        body { background-color: #2f4b9f; }
        /* 결제 폼의 결과 모달 오버레이가 남아 있어도 로그인 화면에서는 항상 숨긴다. */
        #result-modal,
        .result-backdrop {
          display: none !important;
        }
      </style>
    </head>
    <body class="bg-brand-blue text-white font-sans antialiased min-h-screen flex items-center justify-center">
      <div class="bg-white/10 backdrop-blur border border-white/20 rounded-2xl px-8 py-10 max-w-sm w-full shadow-2xl">
        <h1 class="text-xl font-bold mb-2 text-center text-white">SISA Agency Admin</h1>
        <p class="text-xs text-white/80 text-center mb-6">승인된 대행사 전용 어드민 로그인</p>
        <form method="post" class="space-y-4">
          <div>
            <label class="block text-xs font-semibold text-white/80 mb-1">대행사 아이디</label>
            <input name="username" type="text" required class="w-full bg-black/20 border border-white/20 rounded-lg py-2.5 px-3 text-sm text-white placeholder-white/40 focus:outline-none focus:border-blue-300" placeholder="agency id" />
          </div>
          <div>
            <label class="block text-xs font-semibold text-white/80 mb-1">비밀번호</label>
            <input name="password" type="password" required class="w-full bg-black/20 border border-white/20 rounded-lg py-2.5 px-3 text-sm text-white placeholder-white/40 focus:outline-none focus:border-blue-300" placeholder="********" />
          </div>
          {% if error %}
          <p class="text-xs text-red-200">{{ error }}</p>
          {% endif %}
          <button type="submit" class="w-full mt-2 bg-white text-brand-blue font-bold py-2.5 rounded-lg hover:opacity-90 transition" style="color: #2f4b9f;">
            로그인
          </button>
        </form>
      </div>
    </body>
    </html>
    """
    return render_template_string(template, error=error)


@app.route("/hq-admin", methods=["GET", "POST"])
def hq_admin():
    """본사 메인 어드민 대시보드."""
    if not session.get("hq_logged_in"):
        return redirect(url_for("hq_login"))

    history_warnings = cleanup_history_files()

    state = _load_hq_state()
    applications = state.get("applications") or []
    agencies = state.get("agencies") or []
    transactions = state.get("transactions") or []
    message = ""
    crawler_refresh_since = ""

    # 최신 K-VAN 대시보드 스냅샷 1건 조회
    latest_dashboard = None
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM kvan_dashboard ORDER BY captured_at DESC LIMIT 1"
            )
            latest_dashboard = cur.fetchone()
        conn.close()
    except Exception as e:  # noqa: BLE001
        print(f"[WARN] kvan_dashboard 조회 실패: {e}")

    if request.method == "POST":
        action = request.form.get("action", "").strip()
        if action == "approve_application":
            app_id = request.form.get("application_id", "").strip()
            found = None
            for a in applications:
                if str(a.get("id")) == app_id:
                    found = a
                    break
            if found:
                agency_id = datetime.utcnow().strftime("AGY%Y%m%d%H%M%S%f")
                try:
                    conn = get_db()
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO agencies
                            (id, company_name, domain, phone, bank_name, account_number,
                             email_or_sheet, login_id, login_password, fee_percent, created_at, status)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                            """,
                            (
                                agency_id,
                                found.get("company_name", ""),
                                found.get("domain", ""),
                                found.get("phone", ""),
                                found.get("bank_name", ""),
                                found.get("account_number", ""),
                                found.get("email_or_sheet", ""),
                                found.get("login_id", ""),
                                found.get("login_password", ""),
                                found.get("fee_percent", 10),
                                datetime.utcnow().isoformat(),
                                "active",
                            ),
                        )
                        cur.execute("DELETE FROM applications WHERE id = %s", (app_id,))
                    conn.commit()
                    conn.close()
                    message = f"대행사 '{found.get('company_name', '')}' 가 생성되었습니다."
                except Exception as e:  # noqa: BLE001
                    message = f"대행사 승인 처리 중 오류: {e}"
        elif action == "refresh_kvan":
            # 수동으로 K-VAN 크롤링 매크로를 한 번 더 실행
            try:
                trigger_kvan_crawler_refresh()
                crawler_refresh_since = datetime.utcnow().isoformat()
                message = "K-VAN 크롤러 새로고침 신호를 보냈습니다. 로그 박스에서 상세 진행 상태를 확인해 주세요."
            except Exception as e:  # noqa: BLE001
                print(f"[WARN] HQ에서 refresh_kvan 실행 중 오류: {e}")
                message = "K-VAN 크롤링 재실행 중 오류가 발생했습니다. 잠시 후 다시 시도해 주세요."
        elif action == "delete_application":
            app_id = request.form.get("application_id", "").strip()
            if app_id:
                try:
                    conn = get_db()
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM applications WHERE id = %s", (app_id,))
                    conn.commit()
                    conn.close()
                    message = "선택한 대행사 신청이 삭제되었습니다."
                except Exception as e:  # noqa: BLE001
                    message = f"대행사 신청 삭제 중 오류: {e}"
        elif action == "update_fee":
            agency_id = request.form.get("agency_id", "").strip()
            try:
                fee_percent = int(request.form.get("fee_percent", "").strip())
            except ValueError:
                fee_percent = None
            if agency_id and fee_percent is not None:
                try:
                    conn = get_db()
                    with conn.cursor() as cur:
                        cur.execute(
                            "UPDATE agencies SET fee_percent = %s WHERE id = %s",
                            (fee_percent, agency_id),
                        )
                    conn.commit()
                    conn.close()
                    message = "수수료 설정이 저장되었습니다."
                except Exception as e:  # noqa: BLE001
                    message = f"수수료 설정 저장 중 오류: {e}"
        elif action == "update_application_fee":
            app_id = request.form.get("application_id", "").strip()
            try:
                fee_percent = int(request.form.get("fee_percent", "").strip())
            except ValueError:
                fee_percent = None
            if app_id and fee_percent is not None:
                try:
                    conn = get_db()
                    with conn.cursor() as cur:
                        cur.execute(
                            "UPDATE applications SET fee_percent = %s WHERE id = %s",
                            (fee_percent, app_id),
                        )
                    conn.commit()
                    conn.close()
                    message = "대행사 신청 수수료가 저장되었습니다."
                except Exception as e:  # noqa: BLE001
                    message = f"신청 수수료 저장 중 오류: {e}"
        elif action == "bulk_settle":
            tx_ids = request.form.getlist("tx_ids")
            if tx_ids:
                tx_id_set = {str(x).strip() for x in tx_ids if str(x).strip()}
                if tx_id_set:
                    try:
                        conn = get_db()
                        with conn.cursor() as cur:
                            placeholders = ",".join(["%s"] * len(tx_id_set))
                            cur.execute(
                                f"""
                                UPDATE transactions
                                SET settlement_status = '정산완료', settled_at = NOW()
                                WHERE id IN ({placeholders})
                                """,
                                tuple(tx_id_set),
                            )
                        conn.commit()
                        conn.close()
                        message = f"{len(tx_id_set)}건을 정산완료로 표시했습니다."
                    except Exception as e:  # noqa: BLE001
                        message = f"일괄 정산 처리 중 오류: {e}"
        elif action == "update_agency":
            agency_id = request.form.get("agency_id", "").strip()
            do = request.form.get("do", "save").strip()
            if agency_id:
                phone = (request.form.get("phone") or "").strip()
                bank_name = (request.form.get("bank_name") or "").strip()
                account_number = (request.form.get("account_number") or "").strip()
                email_or_sheet = (request.form.get("email_or_sheet") or "").strip()
                login_id_val = (request.form.get("login_id") or "").strip()
                login_pw_val = (request.form.get("login_password") or "").strip()
                kvan_mid_val = (request.form.get("kvan_mid") or "").strip()
                kvan_login_id_val = (request.form.get("kvan_login_id") or "").strip()
                kvan_login_pw_val = (request.form.get("kvan_login_password") or "").strip()
                kvan_login_pin_val = (request.form.get("kvan_login_pin") or "").strip()
                status_val = (request.form.get("status") or "").strip() or "active"
                try:
                    fee_percent = int((request.form.get("fee_percent") or "").strip())
                except ValueError:
                    fee_percent = None
                try:
                    conn = get_db()
                    with conn.cursor() as cur:
                        updates: list[str] = ["status = %s"]
                        params: list = [status_val]
                        if phone:
                            updates.append("phone = %s")
                            params.append(phone)
                        if bank_name:
                            updates.append("bank_name = %s")
                            params.append(bank_name)
                        if account_number:
                            updates.append("account_number = %s")
                            params.append(account_number)
                        if email_or_sheet:
                            updates.append("email_or_sheet = %s")
                            params.append(email_or_sheet)
                        if login_id_val:
                            updates.append("login_id = %s")
                            params.append(login_id_val)
                        if login_pw_val:
                            updates.append("login_password = %s")
                            params.append(login_pw_val)
                        if kvan_mid_val:
                            updates.append("kvan_mid = %s")
                            params.append(kvan_mid_val)
                        if kvan_login_id_val:
                            updates.append("kvan_login_id = %s")
                            params.append(kvan_login_id_val)
                        if kvan_login_pw_val:
                            updates.append("kvan_login_password = %s")
                            params.append(kvan_login_pw_val)
                        if kvan_login_pin_val:
                            updates.append("kvan_login_pin = %s")
                            params.append(kvan_login_pin_val)
                        if fee_percent is not None:
                            updates.append("fee_percent = %s")
                            params.append(fee_percent)
                        params.append(agency_id)
                        cur.execute(
                            f"UPDATE agencies SET {', '.join(updates)} WHERE id = %s",
                            tuple(params),
                        )
                        if do == "settle":
                            cur.execute(
                                """
                                UPDATE transactions
                                SET settlement_status = '정산완료', settled_at = NOW()
                                WHERE agency_id = %s
                                  AND status = 'success'
                                  AND (settlement_status IS NULL OR settlement_status != '정산완료')
                                """,
                                (agency_id,),
                            )
                    conn.commit()
                    conn.close()
                    if do == "settle":
                        message = "선택한 대행사의 미정산 거래를 정산완료로 표시했습니다."
                    else:
                        message = "대행사 정보가 저장되었습니다."
                except Exception as e:  # noqa: BLE001
                    message = f"대행사 정보 저장 중 오류: {e}"
        elif action == "delete_tx":
            tx_id = request.form.get("tx_id", "").strip()
            if tx_id:
                try:
                    conn = get_db()
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM transactions WHERE id = %s", (tx_id,))
                    conn.commit()
                    conn.close()
                    message = "선택한 거래 내역이 삭제되었습니다."
                except Exception as e:  # noqa: BLE001
                    message = f"거래 내역 삭제 중 오류: {e}"
        elif action == "bulk_delete_tx":
            tx_ids = request.form.getlist("tx_ids")
            if tx_ids:
                tx_id_set = {str(x).strip() for x in tx_ids if str(x).strip()}
                if tx_id_set:
                    try:
                        conn = get_db()
                        with conn.cursor() as cur:
                            placeholders = ",".join(["%s"] * len(tx_id_set))
                            cur.execute(
                                f"DELETE FROM transactions WHERE id IN ({placeholders})",
                                tuple(tx_id_set),
                            )
                        conn.commit()
                        conn.close()
                        message = f"{len(tx_id_set)}건의 거래 내역을 삭제했습니다."
                    except Exception as e:  # noqa: BLE001
                        message = f"거래 내역 일괄 삭제 중 오류: {e}"
        elif action == "clear_logs":
            # HQ 어드민에서 버튼으로 로그 파일을 비울 수 있게 한다.
            try:
                ADMIN_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
                with open(ADMIN_LOG_PATH, "w", encoding="utf-8") as lf:
                    lf.write("")
                message = "K-VAN/매크로 로그 파일을 비웠습니다."
            except Exception as e:  # noqa: BLE001
                print(f"[WARN] HQ 로그 파일 삭제 실패: {e}")
                message = "로그 파일 삭제 중 오류가 발생했습니다."
        elif action == "mark_payment_notifications_seen":
            _mark_payment_notifications_seen(agency_id=None)
            message = "결제 완료 알림을 모두 확인 처리했습니다."
        elif action == "mark_expired_with_transactions_seen":
            try:
                if EXPIRED_WITH_TRANSACTIONS_PATH.exists():
                    items = json.loads(EXPIRED_WITH_TRANSACTIONS_PATH.read_text(encoding="utf-8"))
                    if isinstance(items, list):
                        for it in items:
                            it["seen"] = True
                        EXPIRED_WITH_TRANSACTIONS_PATH.write_text(
                            json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8"
                        )
                message = "만료+거래있음 목록을 확인 처리했습니다."
            except Exception as e:
                message = f"확인 처리 중 오류: {e}"

    # POST 처리 후 화면은 항상 DB 기준 최신 상태를 다시 로드한다.
    state = _load_hq_state()
    applications = state.get("applications") or []
    agencies = state.get("agencies") or []
    transactions = state.get("transactions") or []

    # 미확인 결제 알림 건수 (본사는 전체)
    payment_notifications_count = len(_load_payment_notifications(agency_id=None))

    # 만료되었으나 거래 내역이 있는 링크 목록 (크롤러가 저장, 어드민 알림용)
    expired_with_transactions: list = []
    try:
        if EXPIRED_WITH_TRANSACTIONS_PATH.exists():
            raw = EXPIRED_WITH_TRANSACTIONS_PATH.read_text(encoding="utf-8")
            data = json.loads(raw)
            expired_with_transactions = data if isinstance(data, list) else []
    except Exception:
        expired_with_transactions = []
    expired_with_tx_unseen = sum(1 for x in expired_with_transactions if not x.get("seen"))
    expired_with_transactions_reversed = list(reversed(expired_with_transactions))  # 최신순 표시

    # 대행사 관리 페이징 (20개씩)
    try:
        page = int(request.args.get("page", "1"))
    except ValueError:
        page = 1
    page_size = 20
    total_agencies = len(agencies)
    total_pages = (total_agencies + page_size - 1) // page_size if total_agencies else 1
    if page < 1:
        page = 1
    if page > total_pages:
        page = total_pages
    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size
    paged_agencies = agencies[start_idx:end_idx]

    # 전체 거래 기본 날짜(오늘) 문자열
    today_str = datetime.utcnow().strftime("%Y-%m-%d")

    # 최근 HQ 로그 파일 tail (하루치 기준, 최대 1000줄 정도만 표시)
    admin_logs: list[str] = []
    try:
        ADMIN_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        if not ADMIN_LOG_PATH.exists():
            with open(ADMIN_LOG_PATH, "a", encoding="utf-8"):
                pass
        if ADMIN_LOG_PATH.exists():
            with open(ADMIN_LOG_PATH, "r", encoding="utf-8") as lf:
                lines = lf.readlines()
            # 마지막 1000줄만 유지 (사실상 하루치 로그)
            admin_logs = [ln.rstrip("\n") for ln in lines[-1000:]]
    except Exception as e:  # noqa: BLE001
        print(f"[WARN] HQ 로그 파일 읽기 실패: {e}")

    template = """
    <!DOCTYPE html>
    <html lang="ko">
    <head>
      <meta charset="UTF-8" />
      <title>SISA HQ Admin</title>
      <meta name="viewport" content="width=device-width, initial-scale=1.0" id="viewport-meta" />
      <!-- SISA 브랜드 파비콘 -->
      <link rel="icon" href="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'><rect width='100' height='100' rx='22' fill='%232f4b9f'/><circle cx='50' cy='50' r='28' fill='none' stroke='%23ffffff' stroke-width='6'/><ellipse cx='50' cy='50' rx='12' ry='28' fill='none' stroke='%23ffffff' stroke-width='4'/><line x1='22' y1='50' x2='78' y2='50' stroke='%23ffffff' stroke-width='4'/></svg>">
      <script>
        if (screen.width < 1280) {
          var vp = document.getElementById('viewport-meta');
          if (vp) vp.setAttribute('content', 'width=1280');
        }
      </script>
      <script>
        // 본사 어드민은 자동 새로고침을 사용하지 않는다.
        // (수동 새로고침 버튼/크롤링 상태 폴링만 사용)
        // 혹시 이전 페이지에서 남은 오버레이가 DOM에 섞여 있으면 강제로 숨긴다.
        function hideStaleOverlays() {
          var sels = [
            '#result-modal',
            '.result-backdrop',
            "[data-slot='dialog-overlay']",
            '.dialog-overlay',
            '.modal-backdrop'
          ];
          sels.forEach(function (sel) {
            document.querySelectorAll(sel).forEach(function (el) {
              el.style.setProperty('display', 'none', 'important');
              el.style.setProperty('pointer-events', 'none', 'important');
            });
          });
        }
        function runOverlayCleanupBurst() {
          hideStaleOverlays();
          setTimeout(hideStaleOverlays, 300);
          setTimeout(hideStaleOverlays, 1200);
        }
        window.addEventListener('load', runOverlayCleanupBurst);
        // Safari/Chrome의 뒤로가기 캐시(bfcache) 복원 시 load가 실행되지 않을 수 있다.
        window.addEventListener('pageshow', runOverlayCleanupBurst);
        window.addEventListener('DOMContentLoaded', function () {
          var refreshSince = "{{ crawler_refresh_since or '' }}";
          var overlay = document.getElementById("crawler-refresh-overlay");
          var textEl = document.getElementById("crawler-refresh-text");
          if (!refreshSince || !overlay) return;
          var startedAt = Date.now();
          overlay.classList.add("show");
          function stopOverlay(msg) {
            if (textEl && msg) textEl.textContent = msg;
            setTimeout(function () { overlay.classList.remove("show"); }, 400);
          }
          function poll() {
            fetch("/api/crawler-refresh-status?since=" + encodeURIComponent(refreshSince), { cache: "no-store" })
              .then(function (r) { return r.json(); })
              .then(function (d) {
                if (d && d.ok && d.done) {
                  stopOverlay("크롤러 상태 확인이 완료되었습니다.");
                  return;
                }
                if (Date.now() - startedAt > 90000) {
                  stopOverlay("크롤러 확인 시간이 초과되어 표시를 종료합니다.");
                  return;
                }
                setTimeout(poll, 2500);
              })
              .catch(function () {
                if (Date.now() - startedAt > 90000) {
                  stopOverlay("네트워크 확인 시간 초과로 표시를 종료합니다.");
                  return;
                }
                setTimeout(poll, 3000);
              });
          }
          poll();
        });
      </script>
      <style>
        /* 결제 폼에서 사용하던 결과 모달 오버레이가 남아 있어도 HQ 어드민에서는 항상 숨긴다. */
        #result-modal,
        .result-backdrop,
        [data-slot='dialog-overlay'],
        .dialog-overlay,
        .modal-backdrop {
          display: none !important;
          pointer-events: none !important;
        }
        .loading-backdrop { position:fixed; inset:0; background:rgba(2,6,23,0.78); display:none; align-items:center; justify-content:center; z-index:2000; }
        .loading-backdrop.show { display:flex; }
        .loading-card { background:#0f172a; border:1px solid #334155; border-radius:14px; padding:16px 18px; color:#e2e8f0; min-width:240px; text-align:center; box-shadow:0 18px 44px rgba(2,6,23,0.65); }
        .loading-spinner { width:28px; height:28px; border:3px solid #475569; border-top-color:#60a5fa; border-radius:50%; margin:0 auto 10px; animation:spinHQ 0.8s linear infinite; }
        @keyframes spinHQ { to { transform: rotate(360deg); } }
        .overflow-x-auto thead th { position: sticky; top: 0; background: rgba(15,23,42,0.96); z-index: 3; }
        .box-schema { position:sticky; top:72px; z-index:4; margin:6px 0 10px; padding:6px 8px; border-radius:8px; border:1px solid #334155; background:#0b1220; color:#93c5fd; font-size:10px; line-height:1.35; }
        .box-schema code { color:#bfdbfe; }
      </style>
      <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
      <script src="https://cdn.tailwindcss.com"></script>
      <script>
        tailwind.config = {
          theme: {
            extend: {
              fontFamily: { sans: ['Inter', 'sans-serif'] },
              colors: {
                brand: { blue: '#2f4b9f', dark: '#1e326b', accent: '#e6edf7' }
              }
            }
          }
        }
      </script>
    </head>
    <body class="bg-brand-blue text-white font-sans overflow-x-hidden antialiased min-h-screen flex flex-col">
      <header class="fixed top-0 left-0 right-0 z-30 bg-brand-dark/80 backdrop-blur border-b border-white/10">
        <div class="max-w-[96vw] mx-auto px-4 py-3 flex items-center justify-between">
          <div class="flex items-center gap-2">
            <i class="fa-solid fa-shield-halved text-white text-xl"></i>
            <div class="flex flex-col leading-tight">
              <span class="text-sm font-semibold tracking-[0.16em] uppercase text-white/70">SISA HQ</span>
              <span class="text-xs text-white/80">Global Agency & Settlement Admin</span>
            </div>
          </div>
          <div class="flex items-center gap-3 flex-wrap">
            <div class="text-[11px] text-white/70">
              대행사 신청 URL:
              <span class="font-mono bg-white/10 px-2 py-1 rounded-full border border-white/20">
                https://worldsisa.com/agency-register.html
              </span>
            </div>
            <a href="{{ url_for('admin') }}" class="px-3 py-1.5 rounded-lg bg-brand-accent text-brand-blue text-sm font-semibold hover:bg-white transition">
              결제페이지
            </a>
            <a href="{{ url_for('debug_paths') }}" class="px-3 py-1.5 rounded-lg bg-yellow-400 text-black text-sm font-semibold hover:bg-yellow-300 transition">
              경로진단
            </a>
            <a href="{{ url_for('logout') }}" class="px-3 py-1.5 rounded-lg bg-white/10 border border-white/20 text-white text-sm font-medium hover:bg-white/20 transition">
              로그아웃
            </a>
          </div>
        </div>
      </header>
      <div id="crawler-refresh-overlay" class="loading-backdrop" aria-hidden="true">
        <div class="loading-card">
          <div class="loading-spinner"></div>
          <div id="crawler-refresh-text" style="font-size:13px;font-weight:600;">크롤러 작업 진행중...</div>
          <div style="margin-top:4px;font-size:11px;color:#94a3b8;">결과를 확인하면 자동으로 종료됩니다.</div>
        </div>
      </div>
      <main class="flex-grow pt-20 pb-10 px-3 sm:px-4">
        <div class="max-w-[96vw] mx-auto space-y-8">
          {% if history_warnings.warn_7_days or history_warnings.warn_3_days %}
          <script>
            window.addEventListener('load', function () {
              var msg = "";
              {% if history_warnings.warn_7_days %}
              msg += "7일 뒤 삭제 예정 파일: {{ history_warnings.warn_7_days|join(', ') }}\n";
              {% endif %}
              {% if history_warnings.warn_3_days %}
              msg += "3일 뒤 삭제 예정 파일: {{ history_warnings.warn_3_days|join(', ') }}\n";
              {% endif %}
              if (msg) {
                alert("3개월이 지난 세션 JSON 히스토리가 곧 자동 삭제됩니다.\\n다운로드가 필요하면 지금 받아두세요.\\n\\n" + msg);
              }
            });
          </script>
          {% endif %}
          {% if message %}
          <div class="bg-emerald-500/10 border border-emerald-400/40 text-emerald-100 text-sm px-4 py-3 rounded-xl">
            {{ message }}
          </div>
          {% endif %}

          <!-- 0. K-VAN / 매크로 상태 로그 뷰어 -->
          <section class="glass-card rounded-2xl border border-white/20 shadow-xl p-4">
            <div class="flex items-center justify-between mb-2">
              <h2 class="text-sm font-semibold flex items-center gap-2">
                <i class="fa-solid fa-terminal text-brand-accent"></i> K-VAN 크롤링 & 자동결제 로그
              </h2>
              <div class="flex items-center gap-2 text-[10px]">
                <span class="text-white/50">최근 {{ admin_logs|length }}줄</span>
                <form method="post" action="{{ url_for('hq_admin') }}" onsubmit="return confirm('로그 파일을 정말 삭제하시겠습니까?');">
                  <input type="hidden" name="action" value="clear_logs">
                  <button type="submit"
                          class="px-2 py-1 rounded-full bg-white/10 border border-white/30 text-white hover:bg-white/25">
                    로그 삭제
                  </button>
                </form>
              </div>
            </div>
            <p class="text-[10px] text-white/45 mb-2">로그 파일 경로: {{ admin_log_path }}</p>
            {% if admin_logs %}
            <div class="bg-black/40 rounded-xl border border-white/10 p-3 max-h-56 overflow-y-auto text-[11px] font-mono text-white/80 whitespace-pre-wrap">
              {% for line in admin_logs %}
              <div class="leading-tight">{{ line }}</div>
              {% endfor %}
            </div>
            {% else %}
            <p class="text-[11px] text-white/60">아직 기록된 K-VAN/매크로 로그가 없습니다.</p>
            {% endif %}
          </section>

          <!-- 만료되었으나 거래 내역 있음 (크롤러 저장 → 어드민 알림) -->
          <section class="glass-card rounded-2xl border border-white/20 shadow-xl p-5">
            <div class="flex items-center justify-between mb-2">
              <h2 class="text-sm font-semibold flex items-center gap-2">
                <i class="fa-solid fa-link-slash text-amber-400"></i> 만료된 결제 링크 (거래 내역 있음)
              </h2>
              {% if expired_with_transactions %}
              <form method="post" action="{{ url_for('hq_admin') }}">
                <input type="hidden" name="action" value="mark_expired_with_transactions_seen" />
                <button type="submit" class="px-2 py-1 rounded bg-amber-600/80 text-white text-xs hover:bg-amber-600">전체 확인</button>
              </form>
              {% endif %}
            </div>
            <p class="text-[10px] text-white/50 mb-2">만료되었지만 결제가 발생한 링크 목록입니다. DB에 저장되어 있으며 정산 대상에 포함됩니다.</p>
            {% if expired_with_transactions %}
            <div class="overflow-x-auto max-h-48 overflow-y-auto">
              <table class="min-w-full text-sm border-separate border-spacing-y-1">
                <thead class="text-xs text-white/70 sticky top-0 bg-brand-blue">
                  <tr>
                    <th class="px-2 py-1 text-left">세션 ID</th>
                    <th class="px-2 py-1 text-left">제목</th>
                    <th class="px-2 py-1 text-left">대행사 ID</th>
                    <th class="px-2 py-1 text-left">만료 시각</th>
                    <th class="px-2 py-1 text-center">확인</th>
                  </tr>
                </thead>
                <tbody>
                  {% for e in expired_with_transactions_reversed %}
                  <tr class="bg-black/20 hover:bg-black/30 text-[11px] {{ 'opacity-70' if e.seen else '' }}">
                    <td class="px-2 py-1 font-mono text-blue-200">{{ e.session_id or '' }}</td>
                    <td class="px-2 py-1 max-w-[200px] truncate" title="{{ e.title or '' }}">{{ e.title or '-' }}</td>
                    <td class="px-2 py-1 text-white/80">{{ e.agency_id or '본사' }}</td>
                    <td class="px-2 py-1 text-white/60">{{ e.finished_at or '' }}</td>
                    <td class="px-2 py-1 text-center">{{ '확인함' if e.seen else '미확인' }}</td>
                  </tr>
                  {% endfor %}
                </tbody>
              </table>
            </div>
            {% if expired_with_tx_unseen and expired_with_tx_unseen > 0 %}
            <p class="mt-2 text-amber-200 text-xs"><i class="fa-solid fa-bell mr-1"></i> 미확인 {{ expired_with_tx_unseen }}건</p>
            {% endif %}
            {% else %}
            <p class="text-[11px] text-white/60">해당 목록이 없습니다.</p>
            {% endif %}
          </section>

          <!-- 1. 대행사 신청 현황 -->
          <section class="glass-card rounded-2xl border border-white/20 shadow-xl p-5">
            <div class="flex items-center justify-between mb-3">
              <h2 class="text-lg font-semibold flex items-center gap-2">
                <i class="fa-solid fa-file-pen text-brand-accent"></i> 대행사 신청 현황
              </h2>
              <p class="text-[11px] text-white/60">신청서 양식과 동일한 정보가 리스트로 표시됩니다.</p>
            </div>
            <div class="box-schema"><code>applications</code> 항목: <code>created_at, company_name, domain, phone, bank_name, account_number, email_or_sheet, login_id, login_password, fee_percent</code></div>
            {% if applications %}
            <div class="overflow-x-auto">
              <table class="min-w-full text-sm border-separate border-spacing-y-2">
                <thead class="text-xs text-white/70">
                  <tr>
                    <th class="px-3 py-1 text-left">신청일</th>
                    <th class="px-3 py-1 text-left">업체명</th>
                    <th class="px-3 py-1 text-left">도메인(영문)</th>
                    <th class="px-3 py-1 text-left">전화번호</th>
                    <th class="px-3 py-1 text-left">은행/계좌</th>
                    <th class="px-3 py-1 text-left">이메일/구글시트</th>
                    <th class="px-3 py-1 text-left">아이디</th>
                    <th class="px-3 py-1 text-left">비밀번호</th>
                    <th class="px-3 py-1 text-center">수수료%</th>
                    <th class="px-3 py-1 text-center">수수료 저장</th>
                    <th class="px-3 py-1 text-center">승인 및 생성</th>
                    <th class="px-3 py-1 text-center">삭제</th>
                  </tr>
                </thead>
                <tbody>
                  {% for a in applications %}
                  <tr class="bg-black/20 hover:bg-black/30 transition">
                    <td class="px-3 py-2 text-[11px] text-white/70">{{ a.created_at or '' }}</td>
                    <td class="px-3 py-2 font-semibold">{{ a.company_name }}</td>
                    <td class="px-3 py-2 text-[11px] text-white/80">{{ a.domain }}</td>
                    <td class="px-3 py-2 text-[11px] text-white/80">{{ a.phone }}</td>
                    <td class="px-3 py-2 text-[11px] text-white/80">{{ a.bank_name }} / {{ a.account_number }}</td>
                    <td class="px-3 py-2 text-[11px] text-white/70 max-w-[160px] truncate">{{ a.email_or_sheet }}</td>
                    <td class="px-3 py-2 text-[11px] font-mono text-blue-200">{{ a.login_id }}</td>
                    <td class="px-3 py-2 text-[11px] text-white/60">••••••</td>
                    <td class="px-3 py-2 text-center text-[11px]">
                      {{ a.fee_percent or 10 }}%
                    </td>
                    <td class="px-3 py-2 text-center text-[11px]">
                      <form method="post" action="{{ url_for('hq_admin') }}" class="inline-flex items-center gap-1">
                        <input type="hidden" name="action" value="update_application_fee" />
                        <input type="hidden" name="application_id" value="{{ a.id }}" />
                        <input type="number" name="fee_percent" value="{{ a.fee_percent or 10 }}" min="0" max="100"
                               class="w-12 bg-black/40 border border-white/20 rounded px-1 py-0.5 text-[11px] text-center">
                        <span>%</span>
                        <button type="submit" class="text-[10px] px-2 py-0.5 rounded-full bg-white/10 hover:bg-white/20">
                          저장
                        </button>
                      </form>
                    </td>
                    <td class="px-3 py-2 text-center">
                      <form method="post" action="{{ url_for('hq_admin') }}">
                        <input type="hidden" name="action" value="approve_application" />
                        <input type="hidden" name="application_id" value="{{ a.id }}" />
                        <button type="submit" class="px-3 py-1 rounded-full bg-brand-accent text-brand-blue text-[11px] font-semibold hover:bg-white transition">
                          승인 및 생성
                        </button>
                      </form>
                    </td>
                    <td class="px-3 py-2 text-center">
                      <form method="post" action="{{ url_for('hq_admin') }}" onsubmit="return confirm('해당 대행사 신청을 삭제하시겠습니까?');">
                        <input type="hidden" name="action" value="delete_application" />
                        <input type="hidden" name="application_id" value="{{ a.id }}" />
                        <button type="submit" class="px-3 py-1 rounded-full bg-red-500/30 text-red-100 text-[11px] hover:bg-red-500/50">
                          삭제
                        </button>
                      </form>
                    </td>
                  </tr>
                  {% endfor %}
                </tbody>
              </table>
            </div>
            {% else %}
              <p class="text-xs text-white/60">접수된 대행사 신청이 아직 없습니다.</p>
            {% endif %}
          </section>

          <!-- 2. 전체 거래 내역 리스트 -->
          <section class="glass-card rounded-2xl border border-white/20 shadow-xl p-5">
            {% if payment_notifications_count and payment_notifications_count > 0 %}
            <div class="mb-3 p-3 rounded-lg bg-amber-500/20 border border-amber-400/40 flex items-center justify-between gap-2 flex-wrap">
              <span class="text-amber-200 text-sm">
                <i class="fa-solid fa-bell mr-1"></i> 미확인 결제 완료 알림 {{ payment_notifications_count }}건
              </span>
              <form method="post" action="{{ url_for('hq_admin') }}">
                <input type="hidden" name="action" value="mark_payment_notifications_seen" />
                <button type="submit" class="px-2 py-1 rounded bg-amber-600/80 text-white text-xs hover:bg-amber-600">확인</button>
              </form>
            </div>
            {% endif %}
            <div class="flex flex-col md:flex-row items-start md:items-center justify-between mb-3 gap-2">
              <div>
                <h2 class="text-lg font-semibold flex items-center gap-2">
                  <i class="fa-solid fa-list-ul text-brand-accent"></i> 전체 거래 내역
                </h2>
                <p class="text-[11px] text-white/60 hidden sm:block">시간순으로 성사된 주문 결제 건을 확인하고, 정산 상태를 관리합니다.</p>
              </div>
              <div class="flex flex-wrap items-center gap-2 text-[11px]">
                <form method="post" action="{{ url_for('hq_admin') }}">
                  <input type="hidden" name="action" value="refresh_kvan" />
                  <button type="submit"
                          class="px-3 py-1 rounded-full bg-white/10 border border-white/30 text-white hover:bg-white/25 flex items-center gap-1">
                    <i class="fa-solid fa-rotate-right text-xs"></i>
                    <span>새로고침</span>
                  </button>
                </form>
                <div class="flex items-center gap-1">
                  <span class="text-white/70">업체:</span>
                  <select id="txAgencyFilter" onchange="filterTransactions()" class="bg-black/30 border border-white/30 rounded px-2 py-1 text-[11px]">
                    <option value="all">전체</option>
                    {% for ag in agencies %}
                    <option value="{{ ag.id }}">{{ ag.company_name }}</option>
                    {% endfor %}
                  </select>
                </div>
                <div class="flex items-center gap-1">
                  <span class="text-white/70">날짜:</span>
                  <input id="txStartDate" type="date" value="{{ today_str }}" onchange="filterTransactions()" class="bg-black/30 border border-white/30 rounded px-2 py-1 text-[11px]" />
                  <span class="text-white/50">~</span>
                  <input id="txEndDate" type="date" value="{{ today_str }}" onchange="filterTransactions()" class="bg-black/30 border border-white/30 rounded px-2 py-1 text-[11px]" />
                </div>
                <div class="flex items-center gap-1">
                  <span class="text-white/70">상태:</span>
                  <select id="txStatusFilter" onchange="filterTransactions()" class="bg-black/30 border border-white/30 rounded px-2 py-1 text-[11px]">
                    <option value="all">전체</option>
                    <option value="success">성공</option>
                    <option value="fail">실패</option>
                    <option value="other">기타</option>
                  </select>
                </div>
                <a href="{{ url_for('hq_export_excel', scope='transactions') }}"
                   class="ml-auto px-3 py-1 rounded-full bg-white/10 border border-white/30 text-white hover:bg-white/25">
                  엑셀
                </a>
              </div>
            </div>
            <div class="box-schema"><code>transactions</code> 항목: <code>created_at, agency_id, amount, customer_name, status, settlement_status, message, card_type, resident_front, phone_number</code></div>
            {% if transactions %}
            <form method="post" action="{{ url_for('hq_admin') }}" class="space-y-3" onsubmit="return confirm('선택한 거래 내역 처리(삭제/정산완료)를 진행할까요?');">
              <div class="overflow-x-auto">
                <table class="min-w-full text-xs border-separate border-spacing-y-2">
                  <thead class="text-white/70">
                    <tr>
                      <th class="px-3 py-1 text-center"><input type="checkbox" id="tx_check_all" onclick="
                        var cbs = document.querySelectorAll('.tx-check'); 
                        cbs.forEach(function(cb){ cb.checked = this.checked; }.bind(this));
                      "></th>
                      <th class="px-3 py-1 text-left">시간</th>
                      <th class="px-3 py-1 text-left">대행사</th>
                      <th class="px-3 py-1 text-right">금액</th>
                      <th class="px-3 py-1 text-left">구매자</th>
                      <th class="px-3 py-1 text-center">결제상태</th>
                      <th class="px-3 py-1 text-center">정산상태</th>
                    </tr>
                  </thead>
                  <tbody>
                    {% set unsettled_total = 0 %}
                    {% for t in transactions|sort(attribute="created_at", reverse=True) %}
                    {% set ag_name = "" %}
                    {% set ag_fee = 0 %}
                    {% for ag in agencies %}
                      {% if ag.id == t.agency_id %}
                        {% set ag_name = ag.company_name %}
                        {% set ag_fee = ag.fee_percent or 0 %}
                      {% endif %}
                    {% endfor %}
                    {% if not ag_name %}
                      {% set ag_name = "본사" %}
                    {% endif %}
                    {% if t.status == 'success' and t.settlement_status != '정산완료' %}
                      {% set unsettled_total = unsettled_total + (t.amount or 0) %}
                    {% endif %}
                    {% set amount = t.amount or 0 %}
                    <tr class="bg-black/20 hover:bg-black/30 transition align-top"
                        data-tx-row="1"
                        data-agency-id="{{ t.agency_id or '' }}"
                        data-amount="{{ amount }}"
                        data-fee-percent="{{ ag_fee }}"
                        data-date="{{ t.created_at.strftime('%Y-%m-%d') if t.created_at else '' }}"
                        data-status="{{ t.status or '' }}">
                      <td class="px-3 py-2 text-center">
                        <input type="checkbox" class="tx-check" name="tx_ids" value="{{ t.id }}" onclick="updateSelectionSummary()">
                      </td>
                      <td class="px-3 py-2 whitespace-nowrap">{{ t.created_at }}</td>
                      <td class="px-3 py-2 whitespace-nowrap">{{ ag_name }}</td>
                      <td class="px-3 py-2 text-right">{{ amount }} 원</td>
                      <td class="px-3 py-2 whitespace-nowrap">{{ t.customer_name }}</td>
                      <td class="px-3 py-2 text-center">
                        {% if t.status == 'success' %}
                          <span class="px-2 py-1 rounded-full bg-emerald-500/20 text-emerald-200 border border-emerald-500/40 text-[10px]">성공</span>
                        {% elif t.status == 'fail' %}
                          <span class="px-2 py-1 rounded-full bg-red-500/20 text-red-200 border border-red-500/40 text-[10px]">실패</span>
                        {% else %}
                          <span class="px-2 py-1 rounded-full bg-gray-500/20 text-gray-200 border border-gray-500/40 text-[10px]">기타</span>
                        {% endif %}
                      </td>
                      <td class="px-3 py-2 text-center">
                        {% if t.settlement_status == '정산완료' %}
                          <span class="px-2 py-1 rounded-full bg-blue-500/20 text-blue-200 border border-blue-500/40 text-[10px]">정산완료</span>
                        {% else %}
                          <span class="px-2 py-1 rounded-full bg-yellow-500/20 text-yellow-200 border border-yellow-500/40 text-[10px]">미정산</span>
                        {% endif %}
                      </td>
                    </tr>
                    <tr class="bg-black/10" data-tx-detail="1">
                      <td></td>
                      <td colspan="6" class="px-3 pb-3 text-[11px] text-white/70">
                        <div class="flex flex-wrap gap-3">
                          <span><strong>카드구분:</strong> {{ t.card_type }}</span>
                          <span><strong>생년월일(앞 6자리):</strong> {{ t.resident_front }}</span>
                          <span><strong>전화번호(뒷자리):</strong> {{ t.phone_number }}</span>
                          {% if t.message %}
                          <span class="block w-full"><strong>메모:</strong> {{ t.message }}</span>
                          {% endif %}
                          {% if amount == 0 or t.status != 'success' %}
                          <form method="post" action="{{ url_for('hq_admin') }}" class="inline-block ml-auto">
                            <input type="hidden" name="action" value="delete_tx">
                            <input type="hidden" name="tx_id" value="{{ t.id }}">
                            <button type="submit" class="px-2 py-1 rounded-full bg-red-500/30 text-red-100 border border-red-400/60 text-[10px] hover:bg-red-500/50">
                              거래 내역 삭제
                            </button>
                          </form>
                          {% endif %}
                        </div>
                      </td>
                    </tr>
                    {% endfor %}
                  </tbody>
                </table>
              </div>
              <div class="flex flex-col sm:flex-row sm:items-center justify-between mt-3 text-[11px] text-white/80 gap-2">
                <div class="space-y-1">
                  <div>
                    미정산 총 합계 금액:
                    <span class="font-semibold text-brand-accent">{{ unsettled_total }} 원</span>
                  </div>
                  <div>
                    선택 건 현황:
                    총 거래금액 <span id="selTotalAmount" class="font-semibold text-brand-accent">0 원</span>,
                    미정산 금액 <span id="selUnsettledAmount" class="font-semibold text-yellow-200">0 원</span>,
                    입금 예정액 <span id="selNetAmount" class="font-semibold text-emerald-200">0 원</span>
                  </div>
                </div>
                <div class="flex items-center gap-2">
                  <span>선택 건을</span>
                  <button type="submit" name="action" value="bulk_delete_tx"
                          class="px-3 py-1 rounded-full bg-red-500/40 text-red-100 font-semibold hover:bg-red-500/60 transition">
                    삭제
                  </button>
                  <button type="submit" name="action" value="bulk_settle"
                          class="px-3 py-1 rounded-full bg-brand-accent text-brand-blue font-semibold hover:bg-white transition">
                    정산완료 처리
                  </button>
                </div>
              </div>
            </form>
            {% else %}
              <p class="text-xs text-white/60">아직 집계된 거래 내역이 없습니다.</p>
            {% endif %}
            <div class="mt-6 pt-4 border-t border-white/10">
              <h3 class="text-sm font-semibold mb-2 text-white/90">거래 내역 (엑셀형 전체 컬럼)</h3>
              <p class="text-[11px] text-white/50 mb-2">DB transactions 테이블의 모든 컬럼을 리스트로 표시합니다. 가로 스크롤 가능.</p>
              {% if transactions %}
              <div class="overflow-x-auto max-h-[420px] overflow-y-auto border border-white/20 rounded-xl">
                <table class="min-w-max text-[11px] border-collapse">
                  <thead class="text-white/80 bg-black/40 sticky top-0 z-10">
                    <tr>
                      {% for col in tx_excel_columns %}
                      <th class="px-2 py-1.5 text-left whitespace-nowrap border-b border-r border-white/20 font-semibold">{{ col }}</th>
                      {% endfor %}
                    </tr>
                  </thead>
                  <tbody>
                    {% for t in transactions|sort(attribute="created_at", reverse=True) %}
                    <tr class="bg-black/20 hover:bg-black/30 border-b border-white/10">
                      {% for col in tx_excel_columns %}
                      <td class="px-2 py-1.5 whitespace-nowrap border-r border-white/10 text-white/90">
                        {% set val = t.get(col) %}
                        {% if col == 'amount' and val is not none and val != '' %}{{ val }} 원{% elif val is not none and val != '' %}{{ val }}{% else %}-{% endif %}
                      </td>
                      {% endfor %}
                    </tr>
                    {% endfor %}
                  </tbody>
                </table>
              </div>
              {% else %}
              <p class="text-xs text-white/60">표시할 거래가 없습니다.</p>
              {% endif %}
            </div>
          </section>

          <!-- 3. 대행사별 거래 내역 및 정산 시스템 (요약) -->
          <section class="glass-card rounded-2xl border border-white/20 shadow-xl p-5">
            <div class="flex items-center justify-between mb-3">
              <h2 class="text-lg font-semibold flex items-center gap-2">
                <i class="fa-solid fa-building text-brand-accent"></i> 대행사별 거래 내역 및 정산
              </h2>
              <div class="flex items-center gap-2 text-[11px]">
                <p class="text-white/60 hidden sm:block">업체별 수수료 % 설정과 미정산/정산완료 금액을 확인합니다. (수정은 아래 대행사 관리 박스에서 가능합니다.)</p>
                <a href="{{ url_for('hq_export_excel', scope='agency_summary') }}"
                   class="px-3 py-1 rounded-full bg-white/10 border border-white/30 text-white hover:bg-white/25">
                  엑셀
                </a>
              </div>
            </div>
            <div class="box-schema"><code>agencies + transactions(집계)</code> 항목: <code>company_name, domain, login_id, fee_percent, total_amount, unsettled_amount, net_amount, status</code></div>
            {% if agencies %}
            <div class="overflow-x-auto">
              <table class="min-w-full text-sm border-separate border-spacing-y-2">
                <thead class="text-xs text-white/70">
                  <tr>
                    <th class="px-3 py-1 text-left">업체명</th>
                    <th class="px-3 py-1 text-left">도메인</th>
                    <th class="px-3 py-1 text-left">아이디</th>
                    <th class="px-3 py-1 text-center">수수료%</th>
                    <th class="px-3 py-1 text-right">총 거래금액</th>
                    <th class="px-3 py-1 text-right">미정산 금액</th>
                    <th class="px-3 py-1 text-right">입금 예정액</th>
                    <th class="px-3 py-1 text-center">상태</th>
                  </tr>
                </thead>
                <tbody>
                  {% for ag in agencies %}
                  {% set total_amount = 0 %}
                  {% set unsettled_amount = 0 %}
                  {% for t in transactions %}
                    {% set amt = t.amount or 0 %}
                    {% if t.agency_id == ag.id and t.status == 'success' and amt > 0 %}
                      {% set total_amount = total_amount + amt %}
                      {% if t.settlement_status != '정산완료' %}
                        {% set unsettled_amount = unsettled_amount + amt %}
                      {% endif %}
                    {% endif %}
                  {% endfor %}
                  {% set net_amount = unsettled_amount * (100 - (ag.fee_percent or 0)) // 100 %}
                  <tr class="bg-black/20 hover:bg-black/30 transition">
                    <td class="px-3 py-2 font-semibold">{{ ag.company_name }}</td>
                    <td class="px-3 py-2 text-[11px] text-white/80">{{ ag.domain }}</td>
                    <td class="px-3 py-2 text-[11px] font-mono text-blue-200">{{ ag.login_id }}</td>
                    <td class="px-3 py-2 text-center text-[11px] text-white/80">
                      {{ ag.fee_percent or 0 }}%
                    </td>
                    <td class="px-3 py-2 text-right text-[11px] text-white/80">{{ total_amount }} 원</td>
                    <td class="px-3 py-2 text-right text-[11px] text-yellow-200">{{ unsettled_amount }} 원</td>
                    <td class="px-3 py-2 text-right text-[11px] text-emerald-200">{{ net_amount }} 원</td>
                    <td class="px-3 py-2 text-center text-[11px]">
                      {% if ag.status == 'active' %}
                        <span class="px-2 py-1 rounded-full bg-emerald-500/20 text-emerald-200 border border-emerald-500/40 text-[10px]">활성</span>
                      {% else %}
                        <span class="px-2 py-1 rounded-full bg-gray-500/20 text-gray-200 border border-gray-500/40 text-[10px]">중지</span>
                      {% endif %}
                    </td>
                  </tr>
                  {% endfor %}
                </tbody>
              </table>
            </div>
            {% else %}
              <p class="text-xs text-white/60">아직 승인된 대행사가 없습니다.</p>
            {% endif %}
          </section>

          <!-- 4. 대행사 관리 박스 (정보 수정 및 개별 정산) -->
          <section class="glass-card rounded-2xl border border-white/20 shadow-xl p-5">
            <div class="flex items-center justify-between mb-3">
              <h2 class="text-lg font-semibold flex items-center gap-2">
                <i class="fa-solid fa-user-gear text-brand-accent"></i> 대행사 관리
              </h2>
              <div class="flex items-center gap-2 text-[11px]">
                <p class="text-white/60 hidden sm:block">대행사 정보, 수수료 %, 미정산 금액을 확인하고 수정/정산할 수 있습니다.</p>
                <a href="{{ url_for('hq_export_excel', scope='agency_manage') }}"
                   class="px-3 py-1 rounded-full bg-white/10 border border-white/30 text-white hover:bg-white/25">
                  엑셀
                </a>
              </div>
            </div>
            <div class="box-schema"><code>agencies + transactions(관리/정산)</code> 항목: <code>company_name, phone, bank_name, account_number, email_or_sheet, fee_percent, status, kvan_mid, kvan_login_id, kvan_login_password, kvan_login_pin</code></div>
            {% if paged_agencies %}
            <div class="overflow-x-auto">
              <table class="min-w-full text-sm border-separate border-spacing-y-2">
                <thead class="text-xs text-white/70">
                  <tr>
                    <th class="px-3 py-1 text-left">업체명</th>
                    <th class="px-3 py-1 text-left">연락처</th>
                    <th class="px-3 py-1 text-left">은행/계좌</th>
                    <th class="px-3 py-1 text-left">이메일/구글시트</th>
                    <th class="px-3 py-1 text-center">수수료%</th>
                    <th class="px-3 py-1 text-right">총 거래금액</th>
                    <th class="px-3 py-1 text-right">미정산 금액</th>
                    <th class="px-3 py-1 text-right">입금 예정액</th>
                    <th class="px-3 py-1 text-center">상태</th>
                    <th class="px-3 py-1 text-center">관리</th>
                  </tr>
                </thead>
                <tbody>
                  {% for ag in paged_agencies %}
                  {% set total_amount = 0 %}
                  {% set unsettled_amount = 0 %}
                  {% for t in transactions %}
                    {% set amt = t.amount or 0 %}
                    {% if t.agency_id == ag.id and t.status == 'success' and amt > 0 %}
                      {% set total_amount = total_amount + amt %}
                      {% if t.settlement_status != '정산완료' %}
                        {% set unsettled_amount = unsettled_amount + amt %}
                      {% endif %}
                    {% endif %}
                  {% endfor %}
                  {% set net_amount = unsettled_amount * (100 - (ag.fee_percent or 0)) // 100 %}
                  <tr class="bg-black/20 hover:bg-black/30 transition align-top">
                    <td class="px-3 py-2 font-semibold whitespace-nowrap">{{ ag.company_name }}</td>
                    <td class="px-3 py-2 text-[11px] text-white/80 whitespace-nowrap">{{ ag.phone }}</td>
                    <td class="px-3 py-2 text-[11px] text-white/80 whitespace-nowrap">{{ ag.bank_name }} / {{ ag.account_number }}</td>
                    <td class="px-3 py-2 text-[11px] text-white/80 max-w-[160px] truncate">{{ ag.email_or_sheet }}</td>
                    <td class="px-3 py-2 text-center text-[11px] text-white/80">{{ ag.fee_percent or 0 }}%</td>
                    <td class="px-3 py-2 text-right text-[11px] text-white/80">{{ total_amount }} 원</td>
                    <td class="px-3 py-2 text-right text-[11px] text-yellow-200">{{ unsettled_amount }} 원</td>
                    <td class="px-3 py-2 text-right text-[11px] text-emerald-200">{{ net_amount }} 원</td>
                    <td class="px-3 py-2 text-center text-[11px]">
                      {% if ag.status == 'active' %}
                        <span class="px-2 py-1 rounded-full bg-emerald-500/20 text-emerald-200 border border-emerald-500/40 text-[10px]">활성</span>
                      {% else %}
                        <span class="px-2 py-1 rounded-full bg-gray-500/20 text-gray-200 border border-gray-500/40 text-[10px]">중지</span>
                      {% endif %}
                    </td>
                    <td class="px-3 py-2 text-center text-[11px]">
                      <details class="inline-block text-left">
                        <summary class="cursor-pointer text-brand-accent underline">관리</summary>
                        <div class="mt-2 bg-black/40 border border-white/20 rounded-xl p-2 w-64">
                          <form method="post" action="{{ url_for('hq_admin') }}" class="space-y-1">
                            <input type="hidden" name="action" value="update_agency">
                            <input type="hidden" name="agency_id" value="{{ ag.id }}">
                            <div class="flex flex-col gap-1">
                              <input type="text" name="login_id" value="{{ ag.login_id }}" placeholder="로그인 아이디"
                                     class="bg-black/40 border border-white/20 rounded px-2 py-0.5 text-[11px]">
                              <input type="text" name="login_password" value="{{ ag.login_password }}" placeholder="로그인 비밀번호"
                                     class="bg-black/40 border border-white/20 rounded px-2 py-0.5 text-[11px]">
                              <input type="text" name="kvan_mid" value="{{ ag.kvan_mid or '' }}" placeholder="K-VAN MID"
                                     class="bg-black/40 border border-white/20 rounded px-2 py-0.5 text-[11px]">
                              <input type="text" name="kvan_login_id" value="{{ ag.kvan_login_id or '' }}" placeholder="K-VAN 아이디"
                                     class="bg-black/40 border border-white/20 rounded px-2 py-0.5 text-[11px]">
                              <input type="text" name="kvan_login_password" value="{{ ag.kvan_login_password or '' }}" placeholder="K-VAN 비밀번호"
                                     class="bg-black/40 border border-white/20 rounded px-2 py-0.5 text-[11px]">
                              <input type="text" name="kvan_login_pin" value="{{ ag.kvan_login_pin or '' }}" placeholder="K-VAN PIN"
                                     class="bg-black/40 border border-white/20 rounded px-2 py-0.5 text-[11px]">
                              <input type="text" name="phone" value="{{ ag.phone }}" placeholder="전화번호"
                                     class="bg-black/40 border border-white/20 rounded px-2 py-0.5 text-[11px]">
                              <input type="text" name="bank_name" value="{{ ag.bank_name }}" placeholder="은행명"
                                     class="bg-black/40 border border-white/20 rounded px-2 py-0.5 text-[11px]">
                              <input type="text" name="account_number" value="{{ ag.account_number }}" placeholder="계좌번호"
                                     class="bg-black/40 border border-white/20 rounded px-2 py-0.5 text-[11px]">
                              <input type="text" name="email_or_sheet" value="{{ ag.email_or_sheet }}" placeholder="이메일/구글시트"
                                     class="bg-black/40 border border-white/20 rounded px-2 py-0.5 text-[11px]">
                              <div class="flex items-center gap-1">
                                <input type="number" name="fee_percent" value="{{ ag.fee_percent }}" min="0" max="100"
                                       class="w-12 bg-black/40 border border-white/20 rounded px-1 py-0.5 text-[11px] text-center">
                                <span>%</span>
                                <select name="status" class="bg-black/40 border border-white/20 rounded px-1 py-0.5 text-[11px]">
                                  <option value="active" {% if ag.status == 'active' %}selected{% endif %}>활성</option>
                                  <option value="paused" {% if ag.status != 'active' %}selected{% endif %}>중지</option>
                                </select>
                              </div>
                              <div class="flex items-center justify-center gap-1 pt-1">
                                <button type="submit" name="do" value="save"
                                        class="px-2 py-1 rounded-full bg-white/10 hover:bg-white/20 text-white text-[10px]">
                                  정보 저장
                                </button>
                                <button type="submit" name="do" value="settle"
                                        class="px-2 py-1 rounded-full bg-emerald-500/30 hover:bg-emerald-500/50 text-emerald-50 text-[10px]">
                                  미정산 정산완료
                                </button>
                              </div>
                            </div>
                          </form>
                        </div>
                      </details>
                    </td>
                  </tr>
                  {% endfor %}
                </tbody>
              </table>
            </div>
            <div class="mt-3 flex justify-center gap-1 text-[11px]">
              {% for p in range(1, total_pages + 1) %}
                {% if p == page %}
                  <span class="px-2 py-1 rounded-full bg-white/80 text-brand-blue font-semibold">{{ p }}</span>
                {% else %}
                  <a href="{{ url_for('hq_admin', page=p) }}" class="px-2 py-1 rounded-full bg-white/10 text-white hover:bg-white/30">{{ p }}</a>
                {% endif %}
              {% endfor %}
            </div>
            {% else %}
              <p class="text-xs text-white/60">아직 승인된 대행사가 없습니다.</p>
            {% endif %}
          </section>

          <!-- 5. K-VAN 대시보드 요약 (공유용) -->
          <section class="glass-card rounded-2xl border border-white/20 shadow-xl p-5">
            <div class="flex items-center justify-between mb-3">
              <h2 class="text-lg font-semibold flex items-center gap-2">
                <i class="fa-solid fa-chart-line text-brand-accent"></i> K-VAN 대시보드 요약
              </h2>
              <p class="text-[11px] text-white/60">자동 결제 매크로가 수집한 OneQue 대시보드 요약 정보를 공유합니다.</p>
            </div>
            <div class="box-schema"><code>kvan_dashboard</code> 항목: <code>captured_at, monthly_sales_amount, monthly_approved_count, monthly_approved_amount, monthly_canceled_count, monthly_canceled_amount, yesterday_sales_amount, settlement_expected_amount, today_settlement_expected_amount, credit_amount</code></div>
            {% if latest_dashboard %}
            <div class="grid grid-cols-1 md:grid-cols-3 gap-4 text-xs">
              <div class="bg-black/30 rounded-xl border border-white/10 p-4 space-y-2">
                <div class="text-[11px] text-white/70 mb-1">
                  <span class="font-semibold">월 매출</span>
                  <span class="ml-2 text-white/50">({{ latest_dashboard.captured_at }})</span>
                </div>
                <p class="text-lg font-bold text-brand-accent">
                  {{ latest_dashboard.monthly_sales_amount or 0 }} 원
                </p>
                <div class="mt-2 space-y-1">
                  <p class="text-white/70">
                    승인:
                    <span class="font-semibold text-emerald-300">
                      {{ latest_dashboard.monthly_approved_count or 0 }}건 /
                      {{ latest_dashboard.monthly_approved_amount or 0 }} 원
                    </span>
                  </p>
                  <p class="text-white/70">
                    취소:
                    <span class="font-semibold text-red-300">
                      {{ latest_dashboard.monthly_canceled_count or 0 }}건 /
                      {{ latest_dashboard.monthly_canceled_amount or 0 }} 원
                    </span>
                  </p>
                </div>
              </div>
              <div class="bg-black/30 rounded-xl border border-white/10 p-4 space-y-2">
                <div class="text-[11px] text-white/70 mb-1">
                  <span class="font-semibold">전일 매출</span>
                </div>
                <p class="text-lg font-bold text-blue-300">
                  {{ latest_dashboard.yesterday_sales_amount or 0 }} 원
                </p>
                <div class="mt-2 space-y-1">
                  <p class="text-white/70">
                    승인:
                    <span class="font-semibold text-emerald-300">
                      {{ latest_dashboard.yesterday_approved_count or 0 }}건 /
                      {{ latest_dashboard.yesterday_approved_amount or 0 }} 원
                    </span>
                  </p>
                  <p class="text-white/70">
                    취소:
                    <span class="font-semibold text-red-300">
                      {{ latest_dashboard.yesterday_canceled_count or 0 }}건 /
                      {{ latest_dashboard.yesterday_canceled_amount or 0 }} 원
                    </span>
                  </p>
                </div>
              </div>
              <div class="bg-black/30 rounded-xl border border-white/10 p-4 space-y-2">
                <div class="text-[11px] text-white/70 mb-1">
                  <span class="font-semibold">정산 예정 및 크레딧</span>
                </div>
                <p class="text-[11px] text-white/70">
                  정산 예정 금액:
                  <span class="font-semibold text-brand-accent">
                    {{ latest_dashboard.settlement_expected_amount or 0 }} 원
                  </span>
                </p>
                <p class="text-[11px] text-white/70">
                  금일 정산 예정금:
                  <span class="font-semibold text-emerald-300">
                    {{ latest_dashboard.today_settlement_expected_amount or 0 }} 원
                  </span>
                </p>
                <p class="text-[11px] text-white/70">
                  나의 크레딧:
                  <span class="font-semibold text-blue-300">
                    {{ latest_dashboard.credit_amount or 0 }} 원
                  </span>
                </p>
                <div class="mt-2 p-2 bg-black/40 rounded-lg border border-white/10 max-h-24 overflow-y-auto">
                  <p class="text-[10px] text-white/60 whitespace-pre-line">
                    {{ latest_dashboard.recent_tx_summary or "최근 거래 내역 정보가 없습니다." }}
                  </p>
                </div>
              </div>
            </div>
            {% else %}
              <p class="text-xs text-white/60">아직 수집된 K-VAN 대시보드 데이터가 없습니다. 매크로가 한 번 이상 실행되면 자동으로 표시됩니다.</p>
            {% endif %}
          </section>

          <!-- HQ 엑셀 다운로드 -->
          <section class="glass-card rounded-2xl border border-white/20 shadow-xl p-4 flex items-center justify-between text-sm">
            <div class="text-white/70 text-[11px]">
              전체 거래 내역 및 대행사별 정산 정보를 엑셀 파일로 내려받을 수 있습니다.
            </div>
            <a href="{{ url_for('hq_export_excel') }}"
               class="px-4 py-2 rounded-full bg-white text-brand-blue font-semibold text-xs hover:bg-brand-accent transition">
              엑셀 다운받기
            </a>
          </section>
        </div>
      </main>
      <script>
        function filterTransactions() {
          var selAgency = document.getElementById('txAgencyFilter');
          var selStatus = document.getElementById('txStatusFilter');
          var startInput = document.getElementById('txStartDate');
          var endInput = document.getElementById('txEndDate');
          var agencyVal = selAgency ? (selAgency.value || 'all') : 'all';
          var statusVal = selStatus ? (selStatus.value || 'all') : 'all';
          var startDate = startInput && startInput.value ? startInput.value : '';
          var endDate = endInput && endInput.value ? endInput.value : '';

          var rows = document.querySelectorAll('tr[data-tx-row="1"]');
          rows.forEach(function(row) {
            var ag = row.getAttribute('data-agency-id') || '';
            var date = row.getAttribute('data-date') || '';
            var status = (row.getAttribute('data-status') || '').toLowerCase();

            var show = true;
            if (agencyVal !== 'all' && ag !== agencyVal) {
              show = false;
            }
            if (show && startDate && (!date || date < startDate)) {
              show = false;
            }
            if (show && endDate && (!date || date > endDate)) {
              show = false;
            }
            if (show && statusVal !== 'all') {
              if (statusVal === 'other') {
                if (status === 'success' || status === 'fail') show = false;
              } else if (status !== statusVal) {
                show = false;
              }
            }

            row.style.display = show ? '' : 'none';
            var detail = row.nextElementSibling;
            if (detail && detail.getAttribute('data-tx-detail') === '1') {
              detail.style.display = show ? '' : 'none';
            }
          });
          updateSelectionSummary();
        }

        function updateSelectionSummary() {
          var total = 0;
          var unsettled = 0;
          var net = 0;
          var checks = document.querySelectorAll('.tx-check');
          checks.forEach(function(cb) {
            if (cb.checked) {
              var row = cb.closest('tr');
              if (!row || row.style.display === 'none') return;
              var amount = parseInt(row.getAttribute('data-amount') || '0', 10);
              var fee = parseInt(row.getAttribute('data-fee-percent') || '0', 10);
              total += amount;
              unsettled += amount;
              net += Math.floor(amount * (100 - fee) / 100);
            }
          });
          var elTotal = document.getElementById('selTotalAmount');
          var elUn = document.getElementById('selUnsettledAmount');
          var elNet = document.getElementById('selNetAmount');
          if (elTotal) elTotal.textContent = total.toLocaleString('ko-KR') + ' 원';
          if (elUn) elUn.textContent = unsettled.toLocaleString('ko-KR') + ' 원';
          if (elNet) elNet.textContent = net.toLocaleString('ko-KR') + ' 원';
        }
      </script>
    </body>
    </html>
    """
    return render_template_string(
        template,
        applications=applications,
        agencies=agencies,
        transactions=transactions,
        message=message,
        history_warnings=history_warnings,
        admin_logs=admin_logs,
        admin_log_path=str(ADMIN_LOG_PATH),
        payment_notifications_count=payment_notifications_count,
        crawler_refresh_since=crawler_refresh_since,
        expired_with_transactions=expired_with_transactions,
        expired_with_transactions_reversed=expired_with_transactions_reversed,
        expired_with_tx_unseen=expired_with_tx_unseen,
        tx_excel_columns=TX_EXCEL_COLUMNS,
        latest_dashboard=latest_dashboard,
        paged_agencies=paged_agencies,
        page=page,
        total_pages=total_pages,
        today_str=today_str,
    )


@app.route("/agency-admin", methods=["GET", "POST"])
def agency_admin():
    """대행사 전용 결제 세션/거래 대시보드."""
    agency_id = session.get("agency_id")
    if not agency_id:
        return redirect(url_for("agency_login"))

    # 본사에서 저장한 대행사 정보 로드
    state = _load_hq_state()
    agencies = state.get("agencies") or []
    agency = None
    for ag in agencies:
        if str(ag.get("id")) == str(agency_id):
            agency = ag
            break;
    if not agency:
        # 세션에 남아 있지만 HQ 데이터에는 없는 경우 다시 로그인
        session.pop("agency_id", None)
        return redirect(url_for("agency_login"))

    # 세션/히스토리는 admin_state.json 에서 agency_id 기준으로만 필터 (비어있으면 표시 안 함)
    sessions: list[dict] = []
    history: list[dict] = []
    if Path(ADMIN_STATE_PATH).exists():
        try:
            with open(ADMIN_STATE_PATH, "r", encoding="utf-8") as f:
                saved = json.load(f)
            all_sessions = saved.get("sessions") or []
            all_history = saved.get("history") or []
            aid = (agency_id or "").strip()
            sessions = [s for s in all_sessions if aid and str(s.get("agency_id") or "").strip() == aid]
            history = [h for h in all_history if aid and str(h.get("agency_id") or "").strip() == aid]
        except Exception:
            sessions, history = [], []

    # 진행 중(결제중만) vs 완료/종료 구분
    _st = lambda s: (str(s.get("status") or "결제중").strip())
    # K-VAN에서 삭제(deleted_in_kvan=True)되고 거래 내역도 없는 세션은 목록에서 숨긴다.
    _is_visible_ag = lambda s: not (bool(s.get("deleted_in_kvan")) and not bool(s.get("has_transaction")))
    agency_active_sessions = [s for s in sessions if _st(s) == "결제중" and _is_visible_ag(s)]
    agency_completed_sessions = [h for h in history if _is_visible_ag(h)] + [s for s in sessions if _st(s) != "결제중" and _is_visible_ag(s)]

    # DB 기반 거래 내역 (transactions 테이블에서 이 대행사 건만)
    agency_transactions: list[dict] = []
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT *
                FROM transactions
                WHERE agency_id = %s
                ORDER BY created_at DESC
                LIMIT 100
                """,
                (agency_id,),
            )
            agency_transactions = cur.fetchall()
        conn.close()
    except Exception as e:
        print(f"[WARN] agency_admin transactions 조회 실패: {e}")
        agency_transactions = []

    # 결제중 세션인데 주문 JSON이 없으면 자동 재생성 + 매크로 재트리거
    for s in sessions:
        if s.get("status", "결제중") != "결제중":
            continue
        if s.get("kvan_link"):
            continue
        sid = str(s.get("id") or "")
        if not sid:
            continue
        order_file = SESSION_ORDER_DIR / f"{sid}.json"
        if not order_file.exists():
            try:
                _save_session_order_json(sid, str(s.get("amount") or ""), str(s.get("installment") or "일시불"), agency_id=session.get("agency_id"), agency=agency)
                _append_hq_log("WEB", f"[AUTO-HEAL] 대행사 누락 주문 JSON 재생성 session_id={sid}")
                trigger_auto_kvan_async(session_id=sid)
            except Exception as _e:
                _append_hq_log("WEB", f"[AUTO-HEAL][WARN] 대행사 주문 JSON 재생성 실패 session_id={sid}: {_e}")

    message = ""
    crawler_refresh_since = ""
    base_url = request.url_root.rstrip("/")

    if request.method == "POST":
        action = request.form.get("action", "create").strip()
        if action == "create":
            amount = request.form.get("admin_amount", "").strip()
            installment = request.form.get("admin_installment", "일시불").strip()

            # 결제금액이 비어 있으면 세션/링크를 만들지 않고 안내
            if not amount:
                message = "결제 금액을 입력해 주세요. 금액 없이 결제요청 링크를 생성할 수 없습니다."
            else:
                # 이 대행사의 진행 중 세션 수만 카운트
                active_count = sum(
                    1 for s in sessions if s.get("status", "결제중") == "결제중"
                )
                if active_count >= 5:
                    message = "동시에 진행할 수 있는 세션은 최대 5개입니다."
                else:
                    session_id = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")[-12:]
                    new_session = {
                        "id": session_id,
                        "amount": amount,
                        "installment": installment or "",
                        "status": "결제중",
                        "created_at": datetime.utcnow().isoformat(),
                        "agency_id": agency_id,
                    }
                    # 전체 admin_state 에 병합 저장
                    all_sessions = sessions
                    all_history = history
                    if Path(ADMIN_STATE_PATH).exists():
                        try:
                            with open(ADMIN_STATE_PATH, "r", encoding="utf-8") as f:
                                saved = json.load(f)
                            all_sessions = saved.get("sessions") or []
                            all_history = saved.get("history") or []
                        except Exception:
                            all_sessions, all_history = [], []
                    all_sessions.append(new_session)
                    admin_state = {"sessions": all_sessions, "history": all_history}
                    try:
                        with open(ADMIN_STATE_PATH, "w", encoding="utf-8") as f:
                            json.dump(admin_state, f, ensure_ascii=False, indent=2)
                    except Exception as e:  # noqa: BLE001
                        message = f"세션 생성 중 오류가 발생했습니다: {e}"
                    else:
                        try:
                            order_json = _save_session_order_json(session_id, amount, installment, agency_id=session.get("agency_id"), agency=agency)
                            _append_hq_log("WEB", f"세션 주문 JSON 저장 session_id={session_id}, path={order_json}")
                        except Exception as e_order:  # noqa: BLE001
                            _append_hq_log("WEB", f"[WARN] 세션 주문 JSON 저장 실패 session_id={session_id}: {e_order}")
                        _append_hq_log(
                            "WEB",
                            f"AGENCY 세션 생성 agency_id={agency_id}, session_id={session_id}, amount={amount}, installment={installment or '일시불'}",
                        )
                        # 대행사가 링크를 생성한 시점에도 자동 결제 매크로를 준비
                        try:
                            trigger_auto_kvan_async(session_id=session_id)
                        except Exception as e:  # noqa: BLE001
                            print(f"Agency 세션 생성 시 auto_kvan 트리거 실패: {e}")
                        if amount:
                            message = "결제요청 페이지 링크가 생성되었습니다. 링크를 복사하여 고객에게 전달하세요."
                        else:
                            message = "금액이 고정되지 않은 결제요청 링크가 생성되었습니다. 링크를 복사하여 고객에게 전달하세요."
                    # 로컬 세션 리스트도 갱신
                    sessions.append(new_session)
                    # 새로고침 시 중복 생성 방지
                    return redirect(url_for("agency_admin"))
        elif action == "retry_kvan":
            sid = (request.form.get("session_id") or "").strip()
            if sid:
                state_path = Path(ADMIN_STATE_PATH)
                if state_path.exists():
                    try:
                        with open(state_path, "r", encoding="utf-8") as f:
                            saved = json.load(f)
                        for s in saved.get("sessions") or []:
                            if str(s.get("id")) == sid:
                                s["status"] = "결제중"
                                s.pop("error_reason", None)
                                s.pop("failed_at", None)
                                s.pop("kvan_link", None)
                                break
                        with open(state_path, "w", encoding="utf-8") as f:
                            json.dump(saved, f, ensure_ascii=False, indent=2)
                    except Exception as e:  # noqa: BLE001
                        _append_hq_log("WEB", f"[WARN] agency retry 상태 초기화 실패: {e}")
                try:
                    for s in sessions:
                        if str(s.get("id")) == sid:
                            _save_session_order_json(sid, str(s.get("amount") or ""), str(s.get("installment") or "일시불"), agency_id=session.get("agency_id"), agency=agency)
                            break
                    trigger_auto_kvan_async(session_id=sid)
                    _append_hq_log("WEB", f"agency retry_kvan 재요청 session_id={sid}")
                    message = "링크 생성을 다시 요청했습니다."
                except Exception as e:  # noqa: BLE001
                    message = f"재요청 중 오류: {e}"
            return redirect(url_for("agency_admin"))
        elif action == "refresh_kvan":
            # 대행사 어드민에서 수동으로 K-VAN 크롤링 매크로를 한 번 더 실행
            try:
                trigger_kvan_crawler_refresh()
                crawler_refresh_since = datetime.utcnow().isoformat()
                message = "K-VAN 크롤러 새로고침 신호를 보냈습니다. 잠시 후 최신 데이터가 반영됩니다."
            except Exception as e:  # noqa: BLE001
                print(f"[WARN] Agency에서 refresh_kvan 실행 중 오류: {e}")
                message = "K-VAN 크롤링 재실행 중 오류가 발생했습니다. 잠시 후 다시 시도해 주세요."
        elif action == "delete_session":
            sid = (request.form.get("session_id") or "").strip()
            if sid:
                # admin_state.json 전체에서 해당 세션 제거
                if Path(ADMIN_STATE_PATH).exists():
                    try:
                        with open(ADMIN_STATE_PATH, "r", encoding="utf-8") as f:
                            saved = json.load(f)
                        all_sessions = saved.get("sessions") or []
                        all_history = saved.get("history") or []
                        all_sessions = [s for s in all_sessions if str(s.get("id")) != sid]
                        with open(ADMIN_STATE_PATH, "w", encoding="utf-8") as f:
                            json.dump({"sessions": all_sessions, "history": all_history}, f, ensure_ascii=False, indent=2)
                    except Exception:
                        pass
                sessions = [s for s in sessions if str(s.get("id")) != sid]
            return redirect(url_for("agency_admin"))
        elif action == "bulk_delete_agency_tx":
            tx_ids = request.form.getlist("tx_ids")
            tx_id_set = {str(x).strip() for x in tx_ids if str(x).strip()}
            if tx_id_set:
                try:
                    conn = get_db()
                    with conn.cursor() as cur:
                        placeholders = ",".join(["%s"] * len(tx_id_set))
                        cur.execute(
                            f"DELETE FROM transactions WHERE agency_id = %s AND id IN ({placeholders})",
                            (agency_id, *list(tx_id_set)),
                        )
                    conn.commit()
                    conn.close()
                    message = f"{len(tx_id_set)}건의 거래 내역을 삭제했습니다."
                except Exception as e:  # noqa: BLE001
                    message = f"거래 내역 삭제 중 오류: {e}"
            return redirect(url_for("agency_admin"))
        elif action == "mark_payment_notifications_seen":
            _mark_payment_notifications_seen(agency_id=session.get("agency_id"))
            message = "결제 완료 알림을 확인 처리했습니다."
            return redirect(url_for("agency_admin"))

    # 미확인 결제 알림 건수 (현재 로그인한 대행사만)
    _agency_id = session.get("agency_id")
    payment_notifications_count = len(_load_payment_notifications(agency_id=_agency_id))
    has_pending_link = any(
        (s.get("status", "결제중") == "결제중") and not s.get("kvan_link")
        for s in sessions
    )

    template = """
    <!DOCTYPE html>
    <html lang="ko">
    <head>
      <meta charset="UTF-8" />
      <title>SISA 대행사 결제 어드민</title>
      <meta name="viewport" content="width=device-width, initial-scale=1.0" id="viewport-meta" />
      <!-- SISA 브랜드 파비콘 -->
      <link rel="icon" href="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'><rect width='100' height='100' rx='22' fill='%232f4b9f'/><circle cx='50' cy='50' r='28' fill='none' stroke='%23ffffff' stroke-width='6'/><ellipse cx='50' cy='50' rx='12' ry='28' fill='none' stroke='%23ffffff' stroke-width='4'/><line x1='22' y1='50' x2='78' y2='50' stroke='%23ffffff' stroke-width='4'/></svg>">
      <script>
        if (screen.width < 1280) {
          var vp = document.getElementById('viewport-meta');
          if (vp) vp.setAttribute('content', 'width=1280');
        }
        // 결제중이면서 링크가 아직 없는 세션이 있을 때만 7초 후 한 번 자동 새로고침한다.
        window.addEventListener('DOMContentLoaded', function () {
          var hasPending = {{ 'true' if has_pending_link else 'false' }};
          var pendingPopup = document.getElementById("pending-create-popup");
          var pendingBanner = document.getElementById("pending-create-banner");
          if (hasPending) {
            if (pendingPopup) pendingPopup.classList.add("show");
            if (pendingBanner) pendingBanner.classList.add("show");
            setTimeout(function () {
              location.reload();
            }, 7000);
          } else {
            if (pendingPopup) pendingPopup.classList.remove("show");
            if (pendingBanner) pendingBanner.classList.remove("show");
          }

          var refreshSince = "{{ crawler_refresh_since or '' }}";
          var refreshOverlay = document.getElementById("crawler-refresh-overlay");
          var refreshText = document.getElementById("crawler-refresh-text");
          if (!refreshSince || !refreshOverlay) return;
          var startedAt = Date.now();
          refreshOverlay.classList.add("show");
          function stopRefresh(msg) {
            if (refreshText && msg) refreshText.textContent = msg;
            setTimeout(function () { refreshOverlay.classList.remove("show"); }, 400);
          }
          function pollRefresh() {
            fetch("/api/crawler-refresh-status?since=" + encodeURIComponent(refreshSince), { cache: "no-store" })
              .then(function (r) { return r.json(); })
              .then(function (d) {
                if (d && d.ok && d.done) {
                  stopRefresh("크롤러 상태 확인이 완료되었습니다.");
                  return;
                }
                if (Date.now() - startedAt > 90000) {
                  stopRefresh("크롤러 확인 시간이 초과되어 표시를 종료합니다.");
                  return;
                }
                setTimeout(pollRefresh, 2500);
              })
              .catch(function () {
                if (Date.now() - startedAt > 90000) {
                  stopRefresh("네트워크 확인 시간 초과로 표시를 종료합니다.");
                  return;
                }
                setTimeout(pollRefresh, 3000);
              });
          }
          pollRefresh();
        });
        // 결제 페이지의 결과 모달/오버레이가 남아 있는 경우를 대비해 강제로 숨긴다.
        function hideStaleOverlays() {
          var sels = [
            '#result-modal',
            '.result-backdrop',
            "[data-slot='dialog-overlay']",
            '.dialog-overlay',
            '.modal-backdrop'
          ];
          sels.forEach(function (sel) {
            document.querySelectorAll(sel).forEach(function (el) {
              el.style.setProperty('display', 'none', 'important');
              el.style.setProperty('pointer-events', 'none', 'important');
            });
          });
        }
        function runOverlayCleanupBurst() {
          hideStaleOverlays();
          setTimeout(hideStaleOverlays, 300);
          setTimeout(hideStaleOverlays, 1200);
        }
        window.addEventListener('load', runOverlayCleanupBurst);
        // 뒤로가기 캐시 복원 시에도 검은 오버레이를 즉시 정리한다.
        window.addEventListener('pageshow', runOverlayCleanupBurst);
      </script>
      <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
      <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css" />
      <script src="https://cdn.tailwindcss.com"></script>
      <script>
        tailwind.config = {
          theme: {
            extend: {
              fontFamily: { sans: ['Inter', 'sans-serif'] },
              colors: {
                brand: { blue: '#2f4b9f', dark: '#1e326b', accent: '#e6edf7' }
              }
            }
          }
        }
      </script>
      <style>
        body { background-color: #2f4b9f; }
        .glass-card { background: rgba(30, 50, 107, 0.6); backdrop-filter: blur(12px); }
        .loading-backdrop { position:fixed; inset:0; background:rgba(2,6,23,0.78); display:none; align-items:center; justify-content:center; z-index:2000; }
        .loading-backdrop.show { display:flex; }
        .loading-card { background:#0f172a; border:1px solid #334155; border-radius:14px; padding:16px 18px; color:#e2e8f0; min-width:240px; text-align:center; box-shadow:0 18px 44px rgba(2,6,23,0.65); }
        .loading-spinner { width:28px; height:28px; border:3px solid #475569; border-top-color:#60a5fa; border-radius:50%; margin:0 auto 10px; animation:spin2 0.8s linear infinite; }
        @keyframes spin2 { to { transform: rotate(360deg); } }
        .pending-popup { position:fixed; left:50%; top:50%; transform:translate(-50%,-50%); z-index:2100; min-width:260px; max-width:90vw; background:#0f172a; border:1px solid #334155; border-radius:14px; padding:14px 16px; text-align:center; color:#e2e8f0; box-shadow:0 18px 48px rgba(2,6,23,.72); display:none; }
        .pending-popup.show { display:block; }
        .pending-top-banner { position:fixed; top:72px; left:50%; transform:translateX(-50%); z-index:2090; display:none; align-items:center; gap:8px; background:#7c2d12; border:1px solid #fdba74; color:#fff7ed; border-radius:999px; padding:8px 14px; font-size:12px; font-weight:700; box-shadow:0 10px 24px rgba(124,45,18,.45); }
        .pending-top-banner.show { display:flex; animation:pulseBanner2 1.2s ease-in-out infinite; }
        .pending-inline { margin-top:2px; padding:7px 10px; border-radius:10px; background:#7c2d12; border:1px solid #fdba74; color:#fff7ed; font-size:11px; font-weight:700; line-height:1.45; box-shadow:0 8px 16px rgba(124,45,18,.35); animation:pulseBanner2 1.2s ease-in-out infinite; }
        .pending-dot { width:10px; height:10px; border-radius:999px; background:#60a5fa; display:inline-block; animation:pulseDot2 1s infinite ease-in-out; }
        .pending-dot:nth-child(2) { animation-delay:.2s; }
        .pending-dot:nth-child(3) { animation-delay:.4s; }
        @keyframes pulseBanner2 { 0%,100% { opacity:.92; } 50% { opacity:1; } }
        @keyframes pulseDot2 { 0%,100% { opacity:.2; transform:translateY(0);} 50% { opacity:1; transform:translateY(-2px);} }
        .overflow-x-auto thead th { position: sticky; top: 0; background: rgba(15,23,42,0.96); z-index: 3; }
        .box-schema { position:sticky; top:72px; z-index:4; margin:6px 0 10px; padding:6px 8px; border-radius:8px; border:1px solid #334155; background:#0b1220; color:#93c5fd; font-size:10px; line-height:1.35; }
        .box-schema code { color:#bfdbfe; }
        /* 결제 폼에서 사용하던 결과 모달 오버레이가 남아 있어도 대행사 어드민에서는 항상 숨긴다. */
        #result-modal,
        .result-backdrop,
        [data-slot='dialog-overlay'],
        .dialog-overlay,
        .modal-backdrop {
          display: none !important;
          pointer-events: none !important;
        }
      </style>
    </head>
    <body class="bg-brand-blue text-white font-sans overflow-x-hidden antialiased min-h-screen flex flex-col">
      <div id="pending-create-banner" class="pending-top-banner" aria-hidden="true">
        <i class="fa-solid fa-spinner fa-spin"></i>
        <span>K-VAN 링크 생성 중 (1분정도 소요됩니다.)</span>
      </div>
      <div id="pending-create-popup" class="pending-popup" aria-hidden="true">
        <div style="font-size:13px;font-weight:700; margin-bottom:6px;">링크 생성중입니다 (1분정도 소요됩니다.)</div>
        <div style="font-size:11px; color:#94a3b8; margin-bottom:8px;">생성이 완료되면 자동으로 반영됩니다.</div>
        <div style="display:flex; justify-content:center; gap:6px;">
          <span class="pending-dot"></span><span class="pending-dot"></span><span class="pending-dot"></span>
        </div>
      </div>
      <div id="link-loading-overlay" class="loading-backdrop" aria-hidden="true">
        <div class="loading-card">
          <div class="loading-spinner"></div>
          <div id="link-loading-text" style="font-size:13px;font-weight:600;">링크 생성중입니다... (1분정도 소요됩니다.)</div>
          <div style="margin-top:4px;font-size:11px;color:#94a3b8;">잠시만 기다려 주세요.</div>
        </div>
      </div>
      <div id="crawler-refresh-overlay" class="loading-backdrop" aria-hidden="true">
        <div class="loading-card">
          <div class="loading-spinner"></div>
          <div id="crawler-refresh-text" style="font-size:13px;font-weight:600;">크롤러 작업 진행중...</div>
          <div style="margin-top:4px;font-size:11px;color:#94a3b8;">결과를 확인하면 자동으로 종료됩니다.</div>
        </div>
      </div>
      <header class="fixed top-0 left-0 right-0 z-30 bg-brand-dark/80 backdrop-blur border-b border-white/10">
        <div class="max-w-[96vw] mx-auto px-4 py-3 flex items-center justify-between">
          <div class="flex items-center gap-2">
            <i class="fa-solid fa-store text-white text-xl"></i>
            <div class="flex flex-col leading-tight">
              <span class="text-sm font-semibold text-white/80">{{ agency.company_name }}</span>
              <span class="text-[11px] text-white/60">SISA 대행사 결제 어드민</span>
            </div>
          </div>
          <div class="flex items-center gap-3 flex-wrap">
            <div class="text-[11px] text-white/70">
              결제요청 링크 예시:
              <span class="font-mono bg-white/10 px-2 py-1 rounded-full border border-white/20">
                {{ base_url }}/pay/&lt;SESSION_ID&gt;
              </span>
            </div>
            <a href="{{ url_for('logout') }}" class="px-3 py-1.5 rounded-lg bg-white/10 border border-white/20 text-white text-sm font-medium hover:bg-white/20 transition">
              로그아웃
            </a>
          </div>
        </div>
      </header>
      <main class="flex-grow pt-20 pb-10 px-3 sm:px-4">
        <div class="max-w-[96vw] mx-auto space-y-8">
          {% if message %}
          <div class="bg-emerald-500/10 border border-emerald-400/40 text-emerald-100 text-sm px-4 py-3 rounded-xl">
            {{ message }}
          </div>
          {% endif %}
          {% if payment_notifications_count and payment_notifications_count > 0 %}
          <div class="p-3 rounded-lg bg-amber-500/20 border border-amber-400/40 flex items-center justify-between gap-2 flex-wrap">
            <span class="text-amber-200 text-sm">
              <i class="fa-solid fa-bell mr-1"></i> 미확인 결제 완료 알림 {{ payment_notifications_count }}건
            </span>
            <form method="post" action="{{ url_for('agency_admin') }}" onsubmit="return confirm('선택한 거래 내역을 삭제할까요?');">
              <input type="hidden" name="action" value="mark_payment_notifications_seen" />
              <button type="submit" class="px-2 py-1 rounded bg-amber-600/80 text-white text-xs hover:bg-amber-600">확인</button>
            </form>
          </div>
          {% endif %}

          <!-- 세션 생성 -->
          <section class="glass-card rounded-2xl border border-white/20 shadow-xl p-5">
            <h2 class="text-lg font-semibold mb-3 flex items-center gap-2">
              <i class="fa-solid fa-link text-brand-accent"></i> 결제 요청 링크 생성
            </h2>
            <form method="post" class="flex flex-wrap gap-3 items-end text-sm" data-loading-msg="링크 생성중입니다... (1분정도 소요됩니다.)">
              <input type="hidden" name="action" value="create">
              <div>
                <label class="block text-xs mb-1 text-white/70">결제 금액 (선택)</label>
                <input name="admin_amount" type="text" placeholder="예: 550000"
                       class="bg-black/30 border border-white/20 rounded-lg px-3 py-2 text-sm text-white placeholder-white/40 focus:outline-none focus:border-blue-300" />
              </div>
              <div>
                <label class="block text-xs mb-1 text-white/70">할부개월</label>
                <select name="admin_installment"
                        class="bg-black/30 border border-white/20 rounded-lg px-3 py-2 text-sm text-white focus:outline-none focus:border-blue-300">
                  <option value="일시불">일시불</option>
                  <option value="2">2개월</option>
                  <option value="3">3개월</option>
                  <option value="4">4개월</option>
                  <option value="5">5개월</option>
                  <option value="6">6개월</option>
                </select>
              </div>
              <button type="submit"
                      class="h-10 px-5 rounded-full bg-white text-brand-blue font-semibold text-sm hover:bg-brand-accent transition">
                링크 생성
              </button>
            </form>
            <p class="mt-3 text-[11px] text-white/60">
              생성된 세션은 아래 "진행 중인 결제 세션" 목록에 표시되며, 각 세션 ID 를 통해 결제 페이지 링크를 고객에게 전달할 수 있습니다.
            </p>
          </section>

          <!-- 대행사 엑셀 다운로드 -->
          <section class="glass-card rounded-2xl border border-white/20 shadow-xl p-4 flex items-center justify-between text-sm">
            <div class="text-white/70 text-[11px]">
              이 대행사에 해당하는 결제/정산 내역을 엑셀로 내려받을 수 있습니다.
            </div>
            <a href="{{ url_for('agency_export_excel') }}"
               class="px-4 py-2 rounded-full bg-white text-brand-blue font-semibold text-xs hover:bg-brand-accent transition">
              엑셀 다운받기
            </a>
          </section>

          <!-- 진행 중인 세션 -->
          <section class="glass-card rounded-2xl border border-white/20 shadow-xl p-5">
            <h2 class="text-lg font-semibold mb-3 flex items-center gap-2">
              <i class="fa-solid fa-clock text-brand-accent"></i> 진행 중인 결제 세션
              <form method="post" action="{{ url_for('agency_admin') }}" class="ml-2" data-loading-msg="크롤러를 새로고침하는 중입니다...">
                <input type="hidden" name="action" value="refresh_kvan" />
                <button type="submit"
                        class="px-3 py-1 rounded-full bg-white/10 border border-white/30 text-white hover:bg-white/25 flex items-center gap-1 text-xs">
                  <i class="fa-solid fa-rotate-right text-[10px]"></i>
                  <span>새로고침</span>
                </button>
              </form>
            </h2>
            <div class="box-schema">진행 중 = 결제중만 표시. 완료/종료된 세션은 아래 섹션에 표시됩니다.</div>
            {% if agency_active_sessions %}
            <div class="space-y-2 text-sm">
              {% for s in agency_active_sessions %}
              <div class="bg-black/25 border border-white/15 rounded-xl px-3 py-2 flex flex-wrap items-center justify-between gap-2">
                <div class="text-[11px]">
                  <div class="font-mono text-blue-200">SESSION: {{ s.id }}</div>
                  <div class="text-white/80">
                    금액: {{ s.amount or '고객 입력' }} / 할부: {{ s.installment or '일시불' }}
                  </div>
                  <div class="text-white/60">생성일: {{ s.created_at }}</div>
                  <div class="text-white/70">
                    상태: <span class="inline-block px-2 py-0.5 rounded text-[10px] font-semibold bg-emerald-900/50 text-emerald-200 border border-emerald-600/50">결제중</span>
                  </div>
                </div>
                <div class="flex flex-col items-end gap-1 text-[11px]">
                  {% set kvan_link = s.kvan_link %}
                  {% if kvan_link %}
                  <button type="button"
                          onclick="navigator.clipboard && navigator.clipboard.writeText('{{ kvan_link }}'); alert('링크가 복사되었습니다.');"
                          class="px-3 py-1 rounded-full bg-white/10 hover:bg-white/20 border border-white/20">
                    링크 복사
                  </button>
                  <span class="font-mono text-white/70 text-[10px]">{{ kvan_link }}</span>
                  {% elif s.status == '링크생성실패' %}
                  <div class="rounded-lg px-2 py-2 text-[10px]" style="background:#450a0a;border:1px solid #7f1d1d;">
                    <div class="text-red-300 font-semibold">⚠ 링크 생성 실패</div>
                    {% if s.error_reason %}
                    <div class="text-red-200 mt-1">{{ s.error_reason[:80] }}</div>
                    {% endif %}
                    <form method="post" action="{{ url_for('agency_admin') }}" class="mt-2" data-loading-msg="링크 재생성 요청중입니다...">
                      <input type="hidden" name="action" value="retry_kvan">
                      <input type="hidden" name="session_id" value="{{ s.id }}">
                      <button type="submit"
                              class="px-2 py-1 rounded bg-red-600 hover:bg-red-500 text-white text-[10px]">
                        🔄 다시 링크 생성
                      </button>
                    </form>
                  </div>
                  {% else %}
                  <span class="pending-inline">K-VAN 링크를 생성 중입니다. 1분정도 소요됩니다. 잠시 후 새로고침 해 주세요.</span>
                  {% endif %}
                  <form method="post" action="{{ url_for('agency_admin') }}">
                    <input type="hidden" name="action" value="delete_session">
                    <input type="hidden" name="session_id" value="{{ s.id }}">
                    <button type="submit" class="mt-1 px-2 py-1 rounded-full bg-red-500/30 hover:bg-red-500/50 border border-red-400/60 text-red-100 text-[10px]">
                      삭제
                    </button>
                  </form>
                </div>
              </div>
            {% endfor %}
            </div>
            {% else %}
              <p class="text-xs text-white/60">현재 진행 중인 결제 세션이 없습니다.</p>
            {% endif %}
          </section>

          <!-- 과거 세션(요약) -->
          <section class="glass-card rounded-2xl border border-white/20 shadow-xl p-5">
            <h2 class="text-lg font-semibold mb-3 flex items-center gap-2">
              <i class="fa-solid fa-list-check text-brand-accent"></i> 완료/종료된 세션 요약
            </h2>
            <div class="box-schema">크롤링 결과에 따라 링크만료·결제완료·결제 실패 등이 이 섹션에 표시됩니다.</div>
            {% if agency_completed_sessions %}
            <div class="overflow-x-auto">
              <table class="min-w-full text-xs border-separate border-spacing-y-2">
                <thead class="text-white/70">
                  <tr>
                    <th class="px-3 py-1 text-left">세션ID</th>
                    <th class="px-3 py-1 text-right">금액</th>
                    <th class="px-3 py-1 text-left">할부</th>
                    <th class="px-3 py-1 text-left">상태</th>
                    <th class="px-3 py-1 text-left">메모</th>
                  </tr>
                </thead>
                <tbody>
                  {% for h in agency_completed_sessions %}
                  <tr class="bg-black/20 hover:bg-black/30 transition">
                    <td class="px-3 py-2 font-mono text-blue-200">{{ h.id }}</td>
                    <td class="px-3 py-2 text-right">{{ h.amount }}</td>
                    <td class="px-3 py-2">{{ h.installment }}</td>
                    <td class="px-3 py-2 text-[11px] text-white/80">
                      {% set st = (h.status or '') %}
                      {% if st == '만료' or '만료' in st %}
                        <span class="inline-block px-2 py-0.5 rounded text-[10px] font-semibold bg-amber-900/50 text-amber-200 border border-amber-600/50">링크만료</span>
                      {% elif st in ['결제완료','성공','success'] or h.has_transaction %}
                        <span class="inline-block px-2 py-0.5 rounded text-[10px] font-semibold bg-emerald-900/50 text-emerald-200 border border-emerald-600/50">결제완료</span>
                      {% elif st in ['실패','fail','링크생성실패'] %}
                        <span class="inline-block px-2 py-0.5 rounded text-[10px] font-semibold bg-red-900/50 text-red-200 border border-red-600/50">결제 실패</span>
                      {% else %}
                        {{ st or '-' }}
                      {% endif %}
                      {% if h.deleted_in_kvan %}<span class="text-red-300"> · 삭제됨</span>{% endif %}
                    </td>
                    <td class="px-3 py-2 text-[11px] text-white/70 max-w-[200px] truncate">{{ h.result_message }}</td>
                  </tr>
                  {% endfor %}
                </tbody>
              </table>
            </div>
            {% else %}
              <p class="text-xs text-white/60">아직 종료된 세션 기록이 없습니다.</p>
            {% endif %}
          </section>

          <!-- 거래 내역 (본사 DB 연동) -->
          <section class="glass-card rounded-2xl border border-white/20 shadow-xl p-5">
            <div class="flex flex-col sm:flex-row items-start sm:items-center justify-between mb-3 gap-2">
              <h2 class="text-lg font-semibold flex items-center gap-2">
                <i class="fa-solid fa-list-ul text-brand-accent"></i> 거래 내역 (본사 DB 연동)
              </h2>
              <div class="flex flex-wrap items-center gap-2 text-[11px]">
                <div class="flex items-center gap-1">
                  <span class="text-white/70">날짜:</span>
                  <input id="agencyTxStart" type="date" onchange="filterAgencyTransactions()" class="bg-black/30 border border-white/30 rounded px-2 py-1 text-[11px]" />
                  <span class="text-white/50">~</span>
                  <input id="agencyTxEnd" type="date" onchange="filterAgencyTransactions()" class="bg-black/30 border border-white/30 rounded px-2 py-1 text-[11px]" />
                </div>
                <div class="flex items-center gap-1">
                  <span class="text-white/70">상태:</span>
                  <select id="agencyTxStatus" onchange="filterAgencyTransactions()" class="bg-black/30 border border-white/30 rounded px-2 py-1 text-[11px]">
                    <option value="all">전체</option>
                    <option value="success">성공</option>
                    <option value="fail">실패</option>
                    <option value="other">기타</option>
                  </select>
                </div>
              </div>
            </div>
            <div class="box-schema"><code>transactions (agency_id 필터)</code> 항목: <code>created_at, amount, customer_name, status, settlement_status, fee_percent(계산), payable_amount(계산)</code></div>
            {% if agency_transactions %}
            <form method="post" action="{{ url_for('agency_admin') }}">
            <input type="hidden" name="action" value="bulk_delete_agency_tx" />
            <div class="overflow-x-auto">
              <table class="min-w-full text-xs border-separate border-spacing-y-2">
                <thead class="text-white/70">
                  <tr>
                    <th class="px-3 py-1 text-center">
                      <input type="checkbox" id="agency_tx_check_all" onclick="
                        var cbs = document.querySelectorAll('.agency-tx-check');
                        cbs.forEach(function(cb){ cb.checked = this.checked; }.bind(this));
                      ">
                    </th>
                    <th class="px-3 py-1 text-left">시간</th>
                    <th class="px-3 py-1 text-right">금액</th>
                    <th class="px-3 py-1 text-right">수수료율</th>
                    <th class="px-3 py-1 text-right">지급예정금액</th>
                    <th class="px-3 py-1 text-left">구매자</th>
                    <th class="px-3 py-1 text-center">결제상태</th>
                    <th class="px-3 py-1 text-center">정산상태</th>
                  </tr>
                </thead>
                <tbody id="agencyTxBody">
                  {% set fee = agency.fee_percent or 10 %}
                  {% for t in agency_transactions|sort(attribute="created_at", reverse=True) %}
                  {% set amount = t.amount or 0 %}
                  {% set fee_amount = (amount * fee) // 100 %}
                  {% set payable = amount - fee_amount %}
                  <tr class="bg-black/20 hover:bg-black/30 transition align-top"
                      data-date="{{ t.created_at.strftime('%Y-%m-%d') if t.created_at else '' }}"
                      data-status="{{ t.status or '' }}">
                    <td class="px-3 py-2 text-center">
                      <input type="checkbox" class="agency-tx-check" name="tx_ids" value="{{ t.id }}">
                    </td>
                    <td class="px-3 py-2 whitespace-nowrap">{{ t.created_at }}</td>
                    <td class="px-3 py-2 text-right">{{ amount }} 원</td>
                    <td class="px-3 py-2 text-right">{{ fee }}%</td>
                    <td class="px-3 py-2 text-right">{{ payable }} 원</td>
                    <td class="px-3 py-2 whitespace-nowrap">{{ t.customer_name }}</td>
                    <td class="px-3 py-2 text-center">
                      {% if t.status == 'success' %}
                        <span class="px-2 py-1 rounded-full bg-emerald-500/20 text-emerald-200 border border-emerald-500/40 text-[10px]">성공</span>
                      {% elif t.status == 'fail' %}
                        <span class="px-2 py-1 rounded-full bg-red-500/20 text-red-200 border border-red-500/40 text-[10px]">실패</span>
                      {% else %}
                        <span class="px-2 py-1 rounded-full bg-gray-500/20 text-gray-200 border border-gray-500/40 text-[10px]">기타</span>
                      {% endif %}
                    </td>
                    <td class="px-3 py-2 text-center">
                      {% if t.settlement_status == '정산완료' %}
                        <span class="px-2 py-1 rounded-full bg-blue-500/20 text-blue-200 border border-blue-500/40 text-[10px]">정산완료</span>
                      {% else %}
                        <span class="px-2 py-1 rounded-full bg-yellow-500/20 text-yellow-200 border border-yellow-500/40 text-[10px]">미정산</span>
                      {% endif %}
                    </td>
                  </tr>
                  {% endfor %}
                </tbody>
              </table>
            </div>
            <div class="mt-3 text-right">
              <button type="submit" class="px-3 py-1 rounded-full bg-red-500/40 text-red-100 font-semibold hover:bg-red-500/60 transition text-xs">
                선택 거래 삭제
              </button>
            </div>
            </form>
            <div class="mt-6 pt-4 border-t border-white/10">
              <h3 class="text-sm font-semibold mb-2 text-white/90">거래 내역 (엑셀형 전체 컬럼)</h3>
              <p class="text-[11px] text-white/50 mb-2">DB transactions 테이블의 모든 컬럼을 리스트로 표시합니다. 가로 스크롤 가능.</p>
              <div class="overflow-x-auto max-h-[360px] overflow-y-auto border border-white/20 rounded-xl">
                <table class="min-w-max text-[11px] border-collapse">
                  <thead class="text-white/80 bg-black/40 sticky top-0 z-10">
                    <tr>
                      {% for col in tx_excel_columns %}
                      <th class="px-2 py-1.5 text-left whitespace-nowrap border-b border-r border-white/20 font-semibold">{{ col }}</th>
                      {% endfor %}
                    </tr>
                  </thead>
                  <tbody>
                    {% for t in agency_transactions|sort(attribute="created_at", reverse=True) %}
                    <tr class="bg-black/20 hover:bg-black/30 border-b border-white/10">
                      {% for col in tx_excel_columns %}
                      <td class="px-2 py-1.5 whitespace-nowrap border-r border-white/10 text-white/90">
                        {% set val = t.get(col) %}
                        {% if col == 'amount' and val is not none and val != '' %}{{ val }} 원{% elif val is not none and val != '' %}{{ val }}{% else %}-{% endif %}
                      </td>
                      {% endfor %}
                    </tr>
                    {% endfor %}
                  </tbody>
                </table>
              </div>
            </div>
            {% else %}
              <p class="text-xs text-white/60">아직 이 대행사에 대한 거래 내역이 없습니다.</p>
            {% endif %}
          </section>
        </div>
      </main>
      <script>
        function showLinkLoading(msg) {
          var overlay = document.getElementById("link-loading-overlay");
          var textEl = document.getElementById("link-loading-text");
          if (!overlay) return;
          if (textEl && msg) textEl.textContent = msg;
          overlay.classList.add("show");
        }
        (function () {
          var forms = document.querySelectorAll("form[data-loading-msg]");
          forms.forEach(function (f) {
            f.addEventListener("submit", function () {
              var msg = f.getAttribute("data-loading-msg") || "처리중입니다...";
              showLinkLoading(msg);
            });
          });
        })();
      </script>
    </body>
    </html>
    """
    return render_template_string(
        template,
        agency=agency,
        sessions=sessions,
        history=history,
        agency_active_sessions=agency_active_sessions,
        agency_completed_sessions=agency_completed_sessions,
        base_url=base_url,
        message=message,
        agency_transactions=agency_transactions,
        tx_excel_columns=TX_EXCEL_COLUMNS,
        payment_notifications_count=payment_notifications_count,
        has_pending_link=has_pending_link,
        crawler_refresh_since=crawler_refresh_since,
    )


@app.route("/hq-export-excel", methods=["GET"])
def hq_export_excel():
    """본사용 전체 거래/정산 엑셀 다운로드 (섹션별 분기)."""
    if not session.get("hq_logged_in"):
        return redirect(url_for("hq_login"))

    section = request.args.get("scope", "").strip() or "all"

    state = _load_hq_state()
    transactions = state.get("transactions") or []
    agencies = state.get("agencies") or []
    name_map = {str(ag.get("id")): ag.get("company_name", "") for ag in agencies}

    wb = Workbook()

    if section in ("all", "transactions"):
        ws = wb.active
        ws.title = "Transactions"
        headers = [
            "시간",
            "대행사ID",
            "대행사명",
            "금액",
            "이름",
            "카드구분",
            "생년월일(앞6)",
            "전화번호(뒷자리)",
            "결제상태",
            "정산상태",
            "메모",
        ]
        ws.append(headers)
        for t in transactions:
            aid = str(t.get("agency_id") or "")
            ws.append(
                [
                    t.get("created_at", ""),
                    aid,
                    name_map.get(aid, ""),
                    t.get("amount", 0),
                    t.get("customer_name", ""),
                    t.get("card_type", ""),
                    t.get("resident_front", ""),
                    t.get("phone_number", ""),
                    t.get("status", ""),
                    t.get("settlement_status", ""),
                    t.get("message", ""),
                ]
            )
    else:
        # 필요 시 새 워크시트 생성
        ws = wb.active

    if section in ("all", "agency_summary"):
        if section == "all":
            ws2 = wb.create_sheet(title="AgencySummary")
        else:
            ws2 = wb
            ws2.title = "AgencySummary"
        headers2 = [
            "업체ID",
            "업체명",
            "도메인",
            "수수료%",
            "총 거래금액",
            "미정산 금액",
            "입금 예정액",
            "상태",
        ]
        ws2.append(headers2)
        for ag in agencies:
            total_amount = 0
            unsettled_amount = 0
            for t in transactions:
                amt = t.get("amount") or 0
                if t.get("agency_id") == ag.get("id") and t.get("status") == "success" and amt > 0:
                    total_amount += amt
                    if t.get("settlement_status") != "정산완료":
                        unsettled_amount += amt
            fee = ag.get("fee_percent") or 0
            net_amount = unsettled_amount * (100 - fee) // 100
            ws2.append(
                [
                    ag.get("id", ""),
                    ag.get("company_name", ""),
                    ag.get("domain", ""),
                    fee,
                    total_amount,
                    unsettled_amount,
                    net_amount,
                    ag.get("status", ""),
                ]
            )

    if section in ("all", "agency_manage"):
        if section == "all":
            ws3 = wb.create_sheet(title="AgencyManage")
        else:
            ws3 = wb
            ws3.title = "AgencyManage"
        headers3 = [
            "업체ID",
            "업체명",
            "연락처",
            "은행",
            "계좌번호",
            "이메일/구글시트",
            "수수료%",
            "총 거래금액",
            "미정산 금액",
            "입금 예정액",
            "상태",
            "로그인ID",
            "로그인PW",
        ]
        ws3.append(headers3)
        for ag in agencies:
            total_amount = 0
            unsettled_amount = 0
            for t in transactions:
                amt = t.get("amount") or 0
                if t.get("agency_id") == ag.get("id") and t.get("status") == "success" and amt > 0:
                    total_amount += amt
                    if t.get("settlement_status") != "정산완료":
                        unsettled_amount += amt
            fee = ag.get("fee_percent") or 0
            net_amount = unsettled_amount * (100 - fee) // 100
            ws3.append(
                [
                    ag.get("id", ""),
                    ag.get("company_name", ""),
                    ag.get("phone", ""),
                    ag.get("bank_name", ""),
                    ag.get("account_number", ""),
                    ag.get("email_or_sheet", ""),
                    fee,
                    total_amount,
                    unsettled_amount,
                    net_amount,
                    ag.get("status", ""),
                    ag.get("login_id", ""),
                    ag.get("login_password", ""),
                ]
            )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "sisa_hq.xlsx"
    if section == "transactions":
        filename = "sisa_hq_transactions.xlsx"
    elif section == "agency_summary":
        filename = "sisa_hq_agency_summary.xlsx"
    elif section == "agency_manage":
        filename = "sisa_hq_agency_manage.xlsx"

    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/agency-export-excel", methods=["GET"])
def agency_export_excel():
    """대행사 전용 엑셀 다운로드 (자기 거래만)."""
    agency_id = session.get("agency_id")
    if not agency_id:
        return redirect(url_for("agency_login"))

    state = _load_hq_state()
    transactions = state.get("transactions") or []
    agencies = state.get("agencies") or []
    agency = None
    for ag in agencies:
        if str(ag.get("id")) == str(agency_id):
            agency = ag
            break

    filtered = [
        t for t in transactions if str(t.get("agency_id")) == str(agency_id)
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "AgencyTransactions"
    headers = [
        "시간",
        "금액",
        "이름",
        "카드구분",
        "생년월일(앞6)",
        "전화번호(뒷자리)",
        "결제상태",
        "정산상태",
        "메모",
    ]
    ws.append(headers)

    for t in filtered:
        ws.append(
            [
                t.get("created_at", ""),
                t.get("amount", 0),
                t.get("customer_name", ""),
                t.get("card_type", ""),
                t.get("resident_front", ""),
                t.get("phone_number", ""),
                t.get("status", ""),
                t.get("settlement_status", ""),
                t.get("message", ""),
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "sisa_agency_transactions.xlsx"
    if agency:
        filename = f"sisa_{agency.get('company_name','agency')}_transactions.xlsx"

    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

if __name__ == "__main__":
    # 개발용/배포용 서버 실행 (Railway 등)
    port = int(os.environ.get("PORT", "5000"))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)

