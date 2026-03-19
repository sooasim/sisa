from __future__ import annotations

import os
import time
import json
from dataclasses import dataclass
from pathlib import Path
from typing import List
from datetime import datetime, timedelta, timezone
import random
import re
from urllib.parse import urlparse, parse_qs, unquote

from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
import pymysql

from kvan_link_common import (
    build_kvan_transactions_snapshots,
    extract_kvan_session_key_from_url,
    infer_kvan_transaction_header_cell_label,
    parse_amount_won,
    parse_kvan_link_ui_created_at,
    upsert_kvan_link_creation_seed,
)

# 링크 생성 속도 배율은 아래 _is_server_env() 정의 이후 LINK_CREATION_WAIT_FACTOR 로 설정한다.

# 웹 어드민에서 볼 수 있는 간단한 로그 파일 (HQ 어드민에서 tail 형태로 노출)
# auto_kvan.py 는 /app/wsisa/ 에 위치하므로 parent.parent = /app
# Railway/Docker 배포에서는 web_form.py 와 동일한 /app/data 를 공유해야 한다.
BASE_DIR = Path(__file__).resolve().parent.parent  # /app

# 배포 환경에서 SISA_DATA_DIR 이 설정되지 않은 경우,
# auto_kvan.py 가 위치한 wsisa 폴더의 부모(/app) 기준 data 폴더를 사용한다.
# web_form.py 도 동일하게 /app/data 를 사용하므로 파일이 공유된다.
_raw_data_dir = os.environ.get("SISA_DATA_DIR", "").strip()
if _raw_data_dir:
    DATA_DIR = Path(_raw_data_dir)
else:
    # /app/data 가 실제로 존재하면 우선 사용 (Railway Docker 구조)
    _candidate = BASE_DIR / "data"
    DATA_DIR = _candidate

ADMIN_LOG_PATH = DATA_DIR / "hq_logs.log"


def _append_admin_log(source: str, message: str) -> None:
    try:
        ADMIN_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(ADMIN_LOG_PATH, "a", encoding="utf-8") as f:
            ts = datetime.utcnow().isoformat()
            f.write(f"{ts} [{source}] {message}\n")
    except Exception:
        pass

# 서버(Railway 등)에서 실행 시 헤드리스 + 자동 종료
def _is_server_env() -> bool:
    s = str(os.environ.get("SISA_SERVER", "")).strip().lower()
    k = str(os.environ.get("K_VAN_SERVER", "")).strip().lower()
    truthy = ("1", "true", "yes", "y", "on")
    return bool(
        os.environ.get("RAILWAY_ENVIRONMENT")
        or os.environ.get("RUN_HEADLESS")
        or s in truthy
        or k in truthy
    )


def _read_link_creation_wait_factor_env() -> float:
    """환경변수만 읽은 기본 배율 (서버 추가 배율 적용 전)."""
    try:
        return float(
            os.environ.get(
                "K_VAN_LINK_CREATION_WAIT_FACTOR",
                os.environ.get("LINK_CREATION_WAIT_FACTOR", "0.45"),
            )
        )
    except (TypeError, ValueError):
        return 0.45


def _compute_link_creation_wait_factor() -> float:
    """
    링크 생성 대기 배율 (낮을수록 대기 짧음).
    서버 환경에서는 기본적으로 20% 더 빠르게(×0.8) 동작한다.
    """
    base = _read_link_creation_wait_factor_env()
    if _is_server_env():
        base *= 0.8
    return max(0.15, min(1.5, base))


# 링크 생성 속도: 대기 시간 배율. 에러·미검출 시 환경변수로 1.0~0.8 권장.
LINK_CREATION_WAIT_FACTOR = _compute_link_creation_wait_factor()


def _expired_debug(msg: str) -> None:
    if str(os.environ.get("K_VAN_EXPIRED_DEBUG", "")).strip().lower() in ("1", "true", "yes", "y"):
        print(msg)


def _step_start(label: str) -> float:
    """단계별 속도 측정을 위한 헬퍼."""
    t0 = time.perf_counter()
    print(f"[STEP] {label} 시작")
    return t0


def _step_end(label: str, t0: float) -> None:
    dt = time.perf_counter() - t0
    print(f"[STEP] {label} 완료 ({dt:.2f}s)")


SIGN_IN_URL = "https://store.k-van.app/"
FACE_TO_FACE_URL = "https://store.k-van.app/face-to-face-payment"

def _has_payment_links_quick(driver: webdriver.Chrome, retries: int = 5, delay: float = 1.0) -> bool:
    """
    결제링크 관리 화면에 실제 결제 링크 카드가 존재하는지 가볍게 확인한다.

    - '생성된 결제 링크가 없습니다' 문구가 보이면 즉시 False.
    - '거래 내역' 아이콘/버튼이 보이면 True.
    - 위 둘 다 아닌 경우, 짧게 여러 번 재시도 후 결과를 반환한다.
    """
    for attempt in range(retries):
        try:
            # 1) "생성된 결제 링크가 없습니다" 안내 문구가 있으면 링크 없음
            empty_msgs = driver.find_elements(
                By.XPATH,
                "//*[contains(normalize-space(.),'생성된 결제 링크가 없습니다')]",
            )
            if empty_msgs:
                print(f"[EMPTY_CHECK] 결제링크 없음 문구 감지 (attempt={attempt})")
                return False

            # 2) '거래 내역' 버튼/아이콘이 하나라도 있으면 링크가 있다고 판단
            icons = driver.find_elements(
                By.XPATH,
                "//button[@title='거래 내역']"
                " | //button[contains(normalize-space(.),'거래 내역')]"
                " | //button[contains(normalize-space(.),'거래내역')]"
                " | //button[.//svg[contains(@class,'lucide-receipt')]]",
            )
            if icons:
                print(f"[EMPTY_CHECK] 거래 내역 아이콘 감지 → 링크 존재 (attempt={attempt}, count={len(icons)})")
                return True

            # 3) 거래내역 버튼이 없는 UI에서는 KEY 세션 ID 문자열로 링크 존재를 판단
            key_tokens = driver.find_elements(
                By.XPATH,
                "//*[contains(normalize-space(.),'KEY20')]",
            )
            if key_tokens:
                print(f"[EMPTY_CHECK] KEY 세션ID 감지 → 링크 존재 (attempt={attempt}, count={len(key_tokens)})")
                return True

            # 4) 카드 컨테이너(rounded+border) 안에 KEY가 있는 경우 (텍스트가 자식 노드에 분리된 경우)
            card_containers = driver.find_elements(
                By.XPATH,
                "//div[contains(@class,'rounded') and contains(@class,'border')]"
                "[.//*[contains(text(),'KEY') or contains(.,'KEY20')]]",
            )
            if card_containers:
                print(f"[EMPTY_CHECK] 결제링크 카드 컨테이너 감지 → 링크 존재 (attempt={attempt}, count={len(card_containers)})")
                return True
        except Exception as e:
            print(f"[EMPTY_CHECK] 링크 존재 여부 확인 중 예외 (attempt={attempt}): {e}")

        time.sleep(delay)

    print("[EMPTY_CHECK] 여러 번 확인했으나 링크를 찾지 못했습니다 (빈 화면으로 간주).")
    return False


def _go_to_payment_link(driver: webdriver.Chrome, max_attempts: int = 12) -> bool:
    """
    /payment-link 화면으로 안정적으로 진입하기 위한 헬퍼.

    - 단순 driver.get 으로 /dashboard 로 리다이렉트되는 경우가 있어,
      여러 번 재시도 + 대시보드 사이드 메뉴 클릭까지 시도한다.
    - 성공 시 True, 끝까지 실패 시 False 반환.
    """
    url_target = "https://store.k-van.app/payment-link"

    for attempt in range(max_attempts):
        cur = driver.current_url or ""
        if "payment-link" in cur:
            print(f"[NAV] 이미 /payment-link 에 위치 (attempt={attempt}, url={cur})")
            return True

        print(f"[NAV] /payment-link 진입 시도 (attempt={attempt}, current_url={cur})")
        try:
            driver.get(url_target)
        except Exception as e:
            print(f"[NAV] driver.get({url_target}) 중 예외: {e}")

        # URL 이 곧바로 바뀌는지 3초 정도만 기다린다.
        try:
            WebDriverWait(driver, 3).until(EC.url_contains("payment-link"))
            print(f"[NAV] URL 기반 /payment-link 진입 성공 (attempt={attempt}, url={driver.current_url})")
            return True
        except Exception:
            # 여전히 대시보드 등이라면, 메뉴 클릭 방식도 시도해 본다.
            pass

        try:
            # 사이드 메뉴/내비게이션에서 '결제링크' 관련 항목을 찾아 클릭 시도
            nav_btn = None
            candidates = driver.find_elements(
                By.XPATH,
                "//a[contains(@href,'payment-link')]"
                " | //button[contains(@href,'payment-link')]"
                " | //a[contains(normalize-space(.),'결제링크')]"
                " | //a[contains(normalize-space(.),'결제 링크')]"
                " | //button[contains(normalize-space(.),'결제링크')]"
                " | //button[contains(normalize-space(.),'결제 링크')]",
            )
            if candidates:
                nav_btn = candidates[0]

            if nav_btn:
                driver.execute_script(
                    "arguments[0].scrollIntoView({behavior:'instant',block:'center'});",
                    nav_btn,
                )
                time.sleep(0.1)
                driver.execute_script("arguments[0].click();", nav_btn)
                try:
                    WebDriverWait(driver, 3).until(EC.url_contains("payment-link"))
                    print(f"[NAV] 메뉴 클릭으로 /payment-link 진입 성공 (attempt={attempt}, url={driver.current_url})")
                    return True
                except Exception:
                    print("[NAV] 메뉴 클릭 후에도 /payment-link 로 전환되지 않음, 재시도 예정.")
        except Exception as e_nav:
            print(f"[NAV] 메뉴 기반 /payment-link 진입 시도 중 예외: {e_nav}")

        time.sleep(0.5)

    print("[NAV][ERROR] 여러 차례 시도했으나 /payment-link 로 진입하지 못했습니다.")
    return False


def _wait_payment_link_page_ready(driver: webdriver.Chrome, timeout_sec: float = 18.0) -> bool:
    """
    결제링크 관리 페이지가 로드 완료될 때까지 대기한다.
    '권한 확인 중...' 스피너가 사라지고, 빈 화면 문구 또는 카드/KEY/거래내역 버튼이 보이면 준비된 것으로 본다.
    """
    wait = WebDriverWait(driver, max(1.0, timeout_sec))
    try:
        def _page_ready(drv):
            try:
                # 1) 빈 화면 문구가 보이면 준비 완료
                empty = drv.find_elements(
                    By.XPATH,
                    "//*[contains(normalize-space(.),'생성된 결제 링크가 없습니다')]",
                )
                if empty:
                    return True
                # 2) 거래 내역 버튼 또는 KEY 세션ID가 보이면 준비 완료
                icons = drv.find_elements(
                    By.XPATH,
                    "//button[@title='거래 내역']"
                    " | //button[contains(normalize-space(.),'거래 내역')]"
                    " | //button[contains(normalize-space(.),'거래내역')]"
                    " | //*[contains(normalize-space(.),'KEY20')]",
                )
                if icons:
                    return True
                return False
            except Exception:
                return False

        wait.until(_page_ready)
        print("[PAGE_READY] 결제링크 관리 페이지 로드 완료.")
        return True
    except TimeoutException:
        print("[WARN] 결제링크 관리 페이지 준비 대기 타임아웃 - 현재 상태로 진행.")
        return False


# 경로는 파일 상단에서 이미 초기화된 DATA_DIR(/app/data 또는 SISA_DATA_DIR)를 그대로 사용한다.
# (중복 재정의 시 web_form.py 와 경로가 갈라져 lock/heartbeat/wakeup 파일 불일치가 발생할 수 있음)

WAKEUP_FLAG_PATH = DATA_DIR / "crawler_wakeup.flag"

# 로컬 테스트 모드: DB 에는 아무 것도 쓰지 않고, 크롤링/매크로 동작만 확인할 때 사용
# - SISA_LOCAL_TEST 가 명시되면 그 값을 따르고
# - 없으면 "서버 환경이 아니면" 기본적으로 LOCAL_TEST=True 로 동작하게 만든다.
_local_flag = os.environ.get("SISA_LOCAL_TEST")
if _local_flag is None:
    LOCAL_TEST = not _is_server_env()
else:
    LOCAL_TEST = _local_flag.strip().lower() in ("1", "true", "yes", "y")


def signal_crawler_wakeup() -> None:
    """
    K-VAN 크롤러(kvan_crawler.py)에 "즉시 다시 크롤링해 달라"는 신호를 남긴다.

    - DATA_DIR/crawler_wakeup.flag 파일에 타임스탬프를 기록하는 방식으로 구현.
    - 크롤러는 대기(sleep) 중 이 파일을 감지하면 대기를 즉시 종료하고 다음 사이클을 시작한다.
    """
    try:
        WAKEUP_FLAG_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(WAKEUP_FLAG_PATH, "w", encoding="utf-8") as f:
            f.write(time.strftime("%Y-%m-%d %H:%M:%S"))
        print(f"[WAKEUP] 크롤러 깨우기 플래그 생성: {WAKEUP_FLAG_PATH}")
    except Exception as e:
        print(f"[WAKEUP][WARN] 크롤러 깨우기 플래그 생성 실패: {e}")


# 입력 데이터 JSON 파일 (web_form.py 가 생성)
ORDER_JSON_PATH = DATA_DIR / "current_order.json"

# 결제 결과를 저장할 엑셀 / JSON 파일
RESULT_EXCEL_PATH = DATA_DIR / "kvan_results.xlsx"
RESULT_JSON_PATH = DATA_DIR / "last_result.json"
HQ_STATE_PATH = DATA_DIR / "hq_state.json"

# 세션별 주문/결과 디렉토리 (동시 여러 세션용)
SESSION_ORDER_DIR = DATA_DIR / "sessions" / "orders"
SESSION_RESULT_DIR = DATA_DIR / "sessions" / "results"
SESSION_ORDER_DIR.mkdir(parents=True, exist_ok=True)
SESSION_RESULT_DIR.mkdir(parents=True, exist_ok=True)

# 어드민 세션 상태 JSON (본사 + 대행사 공용)
ADMIN_STATE_PATH = DATA_DIR / "admin_state.json"


def _data_dir_candidates() -> list[Path]:
    """배포 경로 차이를 흡수하기 위한 데이터 디렉토리 후보 목록."""
    candidates = [
        Path(DATA_DIR),
        (BASE_DIR / "data"),
        (BASE_DIR.parent / "data"),
        (BASE_DIR / "wsisa" / "data"),
        Path("/app/data"),
    ]
    uniq: list[Path] = []
    seen: set[str] = set()
    for p in candidates:
        k = str(p)
        if k not in seen:
            uniq.append(p)
            seen.add(k)
    return uniq


def _admin_state_candidates() -> list[Path]:
    """환경별 경로 차이를 흡수하기 위해 admin_state.json 후보 경로를 반환."""
    candidates = [d / "admin_state.json" for d in _data_dir_candidates()]
    uniq: list[Path] = []
    seen: set[str] = set()
    for p in candidates:
        k = str(p)
        if k not in seen:
            uniq.append(p)
            seen.add(k)
    return uniq


def _resolved_admin_state_path() -> Path:
    """실제로 존재하는 admin_state.json 경로를 반환.
    없으면 DATA_DIR 기반 기본 경로를 반환한다.
    """
    for p in _admin_state_candidates():
        if p.exists():
            return p
    return Path(ADMIN_STATE_PATH)


def _load_admin_state() -> dict:
    """admin_state.json 을 후보 경로에서 읽어 반환한다."""
    for p in _admin_state_candidates():
        if p.exists():
            try:
                with open(p, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
    return {"sessions": [], "history": []}


def _save_admin_state(state: dict) -> None:
    """admin_state 를 실제 존재하는 경로에 저장한다."""
    st_path = _resolved_admin_state_path()
    st_path.parent.mkdir(parents=True, exist_ok=True)
    with open(st_path, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


def _session_order_path_candidates(session_id: str) -> list[Path]:
    """세션 주문 JSON 후보 경로 목록."""
    candidates = [d / "sessions" / "orders" / f"{session_id}.json" for d in _data_dir_candidates()]
    uniq: list[Path] = []
    seen: set[str] = set()
    for p in candidates:
        k = str(p)
        if k not in seen:
            uniq.append(p)
            seen.add(k)
    return uniq

# 옥션 상품 리스트 (본사 홈페이지 auction.html 기반)
AUCTION_ITEMS: list[dict] = []
AUCTION_LOADED = False


def _parse_session_datetime_auto(ts) -> datetime | None:
    if not ts:
        return None
    try:
        s = str(ts).strip().replace("Z", "+00:00")
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is not None:
            dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
        return dt
    except Exception:
        return None


def _session_considered_terminal_auto(s: dict) -> bool:
    if not isinstance(s, dict):
        return True
    if s.get("deleted") or s.get("has_approval"):
        return True
    st = str(s.get("status") or "").strip()
    if not st:
        return False
    low = st.lower()
    for token in (
        "완료",
        "만료",
        "취소",
        "실패",
        "종료",
        "삭제",
        "expired",
        "취소완료",
        "결제완료",
    ):
        if token in st or token in low:
            return True
    return False


def _has_active_sessions(window_minutes: int = 10) -> bool:
    """
    admin_state.json 기준으로 '지금은 자주 돌아야 하는가'.

    (kvan_crawler.py 와 동일 규칙 — status 미설정을 '결제중'으로 보지 않음)
    """
    try:
        st = _load_admin_state()
        if not st.get("sessions") and not st.get("history"):
            return False
        sessions = st.get("sessions") or []
        history = st.get("history") or []

        cutoff = datetime.utcnow() - timedelta(minutes=window_minutes)

        for s in sessions:
            if _session_considered_terminal_auto(s):
                continue
            if str(s.get("status") or "").strip() == "결제중":
                return True

        for s in sessions:
            if _session_considered_terminal_auto(s):
                continue
            dt = _parse_session_datetime_auto(s.get("created_at"))
            if dt is None:
                continue
            if dt >= cutoff:
                return True

        for h in history:
            if _session_considered_terminal_auto(h):
                continue
            if str(h.get("status") or "").strip() != "결제중":
                continue
            dt = _parse_session_datetime_auto(h.get("created_at"))
            if dt is None:
                continue
            if dt >= cutoff:
                return True

        return False
    except Exception as e:
        print(f"[WARN] _has_active_sessions 검사 실패: {e}")
        return False


def _load_auction_items() -> None:
    """본사 auction.html 에 포함된 csvData 를 파싱하여 상품 리스트를 메모리에 로드."""
    global AUCTION_ITEMS, AUCTION_LOADED
    if AUCTION_LOADED:
        return
    try:
        # BASE_DIR = 프로젝트 루트(wsisa 상위), auction.html 은 그 안에 있음
        auction_path = BASE_DIR / "auction.html"
        if not auction_path.exists():
            AUCTION_LOADED = True
            return
        text = auction_path.read_text(encoding="utf-8", errors="ignore")
        # 첫 번째 백틱 블록(csvData) 추출
        m = re.search(r"const csvData\s*=\s*`([^`]+)`", text, re.DOTALL)
        if not m:
            AUCTION_LOADED = True
            return
        csv_block = m.group(1).strip()
        lines = [ln.strip() for ln in csv_block.splitlines() if ln.strip()]
        if not lines:
            AUCTION_LOADED = True
            return
        # 첫 줄은 헤더
        for row in lines[1:]:
            parts = [p.strip() for p in row.split(",")]
            if not parts or len(parts) < 3:
                continue
            try:
                price = int(parts[0])
            except ValueError:
                continue
            brand = parts[1]
            model = parts[2]
            AUCTION_ITEMS.append({"price": price, "name": f"{brand} {model}"})
    except Exception as e:  # noqa: BLE001
        print(f"[WARN] auction.html 파싱 실패: {e}")
    finally:
        AUCTION_LOADED = True


def _choose_product_name_for_amount(amount: int) -> str:
    """
    신청 금액(amount) 이상인 옥션 금액 구간 중 가장 낮은 구간의 상품명을 선택한다.
    예: 217만원 신청 → 220만원 구간 상품 중 최대 10개에서 랜덤 1개.
    금액 입력란에는 신청 금액(217만원)이 그대로 들어가고, 상품명만 옥션 구간에 맞춘다.
    """
    _load_auction_items()
    if not AUCTION_ITEMS:
        return "SISA 글로벌 옥션 상품"
    # 신청 금액 이상인 옥션 상품만 (같은 금액이 없으면 그 다음 구간 사용)
    candidates = [it for it in AUCTION_ITEMS if (it.get("price") or 0) >= amount]
    if not candidates:
        return "SISA 글로벌 옥션 상품"
    # 그중 가장 낮은 금액 구간(예: 217만 → 220만원 구간)
    min_tier = min(it["price"] for it in candidates)
    tier_items = [it for it in candidates if it["price"] == min_tier]
    if len(tier_items) > 10:
        tier_items = random.sample(tier_items, 10)
    chosen = random.choice(tier_items)
    return chosen.get("name") or "SISA 글로벌 옥션 상품"

# MySQL 환경 변수 (Railway 용) - web_form.py 와 동일하게
DB_HOST = os.environ.get("MYSQLHOST") or os.environ.get("MYSQL_HOST") or "localhost"
DB_PORT = int(os.environ.get("MYSQLPORT") or os.environ.get("MYSQL_PORT") or "3306")
DB_USER = os.environ.get("MYSQLUSER") or os.environ.get("MYSQL_USER") or "root"
DB_PASSWORD = os.environ.get("MYSQLPASSWORD") or os.environ.get("MYSQL_PASSWORD") or ""
DB_NAME = (
    os.environ.get("MYSQL_DATABASE")
    or os.environ.get("MYSQLDATABASE")
    or os.environ.get("MYSQL_DB")
    or "railway"
)
DB_CONNECT_TIMEOUT = int(os.environ.get("MYSQL_CONNECT_TIMEOUT", "5"))
DB_READ_TIMEOUT = int(os.environ.get("MYSQL_READ_TIMEOUT", "10"))
DB_WRITE_TIMEOUT = int(os.environ.get("MYSQL_WRITE_TIMEOUT", "10"))


def get_db():
    return pymysql.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=False,
        connect_timeout=DB_CONNECT_TIMEOUT,
        read_timeout=DB_READ_TIMEOUT,
        write_timeout=DB_WRITE_TIMEOUT,
    )


def _is_retryable_db_error(exc: Exception) -> bool:
    msg = str(exc or "").lower()
    return any(
        k in msg
        for k in (
            "2013",
            "2006",
            "lost connection",
            "server has gone away",
            "connection reset",
            "connection refused",
            "broken pipe",
        )
    )


def _get_db_with_retry(max_attempts: int = 3, delay_sec: float = 0.8):
    last_exc: Exception | None = None
    for attempt in range(1, max_attempts + 1):
        try:
            conn = get_db()
            try:
                conn.ping(reconnect=True)
            except Exception:
                pass
            return conn
        except Exception as e:
            last_exc = e
            if attempt >= max_attempts or not _is_retryable_db_error(e):
                raise
            _append_admin_log("AUTO", f"[WARN] DB 재연결 재시도 {attempt}/{max_attempts}: {e}")
            time.sleep(delay_sec * attempt)
    if last_exc:
        raise last_exc
    raise RuntimeError("DB connection retry failed without exception")

# JSON / 결과 엑셀에서 공통으로 사용하는 필드 목록
HEADERS: List[str] = [
    "login_id",
    "login_password",
    "login_pin",
    "card_type",  # personal / business
    "card_number",
    "expiry_mm",
    "expiry_yy",
    "card_password",
    "installment_months",
    "phone_number",
    "customer_name",
    "resident_front",
    "amount",
    "product_name",
]


# 로그인 페이지 셀렉터
# 제공해주신 로그인 화면 HTML 기준:
# - 아이디 / 비밀번호 input 은 name 이 없고 label 텍스트로만 구분
SIGN_IN_SELECTORS = {
    # 기본: <label>아이디</label> 바로 아래 첫 번째 input
    "id_primary": (
        By.XPATH,
        "//label[normalize-space(text())='아이디']/following::input[1]",
    ),
    # 보조1: placeholder 에 '아이디' 가 포함된 input
    "id_placeholder": (
        By.XPATH,
        "//input[contains(@placeholder,'아이디')]",
    ),
    # 보조2: 로그인 폼 안의 첫 번째 text input (password 제외)
    "id_fallback": (
        By.XPATH,
        "(//input[@type='text' or not(@type)])[1]",
    ),
    # 기본: <label>비밀번호</label> 바로 아래 첫 번째 input
    "password_primary": (
        By.XPATH,
        "//label[normalize-space(text())='비밀번호']/following::input[1]",
    ),
    # 보조: type='password' 인 첫 번째 input
    "password_fallback": (
        By.XPATH,
        "(//input[@type='password'])[1]",
    ),
    # 로그인 버튼 (텍스트가 '로그인' 인 버튼)
    "submit_primary": (
        By.XPATH,
        "//button[contains(normalize-space(.), '로그인')]",
    ),
}

# PIN 팝업 셀렉터 (로그인 이후 2차 인증)
PIN_POPUP_SELECTORS = {
    # "PIN을 입력해 주세요" 문구가 포함된 영역 안의 첫 번째 input
    "input": (
        By.XPATH,
        "//*[contains(text(), 'PIN') and contains(text(), '입력')]/ancestor::div[1]//input",
    ),
    # 확인 버튼
    "confirm": (
        By.XPATH,
        "//button[contains(normalize-space(.), '확인')]",
    ),
}


def _set_session_ttl_to_5min(driver: webdriver.Chrome, max_wait: float = 10.0) -> bool:
    """
    세션 유효시간 Select 컴포넌트( id='session-ttl' )를
    '5분 (일반 결제)' 값으로 변경한다.

    - 1) 트리거 버튼(#session-ttl)을 클릭해 드롭다운을 연다.
    - 2) 드롭다운 안에서 '5분 (일반 결제)' 텍스트가 있는 옵션을 클릭한다.
    - 3) 다시 트리거의 span[data-slot="select-value"] 텍스트가 '5분' 으로
         시작하는지 확인해 실제로 변경되었는지 검증한다.
    - 위 과정을 max_wait 초 동안 반복 시도하고, 성공하면 True, 아니면 False.
    """
    # 이 함수는 최대한 빨리(한 번의 시도 안에서) 3분 트리거를 열고,
    # 5분 또는 60분 옵션을 클릭 시도한 다음, 바로 다음 단계(링크 생성)로 넘어가기 위한 것이다.
    # 사용자가 5분/60분 모두 허용한다고 하였으므로, 여기서는 "최대한 클릭 시도"만 하고
    # 성공 여부에 관계 없이 True 를 반환해 흐름을 빠르게 진행시킨다.

    # 0) 현재 값 로깅 (디버깅용)
    try:
        trigger = driver.find_element(By.ID, "session-ttl")
        value_span = trigger.find_element(By.CSS_SELECTOR, "span[data-slot='select-value']")
        before = (value_span.text or "").strip()
        print(f"[DEBUG] 세션유효시간 초기 값(before)='{before}'")
        if before.startswith("5분") or before.startswith("60분"):
            print(f"[INFO] 세션 유효시간이 이미 '{before}' 로 설정되어 있습니다.")
            return True
    except Exception as e:
        print(f"[DEBUG] 세션유효시간 초기값 읽기 실패: {e}")

    # 1) 트리거 버튼 클릭 (3분 콤보박스 열기)
    try:
        trigger = driver.find_element(By.ID, "session-ttl")
        driver.execute_script("arguments[0].click();", trigger)
        time.sleep(0.15)
    except Exception as e:
        print(f"[DEBUG] 세션유효시간 트리거 클릭 실패: {e}")

    # 2) 드롭다운 안에서 '5분 (일반 결제)' 또는 '5분' 텍스트 옵션 클릭 시도
    try:
        candidates = driver.find_elements(
            By.XPATH,
            "//*[contains(normalize-space(.),'5분 (일반 결제)') or contains(normalize-space(.),'5분')]",
        )
        visible_opts = [el for el in candidates if el.is_displayed()]
        print(f"[DEBUG] 5분 옵션 후보 개수={len(candidates)}, 표시되는 개수={len(visible_opts)}")
        if visible_opts:
            driver.execute_script("arguments[0].click();", visible_opts[0])
            time.sleep(0.2)
    except Exception as e:
        print(f"[DEBUG] 5분 옵션 텍스트 기반 클릭 실패: {e}")

    # 3) 여전히 안 보인다면, 드롭다운 내에서 다시 '3분' 위치를 기준으로 스크롤 후 5분 재시도
    try:
        three_opts = driver.find_elements(
            By.XPATH,
            "//*[contains(normalize-space(.),'3분 (빠른 결제)') or contains(normalize-space(.),'3분')]",
        )
        three_visible = [el for el in three_opts if el.is_displayed()]
        print(f"[DEBUG] 3분 옵션 후보 개수={len(three_opts)}, 표시되는 개수={len(three_visible)}")
        if three_visible:
            three_el = three_visible[0]
            driver.execute_script(
                "arguments[0].scrollIntoView({behavior:'instant', block:'start'});", three_el
            )
            time.sleep(0.1)
            candidates2 = driver.find_elements(
                By.XPATH,
                "//*[contains(normalize-space(.),'5분 (일반 결제)') or contains(normalize-space(.),'5분')]",
            )
            visible2 = [el for el in candidates2 if el.is_displayed()]
            print(f"[DEBUG] 5분 옵션(재검색) 후보 개수={len(candidates2)}, 표시되는 개수={len(visible2)}")
            if visible2:
                driver.execute_script("arguments[0].click();", visible2[0])
                time.sleep(0.2)
    except Exception as e:
        print(f"[DEBUG] 3분 기준 스크롤 후 5분 클릭 시도 실패: {e}")

    # 4) 그래도 안 되면, 3분 클릭 위치에서 위로 30px 좌표 클릭 폴백 (60분 선택 가능)
    try:
        trigger = driver.find_element(By.ID, "session-ttl")
        rect = driver.execute_script(
            """
const el = arguments[0];
const r = el.getBoundingClientRect();
return { x: (r.left + r.right) / 2, y: (r.top + r.bottom) / 2, h: r.bottom - r.top };
""",
            trigger,
        )
        if rect and "x" in rect and "y" in rect:
            from selenium.webdriver.common.action_chains import ActionChains

            actions = ActionChains(driver)
            try:
                actions.move_by_offset(
                    -actions.w3c_actions.pointer_action.x,
                    -actions.w3c_actions.pointer_action.y,
                )
            except Exception:
                pass
            target_x = rect["x"]
            target_y = rect["y"] - 30  # 요청하신 대로 위로 30px 이동
            print(f"[DEBUG] 좌표 기반 5분/60분 클릭 폴백 시도: x={target_x}, y={target_y}")
            actions.move_by_offset(target_x, target_y).click().perform()
            time.sleep(0.2)
    except Exception as e:
        print(f"[DEBUG] 좌표 기반 5분/60분 클릭 폴백 실패: {e}")

    # 5) 최종 값 로그만 남기고, 다음 단계(링크 생성)로 진행
    try:
        trigger = driver.find_element(By.ID, "session-ttl")
        value_span = trigger.find_element(By.CSS_SELECTOR, "span[data-slot='select-value']")
        after = (value_span.text or "").strip()
        print(f"[DEBUG] 세션유효시간 최종 값(after)='{after}'")
    except Exception as e:
        print(f"[DEBUG] 세션유효시간 최종 값 확인 실패: {e}")

    # 성공 여부와 상관 없이, 시간을 더 쓰지 않고 바로 다음 단계로 진행한다.
    print("[INFO] 세션유효시간 변경 시도를 마쳤습니다. 현재 값에 상관없이 링크 생성 단계로 이동합니다.")
    return True


def _find_input_quick(
    driver: webdriver.Chrome,
    css_list: list[str],
    max_wait: float = 3.0,
) -> webdriver.Chrome | None:
    """
    여러 CSS 셀렉터 후보를 짧게(0.2초 간격) 반복해서 검사하며,
    화면에 보이는 첫 번째 input 을 찾는다.
    """
    interval = 0.1 * (2.0 - LINK_CREATION_WAIT_FACTOR)  # 0.05~0.2
    end = time.time() + max_wait
    while time.time() < end:
        for sel in css_list:
            try:
                els = driver.find_elements(By.CSS_SELECTOR, sel)
            except Exception:
                continue
            for el in els:
                try:
                    if el.is_displayed():
                        return el
                except Exception:
                    continue
        time.sleep(interval)
    return None


def _click_button_by_text_retry(
    driver: webdriver.Chrome,
    texts: list[str],
    total_timeout: float,
    description: str,
) -> bool:
    """
    주어진 텍스트 중 하나를 포함하는 버튼/링크를 찾을 때까지 반복 클릭 시도.

    - JS 로 innerText 기반 탐색 후 클릭
    - 안 되면 XPATH 로 button[@aria-label 포함]까지 포함해서 짧게 대기
    - total_timeout 초 안에 성공하면 True, 아니면 False
    """
    end_time = time.time() + total_timeout

    js_code = """
const targets = Array.from(document.querySelectorAll('button,a,div,span'));
for (const el of targets) {
  const t = (el.innerText || '').trim();
  if (!t) continue;
  const norm = t.replace(/\\s+/g, '');
  const target = (arguments[0] || '').trim();
  const targetNorm = target.replace(/\\s+/g, '');
  if (!target) continue;
  if (t.includes(target) || norm.includes(targetNorm)) {
    if (el.offsetParent !== null) {
      el.click();
      return true;
    }
  }
}
return false;
"""

    print(f"[STEP] 버튼 클릭 재시도 시작: {description}")

    while time.time() < end_time:
        # 1차: JS 기반 탐색
        for txt in texts:
            try:
                clicked = driver.execute_script(js_code, txt)
                if clicked:
                    # JS 상으로는 클릭이 실행되었지만,
                    # 실제 상태 변경 여부는 상위 로직에서 따로 검증해야 하므로
                    # 여기서는 '성공'이 아니라 '클릭 시도'로만 기록한다.
                    print(f"[INFO] '{txt}' 텍스트로 {description} 버튼 클릭 시도(JS).")
                    time.sleep(0.3)
                    return True
            except Exception:
                # DOM 로딩 시점 등으로 실패할 수 있으므로 무시하고 다음 시도
                continue

        # 2차: XPATH 기반 짧은 대기
        for txt in texts:
            try:
                wait = WebDriverWait(driver, 0.7)
                # XPath union(|)을 사용해야 하므로 or 대신 | 사용
                xp = (
                    f"//button[contains(normalize-space(.),'{txt}')]"
                    f" | //button[contains(@aria-label,'{txt}')]"
                )
                btn = wait.until(
                    EC.element_to_be_clickable((By.XPATH, xp))
                )
                btn.click()
                print(f"[INFO] '{txt}' 텍스트로 {description} 버튼 클릭 시도(XPATH).")
                time.sleep(0.3)
                return True
            except TimeoutException:
                continue

        time.sleep(0.3)

    print(f"[WARN] {total_timeout:.1f}초 동안 시도했지만 '{description}' 버튼을 찾지 못했습니다.")
    return False


def _click_notice_if_present(driver: webdriver.Chrome) -> None:
    """첫 화면 공지의 '확인 후 로그인' 버튼을 찾아 클릭.

    이전에 실제로 잘 동작하던 단순한 구조로 되돌렸습니다.
    """
    # 1차: XPATH 로 직접 버튼 찾기 (최대 5초)
    try:
        wait = WebDriverWait(driver, 5)
        btn = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, "//button[contains(normalize-space(.),'확인 후 로그인')]")
            )
        )
        btn.click()
        time.sleep(0.3)
        return
    except TimeoutException:
        # 2차: JS로 전체 DOM 을 훑으면서 텍스트 기준으로 재시도
        try:
            driver.execute_script(
                """
const targets = Array.from(document.querySelectorAll('button,a,div,span'));
for (const el of targets) {
  const t = (el.innerText || '').trim();
  if (!t) continue;
  const norm = t.replace(/\\s+/g, '');
  if (t.includes('확인 후 로그인') || norm.includes('확인후로그인')) {
    if (el.offsetParent !== null) {
      el.click();
      return;
    }
  }
}
"""
            )
            time.sleep(0.3)
        except Exception:
            # 공지 팝업이 없거나 이미 닫혀 있으면 조용히 통과
            pass


# 대면결제 페이지 셀렉터 (필요 시 사이트 구조에 맞게 수정)
FACE_TO_FACE_SELECTORS = {
    # 결제 정보
    "card_personal_checkbox": (
        By.XPATH,
        "//span[normalize-space(text())='개인카드']/preceding-sibling::div[1]",
    ),
    "card_business_checkbox": (
        By.XPATH,
        "//span[contains(normalize-space(text()), '사업자')]/preceding-sibling::div[1]",
    ),
    # 카드번호: label 텍스트 기준으로 바로 아래 input
    "card_number": (
        By.XPATH,
        "//label[normalize-space(text())='카드번호']/following::input[1]",
    ),
    # MM / YY 는 label 텍스트 기준으로 select 요소를 찾는다
    "expiry_mm": (
        By.XPATH,
        "//label[normalize-space(text())='MM']/following::select[1]",
    ),
    "expiry_yy": (
        By.XPATH,
        "//label[normalize-space(text())='YY']/following::select[1]",
    ),
    # 비밀번호 앞 두 자리
    "card_password": (
        By.XPATH,
        "//label[normalize-space(text())='비밀번호']/following::input[1]",
    ),
    # 할부개월 select
    "installment_select": (
        By.XPATH,
        "//label[contains(normalize-space(text()), '할부개월')]/following::select[1]",
    ),
    # 연락처 / 이름 / 주민번호 앞자리
    "phone_prefix": (By.XPATH, "//select[contains(@name, 'phone') or contains(., '010')]"),
    "phone_number": (By.XPATH, "//input[@placeholder='연락처를 입력해주세요.']"),
    "customer_name": (
        By.XPATH,
        "//label[normalize-space(text())='이름']/following::input[1]",
    ),
    "resident_front": (
        By.XPATH,
        "//label[contains(normalize-space(text()), '주민등록번호 앞자리')]/following::input[1]",
    ),
    "business_reg_no": (
        By.XPATH,
        "//input[@placeholder='10자리 사업자등록번호를 입력해주세요.']",
    ),
    # 상품정보
    "product_name": (By.XPATH, "//label[contains(., '상품명')]/following::input[1]"),
    "product_price": (By.XPATH, "//label[contains(., '판매가격')]/following::input[1]"),
    # 버튼
    "submit_button": (By.XPATH, "//button[contains(., '결제하기')]"),
}


@dataclass
class PaymentRow:
    login_id: str
    login_password: str
    login_pin: str
    card_type: str
    card_number: str
    expiry_mm: str
    expiry_yy: str
    card_password: str
    installment_months: str
    phone_number: str
    customer_name: str
    resident_front: str
    amount: int
    product_name: str


def load_order_from_json(path: str) -> PaymentRow:
    """web_form.py 가 생성한 JSON 파일에서 1건의 주문 데이터를 읽어온다."""
    if not Path(path).exists():
        raise FileNotFoundError(
            f"{path} 파일을 찾을 수 없습니다. web_form.py 의 웹 폼에서 먼저 데이터를 저장해 주세요."
        )

    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    # ── 필수 필드: amount, login 3종 ─────────────────────────────
    if "amount" not in raw or raw.get("amount") in ("", None, "0"):
        raise ValueError("JSON 데이터에 결제 금액(amount)이 없습니다.")
    if not str(raw.get("product_name") or "").strip():
        raw["product_name"] = "SISA 결제링크"
    if not str(raw.get("login_id") or "").strip():
        raw["login_id"] = os.environ.get("K_VAN_ID", "m3313")
    if not str(raw.get("login_password") or "").strip():
        raw["login_password"] = os.environ.get("K_VAN_PW", "1234")
    if not str(raw.get("login_pin") or "").strip():
        raw["login_pin"] = os.environ.get("K_VAN_PIN", "2424")
    if not str(raw.get("card_type") or "").strip():
        raw["card_type"] = "personal"
    if not str(raw.get("installment_months") or "").strip():
        raw["installment_months"] = "일시불"

    # 카드/고객 정보는 결제링크 생성 전용 모드에서 비어있어도 무방하다 (빈 문자열로 채움)
    for _opt in ("card_number", "expiry_mm", "expiry_yy", "card_password",
                 "phone_number", "customer_name", "resident_front"):
        if _opt not in raw or raw[_opt] is None:
            raw[_opt] = ""
    # ─────────────────────────────────────────────────────────────

    try:
        amount_int = int(str(raw["amount"]).replace(",", "").strip())
    except (TypeError, ValueError) as e:
        raise ValueError(f"amount 값이 숫자가 아닙니다: {raw['amount']!r}") from e

    if amount_int <= 0:
        raise ValueError(f"amount 값이 0 이하입니다: {amount_int}")

    card_type = str(raw.get("card_type", "personal")).strip().lower()
    if card_type not in ("personal", "business"):
        card_type = "personal"

    return PaymentRow(
        login_id=str(raw["login_id"]).strip(),
        login_password=str(raw["login_password"]).strip(),
        login_pin=str(raw["login_pin"]).strip(),
        card_type=card_type,
        card_number=str(raw.get("card_number", "")).strip(),
        expiry_mm=str(raw.get("expiry_mm", "")).strip(),
        expiry_yy=str(raw.get("expiry_yy", "")).strip(),
        card_password=str(raw.get("card_password", "")).strip(),
        installment_months=str(raw.get("installment_months", "일시불")).strip(),
        phone_number=str(raw.get("phone_number", "")).strip(),
        customer_name=str(raw.get("customer_name", "")).strip(),
        resident_front=str(raw.get("resident_front", "")).strip(),
        amount=amount_int,
        product_name=str(raw["product_name"]).strip(),
    )


def _get_chromedriver_path() -> str | None:
    """로컬 tool 폴더 또는 환경변수에 지정된 ChromeDriver 경로."""
    env_path = os.environ.get("CHROMEDRIVER_PATH", "").strip()
    if env_path and Path(env_path).exists():
        return env_path
    base = Path(__file__).resolve().parent / "tool"
    for name in ("chromedriver.exe", "chromedriver"):
        p = base / name
        if p.exists():
            return str(p)
    return None


def create_driver(headless: bool | None = None) -> webdriver.Chrome:
    if headless is None:
        headless = _is_server_env()
    options = webdriver.ChromeOptions()
    if headless:
        options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-setuid-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920,1080")
    else:
        options.add_argument("--start-maximized")
    # Railway 등 서버: 시스템에 설치된 Chrome 사용
    if _is_server_env():
        for binary in ("/usr/bin/google-chrome-stable", "/usr/bin/google-chrome", "/usr/bin/chromium"):
            if Path(binary).exists():
                options.binary_location = binary
                break
    service = None
    driver_path = _get_chromedriver_path()
    if driver_path:
        service = Service(executable_path=driver_path)
    driver = webdriver.Chrome(service=service, options=options) if service else webdriver.Chrome(options=options)
    driver.implicitly_wait(max(3, int(5 * LINK_CREATION_WAIT_FACTOR)))
    return driver


def _parse_amount(text: str) -> int:
    return parse_amount_won(text or "")


def _scrape_dashboard_and_store(driver: webdriver.Chrome) -> None:
    """로그인 후 대시보드 화면에서 매출 요약 정보를 크롤링하여 DB에 저장.

    로컬 테스트 모드(LOCAL_TEST=True)에서는 DB 에 쓰지 않고 바로 리턴한다.
    """
    if LOCAL_TEST:
        print("[LOCAL_TEST] 대시보드 크롤링/DB 저장을 건너뜁니다.")
        return
    try:
        # 대시보드가 렌더링될 시간을 아주 짧게만 준다
        time.sleep(0.5)
        # 너무 오래 기다리지 않도록 짧은 wait 사용
        wait = WebDriverWait(driver, 3)

        # 각 블록은 '월 매출', '전일 매출', '정산 예정 금액', '나의 크레딧' 등의 텍스트를 기준으로 탐색
        def find_block(label_text: str):
            try:
                label_el = wait.until(
                    EC.presence_of_element_located(
                        (By.XPATH, f"//*[normalize-space(text())='{label_text}']")
                    )
                )
                return label_el.find_element(By.XPATH, "./ancestor::div[1]")
            except TimeoutException:
                return None

        monthly_block = find_block("월 매출")
        yesterday_block = find_block("전일 매출")
        settlement_block = find_block("정산 예정 금액")
        credit_block = find_block("나의 크레딧")

        monthly_sales = monthly_approved_cnt = monthly_approved_amt = 0
        monthly_canceled_cnt = monthly_canceled_amt = 0

        if monthly_block:
            # 첫 번째 금액 (월 매출)
            try:
                amt_el = monthly_block.find_element(By.XPATH, ".//*[contains(text(),'원')]")
                monthly_sales = _parse_amount(amt_el.text)
            except Exception:
                pass
            # 승인 / 취소 블록들
            try:
                approve_el = monthly_block.find_element(
                    By.XPATH, ".//*[contains(normalize-space(text()),'승인')]/ancestor::div[1]"
                )
                nums = approve_el.text.splitlines()
                if len(nums) >= 2:
                    monthly_approved_cnt = _parse_amount(nums[0])
                    monthly_approved_amt = _parse_amount(nums[1])
            except Exception:
                pass
            try:
                cancel_el = monthly_block.find_element(
                    By.XPATH, ".//*[contains(normalize-space(text()),'취소')]/ancestor::div[1]"
                )
                nums = cancel_el.text.splitlines()
                if len(nums) >= 2:
                    monthly_canceled_cnt = _parse_amount(nums[0])
                    monthly_canceled_amt = _parse_amount(nums[1])
            except Exception:
                pass

        yesterday_sales = yesterday_approved_cnt = yesterday_approved_amt = 0
        yesterday_canceled_cnt = yesterday_canceled_amt = 0
        if yesterday_block:
            try:
                amt_el = yesterday_block.find_element(By.XPATH, ".//*[contains(text(),'원')]")
                yesterday_sales = _parse_amount(amt_el.text)
            except Exception:
                pass
            try:
                approve_el = yesterday_block.find_element(
                    By.XPATH, ".//*[contains(normalize-space(text()),'승인')]/ancestor::div[1]"
                )
                nums = approve_el.text.splitlines()
                if len(nums) >= 2:
                    yesterday_approved_cnt = _parse_amount(nums[0])
                    yesterday_approved_amt = _parse_amount(nums[1])
            except Exception:
                pass
            try:
                cancel_el = yesterday_block.find_element(
                    By.XPATH, ".//*[contains(normalize-space(text()),'취소')]/ancestor::div[1]"
                )
                nums = cancel_el.text.splitlines()
                if len(nums) >= 2:
                    yesterday_canceled_cnt = _parse_amount(nums[0])
                    yesterday_canceled_amt = _parse_amount(nums[1])
            except Exception:
                pass

        settlement_expected = today_settlement_expected = 0
        if settlement_block:
            try:
                amt_el = settlement_block.find_element(By.XPATH, ".//*[contains(text(),'원')]")
                settlement_expected = _parse_amount(amt_el.text)
            except Exception:
                pass
            # "금일 정산 예정금" 이라는 텍스트가 있으면 그 주변에서 숫자를 찾는다
            try:
                today_el = settlement_block.find_element(
                    By.XPATH, ".//*[contains(normalize-space(text()),'금일 정산 예정금')]/following::div[1]"
                )
                today_settlement_expected = _parse_amount(today_el.text)
            except Exception:
                pass

        credit_amount = 0
        if credit_block:
            try:
                amt_el = credit_block.find_element(By.XPATH, ".//*[contains(text(),'원')]")
                credit_amount = _parse_amount(amt_el.text)
            except Exception:
                pass

        # 최근 거래 내역 텍스트를 간단히 요약 저장 (상세 구조를 몰라서 전체 텍스트 저장)
        recent_summary = ""
        try:
            recent_container = driver.find_element(
                By.XPATH, "//*[contains(normalize-space(text()),'최근 거래 내역')]/ancestor::section[1]"
            )
            recent_summary = recent_container.text.strip()
        except Exception:
            pass

        conn = get_db()
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO kvan_dashboard (
                  captured_at,
                  monthly_sales_amount,
                  monthly_approved_count,
                  monthly_approved_amount,
                  monthly_canceled_count,
                  monthly_canceled_amount,
                  yesterday_sales_amount,
                  yesterday_approved_count,
                  yesterday_approved_amount,
                  yesterday_canceled_count,
                  yesterday_canceled_amount,
                  settlement_expected_amount,
                  today_settlement_expected_amount,
                  credit_amount,
                  recent_tx_summary
                )
                VALUES (NOW(), %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    monthly_sales,
                    monthly_approved_cnt,
                    monthly_approved_amt,
                    monthly_canceled_cnt,
                    monthly_canceled_amt,
                    yesterday_sales,
                    yesterday_approved_cnt,
                    yesterday_approved_amt,
                    yesterday_canceled_cnt,
                    yesterday_canceled_amt,
                    settlement_expected,
                    today_settlement_expected,
                    credit_amount,
                    recent_summary,
                ),
            )
        conn.commit()
        conn.close()
    except Exception as e:
        # 대시보드 크롤링 실패는 치명적이지 않으므로 로그만 남기고 계속 진행
        print(f"[WARN] 대시보드 크롤링/DB 저장 실패: {e}")


def _ensure_kvan_transactions_table() -> None:
    """kvan_transactions 테이블이 없으면 생성."""
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS kvan_transactions (
                  id BIGINT UNSIGNED NOT NULL AUTO_INCREMENT PRIMARY KEY,
                  captured_at DATETIME NOT NULL,
                  merchant_name VARCHAR(255) DEFAULT '',
                  pg_name VARCHAR(100) DEFAULT '',
                  mid VARCHAR(100) DEFAULT '',
                  fee_rate VARCHAR(50) DEFAULT '',
                  tx_type VARCHAR(50) DEFAULT '',
                  amount BIGINT DEFAULT 0,
                  cancel_amount BIGINT DEFAULT 0,
                  payable_amount BIGINT DEFAULT 0,
                  card_company VARCHAR(100) DEFAULT '',
                  card_number VARCHAR(64) DEFAULT '',
                  installment VARCHAR(50) DEFAULT '',
                  approval_no VARCHAR(100) DEFAULT '',
                  registered_at VARCHAR(50) DEFAULT '',
                  raw_text TEXT
                ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;
                """
            )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[WARN] kvan_transactions 테이블 생성 실패: {e}")


def _scrape_transactions_and_store(driver: webdriver.Chrome) -> None:
    """
    K-VAN 결제/취소 거래내역 페이지(/transactions)의 테이블을 크롤링하여
    kvan_transactions 테이블에 저장 (kvan_crawler 와 동일 파싱 규칙).
    """
    if LOCAL_TEST:
        print("[LOCAL_TEST] /transactions 크롤링/DB 저장을 건너뜁니다.")
        if "transactions" in driver.current_url:
            driver.refresh()
        else:
            driver.get("https://store.k-van.app/transactions")
        return

    def _cell_txt(s: str) -> str:
        return re.sub(r"\s+", " ", (s or "").replace("\n", " ").strip())

    try:
        _ensure_kvan_transactions_table()
        if "transactions" in driver.current_url:
            driver.refresh()
        else:
            driver.get("https://store.k-van.app/transactions")
        time.sleep(0.25)
        try:
            WebDriverWait(driver, 16).until(
                EC.presence_of_element_located((By.XPATH, "//table//thead//th"))
            )
            try:
                driver.execute_script(
                    "window.scrollTo(0, Math.max(document.body.scrollHeight, document.documentElement.scrollHeight));"
                )
                time.sleep(0.35)
            except Exception:
                pass
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//table//tbody//tr"))
            )
        except TimeoutException:
            print(
                "[WARN] /transactions 테이블 로딩 타임아웃 — kvan_transactions 는 비우지 않고 유지합니다."
            )
            return

        header_rows = driver.find_elements(By.XPATH, "//table//thead//tr")
        header_candidates: list[list[str]] = []
        for hr in header_rows:
            try:
                cells = hr.find_elements(By.XPATH, ".//th|.//td")
                txts: list[str] = []
                for c in cells:
                    try:
                        html = c.get_attribute("innerHTML") or ""
                        lab = infer_kvan_transaction_header_cell_label(html)
                        if not (lab or "").strip():
                            lab = _cell_txt(c.text)
                        txts.append(lab if (lab or "").strip() else "")
                    except Exception:
                        txts.append("")
                if any((x or "").strip() for x in txts):
                    header_candidates.append(txts)
            except Exception:
                continue

        def _score_header_labels(txts: list[str]) -> int:
            joined = " ".join(txts)
            score = len(txts)
            if "승인번호" in joined:
                score += 80
            if "결제 금액" in joined or "결제금액" in joined:
                score += 40
            if "거래일시" in joined or "등록일" in joined:
                score += 35
            if "거래 유형" in joined or "거래유형" in joined:
                score += 25
            if "MID" in joined:
                score += 10
            return score

        best_headers: list[str] = []
        if header_candidates:
            best_headers = max(header_candidates, key=_score_header_labels)
        if not best_headers:
            print("[WARN] /transactions 헤더 없음 — 저장 생략")
            return

        body_rows: list[list[str]] = []
        for tr in driver.find_elements(By.XPATH, "//table//tbody//tr"):
            try:
                cells = tr.find_elements(By.XPATH, ".//td")
                texts = [_cell_txt(c.text) for c in cells]
                if not any(texts):
                    continue
                body_rows.append(texts)
            except Exception:
                continue

        captured_iso = datetime.utcnow().isoformat()
        snapshot_rows = build_kvan_transactions_snapshots(
            best_headers, body_rows, captured_iso=captured_iso
        )
        if not snapshot_rows and body_rows:
            print(
                f"[WARN] /transactions {len(body_rows)}행 파싱 0건 — 열 매핑 확인 필요. "
                f"헤더={best_headers[:10]} — kvan_transactions DB 유지"
            )
            try:
                print(f"[DEBUG] 첫 행 셀 {len(body_rows[0])}개: {body_rows[0][:12]}")
            except Exception:
                pass
            return
        if not snapshot_rows and not body_rows:
            print("[INFO] /transactions 거래 행 0건 — kvan_transactions 비움")
            conn = get_db()
            with conn.cursor() as cur:
                cur.execute("TRUNCATE TABLE kvan_transactions")
            conn.commit()
            conn.close()
            return

        conn = get_db()
        inserted = 0
        with conn.cursor() as cur:
            cur.execute("TRUNCATE TABLE kvan_transactions")
            for rec in snapshot_rows:
                cur.execute(
                    """
                    INSERT INTO kvan_transactions (
                      captured_at,
                      merchant_name,
                      pg_name,
                      mid,
                      fee_rate,
                      tx_type,
                      amount,
                      cancel_amount,
                      payable_amount,
                      card_company,
                      card_number,
                      installment,
                      approval_no,
                      registered_at,
                      raw_text
                    )
                    VALUES (NOW(), %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                    (
                        rec.get("merchant_name", ""),
                        rec.get("pg_name", ""),
                        rec.get("mid", ""),
                        rec.get("fee_rate", ""),
                        rec.get("tx_type", ""),
                        rec.get("amount", 0),
                        rec.get("cancel_amount", 0),
                        rec.get("payable_amount", 0),
                        rec.get("card_company", ""),
                        rec.get("card_number", ""),
                        rec.get("installment", ""),
                        rec.get("approval_no", ""),
                        rec.get("registered_at", ""),
                        rec.get("raw_text", ""),
                    ),
                )
                inserted += 1
        conn.commit()
        conn.close()
        print(f"[INFO] /transactions 에서 {inserted}건의 거래내역을 kvan_transactions 에 저장했습니다.")
    except Exception as e:
        print(f"[WARN] 거래내역(/transactions) 크롤링/DB 저장 실패: {e}")


def _extract_primary_kvan_key_from_tx_raw(raw_text: str) -> str | None:
    """
    kvan_transactions.raw_text 에서 본 거래에 해당하는 KEY… 세션 토큰 1개를 고른다.
    (첫 KEY 정규식 매칭은 UI에 여러 KEY가 있을 때 오탐이 나기 쉬움)
    """
    rt = (raw_text or "").strip()
    if not rt:
        return None
    for pat in (
        r"sessionId[=:]\s*(KEY[0-9A-Za-z]+)",
        r"sessionid[=:]\s*(KEY[0-9A-Za-z]+)",
    ):
        m = re.search(pat, rt, re.I)
        if m:
            return m.group(1)
    m = re.search(r"/p/(KEY[0-9A-Za-z]+)", rt, re.I)
    if m:
        return m.group(1)
    keys = re.findall(r"(KEY[0-9A-Za-z]+)", rt, re.I)
    if not keys:
        return None
    uniq: list[str] = []
    for k in keys:
        if k not in uniq:
            uniq.append(k)
    return max(uniq, key=len)


def _resolve_agency_id_for_kvan_tx_row(raw_text: str, cur) -> tuple[str | None, str]:
    """
    admin_state 우선, 단일 kvan_links 행이면 DB agency_id 보조.
    반환: (agency_id None/빈=본사, 선택된 KEY 또는 '')
    """
    key = _extract_primary_kvan_key_from_tx_raw(raw_text)
    if not key:
        return None, ""
    aid = _get_agency_id_for_session(key)
    if aid:
        return aid, key
    try:
        cur.execute(
            """
            SELECT agency_id FROM kvan_links
            WHERE kvan_link LIKE %s OR kvan_session_id LIKE %s
               OR internal_session_id = %s
            """,
            (f"%{key}%", f"%{key}%", key),
        )
        rows = cur.fetchall() or []
        if len(rows) == 1:
            db_ag = str(rows[0].get("agency_id") or "").strip()
            return (db_ag if db_ag else None), key
        ags = [
            str(r.get("agency_id") or "").strip()
            for r in rows
            if str(r.get("agency_id") or "").strip()
        ]
        if ags and len(set(ags)) == 1:
            return ags[0], key
    except Exception:
        pass
    return (aid or None), key


def _sync_kvan_to_transactions() -> bool:
    """
    kvan_transactions 에 쌓인 K-VAN 거래내역을
    내부 transactions 테이블과 최대한 매핑하고,
    필요하면 새 거래 레코드를 생성하여 본사/대행사 정산 데이터와 연결한다.

    우선순위:
    1) approval_no(승인번호) 기준으로 기존 transactions 찾기
    2) 없으면 amount + 날짜(+ agency_id) 기준으로 가장 최근 거래 매핑
    3) 그래도 없으면 신규 transactions 레코드 INSERT

    agency_id 매핑:
    - admin_state 의 세션(KEY) 기준 _get_agency_id_for_session 만 사용한다.
    - MID→대행사 폴백은 사용하지 않는다(동일 MID 공유로 구분 불가).
    """
    if LOCAL_TEST:
        print("[LOCAL_TEST] kvan_transactions → transactions 매핑/생성을 건너뜁니다.")
        return False

    updated = 0
    inserted = 0
    try:
        conn = get_db()
        with conn.cursor() as cur:
            # 1) 최근 K-VAN 거래 200건만 사용 (raw_text로 세션 KEY 추출 → 대행사 구분)
            cur.execute(
                """
                SELECT id, captured_at, merchant_name, mid, tx_type,
                       amount, approval_no, registered_at, raw_text
                FROM kvan_transactions
                ORDER BY captured_at DESC
                LIMIT 200
                """
            )
            krows = cur.fetchall()

            for kr in krows:
                amt = kr.get("amount") or 0
                approval = (kr.get("approval_no") or "").strip()
                mid = (kr.get("mid") or "").strip()
                tx_type = (kr.get("tx_type") or "").strip()
                reg = (kr.get("registered_at") or "").strip()
                raw_text = (kr.get("raw_text") or "").strip()
                if not amt or not approval:
                    # 금액/승인번호가 없으면 내부 거래와 매핑하기 어려우므로 건너뜀
                    continue

                # 등록일에서 날짜 부분만 추출 (예: '2026-03-12 10:20:30' -> '2026-03-12')
                reg_date = reg.split(" ")[0] if reg else ""

                # 대행사/본사 구분: raw_text 에서 주 세션 KEY 추정 → admin_state → 필요 시 kvan_links 단일 행
                agency_id, chosen_key = _resolve_agency_id_for_kvan_tx_row(raw_text, cur)
                if chosen_key:
                    print(
                        f"[KVAN-TX-SYNC] approval={approval} key={chosen_key} "
                        f"agency_id={(agency_id or '')!r}"
                    )

                # K-VAN 결제유형 기준으로 내부 status 유추
                tx_status = "other"
                tx_type_text = tx_type or ""
                if "승인" in tx_type_text:
                    tx_status = "success"
                elif "취소" in tx_type_text or "실패" in tx_type_text or "오류" in tx_type_text:
                    tx_status = "fail"

                # 1단계: 승인번호로 기존 거래 찾기
                cur.execute(
                    """
                    SELECT id, agency_id
                    FROM transactions
                    WHERE kvan_approval_no = %s
                    LIMIT 1
                    """,
                    (approval,),
                )
                tx = cur.fetchone()
                if tx:
                    tx_id = tx["id"]
                    # agency_id 가 비어 있고, 이번 K-VAN 에서 대행사를 알 수 있다면 채워 넣는다.
                    fill_agency = (agency_id or "").strip()
                    cur.execute(
                        """
                        UPDATE transactions
                        SET amount = COALESCE(amount, %s),
                            status = %s,
                            kvan_mid = %s,
                            kvan_approval_no = %s,
                            kvan_tx_type = %s,
                            kvan_registered_at = %s,
                            agency_id = COALESCE(NULLIF(TRIM(agency_id), ''), %s)
                        WHERE id = %s
                        """,
                        (amt, tx_status, mid, approval, tx_type, reg, fill_agency, tx_id),
                    )
                    updated += 1
                    continue

                # 2단계: 금액 + 날짜(+ agency_id) 기준으로 기존 거래 찾기
                params: list = [amt, reg_date, reg_date]
                sql = """
                    SELECT id, agency_id
                    FROM transactions
                    WHERE amount = %s
                      AND (%s = '' OR DATE(created_at) = %s)
                """
                if agency_id:
                    sql += " AND agency_id = %s"
                    params.append(agency_id)
                sql += """
                    ORDER BY created_at DESC
                    LIMIT 1
                """
                cur.execute(sql, tuple(params))
                tx = cur.fetchone()
                if tx:
                    tx_id = tx["id"]
                    cur.execute(
                        """
                        UPDATE transactions
                        SET status = %s,
                            kvan_mid = %s,
                            kvan_approval_no = %s,
                            kvan_tx_type = %s,
                            kvan_registered_at = %s
                        WHERE id = %s
                        """,
                        (tx_status, mid, approval, tx_type, reg, tx_id),
                    )
                    updated += 1
                    continue

                # 3단계: 매칭되는 기존 거래가 없으면 새 transactions 레코드 생성
                new_tx_id = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")[-18:]
                message = f"K-VAN {tx_type or '거래'} 자동 연동 (MID={mid}, 승인번호={approval})"
                cur.execute(
                    """
                    INSERT INTO transactions (
                      id,
                      created_at,
                      agency_id,
                      amount,
                      customer_name,
                      phone_number,
                      card_type,
                      resident_front,
                      status,
                      message,
                      settlement_status,
                      settled_at,
                      kvan_mid,
                      kvan_approval_no,
                      kvan_tx_type,
                      kvan_registered_at
                    )
                    VALUES (
                      %s, NOW(), %s, %s,
                      '', '', '', '',
                      %s, %s,
                      '미정산', NULL,
                      %s, %s, %s, %s
                    )
                    """,
                    (
                        new_tx_id,
                        (agency_id or "").strip(),
                        amt,
                        tx_status,
                        message,
                        mid,
                        approval,
                        tx_type,
                        reg,
                    ),
                )
                inserted += 1

        conn.commit()
        conn.close()
        if updated or inserted:
            print(
                f"[INFO] kvan_transactions 기반으로 내부 거래 매핑/생성 완료 "
                f"(updated={updated}, inserted={inserted})"
            )
    except Exception as e:
        print(f"[WARN] K-VAN ↔ 내부 transactions 매핑/생성 중 오류: {e}")
    # 크롤러와 동일: 신규 INSERT 만 '작업 발생'으로 본다(매 사이클 UPDATE 로 빠른 폴링 고착 방지)
    return bool(inserted)


def _ensure_kvan_links_table() -> None:
    """kvan_links 테이블이 없으면 생성."""
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS kvan_links (
                  id BIGINT UNSIGNED NOT NULL AUTO_INCREMENT PRIMARY KEY,
                  captured_at DATETIME NOT NULL,
                  link_created_at DATETIME NULL DEFAULT NULL,
                  title VARCHAR(255) DEFAULT '',
                  amount BIGINT DEFAULT 0,
                  ttl_label VARCHAR(100) DEFAULT '',
                  status VARCHAR(100) DEFAULT '',
                  kvan_link VARCHAR(512) DEFAULT '',
                  mid VARCHAR(100) DEFAULT '',
                  kvan_session_id VARCHAR(100) DEFAULT '',
                  agency_id VARCHAR(64) DEFAULT '',
                  internal_session_id VARCHAR(64) DEFAULT '',
                  raw_text TEXT
                ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;
                """
            )
            try:
                cur.execute(
                    "ALTER TABLE kvan_links ADD COLUMN agency_id VARCHAR(64) DEFAULT ''"
                )
            except Exception:
                pass
            try:
                cur.execute(
                    "ALTER TABLE kvan_links ADD COLUMN internal_session_id VARCHAR(64) DEFAULT ''"
                )
            except Exception:
                pass
            try:
                cur.execute(
                    """
                    SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'kvan_links'
                      AND COLUMN_NAME = 'link_created_at'
                    """
                )
                if not (cur.fetchall() or []):
                    cur.execute(
                        "ALTER TABLE kvan_links ADD COLUMN link_created_at DATETIME NULL DEFAULT NULL"
                    )
            except Exception:
                pass
            try:
                cur.execute(
                    "UPDATE kvan_links SET link_created_at = captured_at WHERE link_created_at IS NULL"
                )
            except Exception:
                pass
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[WARN] kvan_links 테이블 생성 실패: {e}")


def _scrape_payment_links_and_store(driver: webdriver.Chrome) -> None:
    """
    K-VAN 결제링크 관리 페이지(/payment-link)의 리스트를 크롤링하여
    kvan_links 테이블에 저장.

    화면 구조는 React 카드/테이블 형태일 수 있으므로,
    - 링크 URL (https://store.k-van.app/...)
    - 인근 텍스트(상품명, 금액, 유효시간, 상태, MID, 세션ID 등)를 모두 raw_text 로 저장하고
    - 자주 쓰는 필드(제목/금액/유효시간/status/MID/세션ID)는 휴리스틱하게 추출한다.
    """
    if LOCAL_TEST:
        print("[LOCAL_TEST] /payment-link 크롤링/DB 저장을 건너뜁니다.")
        # 항상 실제 새로고침 또는 첫 진입 (진입 실패 시에도 DB 작업은 건너뛰고 종료)
        if not _go_to_payment_link(driver):
            raise RuntimeError("[NAV] LOCAL_TEST 모드에서 /payment-link 로 진입하지 못했습니다.")
        driver.refresh()
        _wait_payment_link_page_ready(driver)
        return

    try:
        _ensure_kvan_links_table()
        # /payment-link 로 안정적으로 진입 후, 항상 새로고침
        if not _go_to_payment_link(driver):
            raise RuntimeError("[NAV] /payment-link 로 진입하지 못해 링크 리스트 크롤링을 중단합니다.")
        driver.refresh()
        _wait_payment_link_page_ready(driver)
        wait = WebDriverWait(driver, 10)

        # 실제 카드/테이블이 렌더링될 때까지 대기:
        # 링크 텍스트 또는 행/아이콘이 보이면 진행한다.
        try:
            wait.until(
                EC.presence_of_element_located(
                    (
                        By.XPATH,
                        "//*[contains(text(),'https://store.k-van.app') "
                        "or contains(@value,'https://store.k-van.app') "
                        "or .//button[@title='거래 내역']]",
                    )
                )
            )
        except TimeoutException:
            print("[WARN] /payment-link 링크 텍스트 대기 타임아웃 - 현재 렌더된 행으로 계속 진행")

        # 각 링크 요소를 기준으로 상위 카드 컨테이너를 찾는다.
        link_elements = driver.find_elements(
            By.XPATH,
            "//*[contains(text(),'https://store.k-van.app') "
            "or contains(@value,'https://store.k-van.app')]",
        )
        # 일부 UI에서는 링크 URL이 텍스트로 노출되지 않고 세션ID(KEY...)만 보인다.
        # 이 경우엔 KEY를 추출해 표준 URL을 구성해서 저장한다.
        fallback_rows: list = []
        if not link_elements:
            fallback_rows = driver.find_elements(
                By.XPATH,
                "//tr[.//*[contains(normalize-space(.),'KEY20')]]"
                " | //*[@role='row'][.//*[contains(normalize-space(.),'KEY20')]]"
                " | //div[contains(@class,'rounded') and .//*[contains(normalize-space(.),'KEY20')]]",
            )
            if not fallback_rows:
                print("[INFO] /payment-link 에 표시된 결제링크가 없습니다.")
                return

        conn = get_db()
        inserted = 0
        with conn.cursor() as cur:
            # 기본 모드: 링크 URL 텍스트 기반 수집
            for el in link_elements:
                try:
                    # 링크 문자열 추출
                    link_text = (el.text or "").strip()
                    if not link_text:
                        link_text = (el.get_attribute("value") or "").strip()
                    if not link_text:
                        link_text = (el.get_attribute("href") or "").strip()
                    if not link_text:
                        continue

                    # 링크 URL의 querystring에서 sessionId를 우선 추출해 매핑 정확도를 높인다.
                    parsed_session_id = ""
                    try:
                        q = parse_qs(urlparse(link_text).query)
                        parsed_session_id = str((q.get("sessionId") or [""])[0] or "").strip()
                    except Exception:
                        parsed_session_id = ""

                    # 카드/행 컨테이너: 가장 가까운 div[role='row'] 또는 카드형 div
                    container = el
                    for _ in range(5):
                        container = container.find_element(By.XPATH, "./parent::*")
                        cls = container.get_attribute("class") or ""
                        if "border" in cls or "rounded" in cls or "shadow" in cls or "row" in cls:
                            break

                    card_text = container.text.strip()

                    # 제목: 상호명(가맹점) → 상품명 → 첫 줄 (K-VAN UI 기준)
                    lines = [ln.strip() for ln in card_text.splitlines() if ln.strip()]
                    title = ""
                    for ln in lines:
                        if "상호명" in ln:
                            title = ln
                            break
                    if not title:
                        for ln in lines:
                            if "상품명" in ln:
                                title = ln
                                break
                    if not title:
                        title = lines[0] if lines else ""

                    # 금액: '원' 포함 줄에서 숫자 추출 (라벨/복수 금액 대응)
                    amount = 0
                    for ln in lines:
                        if "원" in ln:
                            amt = _parse_amount(ln)
                            if amt:
                                amount = amt
                                break

                    # 유효시간/TTL: '분' 텍스트가 있는 줄 추출 (예: '60분 (긴 결제 플로우)')
                    ttl_label = ""
                    for ln in lines:
                        if ("분" in ln and "유효" in ln) or ("세션" in ln):
                            ttl_label = ln
                            break

                    # 상태: '사용중', '만료', '취소' 등 단어가 포함된 줄 추출
                    # 카드가 이미 만료(만료일 경과 또는 배지 '만료')면 status='만료'로 저장해 다음 사이클에서 mark_expired가 DELETE 대상으로 인식
                    status = _extract_status_from_link_lines(lines)
                    if "만료" not in (status or ""):
                        expire_at = _extract_expire_at_from_lines(lines)
                        if expire_at is not None and expire_at < _kvan_now():
                            status = "만료"
                    # resolved_session_id는 아래에서 설정됨

                    # MID / 세션ID: 'MID' 또는 '세션' 텍스트 기반
                    mid = ""
                    kvan_session_id = ""
                    for ln in lines:
                        if "MID" in ln.upper():
                            mid = ln
                        if "세션" in ln or "Session" in ln:
                            kvan_session_id = ln
                    resolved_session_id = parsed_session_id
                    if not resolved_session_id and kvan_session_id:
                        m = re.search(r"(KEY[0-9A-Za-z]+)", kvan_session_id)
                        if m:
                            resolved_session_id = m.group(1)
                    if not resolved_session_id:
                        for ln in lines:
                            m = re.search(r"(KEY[0-9A-Za-z]+)", ln)
                            if m:
                                resolved_session_id = m.group(1)
                                break

                    sid_ag = (resolved_session_id or "").strip()
                    if not sid_ag:
                        m = re.search(r"(KEY[0-9A-Za-z]+)", str(kvan_session_id or ""))
                        if m:
                            sid_ag = m.group(1)
                    if not sid_ag and link_text:
                        m2 = re.search(r"(KEY[0-9A-Za-z]+)", link_text, re.IGNORECASE)
                        if m2:
                            sid_ag = m2.group(1)
                    row_agency_id = (
                        (_get_agency_id_for_session(sid_ag) or "").strip() if sid_ag else ""
                    )

                    cur.execute(
                        "SELECT agency_id, internal_session_id, link_created_at FROM kvan_links WHERE kvan_link = %s LIMIT 1",
                        (link_text,),
                    )
                    prev_kl = cur.fetchone() or {}
                    pres_ag = (prev_kl.get("agency_id") or "").strip()
                    pres_int = (prev_kl.get("internal_session_id") or "").strip()
                    pres_lc = prev_kl.get("link_created_at")
                    try:
                        parsed_ui_created = parse_kvan_link_ui_created_at(card_text)
                    except Exception:
                        parsed_ui_created = None
                    link_created_val = pres_lc or parsed_ui_created
                    if not row_agency_id and pres_ag:
                        row_agency_id = pres_ag
                    internal_sid = pres_int
                    if not internal_sid and sid_ag:
                        internal_sid = _lookup_internal_session_id_for_kvan_key(sid_ag)

                    # 동일 링크가 매 사이클 누적 저장되지 않도록 기존 행을 제거 후 최신 스냅샷 저장
                    cur.execute("DELETE FROM kvan_links WHERE kvan_link = %s", (link_text,))

                    cur.execute(
                        """
                        INSERT INTO kvan_links (
                          captured_at,
                          link_created_at,
                          title,
                          amount,
                          ttl_label,
                          status,
                          kvan_link,
                          mid,
                          kvan_session_id,
                          agency_id,
                          internal_session_id,
                          raw_text
                        )
                        VALUES (NOW(), IFNULL(%s, NOW()), %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        """,
                        (
                            link_created_val,
                            title,
                            amount,
                            ttl_label,
                            status,
                            link_text,
                            mid,
                            resolved_session_id or kvan_session_id,
                            row_agency_id,
                            internal_sid,
                            card_text,
                        ),
                    )
                    inserted += 1
                    if (status or "").strip() == "만료":
                        _expired_debug(f"[EXPIRED_DEBUG] 스크래핑: kvan_links INSERT status='만료' (kvan_session_id={resolved_session_id or kvan_session_id or ''})")

                    # 크롤링으로 새로 얻은 링크를 admin_state.json 에도 즉시 반영
                    # (sessionId 가 있으면 세션 ID 기준으로 우선 매칭)
                    if link_text:
                        try:
                            st_data = _load_admin_state()
                            matched = False
                            for s in st_data.get("sessions") or []:
                                sid = str(s.get("id") or "")
                                if (resolved_session_id and sid and sid == resolved_session_id) or \
                                   (sid and title and sid in title):
                                    if s.get("kvan_link") != link_text:
                                        s["kvan_link"] = link_text
                                        matched = True
                            if matched:
                                _save_admin_state(st_data)
                        except Exception:
                            pass
                except Exception as e_row:
                    print(f"[WARN] 결제링크 카드 파싱/저장 중 오류: {e_row}")
                    continue

            # 폴백 모드: KEY 세션ID 기반 수집
            if not link_elements and fallback_rows:
                for row in fallback_rows:
                    try:
                        row_text = (row.text or "").strip()
                        if not row_text:
                            continue
                        lines = [ln.strip() for ln in row_text.splitlines() if ln.strip()]
                        sid = ""
                        for ln in lines:
                            m = re.search(r"(KEY[0-9A-Za-z]+)", ln)
                            if m:
                                sid = m.group(1)
                                break
                        if not sid:
                            continue
                        link_text = f"https://store.k-van.app/p/{sid}?sessionId={sid}&type=KEYED"
                        title = ""
                        amount = 0
                        status = ""
                        ttl_label = ""
                        mid = ""
                        for ln in lines:
                            if not title and "상호명" in ln:
                                title = ln
                            if not title and "상품명" in ln:
                                title = ln
                            if not amount and "원" in ln:
                                amount = _parse_amount(ln) or 0
                            if not ttl_label and ("분" in ln):
                                ttl_label = ln
                            if not status:
                                status = _extract_status_from_link_lines([ln])
                            if not mid and "MID" in ln.upper():
                                mid = ln
                        if not title:
                            title = lines[0] if lines else ""
                        if "만료" not in (status or ""):
                            expire_at = _extract_expire_at_from_lines(lines)
                            if expire_at is not None and expire_at < _kvan_now():
                                status = "만료"
                        row_agency_fb = (_get_agency_id_for_session(sid) or "").strip()
                        cur.execute(
                            "SELECT agency_id, internal_session_id, link_created_at FROM kvan_links WHERE kvan_link = %s LIMIT 1",
                            (link_text,),
                        )
                        prev_fb = cur.fetchone() or {}
                        if not row_agency_fb:
                            row_agency_fb = (prev_fb.get("agency_id") or "").strip()
                        internal_fb = (prev_fb.get("internal_session_id") or "").strip() or _lookup_internal_session_id_for_kvan_key(sid)
                        pres_lc_fb = prev_fb.get("link_created_at")
                        try:
                            parsed_ui_fb = parse_kvan_link_ui_created_at(row_text)
                        except Exception:
                            parsed_ui_fb = None
                        link_created_fb = pres_lc_fb or parsed_ui_fb
                        cur.execute("DELETE FROM kvan_links WHERE kvan_link = %s", (link_text,))
                        cur.execute(
                            """
                            INSERT INTO kvan_links (
                              captured_at, link_created_at, title, amount, ttl_label, status, kvan_link, mid, kvan_session_id, agency_id, internal_session_id, raw_text
                            )
                            VALUES (NOW(), IFNULL(%s, NOW()), %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                            """,
                            (
                                link_created_fb,
                                title,
                                amount,
                                ttl_label,
                                status,
                                link_text,
                                mid,
                                sid,
                                row_agency_fb,
                                internal_fb,
                                row_text,
                            ),
                        )
                        inserted += 1
                        try:
                            st_data = _load_admin_state()
                            matched = False
                            for s in st_data.get("sessions") or []:
                                if str(s.get("id") or "") == sid:
                                    if s.get("kvan_link") != link_text:
                                        s["kvan_link"] = link_text
                                        matched = True
                            if matched:
                                _save_admin_state(st_data)
                        except Exception:
                            pass
                    except Exception as e_row2:
                        print(f"[WARN] 결제링크 폴백 파싱/저장 중 오류: {e_row2}")
                        continue

        conn.commit()
        conn.close()
        print(f"[INFO] /payment-link 에서 {inserted}건의 결제링크 정보를 kvan_links 에 저장했습니다.")
    except Exception as e:
        print(f"[WARN] 결제링크 관리(/payment-link) 크롤링/DB 저장 실패: {e}")


def _is_expired_link_status(status_text: str) -> bool:
    """
    결제링크 상태 문자열이 '실제 만료/취소'인지 판별한다.

    주의:
    - '취소 가능' 같은 문구는 만료/취소 완료 상태가 아니므로 제외한다.
    """
    s = str(status_text or "").strip()
    if not s:
        return False
    # 테이블 헤더/설명 텍스트(예: "만료일시")는 만료 상태로 보지 않는다.
    if "만료일시" in s:
        return False
    if "만료" in s:
        return True
    if "취소 가능" in s or "취소가능" in s:
        return False
    if s in ("취소", "취소됨", "취소 완료", "취소완료"):
        return True
    return False


def _extract_status_from_link_lines(lines: list[str]) -> str:
    """
    결제링크 카드/행 텍스트에서 실제 상태 문구만 추출한다.
    (헤더 문구의 '만료일시' 같은 단어로 인한 오탐 방지)
    """
    if not lines:
        return ""
    header_markers = (
        "생성일시", "만료일시", "세션ID", "작업", "상호명", "상품명",
        "유효시간", "본인인증", "결제 방식", "PG사", "MID"
    )
    exact_statuses = {
        "사용", "사용중", "사용 중", "대기", "완료", "만료",
        "취소", "취소됨", "취소 완료", "취소완료",
    }
    for raw in lines:
        ln = str(raw or "").strip()
        if not ln:
            continue
        if "취소 가능" in ln or "취소가능" in ln:
            continue
        if any(h in ln for h in header_markers):
            continue
        compact = ln.replace(" ", "")
        if ln in exact_statuses or compact in {x.replace(" ", "") for x in exact_statuses}:
            return ln
        if "상태" in ln and any(k in ln for k in ("사용", "대기", "완료", "만료", "취소")):
            return ln
    return ""


def _extract_expire_at_from_lines(lines: list[str]) -> datetime | None:
    """
    카드 텍스트 라인에서 만료일시를 추출한다.
    예: "만료일: 2026-03-16 16:20:31"
    """
    for raw in lines or []:
        ln = str(raw or "").strip()
        if "만료일" not in ln:
            continue
        m = re.search(r"(20\d{2}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})", ln)
        if not m:
            continue
        ts = m.group(1).strip()
        try:
            return datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
        except Exception:
            continue
    return None


def _kvan_now() -> datetime:
    """
    K-VAN 화면의 시각 기준으로 현재 시간을 계산한다.
    기본값은 KST(+9)이며, 필요 시 K_VAN_TZ_OFFSET_HOURS 로 조정 가능.
    """
    try:
        offset_hours = int(os.environ.get("K_VAN_TZ_OFFSET_HOURS", "9"))
    except Exception:
        offset_hours = 9
    return datetime.utcnow() + timedelta(hours=offset_hours)


def mark_expired_sessions_from_kvan_links() -> None:
    """
    kvan_links 테이블에서 상태가 '만료'인 링크 URL 목록을 조회한 뒤,
    admin_state.json 의 진행 중 세션 중 해당 링크를 가진 세션을 종료 처리한다.

    정책:
    - 만료/취소 링크 데이터(kvan_links)는 DB에서 자동 삭제한다.
    - 거래내역(transactions)은 자동 삭제하지 않는다.
    - admin_state 의 진행중 세션은 즉시 제거하고, history에
      status='만료', deleted_in_kvan=True 로 남긴다.
    """
    _expired_debug("[EXPIRED_DEBUG] mark_expired_sessions_from_kvan_links 진입")
    for attempt in range(1, 4):
        conn = None
        try:
            conn = _get_db_with_retry()
            _expired_debug(f"[EXPIRED_DEBUG] DB 연결 성공 (attempt={attempt})")
            expired_urls: set[str] = set()
            # 만료+거래있음 세션은 삭제 제외 (어드민 알림용으로 유지)
            excluded_sids: set[str] = set()
            if EXPIRED_WITH_TRANSACTIONS_PATH.exists():
                try:
                    data = json.loads(EXPIRED_WITH_TRANSACTIONS_PATH.read_text(encoding="utf-8"))
                    for item in (data if isinstance(data, list) else []):
                        sid = (item.get("session_id") or "").strip()
                        if sid:
                            excluded_sids.add(sid)
                except Exception:
                    pass
            _expired_debug(f"[EXPIRED_DEBUG] 제외 세션(만료+거래있음) 수: {len(excluded_sids)}, sid_sample={list(excluded_sids)[:3]}")

            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT kvan_link, status, kvan_session_id
                    FROM kvan_links
                    WHERE kvan_link IS NOT NULL AND kvan_link != ''
                    """,
                )
                rows = cur.fetchall() or []
            _expired_debug(f"[EXPIRED_DEBUG] kvan_links 전체 행 수: {len(rows)}")
            for i, row in enumerate(rows):
                url = (row.get("kvan_link") or "").strip()
                status_text = str(row.get("status") or "").strip()
                sid = (row.get("kvan_session_id") or "").strip()
                is_exp = _is_expired_link_status(status_text)
                in_excl = sid in excluded_sids
                if i < 5:
                    _expired_debug(f"[EXPIRED_DEBUG]   row[{i}] status={repr(status_text)}, sid={sid[:24] if sid else ''}..., is_expired={is_exp}, excluded={in_excl}")
                if url and is_exp and not in_excl:
                    expired_urls.add(url)
            if len(rows) > 5:
                _expired_debug(f"[EXPIRED_DEBUG]   ... 외 {len(rows)-5}건 (동일 조건으로 expired_urls 집합에 반영)")
            _expired_debug(f"[EXPIRED_DEBUG] 만료로 판별된 URL 수(expired_urls): {len(expired_urls)}, sample={list(expired_urls)[:2]}")
            if not expired_urls:
                _expired_debug("[EXPIRED_DEBUG] expired_urls 비어 있음 → 삭제할 대상 없음, return")
                conn.close()
                return
            st = _load_admin_state()
            sessions = list(st.get("sessions") or [])
            history = list(st.get("history") or [])
            _expired_debug(f"[EXPIRED_DEBUG] admin_state sessions 수: {len(sessions)}, history 수: {len(history)}")
            remaining_sessions: list[dict] = []
            removed_count = 0
            now_iso = datetime.utcnow().isoformat()

            for s in sessions:
                link = (s.get("kvan_link") or "").strip()
                if link and link in expired_urls:
                    removed_count += 1
                    if removed_count <= 3:
                        _expired_debug(f"[EXPIRED_DEBUG]   admin_state 매칭 제거 session_id={s.get('id')}, link_len={len(link)}")
                    sid = str(s.get("id") or "")
                    s["status"] = "만료"
                    s["deleted"] = True
                    s["deleted_in_kvan"] = True
                    s["deleted_at"] = now_iso
                    s["finished_at"] = s.get("finished_at") or now_iso
                    old_msg = str(s.get("result_message") or "").strip()
                    mark_msg = "만료 감지로 K-VAN 링크가 삭제되었습니다."
                    s["result_message"] = f"{old_msg}\n{mark_msg}".strip() if old_msg else mark_msg
                    history = _upsert_history_by_session_id(history, dict(s))
                    _append_admin_log(
                        "AUTO",
                        f"만료 링크 세션 정리 session_id={sid}, link={link[:50]}...",
                    )
                else:
                    remaining_sessions.append(s)

            st["sessions"] = remaining_sessions
            st["history"] = history
            _save_admin_state(st)
            if expired_urls:
                with conn.cursor() as cur:
                    placeholders = ",".join(["%s"] * len(expired_urls))
                    cur.execute(
                        f"DELETE FROM kvan_links WHERE kvan_link IN ({placeholders})",
                        tuple(expired_urls),
                    )
                    deleted_count = cur.rowcount
                _expired_debug(f"[EXPIRED_DEBUG] DELETE FROM kvan_links 실행 완료, 삭제된 행 수: {deleted_count}")
            else:
                _expired_debug("[EXPIRED_DEBUG] expired_urls 비어 있어 DELETE 미실행")
            conn.commit()
            conn.close()
            if removed_count:
                _append_admin_log(
                    "AUTO",
                    f"만료/취소 링크 DB 정리 완료 (세션 {removed_count}건, 링크 {len(expired_urls)}건)",
                )
            return
        except Exception as e:
            try:
                if conn:
                    conn.close()
            except Exception:
                pass
            if attempt < 3 and _is_retryable_db_error(e):
                _append_admin_log("AUTO", f"[WARN] 링크 만료 세션 반영 재시도 {attempt}/3: {e}")
                time.sleep(0.7 * attempt)
                continue
            print(f"[WARN] 링크 만료 세션 반영 실패: {e}")
            return


def _sync_popup_transaction_to_internal(
    session_id: str,
    amount: int,
    approval_no: str,
    card_number: str,
    registered_at: str,
    customer_name: str,
) -> None:
    """
    결제링크 관리 화면의 '거래 내역' 팝업에서 얻은 승인 정보를
    내부 transactions 테이블과 admin_state.json 에 반영한다.

    - session_id 를 기준으로 admin_state.json 에서 agency_id 를 찾는다.
    - kvan_approval_no 가 같은 transactions 가 있으면 업데이트,
      없으면 새 레코드를 INSERT 한다.
    """
    if LOCAL_TEST:
        print("[LOCAL_TEST] 팝업 기반 transactions 동기화를 건너뜁니다.")
        return

    approval_no = (approval_no or "").strip()
    if not approval_no or not amount:
        return

    # session_id(KEY... 또는 우리 세션 id)로 agency_id 찾기. KEY... 는 kvan_link 파싱으로 매칭.
    agency_id: str | None = _get_agency_id_for_session(session_id)

    try:
        conn = get_db()
        with conn.cursor() as cur:
            # 1) 승인번호로 기존 거래 찾기
            cur.execute(
                """
                SELECT id FROM transactions
                WHERE kvan_approval_no = %s
                LIMIT 1
                """,
                (approval_no,),
            )
            row = cur.fetchone()
            if row:
                tx_id = row["id"]
                cur.execute(
                    """
                    UPDATE transactions
                    SET amount = COALESCE(amount, %s),
                        customer_name = COALESCE(customer_name, %s),
                        status = 'success',
                        kvan_registered_at = %s,
                        agency_id = COALESCE(agency_id, %s)
                    WHERE id = %s
                    """,
                    (amount, customer_name or "", registered_at, agency_id, tx_id),
                )
            else:
                # 2) 새 레코드 생성
                new_tx_id = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")[-18:]
                message = (
                    f"K-VAN 결제 승인 (세션ID={session_id}, 승인번호={approval_no}, 카드={card_number})"
                )
                cur.execute(
                    """
                    INSERT INTO transactions (
                      id,
                      created_at,
                      agency_id,
                      amount,
                      customer_name,
                      phone_number,
                      card_type,
                      resident_front,
                      status,
                      message,
                      settlement_status,
                      settled_at,
                      kvan_mid,
                      kvan_approval_no,
                      kvan_tx_type,
                      kvan_registered_at
                    )
                    VALUES (
                      %s, NOW(), %s, %s,
                      %s, '', '', '',
                      'success', %s,
                      '미정산', NULL,
                      '', %s, '결제 승인', %s
                    )
                    """,
                    (
                        new_tx_id,
                        agency_id,
                        amount,
                        customer_name or "",
                        message,
                        approval_no,
                        registered_at,
                    ),
                )
                # 결제 완료 알림 큐에 추가 (본사/대행사 어드민에서 미확인 알림 표시용)
                _append_payment_notification(
                    agency_id=agency_id or "",
                    amount=amount,
                    tx_id=new_tx_id,
                    customer_name=customer_name or "",
                )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[WARN] popup 기반 transactions 동기화 실패: {e}")


def _append_payment_notification(
    agency_id: str,
    amount: int,
    tx_id: str,
    customer_name: str,
) -> None:
    """결제 완료 시 알림 큐 파일에 추가한다. (카카오/문자 연동은 추후 확장)"""
    if not tx_id or amount <= 0:
        return
    try:
        path = DATA_DIR / "payment_notifications.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        items: list[dict] = []
        if path.exists():
            try:
                items = json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                items = []
        items.append({
            "agency_id": agency_id or "",
            "amount": amount,
            "tx_id": tx_id,
            "customer_name": customer_name or "",
            "created_at": datetime.utcnow().isoformat(),
            "seen": False,
        })
        # 최근 500건만 유지
        items = items[-500:]
        path.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        print(f"[WARN] 결제 알림 큐 추가 실패: {e}")


def _close_dialog(driver: webdriver.Chrome, dialog) -> None:
    """
    '거래 내역' 팝업을 안전하게 닫고, 오버레이가 사라질 때까지 짧게 대기한다.
    대기 시간 0.3초로 제한해 리스트 크롤링 속도를 높인다.

    - dialog 내부의 닫기 버튼(data-slot='dialog-close')를 우선 클릭
    - 실패하면 오버레이 클릭 시도
    - dialog가 DOM에서 사라질 때까지 최대 0.3초 대기
    """
    try:
        # 1차: X 버튼 클릭
        try:
            close_btn = dialog.find_element(
                By.XPATH, ".//button[@data-slot='dialog-close']"
            )
            driver.execute_script("arguments[0].click();", close_btn)
        except Exception:
            # 2차: 오버레이 클릭 (배경을 클릭해 닫히는 타입일 수 있음)
            try:
                overlay = driver.find_element(
                    By.XPATH,
                    "//div[@data-slot='dialog-overlay' and @data-state='open']",
                )
                driver.execute_script("arguments[0].click();", overlay)
            except Exception:
                pass

        # 3차: dialog가 사라질 때까지 최대 0.3초만 대기 (기존 2초 → 0.3초)
        try:
            WebDriverWait(driver, 0.3).until_not(
                EC.presence_of_element_located(
                    (
                        By.XPATH,
                        "//div[@role='dialog' and .//h2[normalize-space()='거래 내역']]",
                    )
                )
            )
        except TimeoutException:
            pass
        time.sleep(0.05)
    except Exception:
        pass


def _click_trash_and_confirm(card, wait: WebDriverWait) -> bool:
    """
    카드 안 휴지통 버튼(title='삭제') 클릭 후,
    확인 팝업의 붉은 '삭제' 버튼을 클릭한다.
    """
    try:
        driver = wait._driver  # WebDriverWait 에 연결된 driver

        # 1) 카드 안의 휴지통 아이콘 클릭
        trash_btn = card.find_element(
            By.XPATH,
            ".//button[@title='삭제']"
            " | .//button[.//svg[contains(@class,'lucide-trash') or contains(@class,'lucide-trash-2')]]",
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", trash_btn)
        time.sleep(0.05)
        driver.execute_script("arguments[0].click();", trash_btn)

        # 2) 휴지통 클릭 후 뜨는 경고 다이얼로그(빨간 '삭제' 버튼)가 나타날 때까지 대기
        try:
            alert = wait.until(
                EC.visibility_of_element_located(
                    (
                        By.XPATH,
                        "//div[@role='alertdialog']"
                        " | //div[@data-slot='alert-dialog-content']",
                    )
                )
            )
        except TimeoutException:
            print("[WARN] 휴지통 클릭 후 경고 다이얼로그를 찾지 못했습니다.")
            return False

        # 3) 다이얼로그 안의 붉은 '삭제' 버튼 클릭
        try:
            confirm_btn = alert.find_element(
                By.XPATH,
                ".//button[normalize-space()='삭제']",
            )
        except Exception:
            # fallback: 화면 전체에서라도 '삭제' 버튼을 찾는다
            confirm_btn = wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//button[normalize-space()='삭제']")
                )
            )

        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", confirm_btn)
        time.sleep(0.05)
        driver.execute_script("arguments[0].click();", confirm_btn)

        # 4) alert 다이얼로그/오버레이가 사라질 때까지 잠시 대기
        try:
            WebDriverWait(driver, 2).until_not(
                EC.presence_of_element_located(
                    (
                        By.XPATH,
                        "//div[@data-slot='alert-dialog-overlay' and @data-state='open']",
                    )
                )
            )
        except TimeoutException:
            pass

        time.sleep(0.2)
        return True
    except Exception as e:
        print(f"[WARN] 휴지통/삭제 버튼 처리 중 오류: {e}")
        return False


def _is_session_already_processed(session_id: str) -> bool:
    """
    admin_state.json.history 에서 이미 has_approval 또는 deleted 플래그가 있는 세션이면 True.
    """
    if not session_id:
        return False
    try:
        st = _load_admin_state()
        history = st.get("history") or []
        for h in history:
            if str(h.get("id")) == str(session_id):
                if h.get("has_approval") or h.get("deleted"):
                    return True
        return False
    except Exception as e:
        print(f"[WARN] _is_session_already_processed 실패: {e}")
        return False


def _upsert_history_by_session_id(history: list[dict], entry: dict) -> list[dict]:
    """history에서 동일 session_id를 1건으로 유지하며 upsert."""
    sid = str(entry.get("id") or "").strip()
    if not sid:
        return history
    merged_history: list[dict] = []
    merged_target: dict | None = None
    for h in history or []:
        if str(h.get("id") or "").strip() == sid:
            if merged_target is None:
                merged_target = dict(h)
            else:
                merged_target.update(h)
        else:
            merged_history.append(h)
    if merged_target is None:
        merged_target = {}
    merged_target.update(entry)
    merged_history.append(merged_target)
    return merged_history


def _mark_session_checked(session_id: str, title: str, has_approval: bool) -> None:
    """
    admin_state.json.history 에 has_approval 플래그를 기록해
    다음 크롤링에서 중복 검사하지 않게 한다.
    """
    if not session_id:
        return
    try:
        st = _load_admin_state()
        sessions = list(st.get("sessions") or [])
        history = list(st.get("history") or [])
        now_iso = datetime.utcnow().isoformat()

        if has_approval:
            # 승인 완료된 세션은 진행중 목록에서 제거하고 history로 이동
            remaining_sessions: list[dict] = []
            moved_session: dict | None = None
            for s in sessions:
                if str(s.get("id") or "") == str(session_id):
                    moved_session = dict(s)
                else:
                    remaining_sessions.append(s)

            if moved_session is None:
                moved_session = {"id": session_id}
            moved_session["status"] = "결제완료"
            moved_session["has_approval"] = True
            moved_session["checked_title"] = title
            moved_session["finished_at"] = moved_session.get("finished_at") or now_iso
            history = _upsert_history_by_session_id(history, moved_session)
            st["sessions"] = remaining_sessions
        else:
            # 미승인 체크 결과는 history 신규 생성하지 않고 기존 history 항목만 보강
            for h in history:
                if str(h.get("id") or "") == str(session_id):
                    h["checked_title"] = title or h.get("checked_title") or ""
                    break
            st["sessions"] = sessions

        st["history"] = history
        _save_admin_state(st)
    except Exception as e:
        print(f"[WARN] _mark_session_checked 실패: {e}")


def _mark_session_deleted(session_id: str, title: str) -> None:
    try:
        st = _load_admin_state()
        sessions = list(st.get("sessions") or [])
        history = list(st.get("history") or [])
        now_iso = datetime.utcnow().isoformat()

        remaining_sessions: list[dict] = []
        removed_session: dict | None = None
        for s in sessions:
            if str(s.get("id") or "") == str(session_id):
                removed_session = dict(s)
            else:
                remaining_sessions.append(s)

        if removed_session is None:
            removed_session = {"id": session_id}
        removed_session["status"] = "만료"
        removed_session["deleted"] = True
        removed_session["deleted_in_kvan"] = True
        removed_session["checked_title"] = title
        removed_session["finished_at"] = removed_session.get("finished_at") or now_iso
        old_msg = str(removed_session.get("result_message") or "").strip()
        mark_msg = "만료 감지로 K-VAN 링크가 삭제되었습니다."
        removed_session["result_message"] = f"{old_msg}\n{mark_msg}".strip() if old_msg else mark_msg

        history = _upsert_history_by_session_id(history, removed_session)
        st["sessions"] = remaining_sessions
        st["history"] = history
        _save_admin_state(st)

        # 만료+거래없음 → kvan_links에서 해당 행 삭제 대상으로 표시 (status='만료' 반영 후 mark_expired_sessions_from_kvan_links가 DELETE)
        if not LOCAL_TEST and session_id:
            try:
                _expired_debug(f"[EXPIRED_DEBUG] _mark_session_deleted: 서버 모드, kvan_links UPDATE 시도 session_id={session_id[:24]}...")
                conn = _get_db_with_retry()
                try:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            UPDATE kvan_links
                            SET status = '만료'
                            WHERE (kvan_session_id = %s OR kvan_link LIKE %s)
                            """,
                            (session_id, f"%{session_id}%"),
                        )
                        updated = cur.rowcount
                    conn.commit()
                    _expired_debug(f"[EXPIRED_DEBUG] _mark_session_deleted: kvan_links UPDATE 완료, 반영 행 수: {updated}")
                finally:
                    conn.close()
            except Exception as e_db:
                print(f"[WARN] _mark_session_deleted kvan_links 반영 실패: {e_db}")
        else:
            if LOCAL_TEST:
                _expired_debug(f"[EXPIRED_DEBUG] _mark_session_deleted: LOCAL_TEST=true → kvan_links UPDATE 스킵 session_id={session_id[:24] if session_id else ''}...")
            elif not session_id:
                _expired_debug("[EXPIRED_DEBUG] _mark_session_deleted: session_id 없음 → kvan_links UPDATE 스킵")
    except Exception as e:
        print(f"[WARN] _mark_session_deleted 실패: {e}")


# 만료되었으나 거래 내역이 있는 링크 목록 (어드민 페이지 알림용)
EXPIRED_WITH_TRANSACTIONS_PATH = DATA_DIR / "expired_with_transactions.json"


def _mark_session_expired_with_transactions(
    session_id: str,
    title: str,
    agency_id: str | None = None,
) -> None:
    """
    만료되었지만 거래 내역이 있는 세션: history에 만료로 남기고,
    has_transaction=True, deleted=False 로 표시. 어드민 알림 목록에 추가.
    """
    try:
        st = _load_admin_state()
        sessions = list(st.get("sessions") or [])
        history = list(st.get("history") or [])
        now_iso = datetime.utcnow().isoformat()

        remaining_sessions: list[dict] = []
        moved: dict | None = None
        for s in sessions:
            if str(s.get("id") or "") == str(session_id):
                moved = dict(s)
            else:
                remaining_sessions.append(s)
        if moved is None:
            moved = {"id": session_id}
        moved["status"] = "만료"
        moved["has_transaction"] = True
        moved["deleted"] = False
        moved["deleted_in_kvan"] = False
        moved["checked_title"] = title
        moved["finished_at"] = moved.get("finished_at") or now_iso
        moved["agency_id"] = (agency_id or moved.get("agency_id") or "").strip() or None
        history = _upsert_history_by_session_id(history, moved)
        st["sessions"] = remaining_sessions
        st["history"] = history
        _save_admin_state(st)

        # 어드민 알림용 목록에 추가 (최근 200건 유지)
        try:
            items: list[dict] = []
            if EXPIRED_WITH_TRANSACTIONS_PATH.exists():
                try:
                    items = json.loads(EXPIRED_WITH_TRANSACTIONS_PATH.read_text(encoding="utf-8"))
                except Exception:
                    items = []
            items.append({
                "session_id": session_id,
                "title": (title or "")[:200],
                "agency_id": agency_id or "",
                "finished_at": now_iso,
                "seen": False,
            })
            items = items[-200:]
            EXPIRED_WITH_TRANSACTIONS_PATH.parent.mkdir(parents=True, exist_ok=True)
            EXPIRED_WITH_TRANSACTIONS_PATH.write_text(
                json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            _append_admin_log("CRAWLER", f"만료+거래있음 세션 기록 session_id={session_id} (어드민 알림)")
        except Exception as e:
            print(f"[WARN] 만료+거래있음 목록 저장 실패: {e}")
    except Exception as e:
        print(f"[WARN] _mark_session_expired_with_transactions 실패: {e}")


def _link_matches_kvan_session_id(link: str, session_id: str) -> bool:
    """
    kvan_link(전체 URL)와 팝업/크롤러에서 온 session_id(KEY... 또는 변형)가 같은 세션인지 판별.
    아래 방법 중 하나라도 통과하면 True (오류·인코딩·파싱 차이 보완).
    """
    if not link or not session_id:
        return False
    sid = str(session_id).strip()
    link_raw = link.strip()
    if not sid or not link_raw:
        return False
    sid_l = sid.lower()
    link_l = link_raw.lower()

    def _eq(a: str, b: str) -> bool:
        aa = unquote(str(a).strip())
        bb = unquote(str(b).strip())
        return aa == bb or aa.lower() == bb.lower()

    # 1) 전체 URL 부분 문자열
    if sid in link_raw or sid_l in link_l:
        return True

    # 2) parse_qs 의 sessionId
    try:
        q = parse_qs(urlparse(link_raw).query)
        for v in q.get("sessionId") or []:
            if _eq(v, sid):
                return True
    except Exception:
        pass

    # 3) KEY 직후 ~ &type=KEYED 앞까지 (쿼리에 명시된 형태)
    for pat in (
        r"sessionId=(KEY[^&]+?)&type=KEYED",
        r"sessionid=(KEY[^&]+?)&type=KEYED",
        r"sessionId=(KEY[^&]+?)(?:&|$)",
    ):
        m = re.search(pat, link_raw, re.IGNORECASE)
        if m and _eq(m.group(1), sid):
            return True

    # 4) 경로 /p/KEY...
    m = re.search(r"/p/(KEY[A-Za-z0-9]+)", link_raw, re.IGNORECASE)
    if m and _eq(m.group(1), sid):
        return True

    # 5) URL 안의 모든 KEY… 토큰과 비교 (대소문자 무시)
    for tok in re.findall(r"KEY[A-Za-z0-9]+", link_raw, re.IGNORECASE):
        if _eq(tok, sid):
            return True

    # 6) 팝업 쪽이 KEY 접두어 없이 뒷부분만 온 경우 (KEY 이후 문자열만 동일하면 일치)
    if sid_l.startswith("key") and len(sid_l) > 3:
        suffix = sid_l[3:]
        for tok in re.findall(r"KEY([A-Za-z0-9]+)", link_raw, re.IGNORECASE):
            if tok.lower() == suffix:
                return True
    elif not sid_l.startswith("key") and re.match(r"^[A-Za-z0-9]+$", sid):
        for tok in re.findall(r"KEY([A-Za-z0-9]+)", link_raw, re.IGNORECASE):
            if tok.lower() == sid_l:
                return True

    return False


def _get_agency_id_for_session(session_id: str) -> str | None:
    """
    admin_state.json 에서 session_id에 해당하는 agency_id 반환. 본사/대행사 구분용.
    session_id 는 (1) 우리 쪽 세션 id 또는 (2) K-VAN 링크의 KEY... (거래 내역 팝업에서 옴) 일 수 있음.
    kvan_link 와의 대조는 _link_matches_kvan_session_id 로 여러 방식 중 하나만 통과하면 일치.
    """
    if not session_id:
        return None
    try:
        st = _load_admin_state()
        sid = str(session_id).strip()
        sessions = st.get("sessions") or []
        history = st.get("history") or []

        def _aid_from(s: dict) -> str | None:
            a = (s.get("agency_id") or "").strip()
            return a or None

        # 1) 내부 세션 id 정확 일치 (진행·히스토리 모두, 진행 세션을 먼저)
        for s in sessions + history:
            if str(s.get("id") or "").strip() == sid:
                return _aid_from(s)

        # 2) kvan_link 로만 매칭: 진행 세션 우선 → 히스토리 (오탐 줄임)
        for s in sessions:
            link = (s.get("kvan_link") or "").strip()
            if link and _link_matches_kvan_session_id(link, sid):
                return _aid_from(s)
        for s in history:
            link = (s.get("kvan_link") or "").strip()
            if link and _link_matches_kvan_session_id(link, sid):
                return _aid_from(s)
        return None
    except Exception:
        return None


def _lookup_internal_session_id_for_kvan_key(kvan_key: str) -> str:
    """K-VAN KEY… 에 매칭되는 admin_state 세션의 내부 id (숫자 세션 등)."""
    kk = (kvan_key or "").strip()
    if not kk:
        return ""
    try:
        st = _load_admin_state()
        for bucket in (st.get("sessions") or [], st.get("history") or []):
            for s in bucket:
                link = (s.get("kvan_link") or "").strip()
                if link and _link_matches_kvan_session_id(link, kk):
                    return str(s.get("id") or "").strip()
        return ""
    except Exception:
        return ""


def _scan_payment_link_popups_and_sync(
    driver: webdriver.Chrome,
    allow_popup_for_non_expired: bool = True,
) -> bool:
    """
    결제링크 관리(/payment-link) 화면에서 각 카드의 '거래 내역' 버튼을 클릭해
    팝업의 '결제 승인' 정보를 읽고 내부 DB와 세션에 반영한다.

    반환값: 이번 사이클에서 새로운 승인/삭제/상태 변경이 있으면 True, 아니면 False.
    """
    changed = False
    try:
        t0 = _step_start("결제링크 관리 팝업 기반 동기화")
        wait = WebDriverWait(driver, 5)

        # 항상 결제링크 관리 URL 을 강제로 맞추고,
        # '권한 확인 중...' 스피너가 끝나고 실제 카드 리스트가 나올 때까지
        # 0.5초 간격으로 짧게 여러 번 시도한다 (최대 약 5초).
        if not _go_to_payment_link(driver):
            _step_end("결제링크 관리 팝업 기반 동기화", t0)
            raise RuntimeError("[NAV] /payment-link 로 진입하지 못해 팝업 기반 동기화를 중단합니다.")
        _wait_payment_link_page_ready(driver)

        max_tries = 12  # 0.5초 * 12 ≈ 6초
        icons_found = False
        icons: list = []
        for attempt in range(max_tries):
            icons = driver.find_elements(
                By.XPATH,
                "//button[@title='거래 내역']"
                " | //button[contains(normalize-space(.),'거래 내역')]"
                " | //button[contains(normalize-space(.),'거래내역')]"
                " | //button[.//svg[contains(@class,'lucide-receipt')]]"
                " | //button[contains(@aria-label,'거래') or contains(@aria-label,'내역')]",
            )
            if icons:
                print(
                    f"[POPUP_DEBUG] '거래 내역' 아이콘 감지 (attempt={attempt}, count={len(icons)}, url={driver.current_url})"
                )
                icons_found = True
                break
            else:
                print(
                    f"[POPUP_DEBUG] 아이콘 없음 (attempt={attempt}, url={driver.current_url}) – 0.2초 후 재시도"
                )
            time.sleep(0.2)

        if not icons_found:
            print("[INFO] 결제링크 관리 화면에서 '거래 내역' 아이콘이 없어 팝업 기반 동기화를 건너뜁니다. (transactions 동기화는 계속 진행)")
            _step_end("결제링크 관리 팝업 기반 동기화", t0)
            return False

        # JS 기준으로 버튼/행을 1:1 매핑해 추출한다.
        # 각 버튼에서 시작해 "단일 KEY 세션ID를 가진 최소 조상"만 채택한다.
        row_infos = driver.execute_script(
            """
const list = Array.from(arguments[0] || []);
const out = [];
for (let i = 0; i < list.length; i++) {
  const btn = list[i];
  let cur = btn;
  let picked = null;
  let fallback = null;
  for (let d = 0; d < 9 && cur; d++) {
    const txt = (cur.innerText || "").trim();
    if (txt) {
      const keys = Array.from(new Set(txt.match(/KEY[0-9A-Za-z]+/g) || []));
      if (keys.length === 1 && txt.length <= 1200) {
        const badges = Array.from(cur.querySelectorAll("span[data-slot='badge']")).map(el => (el.innerText || "").trim()).filter(Boolean);
        const hasStatusBadges = badges.length > 0;
        const cand = { sid: keys[0], text: txt, badges: badges, hasStatusBadges: hasStatusBadges };
        if (hasStatusBadges) {
          picked = cand;
          break;
        }
        if (!fallback || txt.length > fallback.text.length) {
          fallback = cand;
        }
      }
    }
    cur = cur.parentElement;
  }
  if (!picked) picked = fallback;
  if (!picked) continue;
  const badges = Array.from(picked.badges || []);
  const lines = picked.text.split("\\n").map(s => s.trim()).filter(Boolean);
  let expiredByLine = lines.some(v => {
    const t = (v || "").replace(/\\s+/g, "");
    return t === "만료" || t === "취소" || t === "취소됨" || t === "취소완료";
  });
  out.push({
    icon_index: i,
    session_id: picked.sid,
    title: lines.length ? lines[0] : "",
    is_expired: badges.some(t => t.includes("만료")) || expiredByLine,
    badge_texts: badges,
    raw_text: picked.text
  });
}
return out;
""",
            icons,
        ) or []
        card_items: list[tuple] = []
        for info in row_infos:
            try:
                idx_js = int(info.get("icon_index"))
                sid = str(info.get("session_id") or "").strip()
                title = str(info.get("title") or "").strip()
                expired = bool(info.get("is_expired"))
                badge_texts = list(info.get("badge_texts") or [])
                raw_text = str(info.get("raw_text") or "")
                if idx_js < 0 or idx_js >= len(icons) or not sid:
                    continue
                card_items.append((icons[idx_js], sid, title, expired, badge_texts, raw_text))
            except Exception:
                continue

        # 아이콘은 찾았지만 버튼→카드 매칭이 안 된 경우: 카드 컨테이너 기준으로 폴백 (카드별 거래내역 버튼 직접 찾기)
        if not card_items and icons:
            print("[POPUP_DEBUG] 버튼→카드 매칭 실패, 카드 컨테이너 기준 폴백 시도")
            fallback_infos = driver.execute_script(
                """
var cards = document.querySelectorAll('div.rounded-lg.border, div[class*="rounded"][class*="border"]');
var out = [];
for (var c = 0; c < cards.length; c++) {
  var card = cards[c];
  var txt = (card.innerText || '').trim();
  var keys = txt.match(/KEY[0-9A-Za-z]+/g);
  if (!keys) continue;
  var uniq = Array.from(new Set(keys));
  if (uniq.length !== 1) continue;
  var sid = uniq[0];
  var badges = Array.from(card.querySelectorAll("span[data-slot='badge']")).map(function(el){ return (el.innerText||'').trim(); }).filter(Boolean);
  var lines = txt.split('\\n').map(function(s){ return s.trim(); }).filter(Boolean);
  var expiredByLine = lines.some(function(v){ var t = (v||'').replace(/\\s+/g,''); return t==='만료'||t==='취소'||t==='취소됨'||t==='취소완료'; });
  var receiptBtn = card.querySelector('button[title="거래 내역"]') || card.querySelector('button[aria-label*="거래"]') || card.querySelector('button[aria-label*="내역"]');
  if (!receiptBtn) {
    var btns = card.querySelectorAll('button');
    for (var b = 0; b < btns.length; b++) {
      var btn = btns[b];
      var label = (btn.getAttribute('aria-label')||'') + (btn.title||'') + (btn.innerText||'');
      if (label.indexOf('거래') >= 0 || label.indexOf('내역') >= 0) { receiptBtn = btn; break; }
    }
  }
  if (!receiptBtn && card.querySelector('div.flex.gap-2')) receiptBtn = card.querySelector('div.flex.gap-2 button');
  if (!receiptBtn) receiptBtn = card.querySelector('button');
  if (!receiptBtn) continue;
  out.push({ session_id: sid, title: lines[0]||'', is_expired: badges.some(function(b){ return b.indexOf('만료')>=0; }) || expiredByLine, badge_texts: badges, raw_text: txt, button: receiptBtn });
}
return out;
""",
            )
            for info in (fallback_infos or []):
                try:
                    sid = str(info.get("session_id") or "").strip()
                    if not sid:
                        continue
                    btn_el = info.get("button")
                    if not btn_el:
                        continue
                    title = str(info.get("title") or "").strip()
                    expired = bool(info.get("is_expired"))
                    badge_texts = list(info.get("badge_texts") or [])
                    raw_text = str(info.get("raw_text") or "")
                    card_items.append((btn_el, sid, title, expired, badge_texts, raw_text))
                except Exception:
                    continue
            if card_items:
                print(f"[POPUP_DEBUG] 카드 폴백으로 {len(card_items)}건 수집")

        if not card_items:
            print("[WARN] 결제링크 카드 매칭 실패: 아이콘은 있으나 카드 정보를 추출하지 못했습니다.")
            _step_end("결제링크 관리 팝업 기반 동기화", t0)
            return False
        # 동일 session_id가 여러 번 잡히는 경우가 있어 세션 단위로 병합한다.
        # 병합 규칙: is_expired=True 가 하나라도 있으면 만료 우선.
        merged_by_sid: dict[str, tuple] = {}
        for item in card_items:
            btn, sid, title, expired, badges, raw_text = item
            prev = merged_by_sid.get(sid)
            if prev is None:
                merged_by_sid[sid] = item
                continue
            prev_btn, _, prev_title, prev_expired, prev_badges, prev_raw = prev
            if expired and not prev_expired:
                merged_by_sid[sid] = item
                continue
            if prev_expired and not expired:
                continue
            merged_title = prev_title if len(str(prev_title or "")) >= len(str(title or "")) else title
            merged_badges = list(prev_badges or [])
            for b in list(badges or []):
                if b not in merged_badges:
                    merged_badges.append(b)
            merged_by_sid[sid] = (
                prev_btn,
                sid,
                merged_title,
                bool(prev_expired or expired),
                merged_badges,
                prev_raw if len(str(prev_raw or "")) >= len(str(raw_text or "")) else raw_text,
            )
        card_items = list(merged_by_sid.values())
        _append_admin_log(
            "CRAWLER",
            f"결제링크 행 스캔 시작 rows={len(card_items)}, allow_popup_for_non_expired={allow_popup_for_non_expired}",
        )

        # 모든 카드를 순차적으로 처리 (세션ID 기준으로 신규만)
        processed_count = 0
        expired_count = 0
        duplicate_count = 0
        no_session_count = 0
        seen_session_ids: set[str] = set()
        for idx, (btn, sid_hint, title_hint, expired_hint, badge_texts, raw_text_hint) in enumerate(card_items, start=1):
            try:
                card_text = str(raw_text_hint or "")
                lines: list[str] = [ln.strip() for ln in card_text.splitlines() if ln.strip()]
                print(f"[CARD_DEBUG] 원시 카드 텍스트 (index={idx}): {lines}")

                is_expired = bool(expired_hint)
                if is_expired:
                    print(f"[TTL_DEBUG] 상태 배지에서 '만료' 감지 → is_expired=True (badges={badge_texts})")
                else:
                    print(f"[TTL_DEBUG] 상태 배지들={badge_texts} → '만료' 없음")

                session_id = sid_hint or ""
                product_title = title_hint or ""
                for ln in lines:
                    if not product_title:
                        product_title = ln  # 첫 줄 정도를 제목으로 사용
                    # 1순위: "세션 ID:" 라벨이 있는 경우 (예전 UI)
                    if "세션 ID" in ln:
                        parts = ln.split("세션 ID:")
                        if len(parts) > 1:
                            session_id = parts[1].strip()
                            continue
                    # 2순위: 문자열에서 KEY... 패턴을 정규식으로 추출
                    if not session_id and "KEY20" in ln:
                        m = re.search(r"(KEY[0-9A-Za-z]+)", ln)
                        if m:
                            session_id = m.group(1)

                # 한 카드/행 안에 세션ID가 여러 개 보이면 비정상 컨테이너로 판단하고 건너뛴다.
                all_keys = list(dict.fromkeys(re.findall(r"(KEY[0-9A-Za-z]+)", card_text)))
                if len(all_keys) > 1:
                    print(f"[CARD_DEBUG] 다중 세션ID 컨테이너 감지(index={idx}, keys={all_keys[:4]}...) → 스킵")
                    continue

                # 배지 만료가 없더라도 만료일시가 현재보다 이전이면 만료로 간주한다.
                # (가상 스크롤/축약 렌더링으로 배지 추출이 누락되는 경우 보정)
                if not is_expired:
                    expire_at = _extract_expire_at_from_lines(lines)
                    now_kvan = _kvan_now()
                    if expire_at and expire_at <= now_kvan:
                        is_expired = True
                        print(
                            "[TTL_DEBUG] 만료일시 경과 감지 → is_expired=True "
                            f"(expire_at={expire_at.isoformat()}, now_kvan={now_kvan.isoformat()})"
                        )

                # 세션 ID 가 없는 행(헤더/틀 행 등)은 실제 결제링크 카드가 아니므로 건너뜀
                if not session_id:
                    no_session_count += 1
                    print(f"[CARD_DEBUG] 세션ID 없음 → 헤더/비카드로 판단, 건너뜀 (index={idx}, title={product_title})")
                    continue

                # 동일 세션을 한 사이클에서 반복 처리하지 않는다.
                if session_id in seen_session_ids:
                    duplicate_count += 1
                    print(f"[CARD_DEBUG] 중복 세션ID 스킵 (session_id={session_id}, index={idx})")
                    continue
                seen_session_ids.add(session_id)
                processed_count += 1

                print(f"[CARD_DEBUG] 인덱스={idx}, 세션ID={session_id}, 제목={product_title}, is_expired={is_expired}")
                # 만료 카드: 거래 내역 유무에 따라 삭제(거래없음) vs DB저장+어드민 알림(거래있음)
                # 다이얼로그 대기/닫기 0.3초로 짧게 해서 리스트 크롤링 속도 확보
                if is_expired:
                    # 거래 내역 버튼 클릭 → 팝업에서 유무 확인 (재시도 시 0.15초만 대기)
                    click_ok = False
                    for _ in range(10):
                        try:
                            driver.execute_script(
                                "arguments[0].scrollIntoView({behavior:'instant',block:'center'});",
                                btn,
                            )
                            time.sleep(0.03)
                            driver.execute_script("arguments[0].click();", btn)
                            click_ok = True
                            break
                        except Exception:
                            time.sleep(0.15)
                    if not click_ok:
                        _mark_session_deleted(session_id, product_title)
                        expired_count += 1
                        changed = True
                        continue
                    # 만료 카드용 팝업 대기: 최대 2초 (5초 대기 축소)
                    try:
                        short_wait = WebDriverWait(driver, 2)
                        dialog = short_wait.until(
                            EC.visibility_of_element_located(
                                (
                                    By.XPATH,
                                    "//div[@role='dialog' and .//h2[normalize-space()='거래 내역']]",
                                )
                            )
                        )
                    except TimeoutException:
                        _mark_session_deleted(session_id, product_title)
                        expired_count += 1
                        changed = True
                        continue
                    popup_text = dialog.text or ""
                    has_no_history = (
                        "거래 내역이 없습니다" in popup_text
                        or "거래 내역 없음" in popup_text
                    )
                    # 테이블에 실제 데이터 행이 없거나, 첫 행이 '없습니다' 문구면 거래없음으로 간주 (K-VAN 문구 차이·로딩 지연 대응)
                    rows = dialog.find_elements(By.XPATH, ".//table//tbody//tr")
                    if not has_no_history and len(rows) == 0:
                        has_no_history = True
                        _expired_debug(f"[EXPIRED_DEBUG] 만료 카드: tbody tr 0개 → 거래없음으로 간주 session_id={session_id[:24]}...")
                    elif not has_no_history and rows:
                        first_row_text = (rows[0].text or "").strip()
                        if "없습니다" in first_row_text or "없음" in first_row_text:
                            has_no_history = True
                            _expired_debug(f"[EXPIRED_DEBUG] 만료 카드: 첫 행이 '없음' 문구 → 거래없음으로 간주 session_id={session_id[:24]}..., first_row={first_row_text[:60]}")
                    if not has_no_history:
                        _expired_debug(f"[EXPIRED_DEBUG] 만료 카드 팝업: popup_text_len={len(popup_text)}, tbody_rows={len(rows)}, has_no_history=False → 만료+거래있음 처리 session_id={session_id[:24]}...")
                    if has_no_history:
                        _close_dialog(driver, dialog)
                        _expired_debug(f"[EXPIRED_DEBUG] 만료+거래없음 → _mark_session_deleted 호출 session_id={session_id}, LOCAL_TEST={LOCAL_TEST}")
                        _mark_session_deleted(session_id, product_title)
                        expired_count += 1
                        _append_admin_log("CRAWLER", f"만료+거래없음 세션 삭제 session_id={session_id}")
                        changed = True
                        continue
                    # 만료 but 거래 내역 있음 → DB 저장 후 어드민 알림 목록에 추가
                    rows = dialog.find_elements(By.XPATH, ".//table//tbody//tr")
                    agency_id = _get_agency_id_for_session(session_id)
                    if rows:
                        try:
                            row = rows[0]
                            tx_type = row.find_element(
                                By.XPATH, ".//span[contains(@data-slot,'badge')]"
                            ).text.strip()
                            amount_text = row.find_element(By.XPATH, ".//td[3]//span").text.strip()
                            amt = _parse_amount(amount_text)
                            approval_no = row.find_element(By.XPATH, ".//td[4]").text.strip()
                            customer_name = row.find_element(By.XPATH, ".//td[5]").text.strip()
                            card_number = row.find_element(By.XPATH, ".//td[6]//span").text.strip()
                            registered_at = row.find_element(By.XPATH, ".//td[7]").text.strip()
                            if "결제 승인" in tx_type and amt:
                                _sync_popup_transaction_to_internal(
                                    session_id=session_id,
                                    amount=amt,
                                    approval_no=approval_no,
                                    card_number=card_number,
                                    registered_at=registered_at,
                                    customer_name=customer_name,
                                )
                        except Exception as e_parse:
                            print(f"[WARN] 만료+거래있음 행 파싱 실패: {e_parse}")
                    _mark_session_expired_with_transactions(session_id, product_title, agency_id)
                    _close_dialog(driver, dialog)
                    expired_count += 1
                    changed = True
                    continue

                if _is_session_already_processed(session_id):
                    print(f"[CARD_DEBUG] 이미 처리된 세션 → 건너뜀 (session_id={session_id})")
                    continue

                if not allow_popup_for_non_expired:
                    _mark_session_checked(session_id, product_title, has_approval=False)
                    continue

                # 거래 내역 버튼 클릭 재시도 시 0.15초 대기 (기존 0.3초 → 0.15초)
                click_ok = False
                for _ in range(10):
                    try:
                        driver.execute_script(
                            "arguments[0].scrollIntoView({behavior:'instant',block:'center'});",
                            btn,
                        )
                        time.sleep(0.03)
                        driver.execute_script("arguments[0].click();", btn)
                        click_ok = True
                        break
                    except Exception as e_click:
                        print(f"[CARD_DEBUG] 거래 내역 버튼 클릭 재시도: {e_click}")
                        time.sleep(0.15)

                if not click_ok:
                    print("[WARN] 거래 내역 버튼 클릭 실패(여러 차례 재시도 후) → 다음 카드로 진행")
                    continue

                # 팝업(dialog) 대기
                try:
                    dialog = wait.until(
                        EC.visibility_of_element_located(
                            (
                                By.XPATH,
                                "//div[@role='dialog' and .//h2[normalize-space()='거래 내역']]",
                            )
                        )
                    )
                except TimeoutException:
                    print("[WARN] '거래 내역' 팝업을 찾지 못했습니다.")
                    continue

                popup_text = dialog.text or ""
                # "거래 내역이 없습니다" / "거래 내역 없음" 등 문구 변화를 모두 허용
                has_no_history = (
                    "거래 내역이 없습니다" in popup_text
                    or "거래 내역 없음" in popup_text
                )

                if has_no_history:
                    _close_dialog(driver, dialog)
                    _mark_session_checked(session_id, product_title, has_approval=False)
                    continue

                # 거래 내역이 하나 이상 있는 경우: 첫 번째 행 기준으로 승인 정보 파싱
                try:
                    # 일부 레이아웃(모바일 카드형)에서는 table 구조가 없을 수 있으므로,
                    # 우선 테이블 행 존재 여부를 확인하고, 없으면 "구조화된 내역 없음" 으로 처리한다.
                    rows = dialog.find_elements(By.XPATH, ".//table//tbody//tr")
                    if not rows:
                        # 구조화된 테이블이 없으면, 승인 여부를 판단할 수 없으므로
                        # 삭제는 절대 하지 않고, 단순히 "확인됨(미승인/기타)" 상태로만 표시한다.
                        _mark_session_checked(session_id, product_title, has_approval=False)
                        continue

                    row = rows[0]
                    tx_type = row.find_element(
                        By.XPATH, ".//span[contains(@data-slot,'badge')]"
                    ).text.strip()
                    amount_text = row.find_element(
                        By.XPATH, ".//td[3]//span"
                    ).text.strip()
                    amt = _parse_amount(amount_text)
                    approval_no = row.find_element(
                        By.XPATH, ".//td[4]"
                    ).text.strip()
                    customer_name = row.find_element(
                        By.XPATH, ".//td[5]"
                    ).text.strip()
                    card_number = row.find_element(
                        By.XPATH, ".//td[6]//span"
                    ).text.strip()
                    registered_at = row.find_element(
                        By.XPATH, ".//td[7]"
                    ).text.strip()

                    # 팝업 헤더에서 세션 ID 다시 추출 시도
                    try:
                        desc_el = dialog.find_element(
                            By.XPATH, ".//p[contains(@data-slot,'dialog-description')]"
                        )
                        desc_text = desc_el.text or ""
                        m = re.search(r"세션 ID:\s*([A-Z0-9]+)", desc_text)
                        if m:
                            session_id = m.group(1)
                    except Exception:
                        pass

                    if "결제 승인" in tx_type and amt:
                        _sync_popup_transaction_to_internal(
                            session_id=session_id,
                            amount=amt,
                            approval_no=approval_no,
                            card_number=card_number,
                            registered_at=registered_at,
                            customer_name=customer_name,
                        )
                        _mark_session_checked(session_id, product_title, has_approval=True)
                        changed = True
                except NoSuchElementException as e_row:
                    # 예상한 테이블/셀 구조가 없으면, 삭제는 하지 않고
                    # 단순히 "확인했지만 구조 불명" 상태로만 표시하고 다음 카드로 넘어간다.
                    print(f"[WARN] '거래 내역' 팝업 파싱 중 요소를 찾지 못했습니다: {e_row}")
                    _mark_session_checked(session_id, product_title, has_approval=False)
                except Exception as e_row:
                    print(f"[WARN] '거래 내역' 팝업 파싱 중 오류: {e_row}")
                finally:
                    _close_dialog(driver, dialog)

            except StaleElementReferenceException as e_card:
                # 카드가 DOM 에서 사라진 경우(이미 삭제되었거나 새로고침됨)는
                # 추가 시도를 중단하고 현재 카드 루프를 빠져나온다.
                print(f"[WARN] 결제링크 카드 처리 중 StaleElement 오류: {e_card}")
                break
            except Exception as e_card:
                print(f"[WARN] 결제링크 카드 처리 중 오류: {e_card}")
                continue

        _append_admin_log(
            "CRAWLER",
            "결제링크 행 스캔 종료 "
            f"processed={processed_count}, expired={expired_count}, duplicate_skipped={duplicate_count}, no_session={no_session_count}, changed={changed}",
        )
        _step_end("결제링크 관리 팝업 기반 동기화", t0)
    except Exception as e:
        print(f"[WARN] 결제링크 팝업 동기화 전반 오류: {e}")
    return changed


def _go_to_payment_link_page(driver: webdriver.Chrome) -> None:
    """좌측 사이드바에서 '결제링크 관리' 메뉴까지만 이동 (리스트 화면)."""
    try:
        t0 = _step_start("결제링크 관리 메뉴 이동")
        wait = WebDriverWait(driver, 5)
        # 사이드바의 '결제링크 관리' 항목 클릭
        link_btn = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, "//*[contains(normalize-space(text()),'결제링크 관리')]")
            )
        )
        link_btn.click()
        time.sleep(0.15 * (1.0 + LINK_CREATION_WAIT_FACTOR))
        _step_end("결제링크 관리 메뉴 이동", t0)
    except Exception as e:
        print(f"[WARN] 결제링크 관리/생성 페이지 이동 실패: {e}")


def _go_to_create_link_page(driver: webdriver.Chrome) -> bool:
    """
    결제링크 관리 화면에서 '+ 생성' 버튼을 눌러
    결제링크 생성 화면(또는 모달)로 진입한다.

    - 단순 텍스트 매칭뿐 아니라,
      '결제링크 관리' 헤더 주변에서 버튼을 위치 기반으로 찾아 클릭하는
      다른 패러다임을 사용한다.
    """
    t0 = _step_start("+ 생성 버튼 클릭 (헤더/위치 기반 탐색)")
    try:
        # 반복 횟수·간격 최적화: 6회 × 0.35초 ≈ 2.1초 (기존 12×0.7≈8.4초 대비 단축)
        n_tries = 6
        retry_sleep = 0.35 * (1.0 + LINK_CREATION_WAIT_FACTOR)
        for _ in range(n_tries):
            try:
                clicked = driver.execute_script(
                    """
const clickCreateButton = () => {
  // 1. 텍스트/aria-label 에 '생성' 이 포함된 버튼/링크를 우선 시도
  const primary = Array.from(document.querySelectorAll('button,a[role="button"]')).find(el => {
    if (!el.offsetParent) return false;
    const t = (el.innerText || '').trim();
    const label = (el.getAttribute('aria-label') || '').trim();
    return t.includes('생성') || label.includes('생성');
  });
  if (primary) {
    primary.click();
    return true;
  }

  // 2. '결제링크 관리' 헤더 근처에서 버튼을 찾는다.
  const headings = Array.from(document.querySelectorAll('h1,h2,h3'));
  const title = headings.find(el => {
    const t = (el.innerText || '').trim();
    return t.includes('결제링크 관리');
  });

  if (title) {
    let container = title.closest('header,section,div') || document.body;
    let btns = Array.from(container.querySelectorAll('button,a[role="button"]'))
      .filter(el => el.offsetParent !== null);

    if (btns.length > 0) {
      // 오른쪽 위에 있을수록, 그리고 크기가 어느 정도 되는 버튼을 우선 선택
      let best = null;
      let bestScore = -Infinity;
      for (const el of btns) {
        const r = el.getBoundingClientRect();
        const width = r.right - r.left;
        const height = r.bottom - r.top;
        // 너무 작은(아이콘 점 등) 요소는 제외
        if (width < 16 || height < 16) continue;
        // 화면의 오른쪽/위쪽에 있을수록 점수가 높도록 가중치 부여
        const score = r.right * 1.0 - r.top * 0.2;
        if (score > bestScore) {
          bestScore = score;
          best = el;
        }
      }
      if (best) {
        best.click();
        return true;
      }
    }
  }

  // 3. 위 방식으로도 찾지 못하면 실패
  return false;
};
return clickCreateButton();
"""
                )
                if clicked:
                    time.sleep(0.25 * (1.0 + LINK_CREATION_WAIT_FACTOR))
                    _step_end("+ 생성 버튼 클릭 (헤더/위치 기반 탐색)", t0)
                    return True
            except Exception:
                pass
            time.sleep(retry_sleep)

        print("[WARN] 여러 번 시도했지만 '+ 생성' 버튼을 찾지 못했습니다.")
        _step_end("+ 생성 버튼 클릭 (헤더/위치 기반 탐색)", t0)
        return False
    except Exception as e:  # noqa: BLE001
        print(f"[WARN] '+ 생성' 버튼 클릭 실패: {e}")
        _step_end("+ 생성 버튼 클릭 (헤더/위치 기반 탐색)", t0)
        return False


def _store_kvan_link_for_session(session_id: str, link: str) -> None:
    """admin_state.json 의 해당 세션에 K-VAN 결제 링크를 매핑해서 저장.

    링크 형식: https://store.k-van.app/p/KEY...?sessionId=KEY...&type=KEYED
    KEY 부터 &type=KEYED 앞까지가 K-VAN 세션 ID이며, 크롤러가 /payment-link 를 스크래핑할 때
    kvan_links.kvan_session_id 로 DB 에 저장한다. 거래 내역 팝업의 session_id(KEY...) 와
    이 kvan_link 를 기준으로 본사/대행사 매칭이 이뤄진다.
    """
    if not session_id or not link:
        return
    session_blob: dict | None = None
    try:
        state_path: Path | None = None
        for candidate in _admin_state_candidates():
            if candidate.exists():
                state_path = candidate
                break
        if state_path is None:
            _append_admin_log("AUTO", f"[WARN] admin_state.json 없음 – 링크 저장 불가 session_id={session_id}")
            return
        with open(state_path, "r", encoding="utf-8") as f:
            state = json.load(f)
        sessions = state.get("sessions") or []
        history = state.get("history") or []
        updated = False
        for s in sessions:
            if str(s.get("id")) == str(session_id):
                s["kvan_link"] = link
                updated = True
                session_blob = dict(s)
                break
        if updated:
            state["sessions"] = sessions
            state["history"] = history
            with open(state_path, "w", encoding="utf-8") as f:
                json.dump(state, f, ensure_ascii=False, indent=2)
            _append_admin_log("AUTO", f"admin_state 링크 저장 완료 session_id={session_id} path={state_path}")
            try:
                upsert_kvan_link_creation_seed(
                    link,
                    str(session_id),
                    session_blob or {},
                    skip_db=LOCAL_TEST,
                )
            except Exception as e_seed:  # noqa: BLE001
                _append_admin_log("AUTO", f"[WARN] kvan_links 시드 DB 저장 실패: {e_seed}")
        else:
            _append_admin_log("AUTO", f"[WARN] admin_state 에서 세션 못 찾음 session_id={session_id}")
    except Exception as e:  # noqa: BLE001
        print(f"[WARN] admin_state 에 K-VAN 링크 저장 실패: {e}")
        _append_admin_log("AUTO", f"[WARN] admin_state 링크 저장 실패 session_id={session_id}: {e}")


def _fill_payment_link_form_and_get_url(
    driver: webdriver.Chrome, row: PaymentRow, session_id: str
) -> str | None:
    """결제링크 생성 페이지에서 폼을 채우고, 생성된 https://store.k-van.app... 링크를 리턴."""
    t0_all = _step_start("결제링크 생성 폼 작성 전체")
    wait = WebDriverWait(driver, max(2, int(3 * LINK_CREATION_WAIT_FACTOR)))
    time.sleep(0.1 * (1.0 + LINK_CREATION_WAIT_FACTOR))

    # 1. 금액 입력
    t0 = _step_start("1. 금액 입력")
    amount_input = None
    amount_xpaths = [
        "//*[contains(normalize-space(text()),'금액')]/following::input[1]",
        "//input[@type='number' or @inputmode='decimal']",
    ]
    for xp in amount_xpaths:
        try:
            amount_input = wait.until(EC.visibility_of_element_located((By.XPATH, xp)))
            break
        except TimeoutException:
            continue
    if amount_input:
        try:
            # 금액 입력창이 화면 중앙에 오도록 스크롤
            driver.execute_script(
                "arguments[0].scrollIntoView({behavior:'smooth',block:'center'});",
                amount_input,
            )
        except Exception:
            pass
        amount_input.clear()
        amount_input.send_keys(str(row.amount))
        time.sleep(0.1 * (1.0 + LINK_CREATION_WAIT_FACTOR))
    _step_end("1. 금액 입력", t0)

    # 2. 상품명: 옥션 리스트에서 amount 에 맞는 상품명 선택
    t0 = _step_start("2. 상품명 입력")
    product_name = _choose_product_name_for_amount(row.amount)
    try:
        name_input = wait.until(
            EC.visibility_of_element_located(
                (By.XPATH, "//*[contains(normalize-space(text()),'상품명')]/following::input[1]")
            )
        )
        name_input.clear()
        name_input.send_keys(product_name)
        time.sleep(0.1 * (1.0 + LINK_CREATION_WAIT_FACTOR))
    except TimeoutException:
        pass
    _step_end("2. 상품명 입력", t0)

    # 3. 상품설명
    t0 = _step_start("3. 상품설명 입력")
    desc_text = "글로벌 중고명품 경매사이트  구매 대행 서비스 즉시구매 결제 및 예치금"
    try:
        desc_input = wait.until(
            EC.visibility_of_element_located(
                (
                    By.XPATH,
                    "//*[contains(normalize-space(text()),'상품설명') or contains(normalize-space(text()),'상품 설명')]/following::textarea[1]",
                )
            )
        )
        desc_input.clear()
        desc_input.send_keys(desc_text)
        time.sleep(0.1 * (1.0 + LINK_CREATION_WAIT_FACTOR))
    except TimeoutException:
        try:
            desc_input = wait.until(
                EC.visibility_of_element_located(
                    (
                        By.XPATH,
                        "//*[contains(normalize-space(text()),'상품설명') or contains(normalize-space(text()),'상품 설명')]/following::input[1]",
                    )
                )
            )
            desc_input.clear()
            desc_input.send_keys(desc_text)
            time.sleep(0.1 * (1.0 + LINK_CREATION_WAIT_FACTOR))
        except TimeoutException:
            pass
    _step_end("3. 상품설명 입력", t0)

    # 폼 아래쪽에 있는 세션유효시간/링크 생성 버튼들이 보이도록 스크롤을 충분히 내린다.
    try:
        driver.execute_script("window.scrollBy(0, 1800);")
        time.sleep(0.15 * (1.0 + LINK_CREATION_WAIT_FACTOR))
    except Exception:
        pass

    # 1-1. 세션유효시간: 콤보박스(#session-ttl)를 '5분 (일반 결제)' 로 변경
    # 4. 세션유효시간은 더 이상 오래 시도하지 않고, 한 번만 빠르게 시도 후 바로 진행
    t0 = _step_start("4. 세션유효시간 선택 (3분→5분, 빠른 시도)")
    try:
        _set_session_ttl_to_5min(driver, max_wait=2.0)
    except Exception:
        # 실패해도 기본값(3분)으로 진행
        pass
    _step_end("4. 세션유효시간 선택 (3분→5분, 빠른 시도)", t0)

    # 5. '링크 생성하기' 버튼 클릭
    t0 = _step_start("5. '링크 생성하기' 버튼 클릭")
    used_copy_button = False
    # 1단계: 반드시 '링크 생성하기' 텍스트 버튼을 먼저 누른다.
    try:
        # 너무 오래 기다리지 않도록, 짧은 반복으로 직접 찾는다.
        clicked = False
        end_ts = time.time() + 2.0
        while time.time() < end_ts and not clicked:
            try:
                create_btn = driver.find_element(
                    By.XPATH, "//button[contains(normalize-space(.),'링크 생성하기')]"
                )
                driver.execute_script(
                    "arguments[0].scrollIntoView({behavior:'instant', block:'center'});",
                    create_btn,
                )
                time.sleep(0.03)
                driver.execute_script("arguments[0].click();", create_btn)
                clicked = True
                break
            except Exception:
                time.sleep(0.08)
        if not clicked:
            print("[WARN] '링크 생성하기' 버튼을 2초 내에 찾지 못했습니다.")
            _step_end("5. '링크 생성하기' 버튼 클릭", t0)
            _step_end("결제링크 생성 폼 작성 전체", t0_all)
            return None
        time.sleep(0.2 * (1.0 + LINK_CREATION_WAIT_FACTOR))
    except Exception:
        print("[WARN] '링크 생성하기' 버튼 클릭 중 예외가 발생했습니다.")
        _step_end("5. '링크 생성하기' 버튼 클릭", t0)
        _step_end("결제링크 생성 폼 작성 전체", t0_all)
        return None

    # 2단계: '링크 생성하기' 가 눌린 뒤에 나타나는 '링크 복사' 버튼이 있으면 눌러준다.
    try:
        copy_btn = WebDriverWait(driver, 2).until(
            EC.element_to_be_clickable(
                (By.XPATH, "//button[contains(normalize-space(.),'링크 복사')]")
            )
        )
        driver.execute_script(
            "arguments[0].scrollIntoView({behavior:'instant', block:'center'});",
            copy_btn,
        )
        time.sleep(0.05)
        driver.execute_script("arguments[0].click();", copy_btn)
        used_copy_button = True
        time.sleep(0.2 * (1.0 + LINK_CREATION_WAIT_FACTOR))
    except TimeoutException:
        print("[DEBUG] '링크 복사' 버튼을 찾지 못했습니다. 생성 버튼만 누른 상태로 진행합니다.")
    _step_end("5. '링크 생성하기' 버튼 클릭", t0)

    # 6. 팝업/페이지 내 '생성하기' 버튼 한 번 깔끔하게, 최대한 빠르게 클릭
    t0 = _step_start("6. 팝업 '생성하기' 버튼 클릭")
    if not used_copy_button:
        try:
            # dialog 안의 '생성하기' 버튼을 우선적으로 찾는다.
            confirm_btn = WebDriverWait(driver, 2).until(
                EC.element_to_be_clickable(
                    (
                        By.XPATH,
                        "//div[@role='dialog' and @data-state='open']"
                        "//button[contains(normalize-space(.),'생성하기') or contains(normalize-space(.),'생성 하기')]",
                    )
                )
            )
        except TimeoutException:
            # dialog 를 못 찾으면 화면 전체에서 '생성하기' 버튼을 찾는다.
            try:
                confirm_btn = WebDriverWait(driver, 2).until(
                    EC.element_to_be_clickable(
                        (
                            By.XPATH,
                            "//button[contains(normalize-space(.),'생성하기') or contains(normalize-space(.),'생성 하기')]",
                        )
                    )
                )
            except TimeoutException:
                print("[WARN] 화면에서 '생성하기' 버튼을 찾지 못했습니다.")
                _step_end("6. 팝업 '생성하기' 버튼 클릭", t0)
                _step_end("결제링크 생성 폼 작성 전체", t0_all)
                return None

        # 한 가지 방법만: scrollIntoView + JS click (가장 안정적인 패턴)
        try:
            driver.execute_script(
                "arguments[0].scrollIntoView({behavior:'instant', block:'center'});",
                confirm_btn,
            )
            time.sleep(0.03)
            driver.execute_script("arguments[0].click();", confirm_btn)
            print("[INFO] '생성하기' 버튼(JS click) 실행.")
        except Exception as e:
            print(f"[WARN] '생성하기' 버튼 클릭 중 오류: {e}")
            _step_end("6. 팝업 '생성하기' 버튼 클릭", t0)
            _step_end("결제링크 생성 폼 작성 전체", t0_all)
            return None
    _step_end("6. 팝업 '생성하기' 버튼 클릭", t0)

    # 7. 생성된 링크 추출 + 링크 복사 아이콘 클릭 + 클립보드 복사
    # 참고: 크롬 왼쪽 상단 '허용' 등 권한 알림은 클릭하지 않아도 됨. 링크는 페이지 input 값에서 읽어 저장하며,
    # 저장 후 admin_state.json 반영 및 signal_crawler_wakeup()으로 어드민/대행사 페이지에 곧 반영된다.
    t0 = _step_start("7. 생성된 링크 추출")
    link_text = None
    try:
        # readonly input.value 형태의 링크를 우선적으로 찾는다.
        url_input = WebDriverWait(driver, max(2, int(3 * LINK_CREATION_WAIT_FACTOR))).until(
            EC.presence_of_element_located(
                (
                    By.XPATH,
                    "//input[@readonly and contains(@value,'https://store.k-van.app')]",
                )
            )
        )
        link_text = (url_input.get_attribute("value") or "").strip()
        if link_text:
            print(f"[INFO] 생성된 결제 링크: {link_text}")

            # (1) 링크 복사 아이콘(버튼)을 클릭해서 K-VAN 내부 복사 로직 실행
            try:
                # 아이콘이 늦게 뜰 수 있으니 짧게 여러 번 재시도 (최대 2초)
                end_copy = time.time() + 2.0
                clicked_copy = False
                while time.time() < end_copy and not clicked_copy:
                    try:
                        copy_btn = WebDriverWait(driver, 0.5).until(
                            EC.element_to_be_clickable(
                                (
                                    By.XPATH,
                                    "//button[.//svg[contains(@class,'copy') or contains(@class,'lucide-copy')]]",
                                )
                            )
                        )
                        driver.execute_script(
                            "arguments[0].scrollIntoView({behavior:'instant', block:'center'});",
                            copy_btn,
                        )
                        time.sleep(0.03)
                        driver.execute_script("arguments[0].click();", copy_btn)
                        clicked_copy = True
                        print("[INFO] K-VAN 화면의 '링크 복사' 아이콘 버튼 클릭.")
                        break
                    except TimeoutException:
                        time.sleep(0.12)

            except TimeoutException:
                print("[WARN] '링크 복사' 아이콘 버튼을 찾지 못했습니다. K-VAN 내부 복사는 생략합니다.")

            # (2) 브라우저 클립보드 API 로도 한 번 더 복사 시도
            try:
                driver.execute_script(
                    """
try {
  const text = arguments[0] || '';
  if (navigator.clipboard && navigator.clipboard.writeText) {
    navigator.clipboard.writeText(text).then(() => {
      console.log('[K-VAN] navigator.clipboard.writeText 성공');
    }).catch(e => {
      console.log('[K-VAN] navigator.clipboard.writeText 실패', e);
    });
  } else {
    const ta = document.createElement('textarea');
    ta.value = text;
    document.body.appendChild(ta);
    ta.select();
    try {
      document.execCommand('copy');
      console.log('[K-VAN] document.execCommand(\"copy\") 성공');
    } catch (e) {
      console.log('[K-VAN] document.execCommand(\"copy\") 실패', e);
    }
    document.body.removeChild(ta);
  }
} catch (e) {
  console.log('[K-VAN] 클립보드 복사 전체 실패', e);
}
""",
                    link_text,
                )
                print("[INFO] 생성된 링크를 클립보드에 복사 시도했습니다.")
            except Exception as e:
                print(f"[WARN] 클립보드 복사 시도 중 오류: {e}")
    except TimeoutException:
        print("[WARN] 생성된 결제 링크 input 을 찾지 못했습니다.")

    if link_text and "https://store.k-van.app" in link_text:
        _store_kvan_link_for_session(session_id, link_text)
        _step_end("7. 생성된 링크 추출", t0)
        _step_end("결제링크 생성 폼 작성 전체", t0_all)
        return link_text

    _step_end("7. 생성된 링크 추출", t0)
    _step_end("결제링크 생성 폼 작성 전체", t0_all)
    return None

def sign_in(driver: webdriver.Chrome, row: PaymentRow) -> None:
    t0_all = _step_start("로그인 전체")
    driver.get(SIGN_IN_URL)
    # 로그인 화면 전체는 비교적 빨리 뜨므로, 글로벌 wait 은 최소한으로만 사용한다.
    wait = WebDriverWait(driver, 3)

    # 공지 팝업(확인 후 로그인 / 로그인 버튼)이 있으면 먼저 처리
    _click_notice_if_present(driver)

    # 아이디 입력창 찾기: 가장 단순한 CSS 셀렉터를 우선 사용
    t0 = _step_start("아이디 입력")
    id_input = _find_input_quick(
        driver,
        [
            "input[placeholder*='아이디']",
            "input[name*='id']",
            "input[type='text']",
        ],
        max_wait=3.0,
    )
    if not id_input:
        # 기존 XPATH 후보들도 한 번 더 시도 (역순으로 – 가장 느린 fallback 먼저)
        id_locators = [
            SIGN_IN_SELECTORS["id_fallback"],
            SIGN_IN_SELECTORS["id_placeholder"],
            SIGN_IN_SELECTORS["id_primary"],
        ]
        for loc in id_locators:
            try:
                id_input = wait.until(EC.visibility_of_element_located(loc))
                break
            except TimeoutException:
                continue

    if not id_input:
        print("[ERROR] 아이디 입력창을 찾지 못했습니다. 로그인 단계를 종료합니다.")
        _step_end("아이디 입력", t0)
        _step_end("로그인 전체", t0_all)
        return

    id_input.clear()
    id_input.send_keys(row.login_id)
    _step_end("아이디 입력", t0)

    # 비밀번호 입력창 찾기
    t0 = _step_start("비밀번호 입력")
    pw_input = _find_input_quick(
        driver,
        [
            "input[type='password']",
            "input[placeholder*='비밀번호']",
        ],
        max_wait=3.0,
    )
    if not pw_input:
        pw_locators = [
            SIGN_IN_SELECTORS["password_fallback"],
            SIGN_IN_SELECTORS["password_primary"],
        ]
        for loc in pw_locators:
            try:
                pw_input = wait.until(EC.visibility_of_element_located(loc))
                break
            except TimeoutException:
                continue

    if not pw_input:
        print("[ERROR] 비밀번호 입력창을 찾지 못했습니다. 로그인 단계를 종료합니다.")
        _step_end("비밀번호 입력", t0)
        _step_end("로그인 전체", t0_all)
        return

    pw_input.clear()
    pw_input.send_keys(row.login_password)
    _step_end("비밀번호 입력", t0)

    # 로그인 버튼 찾기
    t0 = _step_start("로그인 버튼 클릭")
    submit_btn = wait.until(
        EC.element_to_be_clickable(SIGN_IN_SELECTORS["submit_primary"])
    )
    submit_btn.click()
    _step_end("로그인 버튼 클릭", t0)

    # 2차 인증 PIN 팝업 처리 (있으면 입력, 없으면 통과)
    try:
        pin_delay = 0.25 * (1.0 + LINK_CREATION_WAIT_FACTOR)  # 0.25~0.375
        time.sleep(pin_delay)
        pin_wait = WebDriverWait(driver, max(2, int(3 * LINK_CREATION_WAIT_FACTOR)))
        pin_input_container = pin_wait.until(
            EC.visibility_of_element_located(PIN_POPUP_SELECTORS["input"])
        )
        if pin_input_container.tag_name.lower() == "input":
            pin_input = pin_input_container
        else:
            pin_input = pin_input_container.find_element(By.XPATH, ".//input")

        t0_pin = _step_start("PIN 입력 및 확인")
        pin_input.clear()
        pin_input.send_keys(row.login_pin)
        time.sleep(0.15 * (1.0 + LINK_CREATION_WAIT_FACTOR))

        confirm_btn = driver.find_element(*PIN_POPUP_SELECTORS["confirm"])
        confirm_btn.click()
        time.sleep(0.5 * (1.0 + LINK_CREATION_WAIT_FACTOR))
        _step_end("PIN 입력 및 확인", t0_pin)
    except TimeoutException:
        pass

    # 로그인 완료까지 잠시 대기 (홈 또는 다른 보호된 페이지로 진입)
    try:
        # 리다이렉트는 최대 1초만 확인하고, 바로 다음 단계로 진행한다.
        t0 = _step_start("로그인 후 리다이렉트 대기")
        short_wait = WebDriverWait(driver, 1)
        short_wait.until(EC.url_contains("store.k-van.app"))
        _step_end("로그인 후 리다이렉트 대기", t0)
    except TimeoutException:
        print("[WARN] 로그인 후 리다이렉트 URL을 1초 내에 확인하지 못했습니다. 다음 단계로 계속 진행합니다.")
        _step_end("로그인 후 리다이렉트 대기", t0)

    _step_end("로그인 전체", t0_all)


def _click_box(driver: webdriver.Chrome, locator: tuple) -> None:
    """단순 클릭 (div 기반 토글 박스용)."""
    el = driver.find_element(*locator)
    el.click()


def _select_dropdown_any(driver: webdriver.Chrome, locator: tuple, text: str) -> None:
    el = driver.find_element(*locator)
    tag = el.tag_name.lower()

    if tag == "select":
        Select(el).select_by_visible_text(text)
        return

    el.click()
    el.clear()
    el.send_keys(text)
    el.send_keys(Keys.ENTER)


def fill_face_to_face_form(driver: webdriver.Chrome, row: PaymentRow) -> None:
    # 기존 대면결제 URL — 결제링크 생성 플로우가 안정화되면 이 부분을 대체 가능
    driver.get(FACE_TO_FACE_URL)
    wait = WebDriverWait(driver, 30)

    # 페이지가 완전히 렌더링될 수 있도록 사람처럼 잠시 대기
    time.sleep(2.0)

    # 1) 상품명 먼저 입력
    product_input = wait.until(
        EC.visibility_of_element_located(FACE_TO_FACE_SELECTORS["product_name"])
    )
    driver.execute_script(
        "arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
        product_input,
    )
    time.sleep(0.8)
    product_input.clear()
    product_input.send_keys(row.product_name or "잡화")
    time.sleep(0.6)

    # 2) 판매가격(결제금액) 입력
    try:
        price_input = wait.until(
            EC.visibility_of_element_located(FACE_TO_FACE_SELECTORS["product_price"])
        )
    except TimeoutException:
        # 기본 셀렉터로 못 찾으면 몇 가지 대체 XPATH 를 시도
        alt_xpaths = [
            "//*[contains(text(), '판매가격') or contains(text(), '결제금액')]/following::input[1]",
            "//input[@inputmode='decimal' or @type='number']",
        ]
        price_input = None
        for xp in alt_xpaths:
            try:
                price_input = wait.until(
                    EC.visibility_of_element_located((By.XPATH, xp))
                )
                break
            except TimeoutException:
                continue
        if price_input is None:
            raise RuntimeError("판매가격(결제금액) 입력창을 찾지 못했습니다.")

    driver.execute_script(
        "arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
        price_input,
    )
    time.sleep(0.8)
    price_input.clear()
    price_input.send_keys(str(row.amount))
    time.sleep(0.8)

    # 3) 아래 결제 정보 입력창들이 보이도록 스크롤을 조금 더 내림
    try:
        scroll_container = wait.until(
            EC.presence_of_element_located(
                (
                    By.XPATH,
                    "//div[contains(@class,'overflow-y-auto') and contains(@class,'min-h-screen')]",
                )
            )
        )
        driver.execute_script(
            "arguments[0].scrollTo({top: arguments[0].scrollTop + 600, behavior: 'smooth'});",
            scroll_container,
        )
    except TimeoutException:
        # 컨테이너를 못 찾으면 윈도우 스크롤로 폴백
        driver.execute_script("window.scrollBy(0, 600);")
    time.sleep(1.5)

    # 결제 정보 - 카드 종류 체크 (개인 / 법인)
    # 페이지 기본값이 개인카드이므로, 개인카드는 굳이 클릭하지 않는다.
    if row.card_type == "business":
        try:
            _click_box(driver, FACE_TO_FACE_SELECTORS["card_business_checkbox"])
            time.sleep(0.8)
        except Exception:
            pass

    # 카드번호
    card_input = wait.until(
        EC.visibility_of_element_located(FACE_TO_FACE_SELECTORS["card_number"])
    )
    card_input.clear()
    card_input.send_keys(row.card_number)

    # 유효기간 MM / YY
    try:
        _select_dropdown_any(driver, FACE_TO_FACE_SELECTORS["expiry_mm"], row.expiry_mm)
    except Exception:
        mm_input = driver.find_element(*FACE_TO_FACE_SELECTORS["expiry_mm"])
        mm_input.clear()
        mm_input.send_keys(row.expiry_mm)

    try:
        _select_dropdown_any(driver, FACE_TO_FACE_SELECTORS["expiry_yy"], row.expiry_yy)
    except Exception:
        yy_input = driver.find_element(*FACE_TO_FACE_SELECTORS["expiry_yy"])
        yy_input.clear()
        yy_input.send_keys(row.expiry_yy)

    # 비밀번호 앞 두 자리
    pw2_input = wait.until(
        EC.visibility_of_element_located(FACE_TO_FACE_SELECTORS["card_password"])
    )
    pw2_input.clear()
    pw2_input.send_keys(row.card_password)

    # 할부개월
    try:
        _select_dropdown_any(
            driver, FACE_TO_FACE_SELECTORS["installment_select"], row.installment_months
        )
    except Exception:
        pass

    # 연락처
    try:
        _select_dropdown_any(driver, FACE_TO_FACE_SELECTORS["phone_prefix"], "010")
    except Exception:
        pass

    phone_input = wait.until(
        EC.visibility_of_element_located(FACE_TO_FACE_SELECTORS["phone_number"])
    )
    phone_input.clear()
    phone_input.send_keys(row.phone_number)

    # 이름
    name_input = wait.until(
        EC.visibility_of_element_located(FACE_TO_FACE_SELECTORS["customer_name"])
    )
    name_input.clear()
    name_input.send_keys(row.customer_name)

    # 주민등록번호 앞자리 / 사업자등록번호
    if row.card_type == "business":
        # 사업자 카드인 경우: 사업자등록번호 입력칸이 나타남
        try:
            biz_input = wait.until(
                EC.visibility_of_element_located(
                    FACE_TO_FACE_SELECTORS["business_reg_no"]
                )
            )
            biz_input.clear()
            biz_input.send_keys(row.resident_front)
        except Exception:
            # 셀렉터가 맞지 않거나 필드가 안 보이면 조용히 통과
            pass
    else:
        # 개인카드인 경우: 주민등록번호 앞자리 입력
        res_input = wait.until(
            EC.visibility_of_element_located(FACE_TO_FACE_SELECTORS["resident_front"])
        )
        res_input.clear()
        res_input.send_keys(row.resident_front)

    # 버튼이 "결제하기" 로 바뀔 때까지 기다렸다가 클릭
    try:
        submit_btn = wait.until(
            EC.element_to_be_clickable(FACE_TO_FACE_SELECTORS["submit_button"])
        )
    except TimeoutException:
        raise RuntimeError("결제하기 버튼을 찾지 못했습니다. 셀렉터를 확인하세요.")

    submit_btn.click()


def confirm_popup_and_get_result(driver: webdriver.Chrome) -> tuple[str, str]:
    wait = WebDriverWait(driver, 30)

    # 결제 확인 팝업에서 "결제" 또는 "확인" 버튼 클릭 (텍스트 기준으로 찾기)
    clicked_confirm = False
    time.sleep(1.0)

    # 팝업 타이틀(합계 ~ 결제)이 보일 때까지 먼저 대기 (있으면)
    try:
        wait.until(
            EC.visibility_of_element_located(
                (
                    By.XPATH,
                    "//*[contains(normalize-space(.), '합계') and contains(normalize-space(.), '결제')]",
                )
            )
        )
    except TimeoutException:
        # 팝업 타이틀을 못 찾더라도 계속 진행 (환경에 따라 구조가 다를 수 있음)
        pass

    # 결제 버튼 → 확인 버튼 순서로 여러 패턴을 시도
    button_xpaths = [
        # SweetAlert2 팝업 전용 (가장 우선)
        "//div[contains(@class,'swal2-actions')]//button[contains(@class,'swal2-confirm')]",
        "//button[contains(@class,'swal2-confirm') and contains(normalize-space(.), '결제')]",
        # 버튼 자체 텍스트에 "결제"
        "//button[contains(normalize-space(.), '결제')]",
        # 버튼 내부 span 등에 "결제"
        "//button[.//span[contains(normalize-space(normalize-space(.)), '결제')]]",
        # 일부 환경에서는 '확인'만 표시될 수 있음
        "//button[contains(normalize-space(.), '확인')]",
        "//button[.//span[contains(normalize-space(.), '확인')]]",
    ]

    for xp in button_xpaths:
        try:
            btn = wait.until(
                EC.element_to_be_clickable(
                    (
                        By.XPATH,
                        xp,
                    )
                )
            )
            # 버튼이 가려져 있을 수 있으므로 가운데로 스크롤
            try:
                driver.execute_script(
                    "arguments[0].scrollIntoView({behavior:'smooth', block:'center'});",
                    btn,
                )
                time.sleep(0.5)
            except Exception:
                pass

            try:
                btn.click()
            except Exception:
                # 일반 click 이 안 먹히면 JS 로 강제 클릭
                driver.execute_script("arguments[0].click();", btn)

            clicked_confirm = True
            break
        except TimeoutException:
            continue

    if not clicked_confirm:
        print("확인/결제 팝업 버튼을 찾지 못했습니다. 바로 결과 화면이 떴을 수 있습니다.")

    # 결과 화면의 h1 / p 텍스트 읽기
    try:
        title_el = wait.until(
            EC.visibility_of_element_located(
                (By.XPATH, "//h1[contains(., '결제')]")
            )
        )
        title_text = title_el.text.strip()
    except TimeoutException:
        return "unknown", "결과 제목을 찾지 못했습니다."

    try:
        message_el = driver.find_element(
            By.XPATH,
            "//p[contains(@class, 'text-red-500') or contains(@class, 'text-green-500')]",
        )
        message_text = message_el.text.strip()
    except Exception:
        message_text = ""

    status = "success" if "완료" in title_text else "fail" if "실패" in title_text else "unknown"
    return status, f"{title_text} / {message_text}"


def save_result_to_excel(
    path: str, row: PaymentRow, status: str, message: str
) -> None:
    """결제 결과를 별도 엑셀 파일(RESULT_EXCEL_PATH)에 누적 저장.

    기존 버전에서 생성된 엑셀은 card_type 열이 없을 수 있으므로,
    필요하면 결과 시트의 헤더를 최신 HEADERS 구조로 자동 마이그레이션한다.
    """

    def _ensure_latest_header(ws) -> None:
        """기존 결과 시트의 헤더를 최신 HEADERS 구조로 맞춘다."""
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if not any(header_row):
            # 헤더가 비어 있으면 새로 작성
            ws.append(HEADERS + ["result_status", "result_message"])
            return

        # 이전 버전: card_type 이 없던 경우 (열 개수로 판단)
        old_headers = [h for h in HEADERS if h != "card_type"]
        if header_row[: len(old_headers)] == old_headers and len(header_row) == len(
            old_headers
        ) + 2:
            # 4번째 열에 card_type 열 삽입 (기존 데이터는 기본값 없음)
            ws.insert_cols(4)
            ws.cell(row=1, column=4, value="card_type")

    path_obj = Path(path)
    if path_obj.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "results"
        ws.append(HEADERS + ["result_status", "result_message"])

    # 기존 파일인 경우 헤더를 최신 구조로 보정
    _ensure_latest_header(ws)

    data = [getattr(row, key) for key in HEADERS]
    ws.append(data + [status, message])

    try:
        wb.save(path)
    except PermissionError:
        # 파일이 엑셀 등에서 열려 있어서 저장이 안 되는 경우
        print("kvan_results.xlsx 저장에 실패했습니다. 엑셀 파일을 닫은 뒤 다시 실행해 주세요.")


def save_result_to_json(path: str, status: str, message: str) -> None:
    """web_form.py 에서 읽어갈 수 있도록 마지막 결제 결과를 JSON 으로도 저장."""
    payload = {
        "status": status,
        "message": message,
    }
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except OSError:
        # 파일 시스템 오류는 치명적이지 않으므로 조용히 무시
        pass


def append_transaction_to_hq(
    row,
    status: str,
    message: str,
    session_id: str,
) -> None:
    """결제 결과를 본사 어드민용 거래 리스트에 추가."""
    try:
        now = datetime.utcnow()
        tx_id = datetime.utcnow().strftime("TX%Y%m%d%H%M%S%f")

        # 세션 ID 로부터 실제 대행사 ID 를 추적 (admin_state.json 참조)
        agency_id = ""
        if session_id:
            admin_path = DATA_DIR / "admin_state.json"
            if admin_path.exists():
                try:
                    with admin_path.open("r", encoding="utf-8") as f:
                        admin_state = json.load(f)
                except Exception:
                    admin_state = {}
                sessions = admin_state.get("sessions") or []
                history = admin_state.get("history") or []
                for s in sessions:
                    if str(s.get("id")) == str(session_id) and s.get("agency_id"):
                        agency_id = s["agency_id"]
                        break
                if not agency_id:
                    for h in history:
                        if str(h.get("id")) == str(session_id) and h.get("agency_id"):
                            agency_id = h["agency_id"]
                            break

        # 금액은 정수 변환 시도 (실패 시 0)
        try:
            amount_int = int(getattr(row, "amount", 0) or 0)
        except (ValueError, TypeError):
            amount_int = 0

        conn = get_db()
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO transactions
                (id, created_at, agency_id, amount, customer_name, phone_number,
                 card_type, resident_front, status, message, settlement_status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    tx_id,
                    now,
                    agency_id or "",
                    amount_int,
                    getattr(row, "customer_name", ""),
                    getattr(row, "phone_number", ""),
                    getattr(row, "card_type", ""),
                    getattr(row, "resident_front", ""),
                    status,
                    message,
                    "미정산",
                ),
            )
        conn.commit()
        conn.close()
    except Exception:
        # HQ 집계 실패는 결제 자체에는 영향을 주지 않으므로 조용히 무시
        pass


def main() -> None:
    """
    auto_kvan 링크 생성 메인 엔트리 (우선 목적: 결제 링크 생성).

    - 세션 ID 또는 기본 JSON 을 읽어 K-VAN 결제 링크를 생성한다.
    - 전체 실행 시간이 30분(1800초)을 넘으면 안전하게 종료한다.
    - 링크 생성 속도: LINK_CREATION_WAIT_FACTOR(기본 0.5)로 대기 시간을 조절한다.
      불안정하면 1.0으로 올리면 된다.
    """
    import sys

    start_ts = time.time()

    def _check_timeout(stage: str) -> None:
        """30분 타임아웃 체크. 초과 시 TimeoutError 를 발생시킨다."""
        elapsed = time.time() - start_ts
        if elapsed > 1800:
            raise TimeoutError(f"auto_kvan.py 30분 초과 타임아웃 (stage={stage}, elapsed={elapsed:.1f}s)")

    session_id = sys.argv[1].strip() if len(sys.argv) > 1 else ""

    if session_id:
        SESSION_ORDER_DIR.mkdir(parents=True, exist_ok=True)
        SESSION_RESULT_DIR.mkdir(parents=True, exist_ok=True)
        order_candidates = _session_order_path_candidates(session_id)
        order_path = order_candidates[0]
        for p in order_candidates:
            if p.exists():
                order_path = p
                break
        result_json_path = SESSION_RESULT_DIR / f"{session_id}.json"
    else:
        order_path = Path(ORDER_JSON_PATH)
        result_json_path = Path(RESULT_JSON_PATH)

    print("JSON 주문 데이터를 읽습니다...")
    if session_id:
        _append_admin_log(
            "AUTO",
            f"세션 주문 JSON 경로 확인 session_id={session_id}, selected={order_path}, "
            f"candidates={[str(p) for p in _session_order_path_candidates(session_id)]}",
        )
    try:
        row = load_order_from_json(str(order_path))
    except FileNotFoundError as e:
        # 세션 ID 기반 링크 생성 모드에서는, 주문 JSON 이 없어도
        # admin_state.json 에 저장된 세션 정보(금액/할부)를 기반으로
        # 최소한의 PaymentRow 를 구성해 링크 생성을 시도한다.
        if session_id:
            try:
                amount_val = 0
                installment_val = "일시불"
                st_path_used: Path | None = None
                for st_path in _admin_state_candidates():
                    if not st_path.exists():
                        continue
                    with open(st_path, "r", encoding="utf-8") as f:
                        st = json.load(f)
                    sessions = st.get("sessions") or []
                    for s in sessions:
                        if str(s.get("id")) == str(session_id):
                            amt_str = str(s.get("amount") or "").replace(",", "").strip()
                            amount_val = int(amt_str) if amt_str else 0
                            installment_val = str(s.get("installment") or "일시불")
                            st_path_used = st_path
                            break
                    if st_path_used:
                        break
                if amount_val <= 0:
                    _append_admin_log(
                        "AUTO",
                        "[ERROR] 세션 금액이 없어서 링크를 생성할 수 없습니다. "
                        f"session_id={session_id}, order_path={order_path}, "
                        f"admin_state_candidates={[str(p) for p in _admin_state_candidates()]}",
                    )
                    print(e)
                    return
                login_id = os.environ.get("K_VAN_ID", "m3313")
                login_pw = os.environ.get("K_VAN_PW", "1234")
                login_pin = os.environ.get("K_VAN_PIN", "2424")
                _append_admin_log(
                    "AUTO",
                    f"주문 JSON 없이 세션 정보로 링크 생성 시도 "
                    f"session_id={session_id}, amount={amount_val}, installment={installment_val}, "
                    f"admin_state_path={st_path_used or '-'}",
                )
                row = PaymentRow(
                    login_id=login_id,
                    login_password=login_pw,
                    login_pin=login_pin,
                    card_type="personal",
                    card_number="",
                    expiry_mm="",
                    expiry_yy="",
                    card_password="",
                    installment_months=installment_val,
                    phone_number="",
                    customer_name="",
                    resident_front="",
                    amount=amount_val,
                    product_name=f"SISA 세션 {session_id}",
                )
            except Exception as e2:  # noqa: BLE001
                _append_admin_log(
                    "AUTO",
                    f"[ERROR] 세션 기반 기본 주문 데이터 구성 실패 session_id={session_id}: {e2}",
                )
                print(e)
                return
        else:
            _append_admin_log("AUTO", f"주문 JSON 없음 session_id={session_id or '-'} path={order_path}")
            print(e)
            return
    except ValueError as e:
        print(f"입력 데이터 오류: {e}")
        return

    _check_timeout("after_load_order")

    driver: webdriver.Chrome | None = None
    try:
        driver = create_driver()
        _check_timeout("before_sign_in")
        _append_admin_log("AUTO", f"K-VAN 로그인 시작 session_id={session_id or '-'}")
        print("K-VAN 가맹점 페이지에 로그인 중...")
        sign_in(driver, row)
        _check_timeout("after_sign_in")

        _append_admin_log("AUTO", "로그인 완료, 결제링크 관리 페이지로 이동")
        print("로그인 완료. 결제링크 관리 페이지로 이동합니다...")
        _go_to_payment_link_page(driver)
        _check_timeout("after_go_payment_link")

        _append_admin_log("AUTO", "결제링크 생성 페이지 진입 시도")
        print("결제링크 관리 화면에서 '+ 생성' 버튼을 눌러 생성 페이지로 이동합니다...")
        moved = _go_to_create_link_page(driver)
        _check_timeout("after_go_create")
        if not moved:
            _append_admin_log("AUTO", "[ERROR] 링크 생성 페이지 진입 실패 (+ 생성 버튼 동작 안 함)")
            print("[ERROR] '+ 생성' 버튼 클릭 후 생성 페이지로 이동하지 못했습니다. 폼 작성 단계는 건너뜁니다.")
            status = "error"
            msg = "결제링크 생성 페이지 진입 실패(생성 버튼 미동작)."
            save_result_to_excel(RESULT_EXCEL_PATH, row, status, msg)
            save_result_to_json(str(result_json_path), status, msg)
            return

        _append_admin_log("AUTO", "결제링크 생성 폼 작성/전송 시작")
        print("결제링크 생성 페이지에서 폼을 채우고 링크를 생성합니다...")
        link_url = _fill_payment_link_form_and_get_url(driver, row, session_id)
        _check_timeout("after_fill_form")
        if link_url:
            status = "link_created"
            msg = "결제 링크가 생성되었습니다. 고객이 링크로 결제하면 K-VAN 크롤러가 상태를 반영합니다."
            _append_admin_log("AUTO", f"결제 링크 생성 완료 session_id={session_id or '-'} link={link_url}")
            print(f"생성된 결제 링크: {link_url}")
            # admin_state.json 의 해당 세션에 바로 링크를 매핑해서
            # /admin, /agency-admin 화면에서 '링크 복사' 버튼이 즉시 표시되도록 한다.
            if session_id:
                try:
                    _store_kvan_link_for_session(session_id, link_url)
                except Exception as e_store:  # noqa: BLE001
                    print(f"[WARN] admin_state 에 링크 저장 중 오류: {e_store}")
                    _append_admin_log("AUTO", f"[WARN] admin_state 링크 저장 실패 session_id={session_id}: {e_store}")
            # 링크가 새로 생성되었으므로, 크롤러에 즉시 다시 크롤링하도록 신호를 보낸다.
            try:
                signal_crawler_wakeup()
            except Exception as e:
                print(f"[WAKEUP][WARN] 크롤러 깨우기 신호 전송 중 오류: {e}")
            # 서버에서 링크 생성 직후 같은 세션으로 /transactions 를 한 번 읽어
            # 어드민·대행사 화면에 빨리 반영 (크롤러 주기를 기다리지 않음).
            # 끄려면 K_VAN_AFTER_LINK_SCRAPE_TX=0
            _after = os.environ.get("K_VAN_AFTER_LINK_SCRAPE_TX", "1").strip().lower()
            if _is_server_env() and _after not in ("0", "false", "no", "off"):
                try:
                    _append_admin_log(
                        "AUTO",
                        "링크 생성 후 /transactions 즉시 스크랩·transactions 동기화 시도",
                    )
                    _scrape_transactions_and_store(driver)
                    ok = _sync_kvan_to_transactions()
                    _append_admin_log(
                        "AUTO",
                        f"/transactions 스크랩·동기화 완료 (sync_returned={ok})",
                    )
                except Exception as e_tx:  # noqa: BLE001
                    print(f"[WARN] 링크 생성 후 거래내역 스크랩/동기화 실패: {e_tx}")
                    _append_admin_log(
                        "AUTO",
                        f"[WARN] 링크 생성 후 /transactions 스크랩·동기화 실패: {e_tx}",
                    )
        else:
            status = "error"
            msg = "결제 링크 생성에 실패했거나 링크를 찾지 못했습니다."
            _append_admin_log("AUTO", "[ERROR] 결제 링크 생성 실패 또는 링크 미발견")
            print(msg)

        # 결과는 참고용 엑셀/JSON 에만 남기고, 실제 결제/정산 정보는
        # K-VAN 크롤러(kvan_crawler.py)와 DB 동기화를 기준으로 관리한다.
        save_result_to_excel(RESULT_EXCEL_PATH, row, status, msg)
        save_result_to_json(str(result_json_path), status, msg)

        if not _is_server_env():
            input("브라우저를 확인한 뒤, 종료하려면 Enter 키를 누르세요...")
    except TimeoutError as te:
        print(f"[ERROR] {te}")
        _append_admin_log("AUTO", f"[ERROR] {te}")
    except Exception as e:
        print(f"[ERROR] 링크 생성 중 오류: {e}")
        _append_admin_log("AUTO", f"[ERROR] 링크 생성 중 오류: {e}")
        if session_id:
            try:
                save_result_to_json(str(result_json_path), "error", str(e))
            except Exception:
                pass
    finally:
        if driver is not None:
            try:
                driver.quit()
            except Exception as qe:
                print(f"[WARN] driver.quit() 실패: {qe}")


if __name__ == "__main__":
    main()

